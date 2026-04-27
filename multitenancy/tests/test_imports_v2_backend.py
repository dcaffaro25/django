"""Integration tests for the v2 template-import backend (Phase 2).

Covers:
  * Analyze happy path (clean file → session transitions to ``ready``).
  * Analyze detects erp_id conflicts and leaves the session in
    ``awaiting_resolve``.
  * Commit succeeds on a ``ready`` session (mocked write — execute_import_job
    is already tested elsewhere; here we only verify session lifecycle).
  * Commit refuses a session with blocking issues (409).
  * Discard transitions non-terminal sessions to ``discarded``.
  * Cross-tenant session IDs are hidden (404, not 403).

``execute_import_job`` is mocked because it needs full FK fixtures
(Entity, Currency, CostCenter, etc.) that are outside Phase 2's
surface area. Phase 4 tests will exercise the real write.
"""
from __future__ import annotations

import io
from unittest import mock

import pandas as pd
from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.urls import reverse
from rest_framework.test import APIClient

from multitenancy.models import Company, ImportSession

User = get_user_model()


def _build_xlsx(sheets: dict) -> bytes:
    """Build a minimal .xlsx from {sheet_name: [row_dict, ...]}."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        for name, rows in sheets.items():
            df = pd.DataFrame(rows or [{}])
            df.to_excel(xw, sheet_name=name, index=False)
    buf.seek(0)
    return buf.getvalue()


def _upload_file(content: bytes, name: str = "import.xlsx"):
    """Wrap bytes as a Django UploadedFile for multipart POST."""
    from django.core.files.uploadedfile import SimpleUploadedFile
    return SimpleUploadedFile(
        name,
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# URLs we hit. Matching how nord_backend mounts multitenancy.urls under
# ``/<tenant_id>/``. In tests we use the company's id as tenant_id —
# the resolve helper falls back to user.company_id when that's not
# meaningful.
def _url_analyze(tenant_id):
    return f"/{tenant_id}/api/core/imports/v2/analyze/"


def _url_commit(tenant_id, pk):
    return f"/{tenant_id}/api/core/imports/v2/commit/{pk}/"


def _url_session(tenant_id, pk):
    return f"/{tenant_id}/api/core/imports/v2/sessions/{pk}/"


def _url_sessions_list(tenant_id):
    return f"/{tenant_id}/api/core/imports/v2/sessions/"


def _url_sessions_running_count(tenant_id):
    return f"/{tenant_id}/api/core/imports/v2/sessions/running-count/"


class JsonScalarTests(TestCase):
    """Coverage for ``_json_scalar`` -- the per-cell normaliser that
    runs on every Excel cell during template parse. The midnight-only
    branch was added (2026-04-25) after operators reported every
    Transaction row failing with

        '2025-09-16T00:00:00' value has an invalid date format.
        It must be in YYYY-MM-DD format.

    Excel stores date-only cells as ``pd.Timestamp`` at midnight; we
    used to emit them as full ISO datetimes which Django's DateField
    parser rejected. Tests below pin both branches.
    """

    def test_pure_date_emits_yyyy_mm_dd(self):
        import datetime as dt
        from multitenancy.imports_v2.services import _json_scalar
        self.assertEqual(_json_scalar(dt.date(2025, 9, 16)), "2025-09-16")

    def test_midnight_datetime_emits_date_only(self):
        """The bug fix's primary case: an Excel-imported date cell
        arrives as a midnight ``datetime`` and must emit YYYY-MM-DD."""
        import datetime as dt
        from multitenancy.imports_v2.services import _json_scalar
        self.assertEqual(
            _json_scalar(dt.datetime(2025, 9, 16, 0, 0, 0)),
            "2025-09-16",
        )

    def test_midnight_pandas_timestamp_emits_date_only(self):
        """Same as above but for the ``pd.Timestamp`` form pandas
        produces from ``read_excel``."""
        import pandas as pd
        from multitenancy.imports_v2.services import _json_scalar
        self.assertEqual(
            _json_scalar(pd.Timestamp("2025-09-16")),
            "2025-09-16",
        )

    def test_non_midnight_datetime_keeps_full_iso(self):
        """Real datetimes (any non-zero time component) keep the full
        ISO so DateTimeField columns roundtrip without losing time."""
        import datetime as dt
        from multitenancy.imports_v2.services import _json_scalar
        self.assertEqual(
            _json_scalar(dt.datetime(2025, 9, 16, 14, 30, 5)),
            "2025-09-16T14:30:05",
        )

    def test_non_midnight_pandas_timestamp_keeps_full_iso(self):
        import pandas as pd
        from multitenancy.imports_v2.services import _json_scalar
        ts = pd.Timestamp("2025-09-16 14:30:05")
        # pandas' Timestamp.isoformat() may include nanosecond precision;
        # we just need to confirm the time component is preserved.
        out = _json_scalar(ts)
        self.assertTrue(out.startswith("2025-09-16T14:30:05"), out)

    def test_tz_aware_midnight_keeps_full_iso(self):
        """A tz-aware midnight Timestamp could be a different day in the
        local zone, so we conservatively keep the full ISO. Operators
        importing tz-aware data should target a DateTimeField column."""
        import datetime as dt
        from multitenancy.imports_v2.services import _json_scalar
        v = dt.datetime(2025, 9, 16, 0, 0, 0, tzinfo=dt.timezone.utc)
        out = _json_scalar(v)
        # Anything with a tz suffix is fine; the assertion is just that
        # we DIDN'T collapse to date-only.
        self.assertNotEqual(out, "2025-09-16")
        self.assertIn("2025-09-16T00:00:00", out)


class CoerceDateFieldsTests(TestCase):
    """Coverage for ``multitenancy.tasks._coerce_date_fields`` -- the
    defense-in-depth layer that catches any ``execute_import_job``
    input where a DateField column receives an ISO datetime string or
    a datetime object. Complements the v2-only ``_json_scalar`` fix
    so legacy bulk-import / raw API calls don't fail with the
    'invalid date format' error any more.
    """

    def test_datetime_string_with_T_truncates_to_date(self):
        """The case the operator hit: a value like
        ``'2025-09-16T00:00:00'`` arriving on a DateField gets sliced
        to ``'2025-09-16'`` so Django's ``parse_date`` accepts it."""
        from multitenancy.tasks import _coerce_date_fields
        from accounting.models import Transaction  # has a ``date`` DateField

        out = _coerce_date_fields(Transaction, {"date": "2025-09-16T00:00:00"})
        self.assertEqual(out["date"], "2025-09-16")

    def test_datetime_string_without_T_passes_through(self):
        """A pristine ``YYYY-MM-DD`` is left alone."""
        from multitenancy.tasks import _coerce_date_fields
        from accounting.models import Transaction

        out = _coerce_date_fields(Transaction, {"date": "2025-09-16"})
        self.assertEqual(out["date"], "2025-09-16")

    def test_datetime_instance_collapses_to_date(self):
        """Bare ``datetime.datetime`` objects (incl. pd.Timestamp via
        subclassing) get ``.date()``-ed."""
        import datetime as dt
        from multitenancy.tasks import _coerce_date_fields
        from accounting.models import Transaction

        out = _coerce_date_fields(
            Transaction, {"date": dt.datetime(2025, 9, 16, 14, 30)},
        )
        self.assertEqual(out["date"], dt.date(2025, 9, 16))

    def test_pandas_timestamp_collapses_to_date(self):
        import datetime as dt
        import pandas as pd
        from multitenancy.tasks import _coerce_date_fields
        from accounting.models import Transaction

        ts = pd.Timestamp("2025-09-16 14:30:00")
        out = _coerce_date_fields(Transaction, {"date": ts})
        self.assertEqual(out["date"], dt.date(2025, 9, 16))

    def test_unknown_field_passes_through(self):
        """A non-date column with a string value isn't touched."""
        from multitenancy.tasks import _coerce_date_fields
        from accounting.models import Transaction

        out = _coerce_date_fields(Transaction, {"description": "anything"})
        self.assertEqual(out["description"], "anything")

    def test_none_value_passes_through(self):
        from multitenancy.tasks import _coerce_date_fields
        from accounting.models import Transaction

        out = _coerce_date_fields(Transaction, {"date": None})
        self.assertIsNone(out["date"])

    def test_datetime_field_left_alone(self):
        """``DateTimeField`` columns (a subclass of DateField) preserve
        full precision — only plain DateFields get the truncation."""
        from multitenancy.tasks import _coerce_date_fields
        from multitenancy.models import ImportSession

        # ImportSession.created_at is auto_now_add but a normal
        # DateTimeField; the helper must NOT touch it.
        out = _coerce_date_fields(
            ImportSession, {"created_at": "2025-09-16T14:30:00"},
        )
        self.assertEqual(out["created_at"], "2025-09-16T14:30:00")

    def test_space_separated_datetime_string_truncates(self):
        """SQL-style ``2025-09-16 14:30:00`` (space separator instead of
        the ISO ``T``) is a common shape from PostgreSQL ``::text`` casts
        and SQL Server queries — the operator copy-pasting one of those
        into the import tool used to hit the same 'invalid date format'
        wall as the T-form. Slice to YYYY-MM-DD."""
        from multitenancy.tasks import _coerce_date_fields
        from accounting.models import Transaction

        out = _coerce_date_fields(Transaction, {"date": "2025-09-16 14:30:00"})
        self.assertEqual(out["date"], "2025-09-16")

    def test_lowercase_t_separator_truncates(self):
        """Some clients lowercase the ISO ``T`` separator (``2025-09-16t00:00:00``).
        Tolerated."""
        from multitenancy.tasks import _coerce_date_fields
        from accounting.models import Transaction

        out = _coerce_date_fields(Transaction, {"date": "2025-09-16t00:00:00"})
        self.assertEqual(out["date"], "2025-09-16")

    def test_iso_datetime_with_tz_offset_truncates(self):
        """``2025-09-16T00:00:00+03:00`` etc. — anything past the date
        prefix gets dropped on the way to a DateField (the offset is
        meaningless without a time anyway)."""
        from multitenancy.tasks import _coerce_date_fields
        from accounting.models import Transaction

        out = _coerce_date_fields(Transaction, {"date": "2025-09-16T00:00:00+03:00"})
        self.assertEqual(out["date"], "2025-09-16")

    def test_garbage_string_passes_through_unchanged(self):
        """A string that doesn't even look like ``YYYY-MM-DD...`` is left
        alone — ``_coerce_date_fields`` is defensive but not magical;
        Django's own DateField parser will reject it with its standard
        message, which is the better signal for the operator."""
        from multitenancy.tasks import _coerce_date_fields
        from accounting.models import Transaction

        out = _coerce_date_fields(Transaction, {"date": "not-a-date-at-all"})
        self.assertEqual(out["date"], "not-a-date-at-all")

    def test_etl_je_path_imports_coerce_date_fields(self):
        """Smoke-test for commit cb73c13's regression fix:
        ``_import_transactions_with_journal_entries`` (the ETL fast-path
        that bypasses ``execute_import_job`` when auto_create_journal_entries
        is enabled) used to omit ``_coerce_date_fields`` from its imports
        block, leaving Transaction rows with ISO-datetime ``date`` values
        to fail Django's DateField parser at save time. This test just
        confirms the symbol is available in the module's runtime imports
        list — a regression here would mean the fix was reverted."""
        # If the helper isn't importable from the same module the ETL
        # path uses, the ETL-time import inside _import_transactions_…
        # would explode with ImportError on first call. Cheap canary.
        from multitenancy.tasks import _coerce_date_fields  # noqa: F401
        # Directly inspect the source for the call site so the test
        # fails LOUDLY if a future refactor drops it again.
        # ``_import_transactions_with_journal_entries`` is a method on
        # ``ETLPipelineService`` (not a module-level function), so we
        # grab the source through the class.
        import inspect
        from multitenancy.etl_service import ETLPipelineService
        src = inspect.getsource(
            ETLPipelineService._import_transactions_with_journal_entries
        )
        self.assertIn(
            "_coerce_date_fields(model, filtered)",
            src,
            "ETL+JE write path must apply _coerce_date_fields before "
            "instantiating Transaction, or ISO datetime strings will "
            "trip DateField parsing on save.",
        )
        # Also assert the helper is among the imports — the call site
        # uses a function-local ``from multitenancy.tasks import ...``
        # block, so the symbol must be in the import line above the
        # method body.
        self.assertIn("_coerce_date_fields", src)


@override_settings(AUTH_OFF=True)  # sidestep token auth for these tests
class V2AnalyzeEndpointTests(TestCase):
    """analyze/ — file-to-session transitions."""

    @classmethod
    def setUpTestData(cls):
        cls.company = Company.objects.create(name="Acme Co")
        # CustomUser has no direct company_id FK — tenant resolution
        # happens via the URL path (``/<tenant_id>/...``) and
        # ``TenantMiddleware`` in prod. Tests rely on the URL resolving
        # the company, not on a user↔company FK.
        cls.user = User.objects.create_user(username="op", password="x")

    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def test_missing_file_returns_400(self):
        resp = self.client.post(_url_analyze(self.company.id), {}, format="multipart")
        self.assertEqual(resp.status_code, 400, resp.content)
        self.assertIn("file", (resp.json().get("error") or "").lower())

    def test_empty_file_returns_400(self):
        resp = self.client.post(
            _url_analyze(self.company.id),
            {"file": _upload_file(b"", "empty.xlsx")},
            format="multipart",
        )
        self.assertEqual(resp.status_code, 400, resp.content)

    def test_unreadable_file_marks_session_error(self):
        """Garbage bytes → session still created, status=error."""
        resp = self.client.post(
            _url_analyze(self.company.id),
            {"file": _upload_file(b"not a real xlsx", "bad.xlsx")},
            format="multipart",
        )
        self.assertEqual(resp.status_code, 202, resp.content)
        body = resp.json()
        self.assertEqual(body["status"], ImportSession.STATUS_ERROR)
        self.assertEqual(body["is_terminal"], True)
        self.assertIn("error", body.get("result") or {})

    def test_clean_file_transitions_to_ready(self):
        """A file with no transaction-sheet conflicts → ready, no issues.

        Mocks the dry-run with a clean success payload so this test
        stays focused on the analyze-flow contract and doesn't depend
        on the test DB having every FK fixture the real
        ``execute_import_job`` would resolve. Dry-run failures
        (introduced 2026-04-24) translate to ``dry_run_failure``
        blocking issues; we don't want those firing here.
        """
        xlsx = _build_xlsx({
            "Transaction": [
                {"__erp_id": "OMIE-1", "date": "2026-01-01",
                 "description": "x", "amount": "100.00"},
                {"__erp_id": "OMIE-2", "date": "2026-01-02",
                 "description": "y", "amount": "200.00"},
            ],
        })
        clean_imports = {
            "imports": [{
                "model": "Transaction",
                "result": [
                    {"__row_id": "r1", "action": "create", "status": "success"},
                    {"__row_id": "r2", "action": "create", "status": "success"},
                ],
            }],
        }
        with mock.patch(
            "multitenancy.imports_v2.services.execute_import_job",
            return_value=clean_imports,
        ):
            resp = self.client.post(
                _url_analyze(self.company.id),
                {"file": _upload_file(xlsx)},
                format="multipart",
            )
        self.assertEqual(resp.status_code, 202, resp.content)
        body = resp.json()
        self.assertEqual(body["status"], ImportSession.STATUS_READY)
        self.assertEqual(body["is_committable"], True)
        self.assertEqual(body["open_issues"], [])
        self.assertEqual(
            body["summary"],
            {"sheets": {"Transaction": 2}},
        )

    def test_erp_id_conflict_transitions_to_awaiting_resolve(self):
        """Two rows sharing an erp_id but disagreeing on date → one
        blocking ``erp_id_conflict`` issue; session stays in
        ``awaiting_resolve``. Mocks the dry-run with a clean payload
        so the assertion can pin exactly one issue (the conflict)
        rather than two (conflict + dry_run_failure)."""
        xlsx = _build_xlsx({
            "Transaction": [
                {"__erp_id": "OMIE-1", "date": "2026-01-01",
                 "description": "x", "amount": "100.00"},
                {"__erp_id": "OMIE-1", "date": "2026-01-02",
                 "description": "x", "amount": "100.00"},
            ],
        })
        clean_imports = {
            "imports": [{
                "model": "Transaction",
                "result": [
                    {"__row_id": "r1", "action": "create", "status": "success"},
                    {"__row_id": "r2", "action": "create", "status": "success"},
                ],
            }],
        }
        with mock.patch(
            "multitenancy.imports_v2.services.execute_import_job",
            return_value=clean_imports,
        ):
            resp = self.client.post(
                _url_analyze(self.company.id),
                {"file": _upload_file(xlsx)},
                format="multipart",
            )
        self.assertEqual(resp.status_code, 202, resp.content)
        body = resp.json()
        self.assertEqual(body["status"], ImportSession.STATUS_AWAITING_RESOLVE)
        self.assertEqual(body["is_committable"], False)
        self.assertEqual(len(body["open_issues"]), 1)
        issue = body["open_issues"][0]
        self.assertEqual(issue["type"], "erp_id_conflict")
        self.assertEqual(issue["severity"], "error")
        self.assertEqual(issue["location"]["sheet"], "Transaction")
        self.assertEqual(issue["location"]["erp_id"], "OMIE-1")
        self.assertIn("date", issue["context"]["fields"])
        self.assertIn("pick_row", issue["proposed_actions"])

    def test_non_transaction_sheets_do_not_trigger_erp_id_conflict_detection(self):
        """Journal entry or bank sheets can have shared erp_ids without
        the Phase 2 conflict detector firing (manual §11.10c scopes
        grouping to Transactions only). Mocks the dry-run clean so
        ``dry_run_failure`` doesn't sneak in either."""
        xlsx = _build_xlsx({
            "JournalEntry": [
                {"__erp_id": "SHARED", "debit_amount": "100", "credit_amount": 0},
                {"__erp_id": "SHARED", "debit_amount": 0, "credit_amount": "100"},
            ],
        })
        clean_imports = {
            "imports": [{
                "model": "JournalEntry",
                "result": [
                    {"__row_id": "r1", "action": "create", "status": "success"},
                    {"__row_id": "r2", "action": "create", "status": "success"},
                ],
            }],
        }
        with mock.patch(
            "multitenancy.imports_v2.services.execute_import_job",
            return_value=clean_imports,
        ):
            resp = self.client.post(
                _url_analyze(self.company.id),
                {"file": _upload_file(xlsx)},
                format="multipart",
            )
        self.assertEqual(resp.status_code, 202, resp.content)
        body = resp.json()
        self.assertEqual(body["status"], ImportSession.STATUS_READY)
        self.assertEqual(body["open_issues"], [])


@override_settings(AUTH_OFF=True)
class V2TemplateDryRunPreviewTests(TestCase):
    """Template analyze dry-run preview (gated on row count).

    Tests the service-layer behaviour, not the endpoint — patches
    ``execute_import_job`` so we're testing the tally logic +
    threshold gate, not the real write pipeline (which has its own
    tests elsewhere).
    """

    @classmethod
    def setUpTestData(cls):
        cls.company = Company.objects.create(name="Acme Co")
        cls.user = User.objects.create_user(username="op", password="x")

    def _build_file(self, row_count: int) -> bytes:
        """Build a Transaction-only workbook with N identical-ish rows —
        only the ``__erp_id`` varies so we don't trigger erp_id_conflict."""
        rows = [
            {
                "__erp_id": f"OMIE-{i}",
                "date": "2026-01-01",
                "description": "x",
                "amount": "100.00",
            }
            for i in range(row_count)
        ]
        return _build_xlsx({"Transaction": rows})

    def test_small_file_populates_preview_counts(self):
        """Under the threshold → ``execute_import_job`` is called with
        ``commit=False`` and the result is tallied into the preview
        counts the serializer exposes."""
        from multitenancy.imports_v2 import services as svc

        xlsx = self._build_file(3)
        fake_imports = {
            "imports": [
                {
                    "model": "Transaction",
                    "result": [
                        {"__row_id": "r1", "action": "create", "status": "success"},
                        {"__row_id": "r2", "action": "create", "status": "success"},
                        {"__row_id": "r3", "action": "update", "status": "success"},
                    ],
                }
            ]
        }
        with mock.patch.object(svc, "execute_import_job", return_value=fake_imports) as ej:
            session = svc.analyze_template(
                company_id=self.company.id,
                user=self.user,
                file_bytes=xlsx,
                file_name="t.xlsx",
            )
        # execute_import_job was called with commit=False (the dry-run)
        self.assertTrue(ej.called)
        _, kwargs = ej.call_args
        self.assertFalse(kwargs["commit"])
        # Preview counts stored on parsed_payload and mirrored through
        # the serializer via ``session.preview`` (checked through
        # ``parsed_payload['preview']`` directly since this is a
        # service-layer test).
        preview = (session.parsed_payload or {}).get("preview") or {}
        self.assertEqual(preview.get("would_create"), {"Transaction": 2})
        self.assertEqual(preview.get("would_update"), {"Transaction": 1})
        self.assertEqual(preview.get("would_fail", {}), {})
        self.assertEqual(preview.get("total_rows"), 3)

    def test_error_rows_bucket_into_would_fail(self):
        """Rows with ``status=error`` → ``would_fail``, NOT ``would_create``."""
        from multitenancy.imports_v2 import services as svc

        xlsx = self._build_file(2)
        fake_imports = {
            "imports": [
                {
                    "model": "Transaction",
                    "result": [
                        {"__row_id": "r1", "action": "create", "status": "success"},
                        {"__row_id": "r2", "action": "create", "status": "error",
                         "message": "something bad"},
                    ],
                }
            ]
        }
        with mock.patch.object(svc, "execute_import_job", return_value=fake_imports):
            session = svc.analyze_template(
                company_id=self.company.id,
                user=self.user,
                file_bytes=xlsx,
                file_name="t.xlsx",
            )
        preview = (session.parsed_payload or {}).get("preview") or {}
        self.assertEqual(preview.get("would_create"), {"Transaction": 1})
        self.assertEqual(preview.get("would_fail"), {"Transaction": 1})

    def test_above_threshold_skips_dry_run(self):
        """When the file's row count exceeds
        ``TEMPLATE_DRY_RUN_ROW_THRESHOLD`` the dry-run is skipped —
        ``execute_import_job`` is NOT called during analyze. The
        session still lands in ready; preview stays empty."""
        from multitenancy.imports_v2 import services as svc

        # Temporarily lower the threshold so we don't have to build a
        # 5001-row workbook.
        with (
            mock.patch.object(svc, "TEMPLATE_DRY_RUN_ROW_THRESHOLD", 2),
            mock.patch.object(svc, "execute_import_job") as ej,
        ):
            xlsx = self._build_file(5)
            session = svc.analyze_template(
                company_id=self.company.id,
                user=self.user,
                file_bytes=xlsx,
                file_name="t.xlsx",
            )
        ej.assert_not_called()
        self.assertEqual(session.status, ImportSession.STATUS_READY)
        self.assertEqual(
            (session.parsed_payload or {}).get("preview"), {},
        )

    def test_dry_run_exception_swallowed(self):
        """If the dry-run raises, analyze doesn't fail — preview is
        simply empty. The real commit will re-surface the same error
        if it's a genuine data problem."""
        from multitenancy.imports_v2 import services as svc

        xlsx = self._build_file(2)
        with mock.patch.object(
            svc, "execute_import_job", side_effect=RuntimeError("unexpected")
        ):
            session = svc.analyze_template(
                company_id=self.company.id,
                user=self.user,
                file_bytes=xlsx,
                file_name="t.xlsx",
            )
        self.assertEqual(session.status, ImportSession.STATUS_READY)
        self.assertEqual(
            (session.parsed_payload or {}).get("preview"), {},
        )

    def test_empty_sheets_skip_dry_run(self):
        """Parse returned nothing → preview stays empty, no call."""
        from multitenancy.imports_v2 import services as svc

        xlsx = _build_xlsx({"Transaction": []})
        with mock.patch.object(svc, "execute_import_job") as ej:
            session = svc.analyze_template(
                company_id=self.company.id,
                user=self.user,
                file_bytes=xlsx,
                file_name="t.xlsx",
            )
        ej.assert_not_called()
        self.assertEqual(session.status, ImportSession.STATUS_READY)

    # --- row-level preservation + dry_run_failure issue (2026-04-24) ------

    def test_error_rows_preserve_full_messages(self):
        """``row_results`` and ``full_row_results`` must carry the
        per-row ``message`` so the operator can see WHY each row would
        fail. Pre-fix the dry-run discarded everything except the int
        counts — see ``_template_dry_run_preview`` rewrite."""
        from multitenancy.imports_v2 import services as svc

        xlsx = self._build_file(2)
        fake_imports = {
            "imports": [{
                "model": "Transaction",
                "result": [
                    {"__row_id": "r1", "action": "create", "status": "success"},
                    {"__row_id": "r2", "action": "create", "status": "error",
                     "message": "FK entity not found: 'Acme'",
                     "data": {"entity": "Acme"}},
                ],
            }]
        }
        with mock.patch.object(svc, "execute_import_job", return_value=fake_imports):
            session = svc.analyze_template(
                company_id=self.company.id, user=self.user,
                file_bytes=xlsx, file_name="t.xlsx",
            )
        preview = (session.parsed_payload or {}).get("preview") or {}
        row_results = preview.get("row_results") or []
        full_rows = preview.get("full_row_results") or []

        error_rows = [r for r in row_results if r.get("status") == "error"]
        self.assertEqual(len(error_rows), 1)
        self.assertEqual(
            error_rows[0]["message"], "FK entity not found: 'Acme'",
        )
        self.assertEqual(error_rows[0]["__row_id"], "r2")
        # Small input: full list mirrors the display subset row-count.
        self.assertEqual(len(full_rows), 2)

    def test_success_rows_sampled_at_cap(self):
        """When a sheet has more than
        ``TEMPLATE_DRY_RUN_DISPLAY_SUCCESS_SAMPLE`` successful rows the
        display subset caps successes at the sample size while
        ``full_row_results`` still contains every row (for the xlsx
        download)."""
        from multitenancy.imports_v2 import services as svc

        with mock.patch.object(
            svc, "TEMPLATE_DRY_RUN_DISPLAY_SUCCESS_SAMPLE", 5,
        ):
            xlsx = self._build_file(20)
            fake_imports = {
                "imports": [{
                    "model": "Transaction",
                    "result": [
                        {"__row_id": f"r{i}", "action": "create", "status": "success"}
                        for i in range(20)
                    ],
                }]
            }
            with mock.patch.object(
                svc, "execute_import_job", return_value=fake_imports,
            ):
                session = svc.analyze_template(
                    company_id=self.company.id, user=self.user,
                    file_bytes=xlsx, file_name="t.xlsx",
                )
        preview = (session.parsed_payload or {}).get("preview") or {}
        row_results = preview.get("row_results") or []
        full_rows = preview.get("full_row_results") or []
        self.assertEqual(len(row_results), 5)
        self.assertEqual(len(full_rows), 20)
        self.assertTrue(preview.get("display_truncated"))

    def test_would_fail_blocks_commit_via_open_issue(self):
        """Any ``would_fail`` row emits a ``dry_run_failure`` issue and
        flips the session to ``awaiting_resolve`` so commit is blocked.
        Pre-fix: green 'Pronto para importar' coexisted with a
        '409 falhariam' preview — those two states cannot be both
        true at once."""
        from multitenancy.imports_v2 import services as svc
        from multitenancy.imports_v2 import issues as issue_mod

        xlsx = self._build_file(3)
        fake_imports = {
            "imports": [{
                "model": "Transaction",
                "result": [
                    {"__row_id": "r1", "action": "create", "status": "error",
                     "message": "boom-a"},
                    {"__row_id": "r2", "action": "create", "status": "error",
                     "message": "boom-a"},
                    {"__row_id": "r3", "action": "create", "status": "error",
                     "message": "boom-b"},
                ],
            }]
        }
        with mock.patch.object(svc, "execute_import_job", return_value=fake_imports):
            session = svc.analyze_template(
                company_id=self.company.id, user=self.user,
                file_bytes=xlsx, file_name="t.xlsx",
            )
        self.assertEqual(session.status, ImportSession.STATUS_AWAITING_RESOLVE)
        self.assertFalse(session.is_committable())

        dry_run_issues = [
            i for i in (session.open_issues or [])
            if i.get("type") == issue_mod.ISSUE_DRY_RUN_FAILURE
        ]
        self.assertEqual(len(dry_run_issues), 1)
        issue = dry_run_issues[0]
        self.assertEqual(issue["severity"], "error")
        self.assertEqual(issue["context"]["fail_count"], 3)
        self.assertEqual(issue["context"]["model"], "Transaction")
        self.assertEqual(issue["proposed_actions"], ["abort"])
        # Most-frequent distinct message comes first.
        self.assertEqual(issue["context"]["sample_messages"][0], "boom-a")
        self.assertIn("boom-b", issue["context"]["sample_messages"])

    def test_serializer_strips_full_row_results(self):
        """``full_row_results`` is large (up to 5k rows) and must NOT
        ride on the polling GET response. The serializer instead
        exposes ``full_row_count`` + ``has_full_download`` flags so the
        frontend can render 'Baixar Excel completo'."""
        from multitenancy.imports_v2 import services as svc
        from multitenancy.imports_v2.serializers import ImportSessionSerializer

        xlsx = self._build_file(2)
        fake_imports = {
            "imports": [{
                "model": "Transaction",
                "result": [
                    {"__row_id": "r1", "action": "create", "status": "success"},
                    {"__row_id": "r2", "action": "create", "status": "error",
                     "message": "bad"},
                ],
            }]
        }
        with mock.patch.object(svc, "execute_import_job", return_value=fake_imports):
            session = svc.analyze_template(
                company_id=self.company.id, user=self.user,
                file_bytes=xlsx, file_name="t.xlsx",
            )
        data = ImportSessionSerializer(session).data
        self.assertNotIn("full_row_results", data["preview"])
        self.assertEqual(data["preview"]["full_row_count"], 2)
        self.assertTrue(data["preview"]["has_full_download"])
        # The display subset IS served for the on-screen table.
        self.assertEqual(len(data["preview"]["row_results"]), 2)

    def test_preview_download_returns_xlsx(self):
        """Endpoint returns a real .xlsx with one sheet per model + one
        row per dry-run row."""
        from multitenancy.imports_v2 import services as svc
        import openpyxl, io

        xlsx = self._build_file(2)
        fake_imports = {
            "imports": [{
                "model": "Transaction",
                "result": [
                    {"__row_id": "r1", "action": "create", "status": "success",
                     "data": {"amount": "100"}},
                    {"__row_id": "r2", "action": "create", "status": "error",
                     "message": "bad", "data": {"amount": "-5"}},
                ],
            }]
        }
        with mock.patch.object(svc, "execute_import_job", return_value=fake_imports):
            session = svc.analyze_template(
                company_id=self.company.id, user=self.user,
                file_bytes=xlsx, file_name="t.xlsx",
            )

        client = APIClient()
        with override_settings(AUTH_OFF=True):
            resp = client.get(
                f"/{self.company.id}/api/core/imports/v2/sessions/{session.pk}/preview.xlsx"
            )
        self.assertEqual(resp.status_code, 200, resp.content[:200])
        self.assertTrue(resp["Content-Type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml"
        ))
        self.assertIn("attachment", resp["Content-Disposition"])

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        self.assertIn("Transaction", wb.sheetnames)
        ws = wb["Transaction"]
        header = [c.value for c in ws[1]]
        self.assertIn("__row_id", header)
        self.assertIn("status", header)
        self.assertIn("message", header)
        # Header + 2 data rows.
        self.assertEqual(ws.max_row, 3)

    def test_preview_download_404_when_no_preview(self):
        """Session with no preview data returns 404."""
        from multitenancy.imports_v2 import services as svc

        xlsx = _build_xlsx({"Transaction": []})
        with mock.patch.object(svc, "execute_import_job") as ej:
            session = svc.analyze_template(
                company_id=self.company.id, user=self.user,
                file_bytes=xlsx, file_name="t.xlsx",
            )
        ej.assert_not_called()

        client = APIClient()
        with override_settings(AUTH_OFF=True):
            resp = client.get(
                f"/{self.company.id}/api/core/imports/v2/sessions/{session.pk}/preview.xlsx"
            )
        self.assertEqual(resp.status_code, 404)


@override_settings(AUTH_OFF=True)
class V2CommitEndpointTests(TestCase):
    """commit/ — session-lifecycle transitions."""

    @classmethod
    def setUpTestData(cls):
        cls.company = Company.objects.create(name="Acme Co")
        # CustomUser has no direct company_id FK — tenant resolution
        # happens via the URL path (``/<tenant_id>/...``) and
        # ``TenantMiddleware`` in prod. Tests rely on the URL resolving
        # the company, not on a user↔company FK.
        cls.user = User.objects.create_user(username="op", password="x")

    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def _create_ready_session(self) -> ImportSession:
        return ImportSession.objects.create(
            company_id=self.company.id,
            created_by=self.user,
            mode=ImportSession.MODE_TEMPLATE,
            status=ImportSession.STATUS_READY,
            file_name="ok.xlsx",
            file_bytes=b"<pretend xlsx>",
            parsed_payload={"sheets": {"Transaction": [{"amount": "100"}]}},
            open_issues=[],
        )

    def _create_awaiting_resolve_session(self) -> ImportSession:
        return ImportSession.objects.create(
            company_id=self.company.id,
            created_by=self.user,
            mode=ImportSession.MODE_TEMPLATE,
            status=ImportSession.STATUS_AWAITING_RESOLVE,
            file_name="blocked.xlsx",
            file_bytes=b"<pretend xlsx>",
            parsed_payload={"sheets": {"Transaction": [{"amount": "100"}]}},
            open_issues=[{
                "issue_id": "iss-1",
                "type": "erp_id_conflict",
                "severity": "error",
                "location": {"sheet": "Transaction", "erp_id": "X"},
                "context": {"fields": {"date": ["a", "b"]}},
                "proposed_actions": ["pick_row"],
                "message": "conflict",
            }],
        )

    def test_commit_ready_session_calls_execute_import_job_and_finalises(self):
        session = self._create_ready_session()
        fake_result = {"imports": [{"model": "Transaction", "created": 1, "updated": 0}]}
        with mock.patch(
            "multitenancy.imports_v2.services.execute_import_job",
            return_value=fake_result,
        ) as mocked:
            resp = self.client.post(
                _url_commit(self.company.id, session.pk), {}, format="json",
            )
        self.assertEqual(resp.status_code, 202, resp.content)
        body = resp.json()
        self.assertEqual(body["status"], ImportSession.STATUS_COMMITTED)
        # Phase 4A commit wraps the write-backend result with
        # ``substitution_rules_created`` (empty when no rules were staged).
        self.assertEqual(body["result"]["imports"], fake_result["imports"])
        self.assertEqual(body["result"]["substitution_rules_created"], [])
        self.assertIsNotNone(body["committed_at"])
        # And ``execute_import_job`` was called with the session's data.
        call_kwargs = mocked.call_args.kwargs
        self.assertEqual(call_kwargs["company_id"], self.company.id)
        self.assertTrue(call_kwargs["commit"])
        self.assertEqual(
            [s["model"] for s in call_kwargs["sheets"]], ["Transaction"],
        )

        # file_bytes cleared after commit — no hoarding.
        session.refresh_from_db()
        self.assertIsNone(session.file_bytes)

    def test_commit_blocked_session_returns_409(self):
        session = self._create_awaiting_resolve_session()
        resp = self.client.post(
            _url_commit(self.company.id, session.pk), {}, format="json",
        )
        self.assertEqual(resp.status_code, 409, resp.content)
        body = resp.json()
        self.assertIn("not ready", body["error"].lower())
        self.assertEqual(body["status"], ImportSession.STATUS_AWAITING_RESOLVE)

    def test_commit_already_committed_session_returns_409(self):
        session = self._create_ready_session()
        session.status = ImportSession.STATUS_COMMITTED
        session.save(update_fields=["status"])
        resp = self.client.post(
            _url_commit(self.company.id, session.pk), {}, format="json",
        )
        self.assertEqual(resp.status_code, 409, resp.content)

    def test_commit_failure_marks_session_error(self):
        """When ``execute_import_job`` raises, the session should end up
        in ``error`` status with the diagnostic captured on ``result``.

        Phase 6.z: commit is async + runs in a Celery worker. The view
        returns 202 regardless of commit outcome — the frontend polls
        the detail endpoint for the final status. In eager mode the
        worker has already run by the time this returns, so the
        response body already reflects the terminal ``error`` status.
        """
        session = self._create_ready_session()
        with mock.patch(
            "multitenancy.imports_v2.services.execute_import_job",
            side_effect=RuntimeError("boom"),
        ):
            resp = self.client.post(
                _url_commit(self.company.id, session.pk), {}, format="json",
            )
        self.assertEqual(resp.status_code, 202, resp.content)
        body = resp.json()
        self.assertEqual(body["status"], ImportSession.STATUS_ERROR)
        self.assertEqual(body["result"].get("error"), "boom")
        session.refresh_from_db()
        self.assertEqual(session.status, ImportSession.STATUS_ERROR)
        self.assertEqual(session.result.get("error"), "boom")


@override_settings(AUTH_OFF=True)
class V2SessionDetailTests(TestCase):
    """GET + DELETE /sessions/<id>/."""

    @classmethod
    def setUpTestData(cls):
        cls.co1 = Company.objects.create(name="Acme Co")
        cls.co2 = Company.objects.create(name="Globex Co")
        cls.user1 = User.objects.create_user(username="op1", password="x")
        cls.user2 = User.objects.create_user(username="op2", password="x")

    def _make_session(self, company):
        return ImportSession.objects.create(
            company_id=company.id,
            mode=ImportSession.MODE_TEMPLATE,
            status=ImportSession.STATUS_AWAITING_RESOLVE,
            file_name="x",
            parsed_payload={"sheets": {}},
        )

    def test_get_returns_own_session(self):
        session = self._make_session(self.co1)
        client = APIClient()
        client.force_authenticate(self.user1)
        resp = client.get(_url_session(self.co1.id, session.pk))
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()["id"], session.pk)

    def test_get_other_tenants_session_returns_404(self):
        """Cross-tenant reads must return 404 — we don't even confirm the
        id exists in another tenant's space."""
        session = self._make_session(self.co1)
        client = APIClient()
        client.force_authenticate(self.user2)
        resp = client.get(_url_session(self.co2.id, session.pk))
        self.assertEqual(resp.status_code, 404)

    def test_delete_transitions_to_discarded(self):
        session = self._make_session(self.co1)
        client = APIClient()
        client.force_authenticate(self.user1)
        resp = client.delete(_url_session(self.co1.id, session.pk))
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()["status"], ImportSession.STATUS_DISCARDED)

    def test_delete_already_committed_is_noop(self):
        session = self._make_session(self.co1)
        session.status = ImportSession.STATUS_COMMITTED
        session.save(update_fields=["status"])
        client = APIClient()
        client.force_authenticate(self.user1)
        resp = client.delete(_url_session(self.co1.id, session.pk))
        # Terminal sessions don't change state — still committed.
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()["status"], ImportSession.STATUS_COMMITTED)


@override_settings(AUTH_OFF=True)
class V2AsyncTaskTests(TestCase):
    """Phase 6.z — analyze and commit run through Celery.

    These tests hit the task functions directly (no HTTP) to verify
    the worker-only branches: status-gating on re-entry, missing
    session id, and terminal-status translation on unhandled
    exceptions. The happy path is covered by the endpoint tests
    above — in eager mode every endpoint test already goes through
    ``.delay()`` → task body.
    """

    @classmethod
    def setUpTestData(cls):
        cls.company = Company.objects.create(name="Acme Co")

    def _make_session(self, status: str) -> ImportSession:
        return ImportSession.objects.create(
            company_id=self.company.id,
            mode=ImportSession.MODE_TEMPLATE,
            status=status,
            file_name="x.xlsx",
            parsed_payload={"sheets": {"Transaction": []}},
        )

    # --- analyze_session_task ------------------------------------------------

    def test_analyze_task_bails_when_session_missing(self):
        """Missing session → warn + return, no crash."""
        from multitenancy.imports_v2.tasks import analyze_session_task
        # No exception; nothing to assert on — if this raises, the
        # worker would crashloop in production.
        analyze_session_task(9_999_999)

    def test_analyze_task_bails_when_session_not_in_analyzing(self):
        """If the session was discarded / is already terminal, the
        worker must not re-run analyze (would corrupt parsed_payload).
        """
        from multitenancy.imports_v2.tasks import analyze_session_task
        session = self._make_session(ImportSession.STATUS_DISCARDED)
        with mock.patch(
            "multitenancy.imports_v2.services._run_analyze_template"
        ) as body_mock:
            analyze_session_task(session.pk)
        body_mock.assert_not_called()

    def test_analyze_task_marks_error_on_unhandled_exception(self):
        """If the analyze body crashes with something other than the
        ValueError it handles itself, the task wrapper must flip the
        session to ``error`` with a traceback — otherwise the frontend
        poll would hang forever on ``analyzing``."""
        from multitenancy.imports_v2.tasks import analyze_session_task
        session = self._make_session(ImportSession.STATUS_ANALYZING)
        with mock.patch(
            "multitenancy.imports_v2.services._run_analyze_template",
            side_effect=RuntimeError("kaboom"),
        ):
            analyze_session_task(session.pk)
        session.refresh_from_db()
        self.assertEqual(session.status, ImportSession.STATUS_ERROR)
        self.assertEqual(session.result["stage"], "analyze")
        self.assertEqual(session.result["error"], "kaboom")
        self.assertEqual(session.result["type"], "RuntimeError")
        self.assertIn("traceback", session.result)

    def test_analyze_task_dispatches_by_mode(self):
        """ETL-mode session → ``_run_analyze_etl``; template-mode →
        ``_run_analyze_template``. Dispatching on ``session.mode``
        instead of task kwargs keeps the task signature narrow
        (just the pk)."""
        from multitenancy.imports_v2.tasks import analyze_session_task
        etl_session = ImportSession.objects.create(
            company_id=self.company.id,
            mode=ImportSession.MODE_ETL,
            status=ImportSession.STATUS_ANALYZING,
            file_name="x.xlsx",
        )
        with mock.patch(
            "multitenancy.imports_v2.services._run_analyze_etl"
        ) as etl_mock, mock.patch(
            "multitenancy.imports_v2.services._run_analyze_template"
        ) as tpl_mock:
            analyze_session_task(etl_session.pk)
        etl_mock.assert_called_once()
        tpl_mock.assert_not_called()

    # --- commit_session_task -------------------------------------------------

    def test_commit_task_bails_when_session_missing(self):
        from multitenancy.imports_v2.tasks import commit_session_task
        commit_session_task(9_999_999)  # no raise → pass

    def test_commit_task_bails_when_session_not_in_committing(self):
        """If the session was moved back to awaiting_resolve (rare but
        possible via concurrent discard), skip the commit body."""
        from multitenancy.imports_v2.tasks import commit_session_task
        session = self._make_session(ImportSession.STATUS_READY)
        with mock.patch(
            "multitenancy.imports_v2.services._run_commit"
        ) as body_mock:
            commit_session_task(session.pk)
        body_mock.assert_not_called()

    def test_commit_task_safety_net_on_unexpected_exception(self):
        """_run_commit already flips to ``error`` on known exceptions.
        This test patches _run_commit to raise WITHOUT flipping status
        (simulating a bug in the body) and verifies the task wrapper
        catches it and writes the error status as a safety net."""
        from multitenancy.imports_v2.tasks import commit_session_task
        session = self._make_session(ImportSession.STATUS_COMMITTING)

        def raise_without_status_flip(session_arg):
            raise RuntimeError("bug in body")

        with mock.patch(
            "multitenancy.imports_v2.services._run_commit",
            side_effect=raise_without_status_flip,
        ):
            commit_session_task(session.pk)
        session.refresh_from_db()
        self.assertEqual(session.status, ImportSession.STATUS_ERROR)
        self.assertEqual(session.result["stage"], "commit")
        self.assertEqual(session.result["error"], "bug in body")

    def test_commit_task_does_not_overwrite_error_from_body(self):
        """If ``_run_commit`` already wrote a precise error (its own
        diagnostic from execute_import_job failing), the task wrapper
        must NOT clobber it with a generic traceback."""
        from multitenancy.imports_v2.tasks import commit_session_task
        session = self._make_session(ImportSession.STATUS_COMMITTING)

        def simulate_body_writing_error(session_arg):
            session_arg.status = ImportSession.STATUS_ERROR
            session_arg.result = {
                "error": "precise diagnostic",
                "stage": "commit",
                "type": "IntegrityError",
            }
            session_arg.save(update_fields=["status", "result"])
            raise RuntimeError("re-raised after writing status")

        with mock.patch(
            "multitenancy.imports_v2.services._run_commit",
            side_effect=simulate_body_writing_error,
        ):
            commit_session_task(session.pk)
        session.refresh_from_db()
        self.assertEqual(session.status, ImportSession.STATUS_ERROR)
        self.assertEqual(session.result["error"], "precise diagnostic")
        self.assertEqual(session.result["type"], "IntegrityError")
        # The wrapper's generic 'traceback' field must not be present.
        self.assertNotIn("traceback", session.result)


@override_settings(AUTH_OFF=True)
class V2SessionsListEndpointTests(TestCase):
    """Phase 6.z-b — GET /sessions/ and /sessions/running-count/.

    Covers tenant scoping, filters, pagination shape, the
    lightweight serializer's payload, and the cheap count aggregate.
    """

    @classmethod
    def setUpTestData(cls):
        cls.co1 = Company.objects.create(name="Acme Co")
        cls.co2 = Company.objects.create(name="Globex Co")
        cls.user = User.objects.create_user(
            username="op",
            password="x",
            first_name="Olivia",
            last_name="Operator",
        )

    def _make(self, company, status, *, mode="template"):
        return ImportSession.objects.create(
            company_id=company.id,
            created_by=self.user,
            mode=mode,
            status=status,
            file_name=f"{status}.xlsx",
            parsed_payload={"sheets": {"Transaction": [{"x": 1}, {"x": 2}]}},
            open_issues=[] if status in (
                ImportSession.STATUS_READY, ImportSession.STATUS_COMMITTED,
            ) else [{"severity": "error", "type": "erp_id_conflict"}],
        )

    # --- list --------------------------------------------------------------

    def test_list_scopes_to_tenant(self):
        """Sessions in another company must not appear — even without a
        403; just invisible rows."""
        own = self._make(self.co1, ImportSession.STATUS_READY)
        other = self._make(self.co2, ImportSession.STATUS_READY)
        resp = self.client.get(_url_sessions_list(self.co1.id))
        self.assertEqual(resp.status_code, 200, resp.content)
        body = resp.json()
        ids = [r["id"] for r in body["results"]]
        self.assertIn(own.pk, ids)
        self.assertNotIn(other.pk, ids)

    def test_list_payload_is_lightweight(self):
        """The list serializer excludes the heavy blobs — if a future
        dev re-adds parsed_payload / open_issues to the list response,
        this test flags it early."""
        self._make(self.co1, ImportSession.STATUS_READY)
        resp = self.client.get(_url_sessions_list(self.co1.id))
        row = resp.json()["results"][0]
        self.assertIn("id", row)
        self.assertIn("mode", row)
        self.assertIn("status", row)
        self.assertIn("file_name", row)
        self.assertIn("operator_name", row)
        self.assertIn("open_issue_count", row)
        self.assertIn("is_terminal", row)
        # Heavy fields MUST NOT be present — the queue would be slow +
        # memory-hungry with multi-MB payloads per row.
        self.assertNotIn("parsed_payload", row)
        self.assertNotIn("open_issues", row)
        self.assertNotIn("result", row)
        self.assertNotIn("staged_substitution_rules", row)

    def test_list_orders_by_newest_first(self):
        older = self._make(self.co1, ImportSession.STATUS_READY)
        newer = self._make(self.co1, ImportSession.STATUS_AWAITING_RESOLVE)
        resp = self.client.get(_url_sessions_list(self.co1.id))
        ids = [r["id"] for r in resp.json()["results"]]
        self.assertEqual(ids[0], newer.pk)
        self.assertEqual(ids[1], older.pk)

    def test_list_filters_by_status(self):
        analyzing = self._make(self.co1, ImportSession.STATUS_ANALYZING)
        committing = self._make(self.co1, ImportSession.STATUS_COMMITTING)
        committed = self._make(self.co1, ImportSession.STATUS_COMMITTED)
        resp = self.client.get(
            _url_sessions_list(self.co1.id) + "?status=analyzing,committing",
        )
        ids = {r["id"] for r in resp.json()["results"]}
        self.assertIn(analyzing.pk, ids)
        self.assertIn(committing.pk, ids)
        self.assertNotIn(committed.pk, ids)

    def test_list_filters_by_mode(self):
        tmpl = self._make(self.co1, ImportSession.STATUS_READY, mode="template")
        etl = self._make(self.co1, ImportSession.STATUS_READY, mode="etl")
        resp = self.client.get(_url_sessions_list(self.co1.id) + "?mode=etl")
        ids = {r["id"] for r in resp.json()["results"]}
        self.assertIn(etl.pk, ids)
        self.assertNotIn(tmpl.pk, ids)

    def test_list_exposes_operator_and_issue_count(self):
        s = self._make(self.co1, ImportSession.STATUS_AWAITING_RESOLVE)
        resp = self.client.get(_url_sessions_list(self.co1.id))
        row = next(r for r in resp.json()["results"] if r["id"] == s.pk)
        self.assertEqual(row["operator_name"], "Olivia Operator")
        self.assertEqual(row["open_issue_count"], 1)

    def test_list_rejects_invalid_status_silently(self):
        """An unknown ``status=foo`` filter drops to no filter (serving
        all rows) rather than 400 — keeps the frontend resilient to typos
        without throwing the whole queue."""
        self._make(self.co1, ImportSession.STATUS_READY)
        resp = self.client.get(
            _url_sessions_list(self.co1.id) + "?status=bogus",
        )
        self.assertEqual(resp.status_code, 200)
        # All own-tenant rows still present.
        self.assertGreaterEqual(len(resp.json()["results"]), 1)

    # --- running-count -----------------------------------------------------

    def test_running_count_buckets(self):
        self._make(self.co1, ImportSession.STATUS_ANALYZING)
        self._make(self.co1, ImportSession.STATUS_ANALYZING)
        self._make(self.co1, ImportSession.STATUS_COMMITTING)
        self._make(self.co1, ImportSession.STATUS_AWAITING_RESOLVE)
        # Terminal sessions must not count.
        self._make(self.co1, ImportSession.STATUS_COMMITTED)
        self._make(self.co1, ImportSession.STATUS_ERROR)
        self._make(self.co1, ImportSession.STATUS_DISCARDED)
        # A session in another tenant must not count.
        self._make(self.co2, ImportSession.STATUS_ANALYZING)

        resp = self.client.get(_url_sessions_running_count(self.co1.id))
        self.assertEqual(resp.status_code, 200, resp.content)
        body = resp.json()
        self.assertEqual(body["analyzing"], 2)
        self.assertEqual(body["committing"], 1)
        self.assertEqual(body["awaiting_resolve"], 1)
        self.assertEqual(body["total"], 4)

    def test_running_count_zero_when_idle(self):
        self._make(self.co1, ImportSession.STATUS_COMMITTED)
        resp = self.client.get(_url_sessions_running_count(self.co1.id))
        body = resp.json()
        self.assertEqual(body["total"], 0)
        self.assertEqual(body["analyzing"], 0)
        self.assertEqual(body["committing"], 0)
        self.assertEqual(body["awaiting_resolve"], 0)


@override_settings(AUTH_OFF=True)
class V2SubstitutionCacheTests(TestCase):
    """Phase 6.z-d — pre-substitute at analyze, reuse at commit.

    The service-layer cache reuse is orthogonal to the view — these
    tests poke the helpers directly and the commit path through
    mocks so we can assert the ``skip_substitutions`` flag flows or
    not based on the revision + resolutions invariant.
    """

    @classmethod
    def setUpTestData(cls):
        cls.company = Company.objects.create(name="Acme Co")

    # --- revision hash -----------------------------------------------------

    def test_revision_hash_is_stable_when_rules_unchanged(self):
        from multitenancy.imports_v2 import services as svc
        from multitenancy.models import SubstitutionRule

        SubstitutionRule.objects.create(
            company_id=self.company.id,
            model_name="Transaction",
            field_name="description",
            match_type="exact",
            match_value="foo",
            substitution_value="bar",
        )
        a = svc._compute_substitution_revision(self.company.id)
        b = svc._compute_substitution_revision(self.company.id)
        self.assertEqual(a, b)

    def test_revision_hash_flips_when_rule_added(self):
        from multitenancy.imports_v2 import services as svc
        from multitenancy.models import SubstitutionRule

        before = svc._compute_substitution_revision(self.company.id)
        SubstitutionRule.objects.create(
            company_id=self.company.id,
            model_name="Transaction",
            field_name="description",
            match_type="exact",
            match_value="foo",
            substitution_value="bar",
        )
        after = svc._compute_substitution_revision(self.company.id)
        self.assertNotEqual(before, after)

    def test_revision_hash_is_tenant_scoped(self):
        """Rules on another tenant must not perturb this tenant's hash."""
        from multitenancy.imports_v2 import services as svc
        from multitenancy.models import SubstitutionRule

        co2 = Company.objects.create(name="Globex Co")
        before = svc._compute_substitution_revision(self.company.id)
        SubstitutionRule.objects.create(
            company_id=co2.id,
            model_name="Transaction",
            field_name="description",
            match_type="exact",
            match_value="x",
            substitution_value="y",
        )
        after = svc._compute_substitution_revision(self.company.id)
        self.assertEqual(before, after)

    # --- cache-valid predicate --------------------------------------------

    def _make_session(self, *, rev=None, cached=None, resolutions=None):
        payload = {"sheets": {"Transaction": [{"x": 1}]}}
        if rev is not None:
            payload["substitution_revision"] = rev
        if cached is not None:
            payload["sheets_post_substitution"] = cached
        return ImportSession.objects.create(
            company_id=self.company.id,
            mode=ImportSession.MODE_TEMPLATE,
            status=ImportSession.STATUS_READY,
            file_name="x.xlsx",
            parsed_payload=payload,
            resolutions=resolutions or [],
        )

    def test_cache_valid_when_rev_matches_and_no_resolutions(self):
        from multitenancy.imports_v2 import services as svc
        s = self._make_session(
            rev="abc123",
            cached={"Transaction": [{"x": 1}]},
            resolutions=[],
        )
        self.assertTrue(svc._is_substitution_cache_valid(s, "abc123"))

    def test_cache_invalid_when_rev_differs(self):
        from multitenancy.imports_v2 import services as svc
        s = self._make_session(
            rev="abc123",
            cached={"Transaction": [{"x": 1}]},
        )
        self.assertFalse(svc._is_substitution_cache_valid(s, "different"))

    def test_cache_invalid_when_any_resolution_recorded(self):
        """edit_value resolutions mutate raw sheets — the cached
        substituted copy wouldn't reflect that, so we must
        re-substitute at commit."""
        from multitenancy.imports_v2 import services as svc
        s = self._make_session(
            rev="abc123",
            cached={"Transaction": [{"x": 1}]},
            resolutions=[{"issue_id": "iss-1", "action": "edit_value"}],
        )
        self.assertFalse(svc._is_substitution_cache_valid(s, "abc123"))

    def test_cache_invalid_when_payload_missing_cache_key(self):
        from multitenancy.imports_v2 import services as svc
        s = self._make_session()  # no rev, no cache
        self.assertFalse(svc._is_substitution_cache_valid(s, "anything"))

    # --- commit path integration ------------------------------------------

    def test_commit_passes_skip_substitutions_when_cache_valid(self):
        """Happy path: analyze cached fresh rows + rules unchanged →
        commit passes ``skip_substitutions=True`` + the cached rows so
        the pipeline doesn't double-apply the substitution pass."""
        from multitenancy.imports_v2 import services as svc

        rev = svc._compute_substitution_revision(self.company.id)
        cached = {"Transaction": [{"__erp_id": "X", "amount": "100"}]}
        session = self._make_session(rev=rev, cached=cached, resolutions=[])

        fake_result = {"imports": []}
        with mock.patch(
            "multitenancy.imports_v2.services.execute_import_job",
            return_value=fake_result,
        ) as job:
            svc._commit_template_session(session)

        kwargs = job.call_args.kwargs
        self.assertEqual(
            kwargs["import_metadata"].get("import_options", {}).get(
                "skip_substitutions"
            ),
            True,
        )
        # Sheets passed must be the cached copy, not the raw one.
        self.assertEqual(
            [s["model"] for s in kwargs["sheets"]],
            ["Transaction"],
        )
        self.assertEqual(
            kwargs["sheets"][0]["rows"],
            cached["Transaction"],
        )

    def test_write_progress_merges_and_stamps_updated_at(self):
        """``_write_progress`` merges kwargs into ``session.progress``
        and stamps a fresh ``updated_at`` on every call. Merge lets
        callers update one field without clobbering prior ones."""
        from multitenancy.imports_v2 import services as svc

        session = self._make_session()
        svc._write_progress(
            session, stage="detecting", sheets_total=3, sheets_done=0,
        )
        session.refresh_from_db()
        self.assertEqual(session.progress.get("stage"), "detecting")
        self.assertEqual(session.progress.get("sheets_total"), 3)
        self.assertEqual(session.progress.get("sheets_done"), 0)
        self.assertIn("updated_at", session.progress)
        first_stamp = session.progress["updated_at"]

        svc._write_progress(session, sheets_done=2, errors_so_far=1)
        session.refresh_from_db()
        # Merged — stage still there from the earlier call.
        self.assertEqual(session.progress.get("stage"), "detecting")
        self.assertEqual(session.progress.get("sheets_done"), 2)
        self.assertEqual(session.progress.get("errors_so_far"), 1)
        # New stamp (greater-or-equal — test time resolution is ms).
        self.assertGreaterEqual(session.progress["updated_at"], first_stamp)

    def test_write_progress_swallows_save_errors(self):
        """Progress writes must not crash the worker. If the save
        fails mid-atomic, log and carry on — the main work still
        proceeds."""
        from multitenancy.imports_v2 import services as svc

        session = self._make_session()
        with mock.patch.object(
            ImportSession, "save", side_effect=RuntimeError("db blew up"),
        ):
            # Should not raise.
            svc._write_progress(session, stage="writing")

    def test_commit_falls_through_when_cache_stale(self):
        """Rule set changed mid-session → commit must re-substitute
        (no skip flag) so the new rules actually apply."""
        from multitenancy.imports_v2 import services as svc
        from multitenancy.models import SubstitutionRule

        # Cache with a revision that definitely doesn't match current.
        cached = {"Transaction": [{"x": 1}]}
        session = self._make_session(
            rev="stale-hash", cached=cached, resolutions=[],
        )
        # Add a rule after analyze → revision flips.
        SubstitutionRule.objects.create(
            company_id=self.company.id,
            model_name="Transaction",
            field_name="description",
            match_type="exact",
            match_value="a",
            substitution_value="b",
        )

        fake_result = {"imports": []}
        with mock.patch(
            "multitenancy.imports_v2.services.execute_import_job",
            return_value=fake_result,
        ) as job:
            svc._commit_template_session(session)

        kwargs = job.call_args.kwargs
        self.assertIsNone(
            kwargs["import_metadata"].get("import_options", {}).get(
                "skip_substitutions"
            ),
        )
        # Raw sheets passed, not cached — the pipeline will run
        # substitutions fresh.
        self.assertEqual(
            kwargs["sheets"][0]["rows"],
            [{"x": 1}],  # matches ``self._make_session`` default
        )


@override_settings(AUTH_OFF=True)
class V2StaleSessionReaperTests(TestCase):
    """Phase 6.z-f — periodic beat task reaps v2 sessions stuck in
    ``analyzing`` / ``committing`` past the Celery hard time limit.

    Protects against worker SIGKILL / container restart / broker loss
    scenarios where the task dies without flipping session status.
    The frontend polls these sessions indefinitely; the reaper flips
    them to ``error`` so the operator can discard + retry.
    """

    @classmethod
    def setUpTestData(cls):
        cls.company = Company.objects.create(name="Acme Co")

    def _make_stale_session(self, status, *, seconds_ago: int):
        """Create a session whose ``updated_at`` is in the past.

        Django's ``auto_now`` on ``updated_at`` fires on every save,
        so we bypass the ORM helper and update via ``.filter().update()``
        to set the timestamp precisely.
        """
        from django.utils import timezone as tz
        from datetime import timedelta

        s = ImportSession.objects.create(
            company_id=self.company.id,
            mode=ImportSession.MODE_TEMPLATE,
            status=status,
            file_name="stuck.xlsx",
        )
        past = tz.now() - timedelta(seconds=seconds_ago)
        ImportSession.objects.filter(pk=s.pk).update(updated_at=past)
        s.refresh_from_db()
        return s

    def test_reaper_flips_stuck_analyzing_session(self):
        from multitenancy.imports_v2.tasks import reap_stale_sessions_task

        # 30-minute default hard limit + 60s grace = reap at 31+ min old.
        stuck = self._make_stale_session(
            ImportSession.STATUS_ANALYZING, seconds_ago=60 * 60,
        )
        result = reap_stale_sessions_task()
        self.assertEqual(result["reaped"], 1)
        self.assertIn(stuck.pk, result["session_pks"])
        stuck.refresh_from_db()
        self.assertEqual(stuck.status, ImportSession.STATUS_ERROR)
        self.assertEqual(stuck.result["stage"], "timeout")
        self.assertEqual(stuck.result["prior_status"], "analyzing")

    def test_reaper_flips_stuck_committing_session(self):
        from multitenancy.imports_v2.tasks import reap_stale_sessions_task

        stuck = self._make_stale_session(
            ImportSession.STATUS_COMMITTING, seconds_ago=60 * 60,
        )
        reap_stale_sessions_task()
        stuck.refresh_from_db()
        self.assertEqual(stuck.status, ImportSession.STATUS_ERROR)
        self.assertEqual(stuck.result["prior_status"], "committing")

    def test_reaper_ignores_fresh_sessions(self):
        """A session that's only been running 30s must NOT be reaped —
        that would kill legitimate in-flight work."""
        from multitenancy.imports_v2.tasks import reap_stale_sessions_task

        fresh = self._make_stale_session(
            ImportSession.STATUS_ANALYZING, seconds_ago=30,
        )
        result = reap_stale_sessions_task()
        self.assertEqual(result["reaped"], 0)
        fresh.refresh_from_db()
        self.assertEqual(fresh.status, ImportSession.STATUS_ANALYZING)

    def test_reaper_ignores_terminal_sessions(self):
        """Committed / errored / discarded rows stay as-is regardless
        of age — reaper only touches non-terminal sessions."""
        from multitenancy.imports_v2.tasks import reap_stale_sessions_task

        old_committed = self._make_stale_session(
            ImportSession.STATUS_COMMITTED, seconds_ago=60 * 60,
        )
        result = reap_stale_sessions_task()
        self.assertEqual(result["reaped"], 0)
        old_committed.refresh_from_db()
        self.assertEqual(old_committed.status, ImportSession.STATUS_COMMITTED)


class V2TaskRegistrationTests(TestCase):
    """Regression: ensure the v2 ``@shared_task``s land in the worker's
    global Celery registry at boot.

    Why this test exists — the three v2 tasks live at
    ``multitenancy/imports_v2/tasks.py``, a nested module that Celery's
    default ``autodiscover_tasks()`` (scans only each app's top-level
    ``tasks.py``) does NOT reach. Without a transitive import from
    ``multitenancy/tasks.py``, the **web** process registers them
    (via URLConf → views → services → tasks) but the **worker**
    process does not — producing the silent production failure where
    ``.delay()`` and Beat-scheduled tasks are dropped on the worker
    side with a ``Received unregistered task`` log and nothing ever
    runs.
    """

    def test_multitenancy_tasks_transitively_imports_v2_tasks(self):
        """The top-level task module must keep the nested module as an
        attribute so a linter / future refactor can't quietly drop the
        import."""
        import multitenancy.tasks as top_tasks
        from multitenancy.imports_v2 import tasks as nested_tasks

        self.assertIs(
            top_tasks._imports_v2_tasks_register, nested_tasks,
            "multitenancy/tasks.py must import "
            "multitenancy.imports_v2.tasks so Celery autodiscover "
            "reaches it on the worker. See module-level comment for "
            "the full symptom.",
        )

    def test_v2_task_names_registered_on_celery_app(self):
        """The three v2 @shared_task names must be present in the
        global Celery registry so the worker can dispatch them."""
        from nord_backend.celery import app

        required_names = (
            "imports_v2.analyze_session",
            "imports_v2.commit_session",
            "imports_v2.reap_stale_sessions",
        )
        missing = [n for n in required_names if n not in app.tasks]
        self.assertEqual(
            missing, [],
            f"These v2 task names are missing from app.tasks: {missing}. "
            "The worker process cannot execute them — Beat schedules "
            "and .delay() calls will be silently dropped.",
        )


@override_settings(AUTH_OFF=True)
class V2ProgressCallbackTests(TestCase):
    """Phase 6.z-g — intra-atomic row-level progress via Redis channel.

    The import job's ``progress_callback`` fires at sheet boundaries +
    every N rows so the v2 service can publish row-level updates to
    Redis (outside the DB transaction). Tests here exercise the
    callback invocation pattern — the Redis integration is tested
    via the serializer merge tests below with a mocked channel.
    """

    def test_callback_fires_at_sheet_boundary(self):
        """``execute_import_job`` must call the callback at the start
        of each sheet with ``stage=writing``, ``sheet_index``,
        ``current_sheet``, and ``rows_total``."""
        from multitenancy.tasks import execute_import_job

        calls: list = []

        def cb(fields):
            calls.append(dict(fields))

        # Minimal sheets — model must be in MODEL_APP_MAP or the loop
        # short-circuits. Transaction is registered (accounting app).
        # We don't actually care about success here; we care about the
        # callback firing before any per-row work.
        sheets = [
            {"model": "Transaction", "rows": [{"x": 1}, {"x": 2}]},
        ]
        try:
            execute_import_job(
                company_id=1,
                sheets=sheets,
                commit=False,
                progress_callback=cb,
            )
        except Exception:
            # execute_import_job will likely error on a made-up row —
            # that's fine, we just need the sheet-start callback.
            pass

        self.assertTrue(calls, "callback never fired")
        first = calls[0]
        self.assertEqual(first.get("stage"), "writing")
        self.assertEqual(first.get("sheet_index"), 0)
        self.assertEqual(first.get("current_sheet"), "Transaction")
        self.assertEqual(first.get("rows_total"), 2)
        self.assertEqual(first.get("rows_processed"), 0)

    def test_callback_fires_every_100_rows(self):
        """With a 250-row sheet the callback should fire at least at
        rows 0, 100, and 200 (sheet-start + 2 throttled row updates)."""
        from multitenancy.tasks import execute_import_job

        calls: list = []

        def cb(fields):
            calls.append(dict(fields))

        sheets = [
            {"model": "Transaction", "rows": [{"x": i} for i in range(250)]},
        ]
        try:
            execute_import_job(
                company_id=1,
                sheets=sheets,
                commit=False,
                progress_callback=cb,
            )
        except Exception:
            pass

        # Find rows_processed values from calls — should include at
        # least 0, 100, 200.
        processed_values = sorted({
            c.get("rows_processed")
            for c in calls
            if c.get("rows_processed") is not None
        })
        self.assertIn(0, processed_values)
        self.assertIn(100, processed_values)
        self.assertIn(200, processed_values)

    def test_callback_errors_dont_block_import(self):
        """A misbehaving callback must not crash the worker — progress
        is best-effort, the real work always wins."""
        from multitenancy.tasks import execute_import_job

        def cb(fields):
            raise RuntimeError("callback exploded")

        sheets = [
            {"model": "Transaction", "rows": [{"x": 1}]},
        ]
        # Should NOT raise RuntimeError from the callback — it should
        # be swallowed. The ``try/except`` below is just to ignore
        # unrelated errors from the import itself.
        try:
            execute_import_job(
                company_id=1,
                sheets=sheets,
                commit=False,
                progress_callback=cb,
            )
        except RuntimeError as exc:
            self.assertNotIn("callback exploded", str(exc))


@override_settings(AUTH_OFF=True)
class V2ProgressSerializerMergeTests(TestCase):
    """Phase 6.z-g — serializer merges Redis live progress with DB
    snapshot so the frontend sees the freshest data for non-terminal
    sessions and the durable record for terminal ones.
    """

    @classmethod
    def setUpTestData(cls):
        cls.company = Company.objects.create(name="Acme Co")

    def _make_session(self, status, db_progress=None):
        return ImportSession.objects.create(
            company_id=self.company.id,
            mode=ImportSession.MODE_TEMPLATE,
            status=status,
            file_name="x.xlsx",
            progress=db_progress or {},
        )

    def test_non_terminal_prefers_redis_live(self):
        """Live row-level data from Redis overrides the DB stage
        snapshot when the session is still running."""
        from multitenancy.imports_v2.serializers import ImportSessionSerializer

        session = self._make_session(
            ImportSession.STATUS_COMMITTING,
            db_progress={"stage": "writing", "sheets_total": 3},
        )
        live = {
            "stage": "writing",
            "current_sheet": "Transaction",
            "rows_processed": 1500,
            "rows_total": 5000,
        }
        with mock.patch(
            "multitenancy.imports_v2.progress_channel.read",
            return_value=live,
        ):
            data = ImportSessionSerializer(session).data
        # DB field (sheets_total=3) preserved — not in live.
        self.assertEqual(data["progress"]["sheets_total"], 3)
        # Live fields override — row-level detail surfaces.
        self.assertEqual(data["progress"]["rows_processed"], 1500)
        self.assertEqual(data["progress"]["rows_total"], 5000)
        self.assertEqual(data["progress"]["current_sheet"], "Transaction")

    def test_non_terminal_falls_back_to_db_when_redis_empty(self):
        """If Redis has no key (channel disabled, key TTL'd), return
        the DB snapshot verbatim. Graceful degradation to 6.z-e
        behavior."""
        from multitenancy.imports_v2.serializers import ImportSessionSerializer

        session = self._make_session(
            ImportSession.STATUS_COMMITTING,
            db_progress={"stage": "writing", "sheets_total": 3},
        )
        with mock.patch(
            "multitenancy.imports_v2.progress_channel.read",
            return_value=None,
        ):
            data = ImportSessionSerializer(session).data
        self.assertEqual(data["progress"]["stage"], "writing")
        self.assertEqual(data["progress"]["sheets_total"], 3)

    def test_terminal_ignores_redis(self):
        """Terminal sessions read only from the DB — Redis might
        still have a stale key if the worker hasn't cleared it yet,
        but the DB value is canonical once the session is terminal."""
        from multitenancy.imports_v2.serializers import ImportSessionSerializer

        session = self._make_session(
            ImportSession.STATUS_COMMITTED,
            db_progress={"stage": "done"},
        )
        stale_redis = {
            "stage": "writing",
            "rows_processed": 999,
            "rows_total": 1000,
        }
        with mock.patch(
            "multitenancy.imports_v2.progress_channel.read",
            return_value=stale_redis,
        ) as read_mock:
            data = ImportSessionSerializer(session).data
        # Terminal → serializer must not even ask Redis.
        read_mock.assert_not_called()
        self.assertEqual(data["progress"]["stage"], "done")
        self.assertNotIn("rows_processed", data["progress"])


@override_settings(AUTH_OFF=True)
class V2ProgressChannelTests(TestCase):
    """Phase 6.z-g — progress_channel module itself. Tests the
    degradation behavior (no REDIS_URL) and the lazy-import pattern.
    """

    def test_publish_noop_when_redis_url_absent(self):
        """If the settings lookup returns no Redis URL, publish should
        silently do nothing. Dev without a broker → v2 still works,
        just with coarse (stage-only) progress from 6.z-e."""
        from multitenancy.imports_v2 import progress_channel

        with mock.patch.object(
            progress_channel, "_redis_client", return_value=None,
        ):
            # Must not raise.
            progress_channel.publish(999, {"stage": "writing"})

    def test_read_returns_none_when_redis_url_absent(self):
        from multitenancy.imports_v2 import progress_channel

        with mock.patch.object(
            progress_channel, "_redis_client", return_value=None,
        ):
            self.assertIsNone(progress_channel.read(999))

    def test_clear_noop_when_redis_url_absent(self):
        from multitenancy.imports_v2 import progress_channel

        with mock.patch.object(
            progress_channel, "_redis_client", return_value=None,
        ):
            # Must not raise.
            progress_channel.clear(999)
