# api_utils.py
import csv
import io
import pandas as pd
from django.apps import apps
from django.db import transaction
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status

from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from django.forms.models import model_to_dict
from django.utils.timezone import now

from multitenancy.models import Company, Entity, IntegrationRule, SubstitutionRule
from accounting.models import (
    Currency, Bank, BankAccount, Account, CostCenter,
    Transaction, JournalEntry, BankTransaction
)
from core.models import FinancialIndex, IndexQuote, FinancialIndexQuoteForecast
from billing.models import (
    BusinessPartnerCategory, BusinessPartner,
    ProductServiceCategory, ProductService,
    Contract, Invoice, InvoiceLine
)
from hr.models import Employee, Position, TimeTracking, KPI, Bonus, RecurringAdjustment

import json
from datetime import datetime, date, time, timezone
from decimal import Decimal

def _to_bool(v, default=False):
    if isinstance(v, bool):
        return v
    if v is None:
        return default
    return str(v).strip().lower() in {"1","true","t","yes","y","on"}

def _excel_safe(value):
    # None is fine
    if value is None:
        return None

    # Simple scalars
    if isinstance(value, (str, int, float, bool)):
        return value

    # Decimal -> float (or str if you prefer exactness)
    if isinstance(value, Decimal):
        try:
            return float(value)
        except Exception:
            return str(value)

    # Datetime: make naive (Excel/openpyxl dislikes tz-aware)
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.astimezone(timezone.utc).replace(tzinfo=None)
        return value

    # Dates/times are OK as-is
    if isinstance(value, (date, time)):
        return value

    # dict/list/other -> JSON string
    try:
        return json.dumps(value, ensure_ascii=False, indent=2)
    except Exception:
        return str(value)

def _is_missing(v) -> bool:
    # Treat pandas/numpy NA, NaT, plain None/"", and 'NaT' string as missing
    if v is None or v == "":
        return True
    try:
        import pandas as pd  # noqa
        if pd.isna(v):  # type: ignore[attr-defined]
            return True
    except Exception:
        pass
    try:
        import math
        if isinstance(v, float) and math.isnan(v):
            return True
    except Exception:
        pass
    if isinstance(v, str) and v.strip().lower() == "nat":
        return True
    return False

def _to_int_or_none_soft(v):
    if _is_missing(v):
        return None
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        return int(v) if float(v).is_integer() else None
    s = str(v).strip()
    if s.isdigit():
        return int(s)
    try:
        f = float(s)
        return int(f) if f.is_integer() else None
    except Exception:
        return None

# ----- lightweight fingerprint (used by template refs) -----
import hashlib, json, re
from decimal import Decimal, ROUND_HALF_UP
from typing import List, Dict, Any
from django.db import models

_WHITESPACE_RE = re.compile(r"\s+")

IDENTITY_FIELDS_MAP = {}
IGNORE_FIELDS = {"id","created_at","updated_at","created_by","updated_by","is_deleted","is_active"}

_WS_RE = re.compile(r"\s+")

def _norm_row_key(v):
    if v is None:
        return None
    s = str(v).replace("\u00A0", " ").strip()  # collapse NBSP, trim
    return s

def _norm_scalar(val):
    if val is None: return None
    if isinstance(val, (bool, int)): return val
    if isinstance(val, float): return float(str(val))
    if isinstance(val, Decimal): return str(val)
    s = str(val).strip()
    return _WHITESPACE_RE.sub(" ", s)

def _quantize_decimal(val, dp):
    if val in (None, ""): return None
    q = Decimal("1").scaleb(-int(dp))
    return Decimal(str(val)).quantize(q, rounding=ROUND_HALF_UP)

def _canonicalize_row(model, row: Dict[str, Any]) -> Dict[str, Any]:
    """
    Stable identity using "all non-volatile fields present in row & on model".
    Folds *_fk into base name; quantizes DecimalFields; ISO for dates; ints for FK ids.
    Treats NaN/NaT as missing (None).
    """
    field_by = {f.name: f for f in model._meta.get_fields() if hasattr(f, "attname")}
    allowed = set(field_by.keys())

    # fold *_fk
    incoming = {}
    for k, v in row.items():
        if k == "__row_id":
            continue
        base = k[:-3] if k.endswith("_fk") else k
        incoming[base] = v

    ident = {k for k in incoming.keys() if k in allowed and k not in IGNORE_FIELDS}
    out = {}
    for k in sorted(ident):
        v = incoming.get(k)
        f = field_by.get(k)

        # normalize missing-ish first
        if _is_missing(v):
            out[k] = None
            continue

        if isinstance(f, models.DecimalField):
            dp = int(getattr(f, "decimal_places", 0) or 0)
            q = Decimal('1').scaleb(-dp)
            vq = Decimal(str(v)).quantize(q, rounding=ROUND_HALF_UP)
            out[k] = str(vq)
        elif isinstance(f, models.DateField):
            out[k] = str(v)
        elif isinstance(f, models.ForeignKey):
            out[k] = _to_int_or_none_soft(v)
        else:
            out[k] = _norm_scalar(v)
    return out

def row_hash(model, row: Dict[str, Any]) -> str:
    canon = _canonicalize_row(model, row)
    blob  = json.dumps(canon, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()

def table_fingerprint(model, rows: List[Dict[str, Any]], sample_n: int = 200) -> Dict[str, Any]:
    rhashes = [row_hash(model, r) for r in rows]
    unique_sorted = sorted(set(rhashes))
    cols = sorted(_canonicalize_row(model, rows[0]).keys()) if rows else []
    header_blob = json.dumps(cols, separators=(",", ":"), sort_keys=True)
    concat = header_blob + "|" + "|".join(unique_sorted)
    thash  = hashlib.sha256(concat.encode("utf-8")).hexdigest()
    return {
        "row_count": len(rows),
        "colnames": cols,
        "row_hashes": unique_sorted,
        "row_hash_sample": unique_sorted[:sample_n],
        "table_hash": thash,
    }

def jaccard(a: set[str], b: set[str]) -> float:
    if not a and not b: return 1.0
    if not a or not b:  return 0.0
    inter = len(a & b); union = len(a | b)
    return inter / union

# -----------------------
# MPTT PATH SUPPORT (generic)
# -----------------------
PATH_COLS = ("path", "Caminho")
PATH_SEPARATOR = " > "

def _is_mptt_model(model) -> bool:
    return hasattr(model, "_mptt_meta") and any(f.name == "parent" for f in model._meta.fields) and any(f.name == "name" for f in model._meta.fields)

def _get_path_value(row_dict):
    for c in PATH_COLS:
        val = row_dict.get(c)
        if isinstance(val, str) and val.strip():
            return val.strip()
    return None

def _split_path(path_str):
    return [p.strip() for p in str(path_str).split(PATH_SEPARATOR) if p and p.strip()]

def _sort_df_by_path_depth_if_mptt(model, df):
    if not _is_mptt_model(model):
        return df
    if not any(col in df.columns for col in PATH_COLS):
        return df
    df = df.copy()
    def depth(row):
        p = _get_path_value(row.dropna().to_dict())
        return len(_split_path(p)) if p else 0
    df["_depth"] = df.apply(depth, axis=1)
    df.sort_values(by=["_depth"], inplace=True, kind="stable")
    df.drop(columns=["_depth"], inplace=True)
    return df

def _resolve_parent_from_path_chain(model, parts):
    parent = None
    for idx, node_name in enumerate(parts):
        inst = model.objects.filter(name=node_name, parent=parent).first()
        if not inst:
            missing = " > ".join(parts[: idx + 1])
            raise ValueError(f"{model.__name__}: missing ancestor '{missing}'. Provide this row before its children.")
        parent = inst
    return parent

# -----------------------
# Small REST helpers
# -----------------------
def success_response(data, message="Success"):
    return Response({"status": "success", "data": data, "message": message})

def error_response(message, status_code=400):
    return Response({"status": "error", "message": message}, status=status_code)

@transaction.atomic
def generic_bulk_create(viewset, request_data):
    serializer_class = viewset.get_serializer_class()
    serializer = serializer_class(data=request_data, many=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@transaction.atomic
def generic_bulk_update(viewset, request_data):
    serializer_class = viewset.get_serializer_class()
    model = viewset.get_queryset().model
    for item in request_data:
        instance = model.objects.get(id=item['id'])
        serializer = serializer_class(instance, data=item)
        if serializer.is_valid():
            serializer.save()
        else:
            transaction.set_rollback(True)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    return Response({'status': 'bulk update successful'}, status=status.HTTP_200_OK)

@transaction.atomic
def generic_bulk_delete(viewset, ids):
    model = viewset.get_queryset().model
    model.objects.filter(id__in=ids).delete()
    return Response({'status': 'bulk delete successful'}, status=status.HTTP_204_NO_CONTENT)

# Function for CSV response
def create_csv_response(data, delimiter=','):
    csvfile = io.StringIO()
    writer = csv.DictWriter(csvfile, fieldnames=data[0].keys(), delimiter=delimiter)
    writer.writeheader()
    writer.writerows(data)
    csvfile.seek(0)
    return Response(csvfile.getvalue(), content_type='text/csv')

# Function for Excel response
def create_excel_response(data):
    df = pd.DataFrame(data)
    excel_file = io.BytesIO()
    with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Sheet1')
    excel_file.seek(0)
    return Response(excel_file.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# -----------------------
# Import helpers used by Preview/Execute
# -----------------------
def _allowed_keys(model):
    names = set()
    for f in model._meta.fields:
        names.add(f.name)
        att = getattr(f, "attname", None)
        if att:
            names.add(att)  # e.g. entity_id
    fk_aliases = {n + "_fk" for n in names}
    return names | fk_aliases | set(PATH_COLS) | {"__row_id", "id"}

def _filter_unknown(model, row: Dict[str, Any]):
    allowed = _allowed_keys(model)
    filtered = {k: v for k, v in row.items() if k in allowed}
    unknown = sorted([k for k in row.keys() if k not in allowed and k != "__row_id"])
    return filtered, unknown

def _resolve_fk(model, field_name: str, raw_value, row_id_map: Dict[str, Any]):
    """
    NaN-safe FK resolver:
      - '__row_id' indirection (string key in row_id_map)
      - missing-ish (NaN/NaT/None/"") -> None
      - numeric-ish ('3', 3, 3.0, '3.0') -> int id
    """
    # __row_id indirection
    if isinstance(raw_value, str):
        key = _norm_row_key(raw_value)
        if key in row_id_map:
            return row_id_map[key]

    related_field = model._meta.get_field(field_name)
    fk_model = getattr(related_field, "related_model", None)
    if fk_model is None:
        raise ValueError(f"Field '{field_name}' is not a ForeignKey on {model.__name__}")

    # Treat NaN/NaT/None/"" as missing
    if _is_missing(raw_value):
        return None

    # Numeric-ish id (handles '3', 3, 3.0, '3.0')
    fk_id = _to_int_or_none_soft(raw_value)
    if fk_id is None:
        raise ValueError(f"Invalid FK reference format '{raw_value}' for field '{field_name}'")

    try:
        return fk_model.objects.get(id=fk_id)
    except fk_model.DoesNotExist:
        raise ValueError(f"{fk_model.__name__} id={fk_id} not found for field '{field_name}'")

def _quantize_decimal_fields(model, payload: dict) -> dict:
    out = dict(payload)
    for f in model._meta.get_fields():
        if isinstance(f, models.DecimalField):
            name = getattr(f, 'attname', f.name)
            if name in out and out[name] not in (None, ''):
                dp = int(getattr(f, 'decimal_places', 0) or 0)
                q = Decimal('1').scaleb(-dp)
                out[name] = Decimal(str(out[name])).quantize(q, rounding=ROUND_HALF_UP)
    return out

def _coerce_boolean_fields(model, payload: dict) -> dict:
    out = dict(payload)
    for f in model._meta.get_fields():
        if isinstance(f, models.BooleanField):
            name = getattr(f, 'attname', f.name)
            if name in out:
                out[name] = _to_bool(out[name])
    return out

MODEL_APP_MAP = {
    "Account": "accounting",
    "CostCenter": "accounting",
    "Entity": "multitenancy",
    "Company": "multitenancy",
    "IntegrationRule": "multitenancy",
    "SubstitutionRule": "multitenancy",
    "Currency": "accounting",
    "Bank": "accounting",
    "BankAccount": "accounting",
    "Transaction": "accounting",
    "JournalEntry": "accounting",
    "BankTransaction": "accounting",
    "BusinessPartnerCategory": "billing",
    "BusinessPartner": "billing",
    "ProductServiceCategory": "billing",
    "ProductService": "billing",
    "Contract": "billing",
    "Invoice": "billing",
    "InvoiceLine": "billing",
    "FinancialIndex": "core",
    "IndexQuote": "core",
    "FinancialIndexQuoteForecast": "core",
    "Employee": "hr",
    "Position": "hr",
    "TimeTracking": "hr",
    "KPI": "hr",
    "Bonus": "hr",
    "RecurringAdjustment": "hr",
}

# ---- cross-sheet processing order (so Transactions come before JournalEntries) ----
ORDER_HINT = {
    "Company": 10,
    "Currency": 10,
    "Entity": 20,
    "Bank": 20,
    "BankAccount": 30,
    "Account": 40,
    "CostCenter": 40,
    "Transaction": 50,   # parent
    "JournalEntry": 60,  # depends on Transaction
    # others default to 1000
}

def safe_model_dict(instance, exclude_fields=None):
    data = model_to_dict(instance)
    exclude_fields = exclude_fields or []
    for field in exclude_fields:
        data.pop(field, None)
    for field in instance._meta.fields:
        if field.is_relation and field.name in data:
            related_obj = getattr(instance, field.name)
            data[field.name] = related_obj.id if related_obj else None
    return data

# -----------------------
# Preview endpoint (transaction rolled back)
# -----------------------
class BulkImportPreview(APIView):
    def post(self, request, *args, **kwargs):
        try:
            file = request.FILES['file']
            print("[INFO] File received:", file.name)

            xls = pd.read_excel(file, sheet_name=None)
            print(f"[INFO] Loaded {len(xls)} sheets from Excel.")

            preview_data = {}
            errors = []
            row_id_map = {}

            company_id = request.tenant.id if hasattr(request, 'tenant') else None

            with transaction.atomic():
                savepoint = transaction.savepoint()
                print("[DEBUG] Started transaction with savepoint.")
                model_preview = []

                # ---- process in a deterministic, dependency-aware order ----
                ordered_names = sorted(
                    [n for n in xls.keys() if n != "References"],
                    key=lambda n: ORDER_HINT.get(n, 1000)
                )

                for model_name in ordered_names:
                    df = xls[model_name]
                    print(f"\n[INFO] Processing sheet: {model_name}")
                    app_label = MODEL_APP_MAP.get(model_name)

                    if not app_label:
                        msg = f"Unknown model: {model_name}"
                        print("[ERROR]", msg)
                        errors.append({"model": model_name, "row": None, "field": None, "message": msg})
                        continue

                    model = apps.get_model(app_label, model_name)

                    # Keep ancestors first if 'path' given on MPTT models
                    df = _sort_df_by_path_depth_if_mptt(model, df)

                    for i, row in df.iterrows():
                        print(f"[DEBUG] Processing row {i} of model {model_name}")
                        raw = row.dropna().to_dict()
                        row_id = raw.pop('__row_id', None)
                        action = None
                        instance = None
                        try:
                            # 1) drop unknown keys (but keep *_fk for resolution)
                            filtered_input, unknown_cols = _filter_unknown(model, raw)

                            # 2) resolve *_fk
                            fk_fields = {k: v for k, v in filtered_input.items() if k.endswith('_fk')}
                            for fk_field, fk_ref in fk_fields.items():
                                field_name = fk_field[:-3]
                                print(f"  Resolving FK for {field_name} -> {fk_ref}")
                                try:
                                    fk_instance = _resolve_fk(model, field_name, fk_ref, row_id_map)
                                except Exception as e:
                                    error_msg = f"FK reference '{fk_ref}' not found for field '{field_name}' in model {model_name}: {str(e)}"
                                    print("[ERROR]", error_msg)
                                    raise ValueError(error_msg)
                                filtered_input[field_name] = fk_instance
                                del filtered_input[fk_field]

                            # 3) MPTT: derive name/parent from path
                            if _is_mptt_model(model):
                                path_val = _get_path_value(filtered_input)
                                if path_val:
                                    parts = _split_path(path_val)
                                    if not parts:
                                        raise ValueError(f"{model_name}: empty path.")
                                    leaf_name = parts[-1]
                                    parent = None
                                    if len(parts) > 1:
                                        parent = _resolve_parent_from_path_chain(model, parts[:-1])
                                    filtered_input['name'] = filtered_input.get('name', leaf_name) or leaf_name
                                    filtered_input['parent'] = parent
                                    filtered_input.pop('parent_id', None)
                                    filtered_input.pop('parent_fk', None)
                                    for c in PATH_COLS:
                                        filtered_input.pop(c, None)

                            # 4) coerce booleans & decimals
                            payload = _coerce_boolean_fields(model, filtered_input)
                            payload = _quantize_decimal_fields(model, payload)

                            # 5) create/update
                            if 'id' in payload and payload['id']:
                                instance = model.objects.get(id=payload['id'])
                                for field, value in payload.items():
                                    setattr(instance, field, value)
                                action = 'update'
                                print(f"[UPDATE] {model_name} ID {payload['id']}")
                            else:
                                instance = model(**payload)
                                action = 'create'
                                print(f"[CREATE] New {model_name} instance")

                            # Let model validations run:
                            if hasattr(instance, "full_clean"):
                                instance.full_clean()
                            instance.save()
                            print(f"[SAVE] {model_name} row saved successfully.")

                            if row_id:
                                rid = _norm_row_key(row_id)
                                row_id_map[rid] = instance
                                # (optional) also store original to be extra defensive:
                                row_id_map[row_id] = instance
                                #row_id_map[row_id] = instance
                                print(f"[MAP] __row_id '{row_id}' bound to ID {instance.pk}")

                            msg = "ok"
                            if unknown_cols:
                                msg += f" | Ignoring unknown columns: {', '.join(unknown_cols)}"

                            model_preview.append({
                                "model": model_name,
                                "__row_id": row_id,
                                "status": "success",
                                "action": action,
                                "data": model_to_dict(instance, exclude=['created_by', 'updated_by', 'is_deleted', 'is_active']),
                                "message": msg
                            })
                        except Exception as e:
                            error = f"{model_name} row {i}: {str(e)}"
                            print("[ERROR]", error)
                            model_preview.append({
                                "model": model_name,
                                "__row_id": row_id,
                                "status": "error",
                                "action": action,
                                "data": {},
                                "message": str(e)
                            })
                            errors.append({"model": model_name, "row": i, "field": None, "message": str(e)})

                preview_data = model_preview
                transaction.savepoint_rollback(savepoint)
                print("[INFO] Rolled back transaction after preview.")

            return Response({
                "success": not errors,
                "preview": preview_data,
                "errors": errors if errors else []
            })

        except Exception as e:
            print("[FATAL ERROR]", str(e))
            errors.append({"model": locals().get("model_name"), "row": locals().get("i"), "field": None, "message": str(e)})
            return Response({"success": False, "preview": [], "errors": errors})

# -----------------------
# Execute endpoint (persists)
# -----------------------
class BulkImportExecute(APIView):
    def post(self, request, *args, **kwargs):
        try:
            file = request.FILES['file']
            print("[INFO] File received:", file.name)

            xls = pd.read_excel(file, sheet_name=None)
            print(f"[INFO] Loaded {len(xls)} sheets from Excel.")

            row_id_map = {}
            results = []
            errors = []

            with transaction.atomic():
                # ---- process in a deterministic, dependency-aware order ----
                ordered_names = sorted(
                    [n for n in xls.keys() if n != "References"],
                    key=lambda n: ORDER_HINT.get(n, 1000)
                )

                for model_name in ordered_names:
                    df = xls[model_name]
                    print(f"\n[INFO] Processing sheet: {model_name}")
                    app_label = MODEL_APP_MAP.get(model_name)

                    if not app_label:
                        errors.append({"model": model_name, "message": f"Unknown model: {model_name}"})
                        continue

                    model = apps.get_model(app_label, model_name)

                    # Optional: ensure ancestors first when path is present
                    df = _sort_df_by_path_depth_if_mptt(model, df)

                    for i, row in df.iterrows():
                        raw = row.dropna().to_dict()
                        row_id = raw.pop('__row_id', None)
                        action = None
                        try:
                            # 1) filter unknown columns (warn, don't fail)
                            filtered_input, unknown_cols = _filter_unknown(model, raw)

                            # 2) resolve *_fk
                            fk_fields = {k: v for k, v in filtered_input.items() if k.endswith('_fk')}
                            for fk_field, fk_ref in fk_fields.items():
                                field_name = fk_field[:-3]
                                print(f"  Resolving FK for {field_name} -> {fk_ref}")
                                try:
                                    fk_instance = _resolve_fk(model, field_name, fk_ref, row_id_map)
                                except Exception as e:
                                    error_msg = f"FK reference '{fk_ref}' not found for field '{field_name}' in model {model_name}: {str(e)}"
                                    print("[ERROR]", error_msg)
                                    raise ValueError(error_msg)
                                filtered_input[field_name] = fk_instance
                                del filtered_input[fk_field]

                            # 3) MPTT chain
                            if _is_mptt_model(model):
                                path_val = _get_path_value(filtered_input)
                                if path_val:
                                    parts = _split_path(path_val)
                                    if not parts:
                                        raise ValueError(f"{model_name}: empty path.")
                                    leaf_name = parts[-1]
                                    parent = None
                                    if len(parts) > 1:
                                        parent = _resolve_parent_from_path_chain(model, parts[:-1])
                                    filtered_input['name'] = filtered_input.get('name', leaf_name) or leaf_name
                                    filtered_input['parent'] = parent
                                    filtered_input.pop('parent_id', None)
                                    filtered_input.pop('parent_fk', None)
                                    for c in PATH_COLS:
                                        filtered_input.pop(c, None)

                            # 4) coerce booleans & decimals
                            payload = _coerce_boolean_fields(model, filtered_input)
                            payload = _quantize_decimal_fields(model, payload)

                            # 5) create/update & save
                            if 'id' in payload and payload['id']:
                                instance = model.objects.get(id=payload['id'])
                                for field, value in payload.items():
                                    setattr(instance, field, value)
                                action = 'update'
                            else:
                                instance = model(**payload)
                                action = 'create'

                            if hasattr(instance, "full_clean"):
                                instance.full_clean()
                            instance.save()

                            if row_id:
                                rid = _norm_row_key(row_id)
                                row_id_map[rid] = instance
                                # (optional) also store original to be extra defensive:
                                row_id_map[row_id] = instance
                                #row_id_map[row_id] = instance

                            msg = "ok"
                            status_val = "success"
                            if unknown_cols:
                                status_val = "warning"
                                msg += f" | Ignoring unknown columns: {', '.join(unknown_cols)}"

                            results.append({
                                "model": model_name,
                                "__row_id": row_id,
                                "status": status_val,
                                "action": action,
                                "data": safe_model_dict(instance, exclude_fields=['created_by', 'updated_by', 'is_deleted', 'is_active']),
                                "message": msg
                            })

                        except Exception as e:
                            results.append({
                                "model": model_name,
                                "__row_id": row_id,
                                "status": "error",
                                "message": str(e)
                            })
                            errors.append(str(e))

            return Response({
                "success": not errors,
                "results": results,
                "errors": errors
            }, status=200 if not errors else 400)

        except Exception as e:
            return Response({
                "success": False,
                "results": [],
                "errors": [str(e)]
            }, status=400)

# -----------------------
# Helpers for Reference sheet
# -----------------------
def get_dynamic_value(obj, field_name):
    """
    Dynamically fetch either an attribute or a computed value based on the @ convention.
    - If the field starts with '@', call the corresponding get_<field>() method if it exists.
    - Otherwise, return the attribute directly.
    """
    if field_name.startswith('@'):
        method_name = f"get_{field_name[1:]}"  # Remove '@' and prepend 'get_'
        method = getattr(obj, method_name, None)
        if callable(method):
            return method()
        else:
            return None
    else:
        return getattr(obj, field_name, None)

# -----------------------
# Template download
# -----------------------
class BulkImportTemplateDownloadView(APIView):
    def get(self, request, tenant_id):
        wb = Workbook()

        # -------- Main sheets with templates --------
        # NOTE: BankTransaction: removed entity; JournalEntry: added bank_designation_pending
        sheet_defs = {
            "Company": ["__row_id", "name", "subdomain"],
            "Currency": ["__row_id", "code", "name"],
            "Bank": ["__row_id", "name", "bank_code", "country"],
            "BankAccount": ["__row_id", "name", "branch_id", "account_number", "company_fk", "entity_fk", "currency_fk", "bank_fk", "balance_date", "balance"],
            "Account": ["__row_id", "name", "parent_fk", "account_code", "company_fk", "currency_fk", "bank_account_fk", "account_direction", "balance_date", "balance"],
            "CostCenter": ["__row_id", "name", "company_fk", "balance_date", "balance"],
            "Entity": ["__row_id", "name", "company_fk", "parent_fk", "inherit_accounts", "inherit_cost_centers"],
            "BusinessPartnerCategory": ["__row_id", "name", "company_fk", "parent_fk"],
            "BusinessPartner": ["__row_id", "name", "company_fk", "partner_type", "category_fk", "identifier", "address", "city", "state", "zipcode", "country", "email", "phone", "currency_fk", "payment_terms", "is_active"],
            "FinancialIndex": ["__row_id", "name", "index_type", "code", "interpolation_strategy", "description", "quote_frequency", "expected_quote_format", "is_forecastable"],
            "IndexQuote": ["__row_id", "index_fk", "date", "value"],
            "FinancialIndexQuoteForecast": ["__row_id", "index_fk", "date", "estimated_value", "source"],
            "ProductServiceCategory": ["__row_id", "name", "company_fk", "parent_fk"],
            "ProductService": ["__row_id", "name", "company_fk", "code", "category_fk", "description", "item_type", "price", "cost", "currency_fk", "track_inventory", "stock_quantity"],
            "Contract": ["__row_id", "name", "company_fk", "partner_fk", "contract_number", "start_date", "end_date", "recurrence_rule", "base_value", "base_index_date", "adjustment_index_fk", "adjustment_frequency", "adjustment_cap", "description"],
            "Invoice": ["__row_id", "name", "company_fk", "partner_fk", "invoice_type", "invoice_number", "invoice_date", "due_date", "status", "currency", "total_amount", "tax_amount", "discount_amount", "recurrence_rule", "recurrence_start_date", "recurrence_end_date", "description"],
            "InvoiceLine": ["__row_id", "name", "company_fk", "invoice_fk", "product_service_fk", "description", "quantity", "unit_price", "tax_amount"],
            "Transaction": ["__row_id", "company_fk", "date", "entity_fk", "description", "amount", "currency_fk"],
            "JournalEntry": ["__row_id", "date", "company_fk", "transaction_fk", "account_fk", "bank_designation_pending", "cost_center_fk", "debit_amount", "credit_amount"],
            "BankTransaction": ["__row_id", "company_fk", "bank_account_fk", "date", "amount", "description", "currency_fk", "reference_number"],
            "Employee": ["__row_id", "company_fk", "CPF", "name", "position_fk", "hire_date", "salary", "vacation_days", "is_active"],
            "Position": ["__row_id", "company_fk", "title", "description", "department", "hierarchy_level", "min_salary", "max_salary"],
            "TimeTracking": ["__row_id", "company_fk", "employee_fk", "month_date", "total_hours_worked", "total_overtime_hours", "overtime_hours_paid", "days_present", "days_absent", 
                             "leave_days", "effective_hours", "bank_hours_balance", "vacation_start_date", "vacation_end_date", "vacation_days_used", "absence_reason", "status"],
            "KPI": ["__row_id", "company_fk", "employee_fk", "name", "month_date", "value"],
            "Bonus": ["__row_id", "company_fk", "employee_fk", "calculation_formula", "value"],
            "RecurringAdjustment": ["__row_id", "company_fk", "name", "type", "employee_fk", "start_date", "end_date", "base_for_inss", "base_for_fgts", "base_for_irpf", "calculation_formula",
                                    "employer_cost_formula", "priority", "default_account"],
            "IntegrationRule": ["__row_id", "company_fk","name","description","trigger_event","execution_order","filter_conditions","rule","use_celery","is_active","last_run_at","times_executed"],
            "SubstitutionRule": ["__row_id", "company_fk","title","model_name","field_name","column_name","column_index","match_type","match_value","substitution_value","filter_conditions"],
        }

        ws = wb.active
        ws.title = 'Company'
        ws.append(sheet_defs['Company'])

        for sheet_name, columns in sheet_defs.items():
            if sheet_name == 'Company':
                continue
            sheet = wb.create_sheet(sheet_name)
            sheet.append(columns)

        # -------- References sheet --------
        ref_ws = wb.create_sheet("References")
        col_position = 1  # Start at column A

        references = [
            ("Company", Company.objects.all(), ["id", "name", "subdomain"]),
            ('Currency', Currency.objects.all(), ['id', 'code', 'name']),
            ('Bank', Bank.objects.all(), ['id', 'name', 'bank_code']),
            ("BankAccount", BankAccount.objects.all(), ["id", "name", "branch_id", "account_number", "company_id", "entity_id", "currency_id", "bank_id", "balance_date", "balance"]),
            ('Entity', Entity.objects.filter(company_id=tenant_id), ['id', 'name', 'parent_id', '@path']),
            ('CostCenter', CostCenter.objects.filter(company_id=tenant_id), ['id', 'name']),
            ('Account', Account.objects.filter(company_id=tenant_id), ['id', 'name', 'account_code', 'parent_id', '@path', 'account_direction', 'bank_account_id', 'balance_date', 'balance']),
            ('BusinessPartnerCategory', BusinessPartnerCategory.objects.filter(company_id=tenant_id), ['id', 'name', 'parent_id', '@path']),
            ('BusinessPartner', BusinessPartner.objects.filter(company_id=tenant_id), ['id', 'name', 'partner_type']),
            ('ProductServiceCategory', ProductServiceCategory.objects.filter(company_id=tenant_id), ['id', 'name', 'parent_id', '@path']),
            ('ProductService', ProductService.objects.filter(company_id=tenant_id), ['id', 'name', 'code']),
            ('FinancialIndex', FinancialIndex.objects.all(), ['id', 'name', 'code']),
            ('Invoice', Invoice.objects.filter(company_id=tenant_id), ['id', 'invoice_number', 'invoice_date']),
            ('Contract', Contract.objects.filter(company_id=tenant_id), ['id', 'contract_number', 'start_date']),
            ('Transaction', Transaction.objects.filter(company_id=tenant_id), ['id', 'date', 'entity_id', 'description', 'amount', 'state']),
            ('JournalEntry', JournalEntry.objects.filter(company_id=tenant_id), ['id', 'transaction_id', 'account_id', 'bank_designation_pending', 'debit_amount', 'credit_amount', 'date']),
            ('BankTransaction', BankTransaction.objects.filter(company_id=tenant_id), ['id', 'bank_account_id', 'date', 'amount', 'description', "reference_number", "transaction_id", 'status']),
            ("Employee", Employee.objects.filter(company_id=tenant_id), ["id", "company_id", "CPF", "name", "position_id", "hire_date", "salary", "vacation_days", "is_active"]),
            ("Position", Position.objects.filter(company_id=tenant_id), ["id", "company_id", "title", "description", "department", "hierarchy_level", "min_salary", "max_salary"]),
            ("TimeTracking", TimeTracking.objects.filter(company_id=tenant_id), ["id", "company_id", "employee_id", "month_date", "total_hours_worked", "total_overtime_hours", "overtime_hours_paid", "days_present", "days_absent", 
                             "leave_days", "effective_hours", "bank_hours_balance", "vacation_start_date", "vacation_end_date", "vacation_days_used", "absence_reason", "status"]),
            ("KPI", KPI.objects.filter(company_id=tenant_id), ["id", "company_id", "employee_id", "name", "month_date", "value"]),
            ("Bonus", Bonus.objects.filter(company_id=tenant_id), ["id", "company_id", "employee_id", "calculation_formula", "value"]),
            ("RecurringAdjustment", RecurringAdjustment.objects.filter(company_id=tenant_id), ["id", "company_id", "name", "type", "employee_id", "start_date", "end_date", "base_for_inss", "base_for_fgts", "base_for_irpf", "calculation_formula",
                                    "employer_cost_formula", "priority", "default_account"]),
            ("IntegrationRule", IntegrationRule.objects.filter(company_id=tenant_id), ["id", "company_id","name","description","trigger_event","execution_order","filter_conditions","rule","use_celery","is_active","last_run_at","times_executed"]),
            ("SubstitutionRule", SubstitutionRule.objects.filter(company_id=tenant_id), ["id", "company_id","title","model_name","field_name","column_name","column_index","match_type","match_value","substitution_value","filter_conditions"]),
        ]

        for title, queryset, columns in references:
            start_col = col_position

            # Write table title
            ref_ws.cell(row=1, column=start_col, value=title)

            # Write header
            for idx, col_name in enumerate(columns):
                ref_ws.cell(row=2, column=start_col + idx, value=col_name)

            # Write data rows
            for row_idx, obj in enumerate(queryset, start=3):
                for col_idx, field in enumerate(columns):
                    value = get_dynamic_value(obj, field)
                    ref_ws.cell(row=row_idx, column=start_col + col_idx, value=_excel_safe(value))

            col_position += len(columns) + 1  # Move to next table

        # -------- Output --------
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"bulk_import_template_{now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'},
        )
