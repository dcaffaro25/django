# NORD/multitenancy/api_utils.py

import csv
import io
import pandas as pd
from django.apps import apps
from django.db import transaction
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status
from typing import Any, Dict, List, Optional, Tuple
from collections import defaultdict

from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from django.forms.models import model_to_dict
from django.utils.timezone import now
from multitenancy.models import Company, Entity, IntegrationRule, SubstitutionRule
from accounting.models import Currency, Bank, BankAccount, Account, CostCenter, Transaction, JournalEntry, BankTransaction
from core.models import FinancialIndex, IndexQuote, FinancialIndexQuoteForecast
from billing.models import (
    BusinessPartnerCategory, BusinessPartner,
    ProductServiceCategory, ProductService,
    Contract, Invoice, InvoiceLine)
from hr.models import Employee, Position, TimeTracking, KPI, Bonus, RecurringAdjustment

from multitenancy.formula_engine import apply_substitutions

from django.db.models import Model, QuerySet

import re
from decimal import Decimal, InvalidOperation








import json
from datetime import datetime, date, time, timezone
from decimal import Decimal, ROUND_HALF_UP


def success_response(data, message="Success"):
    return Response({"status": "success", "data": data, "message": message})

def error_response(message, status_code=400):
    return Response({"status": "error", "message": message}, status=status_code)


@transaction.atomic
def generic_bulk_create(viewset, request_data):
    serializer_class = viewset.get_serializer_class()
    serializer = serializer_class(data=request_data, many=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@transaction.atomic
def generic_bulk_update(viewset, request_data):
    serializer_class = viewset.get_serializer_class()
    model = viewset.get_queryset().model
    for item in request_data:
        instance = model.objects.get(id=item['id'])
        serializer = serializer_class(instance, data=item)
        if serializer.is_valid():
            serializer.save()
        else:
            transaction.set_rollback(True)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    return Response({'status': 'bulk update successful'}, status=status.HTTP_200_OK)

@transaction.atomic
def generic_bulk_delete(viewset, ids):
    model = viewset.get_queryset().model
    model.objects.filter(id__in=ids).delete()
    return Response({'status': 'bulk delete successful'}, status=status.HTTP_204_NO_CONTENT)

# Function for CSV response
def create_csv_response(data, delimiter=','):
    csvfile = io.StringIO()
    writer = csv.DictWriter(csvfile, fieldnames=data[0].keys(), delimiter=delimiter)
    writer.writeheader()
    writer.writerows(data)
    csvfile.seek(0)
    return Response(csvfile.getvalue(), content_type='text/csv')

# Function for Excel response
def create_excel_response(data):
    df = pd.DataFrame(data)
    excel_file = io.BytesIO()
    with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Sheet1')
    excel_file.seek(0)
    return Response(excel_file.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


MODEL_APP_MAP = {
    "Account": "accounting",
    "CostCenter": "accounting",
    "Entity": "multitenancy",
    "Company": "multitenancy",
    "IntegrationRule": "multitenancy",
    "SubstitutionRule": "multitenancy",
    "Currency": "accounting",
    "Bank": "accounting",
    "BankAccount": "accounting",
    "Transaction": "accounting",
    "JournalEntry": "accounting",
    "BankTransaction": "accounting",
    "BusinessPartnerCategory": "billing",
    "BusinessPartner": "billing",
    "ProductServiceCategory": "billing",
    "ProductService": "billing",
    "Contract": "billing",
    "Invoice": "billing",
    "InvoiceLine": "billing",
    "FinancialIndex": "core",
    "IndexQuote": "core",
    "FinancialIndexQuoteForecast": "core",
    "DocTypeRule": "npl",
    "SpanRule": "npl"
    
    # Add other model-app mappings as needed
}

def _json_safe(obj):
    """Convert arbitrary Python/Django objects into JSON-serializable structures."""
    if isinstance(obj, Model):
        return obj.pk
    if isinstance(obj, QuerySet):
        return [_json_safe(x) for x in obj]
    if isinstance(obj, (list, tuple, set)):
        return [_json_safe(x) for x in obj]
    if isinstance(obj, dict):
        return {k: _json_safe(v) for k, v in obj.items()}
    # pandas Timestamp
    try:
        import pandas as pd  # already present in your file
        if isinstance(obj, pd.Timestamp):
            return obj.to_pydatetime().isoformat()
    except Exception:
        pass
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    if isinstance(obj, Decimal):
        return str(obj)
    return obj

def safe_model_dict(instance, exclude_fields=None):
    data = model_to_dict(instance)
    exclude_fields = exclude_fields or []

    # Remove campos indesejados
    for field in exclude_fields:
        data.pop(field, None)

    # Substitui FKs por seus respectivos IDs
    for field in instance._meta.fields:
        if field.is_relation and field.name in data:
            related_obj = getattr(instance, field.name)
            data[field.name] = related_obj.id if related_obj else None

    return data

def _normalize_row_id_token(v):
    """Remove NBSP, trim e devolve string."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return str(v)
    return str(v).replace("\u00A0", " ").strip()

_ROWID_NUMERIC_RE = re.compile(r"^\d+(?:\.0+)?$")

def _rowid_to_int_strict(v):
    """
    Retorna int se o token do __row_id representar um inteiro puro (ex: 1181 ou 1181.0).
    Caso contrário, retorna None (trataremos como CREATE).
    """
    if v is None:
        return None
    s = _normalize_row_id_token(v)
    if not s:
        return None
    # 1) match de padrão numérico (aceita .0, .00 etc)
    if _ROWID_NUMERIC_RE.fullmatch(s):
        # tenta Decimal primeiro para evitar float weirdness
        try:
            d = Decimal(s)
            return int(d)  # 1181.0 -> 1181
        except (InvalidOperation, ValueError):
            try:
                return int(float(s))
            except Exception:
                return None
    # 2) tipos numéricos já vindos como int/float
    if isinstance(v, (int, float)):
        try:
            f = float(v)
            return int(f) if f.is_integer() else None
        except Exception:
            return None
    return None

# Accepts: "1181", "1181.0", "-1181", "-1181.000"
_ROWID_DECISION_RE = re.compile(r"^(-?\d+)(?:\.0+)?$")

def classify_row_id(token: Any) -> tuple[str, Optional[int]]:
    """
    Returns ('create', None) | ('edit', id) | ('delete', id)
    - strings with any letters → create
    - positive int (or int.0) → edit(id)
    - negative int (or -int.0) → delete(abs(id))
    """
    s = _normalize_row_id_token(token)
    if not s:
        return ("create", None)
    m = _ROWID_DECISION_RE.fullmatch(s)
    if not m:
        # has letters or other chars → create
        return ("create", None)
    val = int(m.group(1))
    if val < 0:
        return ("delete", abs(val))
    return ("edit", val)

def _rowid_to_int(row_id):
    """Usa seu _to_int_or_none_soft para permitir '12', 12.0 etc."""
    return _to_int_or_none_soft(row_id)

def _row_observations(audit_by_rowid: Dict[Any, List[dict]], rid_norm: Any) -> List[str]:
    obs: List[str] = []
    for ch in audit_by_rowid.get(rid_norm, []):
        if ch.get("field") == "__row_id":
            continue
        obs.append(
            f"campo '{ch.get('field')}' alterado de '{ch.get('old')}' para '{ch.get('new')}' (regra id={ch.get('rule_id')})"
        )
    return obs

def _allowed_keys(model) -> set:
    names = set()
    for f in model._meta.fields:
        names.add(f.name)
        att = getattr(f, "attname", None)
        if att:
            names.add(att)  # e.g. entity_id
    fk_aliases = {n + "_fk" for n in names}
    # allow path helper + id + __row_id and company_fk convenience
    return names | fk_aliases | set(PATH_COLS) | {"__row_id", "id", "company_fk"}


def _filter_unknown(model, row: Dict[str, Any]) -> Tuple[Dict[str, Any], List[str]]:
    allowed = _allowed_keys(model)
    filtered = {k: v for k, v in row.items() if k in allowed}
    unknown = sorted([k for k in row.keys() if k not in allowed and k != "__row_id"])
    return filtered, unknown

class BulkImportPreview(APIView):
    def post(self, request, *args, **kwargs):
        
        model_name = None
        i = None
        errors = []           # make available to outer except
        preview_data = {}     # keep same var used below
        
        
        try:
            company_id = request.data.get("company_id") or getattr(request.user, "company_id", None)
            if not company_id:
                return Response({"error": "No company defined"}, status=400)
            
            file = request.FILES['file']
            print("[INFO] File received:", file.name)

            xls = pd.read_excel(
                    file,
                    sheet_name=None,
                    converters={
                        "__row_id": lambda x: None if pd.isna(x) else str(x).replace("\u00A0", " ").strip()
                    },
                )
            print(f"[INFO] Loaded {len(xls)} sheets from Excel.")

            PREFERRED_ORDER = [
                "Company", "Entity", "Currency", "Account", "CostCenter",
                "Transaction", "JournalEntry"
            ]
            sheet_names = [n for n in xls.keys() if n != "References"]
            order_index = {name: i for i, name in enumerate(PREFERRED_ORDER)}
            sheet_names_sorted = sorted(sheet_names, key=lambda n: order_index.get(n, 999))

            row_id_map = {}
            row_id_pk: Dict[str, int] = {}
            
            with transaction.atomic():
                savepoint = transaction.savepoint()
                print("[DEBUG] Started transaction with savepoint.")
                model_preview = []
                for model_name in sheet_names_sorted:
                    df = xls[model_name]

                    print(f"\n[INFO] Processing sheet: {model_name}")
                    app_label = MODEL_APP_MAP.get(model_name)

                    if not app_label:
                        msg = f"Unknown model: {model_name}"
                        print("[ERROR]", msg)
                        errors.append({"model": model_name, "row": None, "field": None, "message": f"Unknown model: {model_name}"})
                        continue

                    model = apps.get_model(app_label, model_name)
                    #model_preview = []
                    
                    raw_rows = df.where(pd.notnull(df), None).to_dict(orient="records")
                    
                    rows, audit = apply_substitutions(
                        raw_rows,
                        company_id=company_id,
                        model_name=model_name,
                        return_audit=True,
                    )

                    
                    audit_by_rowid: Dict[Any, List[dict]] = defaultdict(list)
                    for ch in (audit or []):
                        audit_by_rowid[ch.get("__row_id")].append(ch)
                    
                    
                    
                    for i, row_data in enumerate(rows):
                        print(f"[DEBUG] Processing row {i} of model {model_name}")
                        # row_data already has substitutions applied; normalize Nones and drop NaNs if any
                        row_data = {k: (None if (isinstance(v, float) and pd.isna(v)) else v) for k, v in (row_data or {}).items()}
                        row_id_raw = row_data.pop('__row_id', None)
                        row_id = _normalize_row_id_token(row_id_raw)  # <--- normaliza NBSP/espacos
                        print("  __row_id (raw):", repr(row_id_raw), "-> normalized:", repr(row_id))
                        print("  Raw (substituted) data:", row_data)
                        print("  __row_id:", row_id)
                        
                        # Remova 'id' do payload se vier vazio
                        if not row_data.get('id'):
                            row_data.pop('id', None)
                        
                        row_data, unknown = _filter_unknown(model, row_data)
                        msg = "ok"
                        if unknown:
                            msg += f" | Ignoring unknown columns: {', '.join(unknown)}"
                        
                        instance = None
                        action = 'create'
                        
                        try:
                            fk_mappings = {}
                            # ------- FKs (igual ao seu código) -------
                            fk_fields = {k: v for k, v in row_data.items() if k.endswith('_fk')}
                            for fk_field, fk_ref in fk_fields.items():
                                field_name = fk_field[:-3]
                                related_field = model._meta.get_field(field_name)
                                fk_model = related_field.related_model
                                print(f"  Resolving FK for {field_name} -> {fk_ref}")
                        
                                if fk_ref is None:
                                    if getattr(related_field, 'null', False):
                                        row_data[field_name] = None
                                        fk_mappings[field_name] = {"source": None, "resolved_id": None}
                                        del row_data[fk_field]
                                        continue
                                    raise ValueError(f"Invalid FK reference format: {fk_ref}")
                        
                                if isinstance(fk_ref, str) and fk_ref in row_id_pk:
                                    fk_pk = row_id_pk[fk_ref]
                                    fk_instance = fk_model.objects.get(pk=fk_pk)
                                elif isinstance(fk_ref, (int, float)) or (isinstance(fk_ref, str) and fk_ref.isdigit()):
                                    fk_instance = fk_model.objects.get(pk=int(float(fk_ref)))
                                else:
                                    raise ValueError(
                                        f"Unresolved FK '{field_name}': __row_id '{fk_ref}' not found. "
                                        f"Ensure the referenced sheet/row appears earlier (dependency order)."
                                    )
                        
                                row_data[field_name] = fk_instance
                                del row_data[fk_field]
                                fk_mappings[field_name] = {"source": fk_ref, "resolved_id": fk_instance.pk}
                        
                            # ------- DECISION BY __row_id (create/edit/delete) -------
                            decision, rid = classify_row_id(row_id)
                            
                            if decision == "edit":
                                # __row_id numeric positive → EDIT
                                try:
                                    instance = model.objects.get(pk=rid)
                                except model.DoesNotExist:
                                    raise ValueError(f"__row_id '{row_id}' indicates edit (id={rid}), but record does not exist.")
                                for field, value in row_data.items():
                                    if field == 'id':  # do not overwrite id
                                        continue
                                    setattr(instance, field, value)
                                action = 'edit'
                                print(f"[EDIT] {model_name} ID {rid} (via __row_id)")
                            
                            elif decision == "delete":
                                # __row_id numeric negative → DELETE (preview-only: verify and mark)
                                exists = model.objects.filter(pk=rid).exists()
                                if not exists:
                                    raise ValueError(f"__row_id '{row_id}' indicates delete (id={rid}), but record does not exist.")
                                action = 'delete'
                                instance = model.objects.get(pk=rid)  # only to show data in preview; no delete here
                                print(f"[DELETE:PREVIEW] {model_name} ID {rid}")
                            
                            else:
                                # CREATE
                                # (fallback: if user also filled 'id', ignore it on create)
                                row_data.pop('id', None)
                                instance = model(**row_data)
                                action = 'create'
                                print(f"[CREATE] New {model_name} instance")
                            
                            # Persist (preview runs inside a savepoint and is rolled back later)
                            if action != 'delete':
                                instance.save()
                            else:
                                # For preview of delete, don't call .delete() (we'll do it in Execute)
                                pass
                            
                            # map row_id to pk for later FKs (use string key)
                            if row_id is not None and action != 'delete':
                                row_id_pk[str(row_id)] = int(instance.pk)
                                row_id_map[str(row_id)] = instance
                                print(f"[MAP] __row_id '{row_id}' bound to ID {instance.pk}")
                        
                            print(f"[SAVE] {model_name} row saved successfully.")
                        
                            data_dict = model_to_dict(instance, exclude=['created_by','updated_by','is_deleted','is_active'])
                            data_dict['id'] = instance.pk
                        
                            model_preview.append({
                                'model': model_name,
                                '__row_id': row_id,
                                'status': 'success',
                                'action': action,  # 'edit' ou 'create'
                                'data': _json_safe(data_dict),
                                'row_data': _json_safe(row_data),
                                "message": msg,
                                "observations": _row_observations(audit_by_rowid, row_id),
                                'mappings': _json_safe(fk_mappings),
                                'row_id_map': _json_safe(row_id_map),
                                "row_id_pk": _json_safe(row_id_pk),
                            })
                        
                        except Exception as e:
                            error = f"{model_name} row {i}: {str(e)}"
                            print("[ERROR]", error)
                            # Make model_to_dict safe in error path
                            try:
                                data_payload = model_to_dict(instance, exclude=['created_by', 'updated_by', 'is_deleted', 'is_active']) if instance else row_data
                            except Exception:
                                data_payload = row_data
                            model_preview.append({
                                'model': model_name,
                                '__row_id': row_id,
                                'status': 'error',
                                'action': action,
                                'data': _json_safe(data_payload),
                                "message": str(e),
                                "observations": _row_observations(audit_by_rowid, row_id),
                                'mappings': _json_safe(fk_mappings),
                                'row_id_map': _json_safe(row_id_map),
                                "row_id_pk": _json_safe(row_id_pk),
                            })
                            errors.append({"model": model_name, "row": i, "field": None, "message": str(e)})
                preview_data = model_preview
    
                transaction.savepoint_rollback(savepoint)
                print("[INFO] Rolled back transaction after preview.")

            return Response({
                "success": not errors,
                "preview": preview_data,
                "errors": errors if errors else []
            })#, status=status.HTTP_200_OK if not errors else status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            print("[FATAL ERROR]", str(e))
            safe_model = model_name if model_name is not None else "Unknown"
            safe_row = i if i is not None else None
            # if something blew up before errors existed, keep shape consistent
            if not errors:
                errors = []
            errors.append({"model": safe_model, "row": safe_row, "field": None, "message": str(e)})
            return Response({"success": False, "preview": [], "errors": errors})#, status=status.HTTP_400_BAD_REQUEST)
        

class BulkImportExecute(APIView):
    def post(self, request, *args, **kwargs):
        try:
            file = request.FILES['file']
            print("[INFO] File received:", file.name)

            xls = pd.read_excel(
                    file,
                    sheet_name=None,
                    converters={
                        "__row_id": lambda x: None if pd.isna(x) else str(x).replace("\u00A0", " ").strip()
                    },
                )
            print(f"[INFO] Loaded {len(xls)} sheets from Excel.")

            row_id_map = {}
            results = []
            errors = []

            with transaction.atomic():
                for model_name, df in xls.items():
                    print(f"\n[INFO] Processing sheet: {model_name}")
                    app_label = MODEL_APP_MAP.get(model_name)

                    if not app_label:
                        errors.append({"model": model_name, "message": f"Unknown model: {model_name}"})
                        continue

                    model = apps.get_model(app_label, model_name)

                    for i, row in df.iterrows():
                        row_data = row.dropna().to_dict()
                        row_id_raw = row_data.pop('__row_id', None)
                        row_id = _normalize_row_id_token(row_id_raw)
                        try:
                            # Handle FK fields
                            fk_fields = {k: v for k, v in row_data.items() if k.endswith('_fk')}
                            for fk_field, fk_ref in fk_fields.items():
                                field_name = fk_field[:-3]
                                print(f"  Resolving FK for {field_name} -> {fk_ref}")
                                try:
                                    # Try resolving from __row_id map first
                                    if isinstance(fk_ref, str) and fk_ref in row_id_map:
                                        fk_instance = row_id_map[fk_ref]
                                    # If numeric (int or numeric string), try to fetch the actual object from DB
                                    elif isinstance(fk_ref, (int, float)) or (isinstance(fk_ref, str) and fk_ref.isdigit()):
                                        related_field = model._meta.get_field(field_name)
                                        fk_model = related_field.related_model
                                        fk_instance = fk_model.objects.get(id=int(fk_ref))
                                    else:
                                        raise ValueError(f"Invalid FK reference format: {fk_ref}")
                                except Exception as e:
                                    error_msg = f"FK reference '{fk_ref}' not found for field '{field_name}' in model {model_name}: {str(e)}"
                                    print("[ERROR]", error_msg)
                                    raise ValueError(error_msg)
                            
                                row_data[field_name] = fk_instance
                                del row_data[fk_field]

                            decision, rid = classify_row_id(row_id)

                            if decision == "edit":
                                try:
                                    instance = model.objects.get(pk=rid)
                                except model.DoesNotExist:
                                    raise ValueError(f"__row_id '{row_id}' indicates edit (id={rid}), but record does not exist.")
                                for field, value in row_data.items():
                                    if field == 'id':
                                        continue
                                    setattr(instance, field, value)
                                action = 'edit'
                                instance.save()
                            
                            elif decision == "delete":
                                ok, _ = model.objects.filter(pk=rid).delete()
                                if ok == 0:
                                    raise ValueError(f"__row_id '{row_id}' indicates delete (id={rid}), but record does not exist.")
                                action = 'delete'
                                instance = None  # no instance to return
                            
                            else:
                                # CREATE
                                row_data.pop('id', None)
                                instance = model(**row_data)
                                action = 'create'
                                instance.save()
                            
                            results.append({
                                "model": model_name,
                                "__row_id": row_id,
                                "status": "success",
                                "action": action,
                                "data": None if instance is None else safe_model_dict(instance, exclude_fields=['created_by', 'updated_by', 'is_deleted', 'is_active']),
                                "message": "ok"
                            })

                        except Exception as e:
                            results.append({
                                "model": model_name,
                                "__row_id": row_id,
                                "status": "error",
                                "message": str(e)
                            })
                            errors.append(str(e))

            return Response({
                "success": not errors,
                "results": results,
                "errors": errors
            }, status=200 if not errors else 400)

        except Exception as e:
            return Response({
                "success": False,
                "results": [],
                "errors": [str(e)]
            }, status=400)
        
        
def generate_bulk_import_template():
    wb = Workbook()

    # Ensure sheets are ordered logically
    sheet_defs = {
        "company": ["__row_id", "name", "subdomain"],
        "currency": ["__row_id", "code", "name"],
        "bank": ["__row_id", "name", "bank_code"],
        "bankaccount": ["__row_id", "name", "account_number", "company_fk", "entity_fk", "currency_fk", "bank_fk"],
        "account": ["__row_id", "name", "account_code", "company_fk", "currency_fk", "bank_account_fk", "account_direction", "balance_date", "balance"],
        "costcenter": ["__row_id", "name", "company_fk"],
        "entity": ["__row_id", "name", "company_fk", "parent_fk", "inherit_accounts", "inherit_cost_centers"],
        "BusinessPartnerCategory": ["__row_id", "name", "company_fk", "parent_fk"],
        "BusinessPartner": ["__row_id", "name", "company_fk", "partner_type", "category_fk", "identifier", "address", 
                            "city", "state", "zipcode", "country", "email", "phone", "currency_fk", "payment_terms", "is_active", ""],
        
        "FinancialIndex": ["__row_id", "name", "index_type", "code", "interpolation_strategy", "description", 
                           "quote_frequency", "expected_quote_format", "is_forecastable"],
        
        
        "IndexQuote": ["__row_id", "index_fk", "date", "value"],
        "FinancialIndexQuoteForecast": ["__row_id", "index_fk", "date", "estimated_value", "source"],
        
        
        "ProductServiceCategory": ["__row_id", "name", "company_fk", "parent_fk"],
        
        "ProductService": ["__row_id", "name", "company_fk", "code", "category_fk", "description", 
                           "item_type", "price", "cost", "currency_fk", "track_inventory", "stock_quantity"],
        
        "Contract": ["__row_id", "name", "company_fk", "partner_fk", "contract_number", "start_date", "end_date",
                     "recurrence_rule", "base_value", "base_index_date", "adjustment_index_fk", "adjustment_frequency", 
                     "adjustment_cap", "description"],
        
        "Invoice": ["__row_id", "name", "company_fk", "partner_fk", "invoice_type", "invoice_number", "invoice_date", 
                    "due_date", "status", "currency", "total_amount", "tax_amount", "discount_amount", "recurrence_rule", 
                    "recurrence_start_date", "recurrence_end_date", "description"],
        
        "InvoiceLine": ["__row_id", "name", "company_fk", "invoice_fk", "product_service_fk", "description", 
                        "quantity", "unit_price", "tax_amount"],
        
        
    }



    sample_data = {
        "company": [["c1", "Acme Inc.", "acme"]],
        "currency": [["cur1", "USD", "US Dollar"]],
        "bank": [["b1", "Bank of Nowhere", "999"]],
        "bankaccount": [["ba1", "Main Account", "12345-6", "c1", "e1", "cur1", "b1"]],
        "account": [["a1", "Cash", "1.01.01", "c1", "cur1", "ba1", 1, "2023-12-31", 10000.00]],
        "costcenter": [["cc1", "Operations", "c1"]],
        "entity": [["e1", "Headquarters", "c1", "", True, True]],
        "BusinessPartnerCategory": [["bpc1", "Suppliers", "c1", ""]],
    "BusinessPartner": [["bp1", "Globex Corp.", "c1", "supplier", "bpc1", "123456789", "123 Market St", 
                          "San Francisco", "CA", "94105", "USA", "contact@globex.com", "+14155552671", 
                          "cur1", "30 days", True, ""]],
    
    "FinancialIndex": [["fi1", "CPI", "inflation", "CPI", "linear", "Consumer Price Index", 
                         "monthly", "accumulated", True]],

    "IndexQuote": [["iq1", "fi1", "2024-01-01", 110.5],
                   ["iq2", "fi1", "2024-02-01", 111.0],
                   ["iq3", "fi1", "2024-03-01", 111.7]],

    "FinancialIndexQuoteForecast": [["iqf1", "fi1", "2024-04-01", 112.3, "internal forecast"],
                                    ["iqf2", "fi1", "2024-05-01", 113.0, "internal forecast"]],

    "ProductServiceCategory": [["psc1", "Office Supplies", "c1", ""]],
    "ProductService": [["ps1", "Printer Paper", "c1", "PP001", "psc1", "A4 white printer paper", 
                         "product", 5.00, 2.00, "cur1", True, 500]],

    "Contract": [["ct1", "Office Supplies Agreement", "c1", "bp1", "CT2024-001", "2024-01-01", "2024-12-31",
                  "FREQ=MONTHLY", 1000.00, "2024-01-01", "", "monthly", "", "Contract for monthly supply of office items."]],

    "Invoice": [["inv1", "January Invoice", "c1", "bp1", "purchase", "INV2024-001", "2024-01-01", 
                 "2024-01-30", "issued", "cur1", 500.00, 50.00, 0.00, "", "", "", "Monthly supply invoice."]],

    "InvoiceLine": [["il1", "Paper Order", "c1", "inv1", "ps1", "A4 white printer paper", 
                     100, 5.00, 50.00]],


    }

    for sheet_name, columns in sheet_defs.items():
        if sheet_name == "company":
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)

        ws.append(columns)
        for row in sample_data.get(sheet_name, []):
            ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


class BulkImportTemplateDownloadView2(APIView):
    #permission_classes = [IsAuthenticated]  # (opcional) restrinja para usuários autenticados

    def get(self, request):
        wb = Workbook()

        sheet_defs = {
            "company": ["__row_id", "name", "subdomain"],
            "currency": ["__row_id", "code", "name"],
            "bank": ["__row_id", "name", "bank_code"],
            "bankaccount": ["__row_id", "name", "account_number", "company_fk", "entity_fk", "currency_fk", "bank_fk"],
            "account": ["__row_id", "name", "account_code", "company_fk", "currency_fk", "bank_account_fk", "account_direction", "balance_date", "balance"],
            "costcenter": ["__row_id", "name", "company_fk"],
            "entity": ["__row_id", "name", "company_fk", "parent_fk", "inherit_accounts", "inherit_cost_centers"],
            "businesspartnercategory": ["__row_id", "name", "company_fk", "parent_fk"],
            "businesspartner": ["__row_id", "name", "company_fk", "partner_type", "category_fk", "identifier", "address", "city", "state", "zipcode", "country", "email", "phone", "currency_fk", "payment_terms", "is_active", ""],
            "financialindex": ["__row_id", "name", "index_type", "code", "interpolation_strategy", "description", "quote_frequency", "expected_quote_format", "is_forecastable"],
            "indexquote": ["__row_id", "index_fk", "date", "value"],
            "financialindexquoteforecast": ["__row_id", "index_fk", "date", "estimated_value", "source"],
            "productservicecategory": ["__row_id", "name", "company_fk", "parent_fk"],
            "productservice": ["__row_id", "name", "company_fk", "code", "category_fk", "description", "item_type", "price", "cost", "currency_fk", "track_inventory", "stock_quantity"],
            "contract": ["__row_id", "name", "company_fk", "partner_fk", "contract_number", "start_date", "end_date", "recurrence_rule", "base_value", "base_index_date", "adjustment_index_fk", "adjustment_frequency", "adjustment_cap", "description"],
            "invoice": ["__row_id", "name", "company_fk", "partner_fk", "invoice_type", "invoice_number", "invoice_date", "due_date", "status", "currency", "total_amount", "tax_amount", "discount_amount", "recurrence_rule", "recurrence_start_date", "recurrence_end_date", "description"],
            "invoiceline": ["__row_id", "name", "company_fk", "invoice_fk", "product_service_fk", "description", "quantity", "unit_price", "tax_amount"],
        }

        sample_data = {
            "company": [["c1", "Acme Inc.", "acme"]],
            "currency": [["cur1", "USD", "US Dollar"]],
            "bank": [["b1", "Bank of Nowhere", "999"]],
            "bankaccount": [["ba1", "Main Account", "12345-6", "c1", "e1", "cur1", "b1"]],
            "account": [["a1", "Cash", "1.01.01", "c1", "cur1", "ba1", 1, "2023-12-31", 10000.00]],
            "costcenter": [["cc1", "Operations", "c1"]],
            "entity": [["e1", "Headquarters", "c1", "", True, True]],
            "BusinessPartnerCategory": [["bpc1", "Suppliers", "c1", ""]],
            "BusinessPartner": [["bp1", "Globex Corp.", "c1", "supplier", "bpc1", "123456789", "123 Market St", "San Francisco", "CA", "94105", "USA", "contact@globex.com", "+14155552671", "cur1", "30 days", True, ""]],
            "FinancialIndex": [["fi1", "CPI", "inflation", "CPI", "linear", "Consumer Price Index", "monthly", "accumulated", True]],
            "IndexQuote": [["iq1", "fi1", "2024-01-01", 110.5], ["iq2", "fi1", "2024-02-01", 111.0], ["iq3", "fi1", "2024-03-01", 111.7]],
            "FinancialIndexQuoteForecast": [["iqf1", "fi1", "2024-04-01", 112.3, "internal forecast"], ["iqf2", "fi1", "2024-05-01", 113.0, "internal forecast"]],
            "ProductServiceCategory": [["psc1", "Office Supplies", "c1", ""]],
            "ProductService": [["ps1", "Printer Paper", "c1", "PP001", "psc1", "A4 white printer paper", "product", 5.00, 2.00, "cur1", True, 500]],
            "Contract": [["ct1", "Office Supplies Agreement", "c1", "bp1", "CT2024-001", "2024-01-01", "2024-12-31", "FREQ=MONTHLY", 1000.00, "2024-01-01", "", "monthly", "", "Contract for monthly supply of office items."]],
            "Invoice": [["inv1", "January Invoice", "c1", "bp1", "purchase", "INV2024-001", "2024-01-01", "2024-01-30", "issued", "cur1", 500.00, 50.00, 0.00, "", "", "", "Monthly supply invoice."]],
            "InvoiceLine": [["il1", "Paper Order", "c1", "inv1", "ps1", "A4 white printer paper", 100, 5.00, 50.00]],
        }

        for sheet_name, columns in sheet_defs.items():
            if sheet_name == "company":
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(sheet_name)

            ws.append(columns)
            for row in sample_data.get(sheet_name, []):
                ws.append(row)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"bulk_import_template_{now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'},
        )
    

def get_dynamic_value(obj, field_name):
    """
    Dynamically fetch either an attribute or a computed value based on the @ convention.
    - If the field starts with '@', call the corresponding get_<field>() method if it exists.
    - Otherwise, return the attribute directly.
    """
    if field_name.startswith('@'):
        method_name = f"get_{field_name[1:]}"  # Remove '@' and prepend 'get_'
        method = getattr(obj, method_name, None)
        if callable(method):
            return method()
        else:
            return None  # or raise Exception(f"Method {method_name} not found")
    else:
        return getattr(obj, field_name, None)

class BulkImportTemplateDownloadView(APIView):
    def get(self, request, tenant_id):
        wb = Workbook()

        # -------- Main sheets with templates --------
        sheet_defs = {
            "Company": ["__row_id", "name", "subdomain"],
            "Currency": ["__row_id", "code", "name"],
            "Bank": ["__row_id", "name", "bank_code", "country"],
            "BankAccount": ["__row_id", "name", "branch_id", "account_number", "company_fk", "entity_fk", "currency_fk", "bank_fk", "balance_date", "balance"],
            "Account": ["__row_id", "name", "parent_fk", "account_code", "company_fk", "currency_fk", "bank_account_fk", "account_direction", "balance_date", "balance"],
            "CostCenter": ["__row_id", "name", "company_fk"],
            "Entity": ["__row_id", "name", "company_fk", "parent_fk", "inherit_accounts", "inherit_cost_centers"],
            "BusinessPartnerCategory": ["__row_id", "name", "company_fk", "parent_fk"],
            "BusinessPartner": ["__row_id", "name", "company_fk", "partner_type", "category_fk", "identifier", "address", "city", "state", "zipcode", "country", "email", "phone", "currency_fk", "payment_terms", "is_active"],
            "FinancialIndex": ["__row_id", "name", "index_type", "code", "interpolation_strategy", "description", "quote_frequency", "expected_quote_format", "is_forecastable"],
            "IndexQuote": ["__row_id", "index_fk", "date", "value"],
            "FinancialIndexQuoteForecast": ["__row_id", "index_fk", "date", "estimated_value", "source"],
            "ProductServiceCategory": ["__row_id", "name", "company_fk", "parent_fk"],
            "ProductService": ["__row_id", "name", "company_fk", "code", "category_fk", "description", "item_type", "price", "cost", "currency_fk", "track_inventory", "stock_quantity"],
            "Contract": ["__row_id", "name", "company_fk", "partner_fk", "contract_number", "start_date", "end_date", "recurrence_rule", "base_value", "base_index_date", "adjustment_index_fk", "adjustment_frequency", "adjustment_cap", "description"],
            "Invoice": ["__row_id", "name", "company_fk", "partner_fk", "invoice_type", "invoice_number", "invoice_date", "due_date", "status", "currency", "total_amount", "tax_amount", "discount_amount", "recurrence_rule", "recurrence_start_date", "recurrence_end_date", "description"],
            "InvoiceLine": ["__row_id", "name", "company_fk", "invoice_fk", "product_service_fk", "description", "quantity", "unit_price", "tax_amount"],
            "Transaction": ["__row_id", "company_fk", "date", "entity_fk", "description", "amount", "currency_fk"],
            "JournalEntry": ["__row_id", "date", "company_fk", "transaction_fk", "account_fk", "cost_center_fk", "debit_amount", "credit_amount"],
            "BankTransaction": ["__row_id", "company_fk", "entity_fk", "bank_account_fk", "date", "amount", "description", "currency_fk", "transaction_type", "check_number", "reference_number", "payee", "memo", "account_number", "routing_number", "transaction_id"],
            "IntegrationRule": ["__row_id", "company_fk","name","description","trigger_event","execution_order","filter_conditions","rule","use_celery","is_active","last_run_at","times_executed"],
            "SubstitutionRule": ["__row_id", "company_fk","title","model_name","field_name","column_name","column_index","match_type","match_value","substitution_value","filter_conditions"],
        }

        ws = wb.active
        ws.title = 'Company'
        ws.append(sheet_defs['Company'])

        for sheet_name, columns in sheet_defs.items():
            if sheet_name == 'Company':
                continue
            sheet = wb.create_sheet(sheet_name)
            sheet.append(columns)

        # -------- References sheet --------
        ref_ws = wb.create_sheet("References")
        col_position = 1  # Start at column A

        references = [
            ("Company", Company.objects.all(), ["id", "name", "subdomain"]),
            ('Currency', Currency.objects.all(), ['id', 'code', 'name']),
            ('Bank', Bank.objects.all(), ['id', 'name', 'bank_code']),
            ("BankAccount", BankAccount.objects.all(), ["id", "name", "branch_id", "account_number", "company_id", "entity_id", "currency_id", "bank_id", "balance_date", "balance"],),
            ('Entity', Entity.objects.filter(company_id=tenant_id), ['id', 'name', 'parent_id', '@path']),
            ('CostCenter', CostCenter.objects.filter(company_id=tenant_id), ['id', 'name']),
            ('Account', Account.objects.filter(company_id=tenant_id), ['id', 'name', 'account_code', 'parent_id', '@path', 'account_direction', 'bank_account_id', 'balance_date', 'balance']),
            ('BusinessPartnerCategory', BusinessPartnerCategory.objects.filter(company_id=tenant_id), ['id', 'name', 'parent_id', '@path']),
            ('BusinessPartner', BusinessPartner.objects.filter(company_id=tenant_id), ['id', 'name', 'partner_type']),
            ('ProductServiceCategory', ProductServiceCategory.objects.filter(company_id=tenant_id), ['id', 'name', 'parent_id', '@path']),
            ('ProductService', ProductService.objects.filter(company_id=tenant_id), ['id', 'name', 'code']),
            ('FinancialIndex', FinancialIndex.objects.all(), ['id', 'name', 'code']),
            ('Invoice', Invoice.objects.filter(company_id=tenant_id), ['id', 'invoice_number', 'invoice_date']),
            ('Contract', Contract.objects.filter(company_id=tenant_id), ['id', 'contract_number', 'start_date']),
            ('Transaction', Transaction.objects.filter(company_id=tenant_id), ['id', 'date', 'entity_id', 'description', 'amount', 'state']),
            ('JournalEntry', JournalEntry.objects.filter(company_id=tenant_id), ['id', 'transaction_id', 'account_id', 'debit_amount', 'credit_amount', 'date']),
            ('BankTransaction', BankTransaction.objects.filter(company_id=tenant_id), ['id', 'entity_id', 'bank_account_id', 'date', 'amount', 'description', 'transaction_type', 'status']),
            ("IntegrationRule", IntegrationRule.objects.filter(company_id=tenant_id), ["id", "company_id","name","description","trigger_event","execution_order","filter_conditions","rule","use_celery","is_active","last_run_at","times_executed"]),
            ("SubstitutionRule", SubstitutionRule.objects.filter(company_id=tenant_id), ["id", "company_id","title","model_name","field_name","column_name","column_index","match_type","match_value","substitution_value","filter_conditions"]),
        ]

        for title, queryset, columns in references:
            start_col = col_position

            # Write table title
            ref_ws.cell(row=1, column=start_col, value=title)

            # Write header
            for idx, col_name in enumerate(columns):
                ref_ws.cell(row=2, column=start_col + idx, value=col_name)

            # Write data rows
            for row_idx, obj in enumerate(queryset, start=3):
                for col_idx, field in enumerate(columns):
                    value = get_dynamic_value(obj, field)
                    ref_ws.cell(row=row_idx, column=start_col + col_idx, value=_excel_safe(value))

            col_position += len(columns) + 1  # Move to next table

        # -------- Output --------
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"bulk_import_template_{now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'},
        )
    
    
    
    
# ----------------------------------------------------------------------
# Small helpers
# ----------------------------------------------------------------------
def _to_int_or_none(x):
    if x in ("", None):
        return None
    try:
        return int(float(x))
    except Exception:
        return None

def _parse_json_or_empty(v):
    if v in ("", None):
        return {}
    if isinstance(v, dict):
        return v
    try:
        return json.loads(v)
    except Exception:
        return {}

def _to_bool(val, default=False):
    if isinstance(val, bool):
        return val
    if val is None:
        return default
    return str(val).strip().lower() in {"1", "true", "t", "yes", "y", "on"}

# Normalize string-like row ids/FK tokens (kills NBSP, trims)

def _norm_row_key2(v):
    if v is None:
        return None
    s = str(v).replace("\u00A0", " ").strip()

    return s

def _norm_row_key(key: Any) -> Any:
    """
    Normalize a row key from an import sheet.  Strings are stripped of surrounding
    whitespace and converted to lower case.  Non-strings are returned unchanged.
    """
    if isinstance(key, str):
        return key.strip().lower()
    return key

def _path_depth(row: Dict[str, Any]) -> int:
    """Used to sort MPTT rows so parents come first."""
    for c in PATH_COLS:
        if c in row and row[c]:
            return len(str(row[c]).strip().replace(" > ", "\\").split("\\"))
    return 0


def _excel_safe(value):
    # None is fine
    if value is None:
        return None

    # Simple scalars
    if isinstance(value, (str, int, float, bool)):
        return value

    # Decimal -> float (or str if you prefer exactness)
    if isinstance(value, Decimal):
        try:
            return float(value)
        except Exception:
            return str(value)

    # Datetime: make naive
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.astimezone(timezone.utc).replace(tzinfo=None)
        return value

    # Dates/times are OK as-is
    if isinstance(value, (date, time)):
        return value

    # dict/list/other -> JSON string
    try:
        return json.dumps(value, ensure_ascii=False, indent=2)
    except Exception:
        return str(value)

def _is_missing(v) -> bool:
    if v is None or v == "":
        return True
    try:
        import pandas as pd  # noqa
        if pd.isna(v):  # type: ignore[attr-defined]
            return True
    except Exception:
        pass
    try:
        import math
        if isinstance(v, float) and math.isnan(v):
            return True
    except Exception:
        pass
    if isinstance(v, str) and v.strip().lower() == "nat":
        return True
    return False

def _to_int_or_none_soft(v):
    if _is_missing(v):
        return None
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        return int(v) if float(v).is_integer() else None
    s = str(v).strip()
    if s.isdigit():
        return int(s)
    try:
        f = float(s)
        return int(f) if f.is_integer() else None
    except Exception:
        return None

# ----- lightweight fingerprint (used by template refs) -----
import hashlib, json as _json, re as _re
from decimal import Decimal as _D, ROUND_HALF_UP as _RHU
from typing import List as _List, Dict as _Dict, Any as _Any
from django.db import models

_WHITESPACE_RE = _re.compile(r"\s+")

IDENTITY_FIELDS_MAP = {}
IGNORE_FIELDS = {"id","created_at","updated_at","created_by","updated_by","is_deleted","is_active"}

def _norm_scalar(val):
    if val is None: return None
    if isinstance(val, (bool, int)): return val
    if isinstance(val, float): return float(str(val))
    if isinstance(val, Decimal): return str(val)
    s = str(val).strip()
    return _WHITESPACE_RE.sub(" ", s)

def _quantize_decimal(val, dp):
    if val in (None, ""): return None
    q = _D("1").scaleb(-int(dp))
    return _D(str(val)).quantize(q, rounding=_RHU)

def _canonicalize_row(model, row: _Dict[str, _Any]) -> _Dict[str, _Any]:
    """
    Stable identity using "all non-volatile fields present in row & on model".
    Folds *_fk into base name; quantizes DecimalFields; ISO for dates; ints for FK ids.
    Treats NaN/NaT as missing (None).
    """
    field_by = {f.name: f for f in model._meta.get_fields() if hasattr(f, "attname")}
    allowed = set(field_by.keys())

    incoming = {}
    for k, v in row.items():
        if k == "__row_id":
            continue
        base = k[:-3] if k.endswith("_fk") else k
        incoming[base] = v

    ident = {k for k in incoming.keys() if k in allowed and k not in IGNORE_FIELDS}
    out = {}
    for k in sorted(ident):
        v = incoming.get(k)
        f = field_by.get(k)

        if _is_missing(v):
            out[k] = None
            continue

        if isinstance(f, models.DecimalField):
            dp = int(getattr(f, "decimal_places", 0) or 0)
            q  = _D("1").scaleb(-dp)
            vq = _D(str(v)).quantize(q, rounding=_RHU)
            out[k] = str(vq)
        elif isinstance(f, models.DateField):
            out[k] = str(v)
        elif isinstance(f, models.ForeignKey):
            out[k] = _to_int_or_none_soft(v)
        else:
            out[k] = _norm_scalar(v)
    return out

def row_hash(model, row: _Dict[str, _Any]) -> str:
    canon = _canonicalize_row(model, row)
    blob  = _json.dumps(canon, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()

def table_fingerprint(model, rows: _List[_Dict[str, _Any]], sample_n: int = 200) -> _Dict[str, _Any]:
    rhashes = [row_hash(model, r) for r in rows]
    unique_sorted = sorted(set(rhashes))
    cols = sorted(_canonicalize_row(model, rows[0]).keys()) if rows else []
    header_blob = _json.dumps(cols, separators=(",", ":"), sort_keys=True)
    concat = header_blob + "|" + "|".join(unique_sorted)
    thash  = hashlib.sha256(concat.encode("utf-8")).hexdigest()
    return {
        "row_count": len(rows),
        "colnames": cols,
        "row_hashes": unique_sorted,
        "row_hash_sample": unique_sorted[:sample_n],
        "table_hash": thash,
    }

def jaccard(a: set[str], b: set[str]) -> float:
    if not a and not b: return 1.0
    if not a or not b:  return 0.0
    inter = len(a & b); union = len(a | b)
    return inter / union

# -----------------------
# MPTT PATH SUPPORT (generic)
# -----------------------
PATH_COLS = ("path", "Caminho")
PATH_SEPARATOR = " > "

def _is_mptt_model(model) -> bool:
    return hasattr(model, "_mptt_meta") and any(f.name == "parent" for f in model._meta.fields) and any(f.name == "name" for f in model._meta.fields)

def _get_path_value(row_dict):
    for c in PATH_COLS:
        val = row_dict.get(c)
        if isinstance(val, str) and val.strip():
            return val.strip()
    return None

def _split_path(path_str):
    return [p.strip() for p in str(path_str).split(PATH_SEPARATOR) if p and p.strip()]

def _sort_df_by_path_depth_if_mptt(model, df):
    if not _is_mptt_model(model):
        return df
    if not any(col in df.columns for col in PATH_COLS):
        return df
    df = df.copy()
    def depth(row):
        p = _get_path_value(row.dropna().to_dict())
        return len(_split_path(p)) if p else 0
    df["_depth"] = df.apply(depth, axis=1)
    df.sort_values(by=["_depth"], inplace=True, kind="stable")
    df.drop(columns=["_depth"], inplace=True)
    return df

def _resolve_parent_from_path_chain(model, parts):
    parent = None
    for idx, node_name in enumerate(parts):
        inst = model.objects.filter(name=node_name, parent=parent).first()
        if not inst:
            missing = " > ".join(parts[: idx + 1])
            raise ValueError(f"{model.__name__}: missing ancestor '{missing}'. Provide this row before its children.")
        parent = inst
    return parent

# -----------------------
# Small REST helpers
# -----------------------
def success_response(data, message="Success"):
    return Response({"status": "success", "data": data, "message": message})

def error_response(message, status_code=400):
    return Response({"status": "error", "message": message}, status=status_code)