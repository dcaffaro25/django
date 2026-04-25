"""Business logic for the v2 analyze → commit flow (template mode).

Phase 2 scope — happy path + erp_id conflict detection. Phase 3 adds
the ETL parse front-half; Phase 4 adds the resolve step plus the
remaining issue types (unmatched references, JE balance mismatch,
bad dates, ambiguous FKs, ETL missing parameters).

Keep this module thin: it owns the session lifecycle and delegates
the heavy parse/write work to the legacy helpers in
``multitenancy.tasks``. That means a v2 analyze → commit of a clean
file writes exactly what the legacy bulk-import would have written —
the only difference is that v2 inserts an issue-detection gate
between parse and write.
"""
from __future__ import annotations

import datetime
import decimal
import hashlib
import io
from datetime import timedelta
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from django.db import transaction
from django.utils import timezone

from multitenancy.etl_service import ETLPipelineService
from multitenancy.models import ImportSession, ImportTransformationRule, SubstitutionRule
from multitenancy.tasks import (
    _group_transaction_rows_by_erp_id,
    execute_import_job,
)

from . import issues as issue_mod
from . import resolve_handlers as _resolve_handlers


# --- session TTL -----------------------------------------------------------

SESSION_TTL = timedelta(hours=24)


# --- Excel parsing ---------------------------------------------------------

# Sheets we skip even when present in the uploaded file — they're
# documentation, not data. Mirrors the legacy bulk-import behaviour at
# ``multitenancy.views.BulkImportAPIView``.
_RESERVED_SHEETS = frozenset({"References", "ImportHelp"})


def _json_scalar(v: Any) -> Any:
    """Coerce a single cell value into a JSON-safe scalar.

    pandas auto-parses date-looking columns to ``Timestamp`` and numeric
    columns to ``numpy.int64`` / ``numpy.float64``; none of those
    serialize through Django's JSONField without help. Normalise to:
      * ``None`` for NaN / None
      * ``YYYY-MM-DD`` (date-only ISO) for ``datetime.date`` AND for
        ``datetime.datetime`` / ``pd.Timestamp`` whose time component
        is exactly midnight (00:00:00.000000). Excel stores date-only
        cells as Timestamps at midnight; emitting the full
        ``YYYY-MM-DDTHH:MM:SS`` form for those tripped Django's
        ``DateField`` parser (it only accepts the date-only form).
        Real datetimes (any non-midnight time component) keep the
        full ISO so DateTimeFields receive their full precision.
      * ``str`` for Decimal (preserving precision)
      * ``int`` for numpy ints; ``float`` for numpy floats (with NaN →
        None)
      * str for everything else that isn't already a JSON primitive.
    """
    if v is None:
        return None
    # pandas NaN / NaT (both are float('nan') equivalents — the `!=` trick
    # catches both without importing math.isnan).
    if isinstance(v, float) and v != v:
        return None
    # Pure date (no time component on the type itself) → date-only ISO.
    # NB: must be checked BEFORE datetime/Timestamp because
    # ``isinstance(date_obj, datetime)`` is False but ``isinstance(dt,
    # date)`` is True — the type hierarchy goes date <- datetime, so
    # we reach this branch only for actual ``datetime.date`` instances.
    if isinstance(v, datetime.date) and not isinstance(v, datetime.datetime):
        return v.isoformat()
    if isinstance(v, (datetime.datetime, pd.Timestamp)):
        # Midnight-only Timestamps come from Excel cells formatted as
        # date (no time entered). Emit YYYY-MM-DD so DateField parsers
        # accept the value. Anything with a real time component keeps
        # the full ISO form so DateTimeField roundtrips work.
        if (
            v.hour == 0 and v.minute == 0
            and v.second == 0 and v.microsecond == 0
            and getattr(v, "tzinfo", None) is None
        ):
            # ``v.date()`` works for both pd.Timestamp and datetime.
            return v.date().isoformat()
        return v.isoformat()
    if isinstance(v, decimal.Decimal):
        return str(v)
    # numpy scalars carry an ``.item()`` method returning a Python
    # primitive. Safer than relying on implicit conversions.
    if hasattr(v, "item") and callable(v.item):
        try:
            py = v.item()
        except (ValueError, TypeError):
            return str(v)
        if isinstance(py, float) and py != py:
            return None
        return py
    if isinstance(v, (bool, int, float, str)):
        return v
    return str(v)


def _parse_template_file(file_bytes: bytes) -> Dict[str, List[Dict[str, Any]]]:
    """Parse an uploaded Excel template into ``{sheet_name: [row_dict, ...]}``.

    Reproduces the legacy parser's quirks (NaN/inf → None, skip
    reserved sheets) so a v2 session run against the same file would
    produce the same parsed payload as the legacy endpoint.

    Raises ``ValueError`` on unreadable files — caller converts to a
    session error.
    """
    try:
        dfs = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None)
    except Exception as exc:  # pragma: no cover - pandas wrap
        raise ValueError(f"could not read workbook: {exc}") from exc

    out: Dict[str, List[Dict[str, Any]]] = {}
    for sheet_name, df in dfs.items():
        if sheet_name in _RESERVED_SHEETS:
            continue
        # Replace pandas sentinels with None so downstream code sees
        # plain Python types. ``where(pd.notna(df), None)`` replaces
        # NaN/NaT; inf values are independent and need a separate pass.
        df = df.where(pd.notna(df), None)
        rows = df.to_dict(orient="records")
        cleaned: List[Dict[str, Any]] = []
        for r in rows:
            fixed = {}
            for k, v in r.items():
                # ``_json_scalar`` handles Timestamp/Decimal/numpy
                # primitives + ±inf floats. Keeps the payload
                # round-trippable through JSONField.
                if isinstance(v, float) and (
                    v == float("inf") or v == float("-inf")
                ):
                    fixed[k] = None
                else:
                    fixed[k] = _json_scalar(v)
            cleaned.append(fixed)
        out[sheet_name] = cleaned
    return out


# --- issue detection -------------------------------------------------------

# Shared fields that must agree across rows of the same erp_id group on
# the Transaction sheet. Mirrors manual §11.10c.1 (Mode 1).
_TRANSACTION_SHARED_FIELDS = (
    "date",
    "entity",
    "entity_id",
    "entity_erp_id",
    "currency",
    "currency_id",
    "currency_erp_id",
    "description",
    "amount",
)


def _detect_transaction_erp_id_conflicts(
    rows: List[Dict[str, Any]],
    sheet_name: str,
) -> List[Dict[str, Any]]:
    """Run the Phase 0 grouping helper against one Transaction sheet and
    convert every conflict into an ``Issue`` dict.
    """
    _groups, conflicts = _group_transaction_rows_by_erp_id(
        rows, shared_fields=_TRANSACTION_SHARED_FIELDS,
    )
    out: List[Dict[str, Any]] = []
    for c in conflicts:
        # Convert the ``fields`` map into human-readable summaries for the
        # ``message`` field. Frontend also gets the raw shape in ``context``.
        field_summaries = []
        for fname, values in c["fields"].items():
            rendered = " vs ".join(
                "null" if v is None else str(v) for v in values
            )
            field_summaries.append(f"{fname} ({rendered})")
        msg = (
            f"{sheet_name}: erp_id={c['erp_id']!r} shared by "
            f"{len(c['row_ids'])} rows but disagrees on: "
            + ", ".join(field_summaries)
        )
        out.append(issue_mod.make_issue(
            type=issue_mod.ISSUE_ERP_ID_CONFLICT,
            severity=issue_mod.SEVERITY_ERROR,
            location={"sheet": sheet_name, "erp_id": c["erp_id"], "row_ids": c["row_ids"]},
            context={"fields": c["fields"]},
            proposed_actions=[
                issue_mod.ACTION_PICK_ROW,
                issue_mod.ACTION_SKIP_GROUP,
                issue_mod.ACTION_ABORT,
            ],
            message=msg,
        ))
    return out


# --- Phase 4B detectors: bad_date_format / negative_amount / unmatched_reference -----

# Field names the ``bad_date_format`` detector scans on every sheet. Keeping
# this list narrow avoids false positives on columns that look like dates
# but aren't meant to parse (e.g. free-text description columns).
_DATE_LIKE_FIELDS = ("date", "due_date", "payment_date", "start_date", "end_date")


def _tryparse_date(v: Any) -> Optional[datetime.date]:
    """Best-effort date parse. Returns ``date`` on success, ``None`` on
    anything it can't interpret (empty, non-date string, garbled input).

    Covers:
      * Python ``date`` / ``datetime`` / pandas ``Timestamp`` (already
        normalised to ISO via ``_json_scalar`` during parse, but we
        still accept the native types for safety).
      * ISO 8601 date prefixes (``YYYY-MM-DD``).
      * pt-BR ``DD/MM/YYYY``.
      * US ``MM/DD/YYYY`` — only when the BR interpretation would fail
        (heuristic, deliberately conservative).

    Deliberately *not* falling back to ``dateutil.parse`` — that accepts
    so much garbage that the detector would miss real bad values.
    """
    if v is None or v == "":
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, pd.Timestamp):
        return v.date()
    if not isinstance(v, str):
        return None
    s = v.strip()
    if not s:
        return None
    # ISO 8601 first (matches our own _json_scalar output)
    try:
        return datetime.date.fromisoformat(s[:10])
    except ValueError:
        pass
    # DD/MM/YYYY
    import re as _re
    m = _re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        d, mo, y = (int(x) for x in m.groups())
        try:
            return datetime.date(y, mo, d)
        except ValueError:
            # BR parse failed; try US ordering as a fallback.
            try:
                return datetime.date(y, d, mo)
            except ValueError:
                return None
    return None


def _detect_bad_date_format(
    rows: List[Dict[str, Any]],
    sheet_name: str,
) -> List[Dict[str, Any]]:
    """Emit one ``bad_date_format`` issue per row × date-column with an
    unparseable value. Skips empty / null values — those are a
    different problem (required-field validation) handled elsewhere.
    """
    out: List[Dict[str, Any]] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        row_id = row.get("__row_id")
        for field in _DATE_LIKE_FIELDS:
            if field not in row:
                continue
            val = row[field]
            if val is None or val == "":
                continue
            if _tryparse_date(val) is None:
                out.append(issue_mod.make_issue(
                    type=issue_mod.ISSUE_BAD_DATE_FORMAT,
                    severity=issue_mod.SEVERITY_ERROR,
                    location={"sheet": sheet_name, "row_id": row_id, "field": field},
                    context={"value": val},
                    proposed_actions=[
                        issue_mod.ACTION_EDIT_VALUE,
                        issue_mod.ACTION_IGNORE_ROW,
                        issue_mod.ACTION_ABORT,
                    ],
                    message=(
                        f"{sheet_name} row {row_id!r}: value {val!r} in "
                        f"column {field!r} is not a parseable date (expected "
                        f"ISO YYYY-MM-DD or DD/MM/YYYY)"
                    ),
                ))
    return out


def _detect_negative_amounts(
    rows: List[Dict[str, Any]],
    sheet_name: str,
    *,
    rule: Optional[ImportTransformationRule] = None,
) -> List[Dict[str, Any]]:
    """Emit ``negative_amount`` issues for rows whose value in a column
    flagged ``positive_only=True`` (via ``ImportTransformationRule.column_options``)
    is negative.

    Only runs when a rule is passed (ETL mode) — template mode has no
    column_options to consult. Non-numeric values are skipped silently
    (they surface as a different issue type, not this one).
    """
    if rule is None:
        return []
    column_options = getattr(rule, "column_options", None) or {}
    if not isinstance(column_options, dict):
        return []
    positive_only_fields = {
        name for name, opts in column_options.items()
        if isinstance(opts, dict) and opts.get("positive_only") is True
    }
    if not positive_only_fields:
        return []

    out: List[Dict[str, Any]] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        row_id = row.get("__row_id")
        for field in positive_only_fields:
            if field not in row:
                continue
            val = row[field]
            if val is None or val == "":
                continue
            # Accept BR-style decimals (``1.234,56``) and plain floats.
            try:
                num = float(str(val).replace(".", "").replace(",", "."))
            except (ValueError, TypeError):
                continue
            if num < 0:
                out.append(issue_mod.make_issue(
                    type=issue_mod.ISSUE_NEGATIVE_AMOUNT,
                    severity=issue_mod.SEVERITY_ERROR,
                    location={"sheet": sheet_name, "row_id": row_id, "field": field},
                    context={"value": val, "parsed": num},
                    proposed_actions=[
                        issue_mod.ACTION_EDIT_VALUE,
                        issue_mod.ACTION_IGNORE_ROW,
                        issue_mod.ACTION_ABORT,
                    ],
                    message=(
                        f"{sheet_name} row {row_id!r}: column {field!r} is "
                        f"flagged positive-only but got {val!r} ({num})"
                    ),
                ))
    return out


# Mapping of reference field (column name in the import payload) → info
# needed to look up DB candidates for ``unmatched_reference`` detection.
# Only the narrow subset we can resolve confidently today: (name-lookup,
# same company). Account.path lookups and cost-center paths get similar
# treatment when they cross the "only check-by-name" line.
_REFERENCE_FIELD_LOOKUPS = (
    # (sheet-hint, column, app_label, model_name, lookup_field)
    #
    # ``sheet-hint`` filters which sheets we scan this column on —
    # ``None`` means any sheet. Narrow this list if false positives
    # surface (e.g. a Description column named "entity" in some template).
    #
    # Only tenant-scoped models (those with ``company_id``) belong here —
    # global models like ``Currency`` would FieldError on the company
    # filter. Extending: add any TenantAwareBaseModel-derived FK target.
    (None, "entity", "multitenancy", "Entity", "name"),
)


def _detect_unmatched_references(
    rows: List[Dict[str, Any]],
    sheet_name: str,
    *,
    company_id: int,
) -> List[Dict[str, Any]]:
    """For each reference column on the sheet, check that every distinct
    string value resolves to a DB row in the current company. Unresolved
    values → ``unmatched_reference`` issue. Ambiguous (>1 match) values
    → ``fk_ambiguous``. Matching values are ignored.

    Pragmatic scope: two well-known string references (``entity`` →
    ``Entity.name``, ``currency`` → ``Currency.code``). Extending the
    lookup table is cheap — add a row to ``_REFERENCE_FIELD_LOOKUPS``
    and the detector picks it up.
    """
    from django.apps import apps as _apps

    issues: List[Dict[str, Any]] = []
    # Collect distinct (field, value) pairs that actually appear.
    seen: Dict[str, set] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        for _hint, field, _app, _model, _lookup in _REFERENCE_FIELD_LOOKUPS:
            if _hint is not None and _hint != sheet_name:
                continue
            if field not in row:
                continue
            val = row[field]
            if val is None or val == "":
                continue
            if not isinstance(val, str):
                continue
            seen.setdefault(field, set()).add(val)

    if not seen:
        return issues

    for hint, field, app, model_name, lookup_field in _REFERENCE_FIELD_LOOKUPS:
        if hint is not None and hint != sheet_name:
            continue
        values = seen.get(field)
        if not values:
            continue
        try:
            model_cls = _apps.get_model(app, model_name)
        except LookupError:
            continue

        for val in values:
            qs = model_cls.objects.filter(
                company_id=company_id, **{lookup_field: val},
            )
            count = qs.count()
            if count == 0:
                issues.append(issue_mod.make_issue(
                    type=issue_mod.ISSUE_UNMATCHED_REFERENCE,
                    severity=issue_mod.SEVERITY_ERROR,
                    location={
                        "sheet": sheet_name,
                        "field": field,
                        "value": val,
                    },
                    context={
                        "value": val,
                        "related_model": model_name,
                        "related_app": app,
                        "lookup_field": lookup_field,
                    },
                    proposed_actions=[
                        issue_mod.ACTION_MAP_TO_EXISTING,
                        issue_mod.ACTION_IGNORE_ROW,
                        issue_mod.ACTION_ABORT,
                    ],
                    message=(
                        f"{sheet_name}: {field}={val!r} does not resolve to "
                        f"any {model_name} row in this company "
                        f"(lookup by {lookup_field})"
                    ),
                ))
            elif count > 1:
                issues.append(issue_mod.make_issue(
                    type=issue_mod.ISSUE_FK_AMBIGUOUS,
                    severity=issue_mod.SEVERITY_ERROR,
                    location={
                        "sheet": sheet_name,
                        "field": field,
                        "value": val,
                    },
                    context={
                        "value": val,
                        "related_model": model_name,
                        "related_app": app,
                        "lookup_field": lookup_field,
                        "candidate_ids": list(qs.values_list("pk", flat=True)[:10]),
                        "match_count": count,
                    },
                    proposed_actions=[
                        issue_mod.ACTION_MAP_TO_EXISTING,
                        issue_mod.ACTION_IGNORE_ROW,
                        issue_mod.ACTION_ABORT,
                    ],
                    message=(
                        f"{sheet_name}: {field}={val!r} matches {count} "
                        f"{model_name} rows by {lookup_field}; operator must "
                        f"pick one"
                    ),
                ))
    return issues


# --- public: analyze -------------------------------------------------------


# --- template dry-run at analyze -------------------------------------------

# Above this total row count the template analyze skips the dry-run step —
# running ``execute_import_job(commit=False)`` on a 10k-row file means
# rolling-back-writing 10k rows just to count them, which roughly doubles
# analyze cost. The operator still gets the per-sheet row counts in
# ``summary.sheets`` and full diagnostics; they just don't get
# ``would_create`` / ``would_update`` / ``would_fail`` counts on the
# preview panel.
#
# Threshold is deliberately conservative (5000). A follow-up can expose
# this as a per-session knob (e.g. "rodar prévia detalhada" explicit
# button) if operators want counts on bigger files.
TEMPLATE_DRY_RUN_ROW_THRESHOLD = 5000

# Display subset: how many SUCCESS rows per model the on-screen preview
# table includes. All errors are always shown regardless of count; the
# successes are sampled at evenly-spaced indices so the preview shows
# variety rather than the first 100. The complete row list lives in
# ``full_row_results`` (xlsx download only — too big for the polling
# JSON response, see ``ImportSessionSerializer.get_preview``).
TEMPLATE_DRY_RUN_DISPLAY_SUCCESS_SAMPLE = 100

# How many distinct error messages to surface in the open_issue context
# so the operator gets a quick read on what's wrong without bloating
# the issues list.
TEMPLATE_DRY_RUN_ISSUE_SAMPLE_MESSAGES = 5


def _sample_evenly(items: List[Any], k: int) -> List[Any]:
    """Pick up to ``k`` items at evenly-spaced indices (no randomness —
    deterministic for snapshot tests). ``k >= len(items)`` returns the
    whole list. The first and last items are always included so the
    sample spans the full range."""
    n = len(items)
    if k <= 0 or n == 0:
        return []
    if k >= n:
        return list(items)
    return [items[int(i * (n - 1) / (k - 1))] for i in range(k)]


def _template_dry_run_preview(
    company_id: int,
    sheets_dict: Dict[str, List[Dict[str, Any]]],
    session_pk: int,
    filename: str,
) -> Dict[str, Any]:
    """Run ``execute_import_job(commit=False)`` against the analyzed rows
    and tally per-model would_create / would_update / would_fail counts
    for the ``AnalyzePreviewPanel``.

    Exceptions from the underlying import job are swallowed and logged —
    analyze should never fail just because the preview step hit trouble.
    The real commit will re-raise the same issue with full context if
    it's a genuine bug in the input.

    Returns a dict on ``parsed_payload['preview']``:
    ``{would_create, would_update, would_fail, total_rows, row_results,
    full_row_results, display_truncated}``.

    ``row_results`` is the on-screen subset (every error + a sample of
    up to ``TEMPLATE_DRY_RUN_DISPLAY_SUCCESS_SAMPLE`` successes per
    model). ``full_row_results`` is the complete list, consumed only by
    the xlsx download endpoint. Both share the same per-row schema:
    ``{__row_id, model, status, action, message, data}``.
    """
    sheets_for_job = [
        {"model": sheet_name, "rows": rows}
        for sheet_name, rows in sheets_dict.items()
        if isinstance(rows, list)
    ]
    if not sheets_for_job:
        return {}

    try:
        # PERF (Phase 6.z-f): pre-load FK caches so the dry-run doesn't
        # fire one DB query per FK per row. Pattern mirrors
        # ``ETLPipelineService.__init__`` where the cache gets shared
        # across the run. Typical 3-5k row file with 3-5 FK fields per
        # row goes from thousands of queries to dozens.
        from multitenancy.lookup_cache import LookupCache
        lookup_cache = LookupCache(company_id)
        lookup_cache.load()
        with transaction.atomic():
            result = execute_import_job(
                company_id=company_id,
                sheets=sheets_for_job,
                commit=False,
                import_metadata={
                    "source": "v2_template_dry_run",
                    "function": "imports_v2.services.analyze_template",
                    "session_id": session_pk,
                    "filename": filename,
                },
                lookup_cache=lookup_cache,
            )
    except Exception:  # pragma: no cover - defensive; see docstring
        return {}

    would_create: Dict[str, int] = {}
    would_update: Dict[str, int] = {}
    would_fail: Dict[str, int] = {}
    total_rows = 0
    # Per-sheet buckets so the success sampling is even WITHIN each
    # model. Merging into one flat list would oversample big sheets and
    # hide small ones.
    errors_by_model: Dict[str, List[Dict[str, Any]]] = {}
    successes_by_model: Dict[str, List[Dict[str, Any]]] = {}

    for sheet_result in (result or {}).get("imports", []) or []:
        model = sheet_result.get("model")
        if not model:
            continue
        for row in sheet_result.get("result", []) or []:
            total_rows += 1
            status_val = (row.get("status") or "").lower()
            action_val = (row.get("action") or "").lower()
            # The success-row ``data`` blob comes from
            # ``_safe_model_dict(instance)`` which carries raw datetime
            # / date / Decimal objects -- fine for in-memory use but
            # NOT JSON-safe. ``parsed_payload`` is a JSONField, so
            # walk every cell through ``_json_scalar`` (the same
            # normaliser ``_parse_template_file`` already uses for
            # input cells). Without this, sessions with success rows
            # in the dry-run hit ``TypeError: Object of type datetime
            # is not JSON serializable`` at save time.
            raw_data = row.get("data") or {}
            data = {
                k: _json_scalar(v) for k, v in raw_data.items()
            } if isinstance(raw_data, dict) else raw_data
            normalised = {
                "__row_id": row.get("__row_id"),
                "model": model,
                "status": status_val or "ok",
                "action": action_val or None,
                "message": row.get("message"),
                "data": data,
            }
            if status_val == "error":
                would_fail[model] = would_fail.get(model, 0) + 1
                errors_by_model.setdefault(model, []).append(normalised)
                continue
            if action_val == "create":
                would_create[model] = would_create.get(model, 0) + 1
            elif action_val == "update":
                would_update[model] = would_update.get(model, 0) + 1
            # rows with action="delete" / "skipped_duplicate" / etc. are
            # not counted in any bucket but still kept in the success
            # sample so the operator sees them in the preview table.
            successes_by_model.setdefault(model, []).append(normalised)

    # Full list (xlsx download): every row, errors before successes
    # within each model so an operator scanning the file sees failures
    # first.
    ordered_models = list(errors_by_model.keys()) + [
        m for m in successes_by_model if m not in errors_by_model
    ]
    full_row_results: List[Dict[str, Any]] = []
    for model in ordered_models:
        full_row_results.extend(errors_by_model.get(model, []))
        full_row_results.extend(successes_by_model.get(model, []))

    # Display subset (on-screen): all errors + up to N evenly-sampled
    # successes per sheet. ``data`` is dropped from displayed successes
    # to keep the JSON small — the xlsx still has the full row.
    row_results: List[Dict[str, Any]] = []
    total_successes = 0
    displayed_successes = 0
    for model in ordered_models:
        row_results.extend(errors_by_model.get(model, []))
        success_rows = successes_by_model.get(model, [])
        total_successes += len(success_rows)
        sampled = _sample_evenly(
            success_rows, TEMPLATE_DRY_RUN_DISPLAY_SUCCESS_SAMPLE,
        )
        displayed_successes += len(sampled)
        for r in sampled:
            row_results.append({**r, "data": None})
    display_truncated = displayed_successes < total_successes

    return {
        "would_create": would_create,
        "would_update": would_update,
        "would_fail": would_fail,
        "total_rows": total_rows,
        "row_results": row_results,
        "full_row_results": full_row_results,
        "display_truncated": display_truncated,
    }


def _emit_dry_run_failure_issues(
    preview: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """Translate ``would_fail`` per-model counts into one
    ``dry_run_failure`` Issue per affected sheet. Operators can only
    abort on these (no inline fix — the error is in the source file,
    not in row-level metadata we can rewrite via resolve handlers).
    """
    would_fail = (preview or {}).get("would_fail") or {}
    if not would_fail:
        return []

    # Group error messages by model so the open_issue context can carry
    # a sample of the most common ones for that sheet.
    from collections import Counter
    errors_by_model: Dict[str, List[str]] = {}
    for row in (preview.get("full_row_results") or []):
        if (row.get("status") or "").lower() != "error":
            continue
        model = row.get("model") or "?"
        msg = row.get("message") or "(sem mensagem)"
        errors_by_model.setdefault(model, []).append(msg)

    issues: List[Dict[str, Any]] = []
    for model, fail_count in would_fail.items():
        if not fail_count:
            continue
        freq = Counter(errors_by_model.get(model, []))
        sample_messages = [m for m, _ in freq.most_common(
            TEMPLATE_DRY_RUN_ISSUE_SAMPLE_MESSAGES,
        )]
        issues.append(issue_mod.make_issue(
            type=issue_mod.ISSUE_DRY_RUN_FAILURE,
            severity=issue_mod.SEVERITY_ERROR,
            location={"sheet": model},
            context={
                "model": model,
                "fail_count": fail_count,
                "sample_messages": sample_messages,
            },
            proposed_actions=[issue_mod.ACTION_ABORT],
            message=(
                f"{fail_count} linha(s) falhariam ao importar em "
                f"'{model}'. Corrija o arquivo e reenvie."
            ),
        ))
    return issues


def _emit_etl_dry_run_failure_issues(
    etl_errors: Dict[str, List[Any]],
) -> List[Dict[str, Any]]:
    """ETL-mode equivalent of ``_emit_dry_run_failure_issues``.

    The ETL pipeline already exposes errors in four buckets
    (``python_errors``, ``database_errors``, ``substitution_errors``,
    ``warnings``); this helper collapses the first three into a single
    blocking issue so commit is gated until the operator aborts and
    reuploads a corrected file. Warnings are advisory and don't block.
    """
    python_errors = etl_errors.get("python_errors") or []
    database_errors = etl_errors.get("database_errors") or []
    substitution_errors = etl_errors.get("substitution_errors") or []
    total = len(python_errors) + len(database_errors) + len(substitution_errors)
    if total == 0:
        return []

    sample_messages: List[str] = []
    # Database errors first (most actionable — usually a constraint
    # violation tied to a specific row), then python, then substitution.
    for bucket in (database_errors, python_errors, substitution_errors):
        for err in bucket:
            if not isinstance(err, dict):
                continue
            msg = (
                err.get("message")
                or err.get("error")
                or err.get("detail")
                or str(err)
            )
            if msg and msg not in sample_messages:
                sample_messages.append(msg)
            if len(sample_messages) >= TEMPLATE_DRY_RUN_ISSUE_SAMPLE_MESSAGES:
                break
        if len(sample_messages) >= TEMPLATE_DRY_RUN_ISSUE_SAMPLE_MESSAGES:
            break

    return [issue_mod.make_issue(
        type=issue_mod.ISSUE_DRY_RUN_FAILURE,
        severity=issue_mod.SEVERITY_ERROR,
        location={"sheet": None},  # ETL errors span the whole pipeline
        context={
            "model": None,
            "fail_count": total,
            "python_errors": len(python_errors),
            "database_errors": len(database_errors),
            "substitution_errors": len(substitution_errors),
            "sample_messages": sample_messages,
        },
        proposed_actions=[issue_mod.ACTION_ABORT],
        message=(
            f"{total} erro(s) reportados pela pipeline de ETL. "
            "Veja 'Erros da ETL' abaixo, corrija o arquivo e reenvie."
        ),
    )]


def build_preview_xlsx(session: ImportSession) -> Optional[bytes]:
    """Render the full dry-run row results to a multi-sheet .xlsx.

    Template mode: one sheet per model, columns =
    ``__row_id, status, action, message`` plus one column per data
    field observed across that sheet's rows (preserving the original
    input verbatim so an operator can grep for the failing row in the
    source file).

    ETL mode: one flat ``ETL errors`` sheet flattening the four
    pipeline error buckets with a ``bucket`` discriminator column.

    Returns ``None`` when the session has no preview data — the view
    translates that to a 404.
    """
    from io import BytesIO
    from openpyxl import Workbook

    payload = session.parsed_payload or {}
    preview = payload.get("preview") or {}
    full_rows = preview.get("full_row_results") or []

    # ETL fallback: no template-style row_results → flatten etl_errors.
    if not full_rows:
        etl_errors = payload.get("etl_errors") or {}
        if not any(
            etl_errors.get(k)
            for k in ("python_errors", "database_errors",
                      "substitution_errors", "warnings")
        ):
            return None
        wb = Workbook()
        _populate_etl_errors_sheet(wb, etl_errors)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    wb = Workbook()
    wb.remove(wb.active)  # ditch default blank sheet

    # Group rows by model preserving insertion order.
    rows_by_model: Dict[str, List[Dict[str, Any]]] = {}
    for r in full_rows:
        rows_by_model.setdefault(r.get("model") or "?", []).append(r)

    for model, rows in rows_by_model.items():
        # Excel sheet names: 31-char max, can't contain a few punctuation
        # characters. Defensive sanitise.
        safe_name = str(model)[:31].replace("/", "_").replace("\\", "_")
        ws = wb.create_sheet(title=safe_name or "sheet")

        # Union of data keys in first-seen order.
        data_keys: List[str] = []
        seen = set()
        for r in rows:
            for k in (r.get("data") or {}).keys():
                if k not in seen:
                    seen.add(k)
                    data_keys.append(k)

        header = ["__row_id", "status", "action", "message"] + data_keys
        ws.append(header)
        for r in rows:
            data = r.get("data") or {}
            ws.append([
                r.get("__row_id"),
                r.get("status"),
                r.get("action"),
                r.get("message"),
            ] + [
                _excel_cell_value(data.get(k)) for k in data_keys
            ])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _excel_cell_value(v: Any) -> Any:
    """Normalise a value for openpyxl: dicts/lists → JSON repr, primitives
    pass through. Nested structures would otherwise raise."""
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    import json
    try:
        return json.dumps(v, ensure_ascii=False, default=str)
    except Exception:
        return str(v)


def _populate_etl_errors_sheet(wb, etl_errors: Dict[str, List[Any]]) -> None:
    """Helper for ``build_preview_xlsx``: flatten the four ETL error
    buckets into one sheet with a ``bucket`` discriminator column."""
    ws = wb.active
    ws.title = "ETL errors"
    all_keys: List[str] = []
    seen = set()
    flat: List[Dict[str, Any]] = []
    for bucket in ("python_errors", "database_errors",
                   "substitution_errors", "warnings"):
        for err in etl_errors.get(bucket, []) or []:
            row = {"bucket": bucket}
            if isinstance(err, dict):
                row.update(err)
            else:
                row["message"] = str(err)
            for k in row.keys():
                if k not in seen:
                    seen.add(k)
                    all_keys.append(k)
            flat.append(row)

    header = ["bucket"] + [k for k in all_keys if k != "bucket"]
    ws.append(header)
    for row in flat:
        ws.append([_excel_cell_value(row.get(k)) for k in header])


# --- Phase 6.z-e: live progress writes -------------------------------------

# Stage keys — kept short + string-y so the frontend can map them to
# Portuguese labels without negotiating an enum. Any stage that isn't
# in this list renders as the raw key (so future additions don't break
# the UI — they just display verbatim until the label map is updated).
PROGRESS_STAGE_PARSING = "parsing"
PROGRESS_STAGE_DETECTING = "detecting"
PROGRESS_STAGE_DRY_RUN = "dry_run"
PROGRESS_STAGE_MATERIALIZING_RULES = "materializing_rules"
PROGRESS_STAGE_WRITING = "writing"
PROGRESS_STAGE_DONE = "done"


def _write_progress(session: ImportSession, **fields: Any) -> None:
    """Merge ``fields`` into ``session.progress`` and persist just that column.

    Called at stage boundaries OUTSIDE any ``transaction.atomic()`` block
    so the polling frontend can observe updates before the session
    commit fires. The function tolerates being called inside an atomic
    block too — the save still happens, it just won't be externally
    visible until the block commits.

    Always stamps ``updated_at`` on the progress dict (ISO-8601 UTC)
    so the frontend can show "atualizado há 3s" style staleness cues
    without needing a separate field.

    Keeps ``update_fields=["progress", "updated_at"]`` so we don't
    accidentally race with other fields the worker is updating in a
    different scope.
    """
    current = dict(session.progress or {})
    current.update(fields)
    current["updated_at"] = timezone.now().isoformat()
    session.progress = current
    try:
        session.save(update_fields=["progress", "updated_at"])
    except Exception:
        # Progress writes must not crash the worker. If the save fails
        # (e.g. connection blew up mid-atomic), log and carry on — the
        # main work is what matters.
        import logging
        logging.getLogger(__name__).warning(
            "progress write failed for session #%s", session.pk,
            exc_info=True,
        )


# --- Phase 6.z-d: substitution cache reuse ---------------------------------


def _compute_substitution_revision(company_id: int) -> str:
    """Stable fingerprint of the active substitution rules for a company.

    Used to detect whether the cached ``sheets_post_substitution`` a
    session stashed at analyze is still valid at commit time. Any
    change — rule added, edited, deactivated — flips the hash and
    forces the commit to re-run the substitution pass.

    Keeps the hash space small: only fields that actually drive
    ``apply_substitutions`` output (match_type, match_value,
    substitution_value, filter_conditions, field_name, model_name)
    feed in. Re-ordering rules doesn't change the hash because the
    iteration is sorted by pk.

    Returns a hex digest (16 chars — SHA256 truncated is fine for
    invalidation; collisions are harmless, they just trigger a
    safe re-substitute).
    """
    from multitenancy.models import SubstitutionRule

    rows = (
        SubstitutionRule.objects.filter(company_id=company_id)
        .order_by("pk")
        .values_list(
            "pk",
            "model_name",
            "field_name",
            "match_type",
            "match_value",
            "substitution_value",
            "filter_conditions",
        )
    )
    # Serialize deterministically — ``str`` on the tuple is enough
    # because the fields are primitive (+ filter_conditions may be
    # dict/None; dicts aren't deterministic across insertion order
    # but our rules create filter_conditions once and never mutate,
    # and Django's JSONField round-trips preserve insertion order
    # in practice). Hashing the full list of tuples.
    h = hashlib.sha256()
    for row in rows:
        h.update(repr(row).encode("utf-8"))
    return h.hexdigest()[:16]


def _apply_sheets_substitutions(
    company_id: int,
    sheets: Dict[str, List[Dict[str, Any]]],
) -> Dict[str, List[Dict[str, Any]]]:
    """Run ``apply_substitutions`` on each sheet + return the result.

    Scoped per model_name so rules match their target sheet only —
    same scoping ``execute_import_job`` uses internally. Idempotent
    on its own output (rules match against ``match_value``, not the
    substituted value), so running it twice is safe; we just don't
    want to pay the cost twice.
    """
    from multitenancy.formula_engine import apply_substitutions

    out: Dict[str, List[Dict[str, Any]]] = {}
    for sheet_name, rows in sheets.items():
        if not isinstance(rows, list):
            out[sheet_name] = rows  # shouldn't happen; pass through
            continue
        try:
            substituted = apply_substitutions(
                rows,
                company_id=company_id,
                model_name=sheet_name,
                return_audit=False,
            )
            out[sheet_name] = list(substituted)
        except Exception:
            # Mirror execute_import_job's behaviour — on failure use
            # the raw rows. Commit will re-try (without skip flag) so
            # the real error surfaces in the session result if it
            # genuinely breaks the import.
            out[sheet_name] = rows
    return out


def _is_substitution_cache_valid(
    session: ImportSession, current_revision: str,
) -> bool:
    """True if the cached ``sheets_post_substitution`` is safe to reuse.

    Two invariants:
      1. The substitution-rule set hasn't changed since analyze
         (``current_revision`` matches the cached one).
      2. No resolutions have mutated individual rows (``edit_value``
         resolutions overwrite ``parsed_payload['sheets']`` — the
         pre-substituted copy wouldn't reflect that).

    When either breaks, commit falls through to the raw sheets + a
    fresh substitution pass. Correct by construction.
    """
    payload = session.parsed_payload or {}
    cached_rev = payload.get("substitution_revision")
    cached_sheets = payload.get("sheets_post_substitution")
    if not cached_rev or not isinstance(cached_sheets, dict):
        return False
    if cached_rev != current_revision:
        return False
    # Any resolution implies the operator may have edited a row
    # value. Safest to re-substitute.
    if session.resolutions:
        return False
    return True


def _create_template_session(
    *,
    company_id: int,
    user,
    file_bytes: bytes,
    file_name: str,
    config: Optional[Dict[str, Any]] = None,
) -> ImportSession:
    """Persist a fresh template-mode session in ``ANALYZING`` state.

    No parsing, no detectors — just the row creation so the caller can
    either run the analyze body inline (sync) or enqueue it via Celery
    (async, Phase 6.z). File bytes + config are stored so the worker
    has everything it needs from just the session pk.
    """
    file_hash = hashlib.sha256(file_bytes).hexdigest()
    now = timezone.now()
    return ImportSession.objects.create(
        company_id=company_id,
        created_by=user if (user and getattr(user, "is_authenticated", False)) else None,
        mode=ImportSession.MODE_TEMPLATE,
        status=ImportSession.STATUS_ANALYZING,
        file_name=file_name or "upload.xlsx",
        file_hash=file_hash,
        file_bytes=file_bytes,
        config=dict(config or {}),
        expires_at=now + SESSION_TTL,
    )


def _run_analyze_template(session: ImportSession) -> ImportSession:
    """Analyze body for a template-mode session already in ``ANALYZING``.

    Reads ``session.file_bytes`` + ``session.company_id``, runs detectors
    and the dry-run preview, then flips the session to its terminal
    analyze status (``ready``, ``awaiting_resolve``, or ``error`` on a
    parse failure). Callable from either the sync entry point or from
    a Celery worker — the contract is the same.

    Unexpected exceptions propagate so the caller (worker wrapper) can
    flip the session to ``error`` with a traceback; ``ValueError`` from
    the parser is handled inline since that's a known-bad-file signal,
    not a crash.
    """
    company_id = session.company_id
    file_bytes = bytes(session.file_bytes or b"")

    # Phase 6.z-e — stage progress. Writes happen outside the dry-run's
    # atomic block so the polling frontend sees them live.
    _write_progress(session, stage=PROGRESS_STAGE_PARSING)

    try:
        sheets = _parse_template_file(file_bytes)
    except ValueError as exc:
        session.status = ImportSession.STATUS_ERROR
        session.result = {"error": str(exc), "stage": "parse"}
        session.save(update_fields=["status", "result", "updated_at"])
        return session

    sheet_names = list(sheets.keys())
    sheets_total = len(sheet_names)
    detected_issues: List[Dict[str, Any]] = []
    _write_progress(
        session,
        stage=PROGRESS_STAGE_DETECTING,
        sheets_total=sheets_total,
        sheets_done=0,
        errors_so_far=0,
    )
    for idx, sheet_name in enumerate(sheet_names):
        rows = sheets[sheet_name]
        # Per-sheet progress — detectors don't wrap in atomic, so these
        # writes are visible to the polling frontend in real time.
        _write_progress(
            session,
            stage=PROGRESS_STAGE_DETECTING,
            current_sheet=sheet_name,
            sheets_done=idx,
            sheets_total=sheets_total,
            errors_so_far=len(detected_issues),
        )
        # Only Transaction sheets carry the grouping semantics in Phase 2.
        # Other sheets (JournalEntry, BankTransaction, etc.) are passed
        # through to the commit step unchanged.
        if sheet_name == "Transaction":
            detected_issues.extend(
                _detect_transaction_erp_id_conflicts(rows, sheet_name)
            )
        # Phase 4B detectors — run on every sheet.
        detected_issues.extend(_detect_bad_date_format(rows, sheet_name))
        detected_issues.extend(_detect_unmatched_references(
            rows, sheet_name, company_id=company_id,
        ))
    _write_progress(
        session,
        stage=PROGRESS_STAGE_DETECTING,
        sheets_done=sheets_total,
        sheets_total=sheets_total,
        errors_so_far=len(detected_issues),
    )

    # Dry-run preview — gated on row count to keep analyze cheap on big
    # imports. For files under the threshold, runs
    # ``execute_import_job(commit=False)`` and tallies per-model
    # would_create / would_update / would_fail so the
    # ``AnalyzePreviewPanel`` can render "what commit would write".
    # Above the threshold we skip — operator still sees sheet-level row
    # counts via ``summary.sheets`` plus the open_issues list, just not
    # the bottom-line create/update/fail tallies.
    total_rows = sum(
        len(rows) if isinstance(rows, list) else 0
        for rows in sheets.values()
    )
    preview: Dict[str, Any] = {}
    if 0 < total_rows <= TEMPLATE_DRY_RUN_ROW_THRESHOLD:
        _write_progress(
            session,
            stage=PROGRESS_STAGE_DRY_RUN,
            sheets_total=sheets_total,
            errors_so_far=len(detected_issues),
        )
        preview = _template_dry_run_preview(
            company_id=company_id,
            sheets_dict=sheets,
            session_pk=session.pk,
            filename=session.file_name,
        )
        # Translate any per-row failures from the dry-run into blocking
        # ``dry_run_failure`` issues so commit stops on
        # ``awaiting_resolve`` instead of marching forward to write
        # known-bad rows. Pre-fix the UI showed "Pronto para importar"
        # next to "409 falhariam" — those two states cannot coexist.
        detected_issues.extend(_emit_dry_run_failure_issues(preview))

    # Phase 6.z-d — pre-substitute once at analyze and stash the result
    # plus a hash of the active substitution-rule set. Commit will check
    # the hash + resolutions; if both still match, it reuses the cached
    # rows with ``import_options.skip_substitutions`` so the pipeline
    # doesn't pay the substitution cost a second time.
    #
    # ``apply_substitutions`` is idempotent on its own output so even if
    # this cache becomes stale in a way the invariant check misses, the
    # downstream commit re-runs it without harm — worst case we waste a
    # pass, not corrupt data.
    substitution_revision = _compute_substitution_revision(company_id)
    sheets_post_substitution = _apply_sheets_substitutions(company_id, sheets)

    # Freeze the parsed payload on the session so commit/resolve can read
    # it without re-parsing. The raw bytes stay on ``file_bytes`` too
    # (redundant but intentional — resolve may re-parse with different
    # config and we want both versions available).
    session.parsed_payload = {
        "sheets": sheets,
        "sheets_post_substitution": sheets_post_substitution,
        "substitution_revision": substitution_revision,
        "preview": preview,
    }
    session.open_issues = detected_issues
    if issue_mod.has_blocking_issues(detected_issues):
        session.status = ImportSession.STATUS_AWAITING_RESOLVE
    else:
        session.status = ImportSession.STATUS_READY
    session.save(update_fields=[
        "parsed_payload", "open_issues", "status", "updated_at",
    ])
    # Final progress mark — frontend strip flips to "concluído" then
    # fades out once status leaves "analyzing".
    _write_progress(
        session,
        stage=PROGRESS_STAGE_DONE,
        sheets_done=sheets_total,
        sheets_total=sheets_total,
        errors_so_far=len(detected_issues),
    )
    return session


def analyze_template(
    *,
    company_id: int,
    user,
    file_bytes: bytes,
    file_name: str,
    config: Optional[Dict[str, Any]] = None,
) -> ImportSession:
    """Create a new ``ImportSession`` for template mode and populate it.

    Synchronous entry point — runs parse, detectors, and dry-run preview
    in-process and returns the fully-populated session. Still used by
    tests and by the async wrapper as its "do the work" callee.

    The v2 HTTP views use :func:`analyze_template_async` instead, which
    persists the session and hands the body off to Celery so gunicorn's
    timeout ceiling can't bite on large files.
    """
    session = _create_template_session(
        company_id=company_id, user=user, file_bytes=file_bytes,
        file_name=file_name, config=config,
    )
    return _run_analyze_template(session)


def analyze_template_async(
    *,
    company_id: int,
    user,
    file_bytes: bytes,
    file_name: str,
    config: Optional[Dict[str, Any]] = None,
) -> ImportSession:
    """Create the session in ``ANALYZING`` state and enqueue the worker.

    Returns immediately with the session still in ``analyzing``. The
    frontend polls ``GET /sessions/<id>/`` until ``status`` leaves that
    state. In eager mode (no ``REDIS_URL`` — tests + dev) the worker
    runs inline so the returned session is already terminal; production
    always goes through the broker.
    """
    session = _create_template_session(
        company_id=company_id, user=user, file_bytes=file_bytes,
        file_name=file_name, config=config,
    )
    # Local import to avoid a circular import at module load: tasks.py
    # imports this module to call ``_run_analyze_template``.
    from .tasks import analyze_session_task
    analyze_session_task.delay(session.pk)
    session.refresh_from_db()
    return session


# --- public: commit --------------------------------------------------------


class CommitNotReady(Exception):
    """Raised when ``commit_session`` is called on a non-ready session."""


def _check_commit_gate(session: ImportSession) -> None:
    """Raise ``CommitNotReady`` unless the session is in ``READY`` state.

    Factored out so the sync and async entry points share exactly the
    same gate — callers that want to enqueue still need to fail fast
    on a non-ready session (409) rather than enqueueing and
    disappointing the operator via an error status.

    ETL mode additionally requires ``file_bytes`` because the commit
    re-runs ``ETLPipelineService`` against the original upload. A
    session whose bytes were cleared (committed / discarded / TTL
    swept) can't be committed — catch that here so the view returns
    a 409 instead of enqueueing a doomed task.
    """
    if not session.is_committable():
        raise CommitNotReady(
            f"session #{session.pk} is {session.status}, not ready to commit"
        )
    if session.is_terminal():
        raise CommitNotReady(
            f"session #{session.pk} is already terminal ({session.status})"
        )
    if session.mode == ImportSession.MODE_ETL and not session.file_bytes:
        raise CommitNotReady(
            f"session #{session.pk} has no file_bytes (expired?) — re-upload"
        )


def _run_commit(session: ImportSession) -> ImportSession:
    """Commit body for a session already in ``COMMITTING`` state.

    Materialises staged substitution rules, dispatches to the mode's
    write backend, and sets the final ``committed`` / ``error`` status.
    Callable from either the sync entry point or a Celery worker.

    Write errors flip the session to ``error`` with a diagnostic and
    re-raise so the caller can decide how to report back (sync view
    returns 500; worker just logs since the UI already polls the
    status).
    """
    # Phase 6.z-e — stage progress before the atomic block. Intra-write
    # row-level progress isn't observable from here (the atomic wraps
    # everything); a follow-up with a separate DB connection or Redis
    # progress store can deliver that.
    payload = session.parsed_payload or {}
    sheets_total = 0
    if isinstance(payload, dict):
        for key in ("sheets", "transformed_data"):
            candidate = payload.get(key)
            if isinstance(candidate, dict):
                sheets_total = len(candidate)
                break
    _write_progress(
        session,
        stage=PROGRESS_STAGE_MATERIALIZING_RULES,
        sheets_total=sheets_total,
    )

    try:
        # Wrap BOTH the staged-rule materialisation AND the write backend
        # in one outer atomic block so any error (rule clashes, write
        # failures) rolls both back together. The write backends have
        # their own inner atomic blocks — Django nests them as savepoints.
        with transaction.atomic():
            created_rule_pks = _materialise_staged_rules(session)
            # Progress update inside atomic — visible externally only
            # after commit, but useful for post-mortem debugging via
            # the audit panel.
            _write_progress(
                session,
                stage=PROGRESS_STAGE_WRITING,
                sheets_total=sheets_total,
            )
            if session.mode == ImportSession.MODE_ETL:
                result = _commit_etl_session(session)
            else:
                result = _commit_template_session(session)
    except Exception as exc:  # pragma: no cover - logged via session.result
        session.status = ImportSession.STATUS_ERROR
        session.result = {
            "error": str(exc),
            "stage": "commit",
            "type": type(exc).__name__,
        }
        session.save(update_fields=["status", "result", "updated_at"])
        raise

    # Surface the created rule pks in the commit response so the
    # frontend can link to them in the "Regras criadas" panel. Shallow-
    # copy first so we don't mutate the write-backend's return value.
    if isinstance(result, dict):
        result = {**result, "substitution_rules_created": created_rule_pks}
    else:
        result = {"write_result": result,
                  "substitution_rules_created": created_rule_pks}

    session.status = ImportSession.STATUS_COMMITTED
    session.committed_at = timezone.now()
    session.result = result
    # Clear file_bytes now that we're committed — audit stays in
    # file_hash + file_name + result, but we don't hoard the payload.
    session.file_bytes = None
    session.save(update_fields=[
        "status", "committed_at", "result", "file_bytes", "updated_at",
    ])
    # Final progress marker — frontend strip flips to "concluído" and
    # fades out once status leaves "committing".
    _write_progress(session, stage=PROGRESS_STAGE_DONE)
    return session


def commit_session(session: ImportSession) -> ImportSession:
    """Mode-aware commit: dispatches to the right write backend.

    Template-mode sessions delegate to the legacy ``execute_import_job``;
    ETL-mode sessions re-run ``ETLPipelineService`` with ``commit=True``
    on the original file bytes so the commit reproduces exactly what
    was shown in the analyze step (transformations, substitutions,
    auto-JE creation).

    Synchronous — runs the commit body in-process. The v2 HTTP views
    use :func:`commit_session_async` instead so the 300s gunicorn
    timeout can't bite on large imports.

    Requires ``session.status == READY``. Raises ``CommitNotReady``
    otherwise — the view layer translates that into a 409.

    Write errors flip the session to ``error`` with a diagnostic in
    ``result`` and re-raise; the view layer returns 500.
    """
    _check_commit_gate(session)
    session.status = ImportSession.STATUS_COMMITTING
    session.save(update_fields=["status", "updated_at"])
    return _run_commit(session)


def commit_session_async(session: ImportSession) -> ImportSession:
    """Flip the session to ``COMMITTING`` and enqueue the worker.

    Gate check happens before the flip so a non-ready session never
    leaves ``READY``. In eager mode (tests / dev without Redis) the
    worker runs inline so the returned session is already terminal.
    """
    _check_commit_gate(session)
    session.status = ImportSession.STATUS_COMMITTING
    session.save(update_fields=["status", "updated_at"])
    from .tasks import commit_session_task
    commit_session_task.delay(session.pk)
    session.refresh_from_db()
    return session


def _materialise_staged_rules(session: ImportSession) -> List[int]:
    """Create one ``SubstitutionRule`` per entry in
    ``session.staged_substitution_rules``.

    Each entry is a dict staged by a resolve action (Phase 4B's
    ``map_to_existing`` populates these — in Phase 4A the list is
    expected to be empty; this helper is the plumbing both phases need).

    Runs inside the commit's atomic block: if rule creation fails (e.g.
    ``unique_together`` collision) the entire commit rolls back and the
    session flips to ``error``.

    Returns the list of created rule pks so the commit response can
    surface them to the client under ``substitution_rules_created``.
    """
    staged = session.staged_substitution_rules or []
    if not staged:
        return []
    created: List[int] = []
    for entry in staged:
        if not isinstance(entry, dict):
            raise ValueError(
                f"malformed entry in staged_substitution_rules: {entry!r}"
            )
        rule = SubstitutionRule.objects.create(
            company_id=session.company_id,
            model_name=entry["model_name"],
            field_name=entry["field_name"],
            match_type=entry.get("match_type", "exact"),
            match_value=entry["match_value"],
            substitution_value=entry["substitution_value"],
            filter_conditions=entry.get("filter_conditions"),
            title=entry.get("title"),
            source=SubstitutionRule.SOURCE_IMPORT_SESSION,
            source_session=session,
        )
        created.append(rule.pk)
    return created


def _commit_template_session(session: ImportSession) -> Dict[str, Any]:
    """Template-mode commit: replay parsed rows through
    ``execute_import_job`` (the legacy entry point). Atomic; rollback
    on any write failure.

    Phase 6.z-d: if the session's ``sheets_post_substitution`` cache
    is still valid (rule set unchanged + no resolutions), commit reuses
    those pre-substituted rows and tells ``execute_import_job`` to skip
    the substitution pass via ``import_options.skip_substitutions``.
    Falls through to raw sheets + fresh substitution on any cache miss —
    correct by construction.
    """
    payload = session.parsed_payload or {}
    raw_sheets: Dict[str, List[Dict[str, Any]]] = payload.get("sheets", {}) or {}

    current_rev = _compute_substitution_revision(session.company_id)
    cache_hit = _is_substitution_cache_valid(session, current_rev)
    if cache_hit:
        sheets_dict = payload.get("sheets_post_substitution") or raw_sheets
    else:
        sheets_dict = raw_sheets

    # ``execute_import_job`` expects sheet dicts shaped like
    # ``{"model": "...", "rows": [...]}``. The parse step stored
    # ``{sheet_name: [row_dict]}`` — model name was implicit from the
    # sheet name (convention used by the legacy endpoint). Reshape here
    # rather than at parse time so the stored payload stays close to
    # the raw workbook.
    sheets_for_job = [
        {"model": sheet_name, "rows": rows}
        for sheet_name, rows in sheets_dict.items()
    ]
    import_metadata: Dict[str, Any] = {
        "source": "v2_template",
        "function": "imports_v2.services._commit_template_session",
        "session_id": session.pk,
        "filename": session.file_name,
    }
    if cache_hit:
        import_metadata["import_options"] = {"skip_substitutions": True}
    # PERF (Phase 6.z-f): pre-load FK caches so the commit doesn't fire
    # one DB query per FK per row. A 3-5k row template commit without
    # this was doing 10k+ round-trips to Postgres and hitting the
    # 10 min Celery hard-kill on remote DBs (Railway). ETL mode has
    # been doing this since day one via ETLPipelineService.
    from multitenancy.lookup_cache import LookupCache
    lookup_cache = LookupCache(session.company_id)
    lookup_cache.load()

    # Phase 6.z-g — intra-atomic row-level progress via Redis. The
    # Redis write bypasses the DB transaction so the polling
    # frontend sees per-batch updates while the atomic block is
    # still open. Falls back to a no-op when REDIS_URL is absent
    # (dev/tests) — in that case the stage-level DB progress from
    # 6.z-e is still the only signal the UI gets.
    from . import progress_channel
    session_pk = session.pk

    def _on_progress(fields: Dict[str, Any]) -> None:
        # Caller-side enrichment: add stage + pass through everything
        # the importer published. Stage is hard-coded to ``writing``
        # because that's the only window this callback fires from.
        progress_channel.publish(
            session_pk,
            {"stage": "writing", **fields},
        )

    with transaction.atomic():
        return execute_import_job(
            company_id=session.company_id,
            sheets=sheets_for_job,
            commit=True,
            import_metadata=import_metadata,
            lookup_cache=lookup_cache,
            progress_callback=_on_progress,
        )


def _commit_etl_session(session: ImportSession) -> Dict[str, Any]:
    """ETL-mode commit: re-run ``ETLPipelineService`` on the original
    file bytes with ``commit=True``.

    Why re-run instead of replaying session.parsed_payload through
    execute_import_job? The ETL pipeline has side-effects (FK
    resolution via lookup_cache, post-processing for JournalEntry
    debit/credit, auto_create_journal_entries hook, integration-rule
    triggers) that aren't reproducible from just the transformed rows.
    Re-running against the same file bytes + same transformation rule
    within the 24h session TTL reproduces what we showed in analyze.
    """
    file_bytes = session.file_bytes
    if not file_bytes:
        raise CommitNotReady(
            f"session #{session.pk} has no file_bytes (expired?) — re-upload"
        )

    # Reconstruct an UploadedFile-like object for ETLPipelineService.
    # The service only reads ``.name`` and pandas-reads the file itself.
    from django.core.files.uploadedfile import SimpleUploadedFile
    file_like = SimpleUploadedFile(
        session.file_name or "import.xlsx",
        bytes(file_bytes),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    config = dict(session.config or {})
    auto_je = config.get("auto_create_journal_entries") or {}
    row_limit = config.get("row_limit")

    service = ETLPipelineService(
        company_id=session.company_id,
        file=file_like,
        commit=True,
        auto_create_journal_entries=auto_je,
        row_limit=row_limit if row_limit is not None else 0,  # 0 = all rows
    )
    # ``execute()`` handles its own transaction. On any failure it
    # raises, which our caller catches to flip session to ``error``.
    return service.execute()


# --- public: ETL analyze ---------------------------------------------------


def _detect_missing_etl_parameters(
    auto_je: Optional[Dict[str, Any]],
    transformed_data: Dict[str, List[Dict[str, Any]]],
    *,
    rule: Optional[ImportTransformationRule] = None,
) -> List[Dict[str, Any]]:
    """Check that every column the auto-JE config references is actually
    present in the transformed data.

    Example: config has ``auto_create_journal_entries.bank_account_field
    = 'bank_account_id'`` but no Transaction row exposes that key
    (operator forgot to map it in ``column_mappings``). Without this
    check the ETL pipeline would silently fail to create bank legs.

    ``auto_je`` is the dict the operator passes per-request (same
    semantics as the legacy ``/api/core/etl/execute/`` endpoint, which
    accepts the config in the POST body rather than on the rule). The
    optional ``rule`` is only used to populate the issue's context for
    the diagnostics panel.

    Returns zero or more blocking issues. Shown to the operator in the
    diagnostics panel per manual §11.10e "Parâmetros ausentes".
    """
    if not auto_je or not isinstance(auto_je, dict) or not auto_je.get("enabled"):
        return []

    # Fields that the auto-JE flow reads from each Transaction row.
    # ``opposing_account_field`` has a default of ``account_path`` in
    # the service; we treat explicit config as the source of truth.
    checked_fields = {
        "bank_account_field": auto_je.get("bank_account_field") or "bank_account_id",
        "opposing_account_field": auto_je.get("opposing_account_field") or "account_path",
    }
    cost_center_field = auto_je.get("cost_center_field")
    if cost_center_field:
        checked_fields["cost_center_field"] = cost_center_field

    # Look at the first row of each transformed model's data to inspect
    # which keys exist. If transformed_data is empty the analyze stage
    # has other, more fundamental errors — no point emitting missing-
    # param issues on top.
    tx_rows = transformed_data.get("Transaction") or []
    if not tx_rows:
        return []
    present_keys = set()
    for row in tx_rows[:5]:  # sample first few; uniform shape in practice
        if isinstance(row, dict):
            present_keys.update(row.keys())

    issues: List[Dict[str, Any]] = []
    for role, field_name in checked_fields.items():
        if not field_name:
            continue
        if field_name not in present_keys:
            issues.append(issue_mod.make_issue(
                type=issue_mod.ISSUE_MISSING_ETL_PARAMETER,
                severity=issue_mod.SEVERITY_ERROR,
                location={
                    "sheet": "Transaction",
                    "expected_column": field_name,
                    "role": role,
                },
                context={
                    "rule_id": rule.pk if rule is not None else None,
                    "rule_name": getattr(rule, "name", None) if rule is not None else None,
                    "auto_create_journal_entries": auto_je,
                    "present_columns": sorted(present_keys),
                },
                proposed_actions=[issue_mod.ACTION_ABORT],
                message=(
                    f"A configuração ``auto_create_journal_entries`` da "
                    f"regra espera a coluna {field_name!r} ({role}) na aba "
                    f"Transaction, mas ela não existe nas linhas "
                    f"transformadas. Ajuste o ``column_mappings`` da regra "
                    f"para produzir essa coluna, ou desabilite "
                    f"auto_create_journal_entries."
                ),
            ))
    return issues


def _create_etl_session(
    *,
    company_id: int,
    user,
    file_bytes: bytes,
    file_name: str,
    transformation_rule_id: Optional[int] = None,
    config: Optional[Dict[str, Any]] = None,
) -> ImportSession:
    """Persist a fresh ETL-mode session in ``ANALYZING`` state.

    Resolves the transformation rule FK (or stores ``None`` if the id
    doesn't match — the body function will surface that as a terminal
    ``error``). The rule is persisted on the session so the worker can
    load it without re-passing ``transformation_rule_id`` through the
    task signature.
    """
    file_hash = hashlib.sha256(file_bytes).hexdigest()
    now = timezone.now()
    cfg = dict(config or {})
    # Freeze the auto-JE config — passed in ``config`` by the caller
    # (matches legacy ETL semantics where the config lives in the POST
    # body, not on the rule). Defaulting to {} means "no auto-JE" which
    # is the safe default.
    cfg.setdefault("auto_create_journal_entries", {})

    rule = None
    if transformation_rule_id is not None:
        try:
            rule = ImportTransformationRule.objects.get(
                pk=transformation_rule_id, company_id=company_id,
            )
        except ImportTransformationRule.DoesNotExist:
            rule = None  # surface as a lookup-stage error below

    session = ImportSession.objects.create(
        company_id=company_id,
        created_by=user if (user and getattr(user, "is_authenticated", False)) else None,
        mode=ImportSession.MODE_ETL,
        status=ImportSession.STATUS_ANALYZING,
        transformation_rule=rule,
        file_name=file_name or "upload.xlsx",
        file_hash=file_hash,
        file_bytes=file_bytes,
        config=cfg,
        expires_at=now + SESSION_TTL,
    )
    # Remember the operator-supplied rule id so the body can distinguish
    # "rule not found" (terminal error) from "no rule requested" (legal
    # in some test paths). Stashed in config so it roundtrips through
    # the Celery worker.
    if transformation_rule_id is not None and rule is None:
        session.config = {**cfg, "_requested_rule_id": transformation_rule_id}
        session.save(update_fields=["config", "updated_at"])
    return session


def _run_analyze_etl(session: ImportSession) -> ImportSession:
    """Analyze body for an ETL-mode session already in ``ANALYZING``.

    Runs ``ETLPipelineService.execute(commit=False)`` against the
    persisted file bytes + rule, then layers on the v2 issue detectors
    and persists the result. Sync and worker callers share this path.

    Unexpected exceptions propagate so the worker wrapper can mark the
    session as ``error``; known failure modes (rule-not-found, ETL
    service error) are handled inline.
    """
    from django.core.files.uploadedfile import SimpleUploadedFile

    company_id = session.company_id
    file_bytes = bytes(session.file_bytes or b"")
    file_name = session.file_name or "upload.xlsx"
    cfg = dict(session.config or {})

    rule = session.transformation_rule
    requested_rule_id = cfg.pop("_requested_rule_id", None)
    if rule is None and requested_rule_id is not None:
        session.status = ImportSession.STATUS_ERROR
        session.result = {
            "error": f"transformation rule #{requested_rule_id} not found for this company",
            "stage": "lookup",
        }
        session.config = cfg  # strip the internal marker on the persisted config
        session.save(update_fields=["status", "result", "config", "updated_at"])
        return session

    # Build a fresh UploadedFile for ETLPipelineService — it reads the
    # stream + ``.name``. We don't give it the raw bytes directly.
    file_like = SimpleUploadedFile(
        file_name,
        file_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    service = ETLPipelineService(
        company_id=company_id,
        file=file_like,
        commit=False,
        auto_create_journal_entries=cfg.get("auto_create_journal_entries") or {},
        # row_limit=0 means "process all rows"; cfg can override for
        # sampled previews.
        row_limit=cfg.get("row_limit", 0),
    )

    try:
        etl_result = service.execute()
    except Exception as exc:
        session.status = ImportSession.STATUS_ERROR
        session.result = {
            "error": str(exc),
            "stage": "etl_execute",
            "type": type(exc).__name__,
        }
        session.save(update_fields=["status", "result", "updated_at"])
        return session

    # ETLPipelineService surfaces a ``transformed_data`` dict shaped as
    # ``{model_name: {"row_count": N, "rows": [...], "sample_columns": [...]}}``.
    # Flatten to ``{model_name: [rows]}`` for our own consumption. Keep
    # the full service result under ``etl_result`` for debugging/audit.
    transformed_flat: Dict[str, List[Dict[str, Any]]] = {}
    raw_td = (etl_result or {}).get("transformed_data") or {}
    for model_name, payload in raw_td.items():
        if isinstance(payload, dict) and "rows" in payload:
            transformed_flat[model_name] = list(payload.get("rows") or [])
        elif isinstance(payload, list):
            transformed_flat[model_name] = payload
        else:
            transformed_flat[model_name] = []

    detected_issues: List[Dict[str, Any]] = []
    # erp_id_conflict — same detector as template mode, applied to the
    # transformed Transaction rows (so substitutions have already had a
    # chance to fix mis-keyed rows).
    if "Transaction" in transformed_flat:
        detected_issues.extend(
            _detect_transaction_erp_id_conflicts(
                transformed_flat["Transaction"], "Transaction",
            )
        )
    # missing_etl_parameter — config-driven column check. Rule is
    # passed for diagnostic-panel context only; the actual check runs
    # against the auto-JE config the caller provided.
    detected_issues.extend(_detect_missing_etl_parameters(
        cfg.get("auto_create_journal_entries"),
        transformed_flat,
        rule=rule,
    ))
    # Phase 4B detectors — run on every transformed sheet.
    for model_name, rows in transformed_flat.items():
        detected_issues.extend(_detect_bad_date_format(rows, model_name))
        detected_issues.extend(_detect_negative_amounts(
            rows, model_name, rule=rule,
        ))
        detected_issues.extend(_detect_unmatched_references(
            rows, model_name, company_id=company_id,
        ))

    # Surface the ETL service's own error buckets as ``parsed_payload['etl_errors']``
    # (the frontend already renders them; we don't translate each one
    # into an Issue in Phase 3 — that's Phase 4 when unmatched-reference
    # resolution ships).
    #
    # ``would_create`` / ``would_fail`` / ``total_rows`` come from
    # ``ETLPipelineService.execute(commit=False)`` — they're the dry-run
    # counts legacy ETL preview already exposes. We pass them through
    # as ``parsed_payload["preview"]`` so the serializer's ``preview``
    # field can surface them to the frontend's "Prévia da importação"
    # panel on a clean analyze.
    etl_errors_payload = {
        "python_errors": etl_result.get("python_errors") or [],
        "database_errors": etl_result.get("database_errors") or [],
        "substitution_errors": etl_result.get("substitution_errors") or [],
        "warnings": etl_result.get("warnings") or [],
    }
    session.parsed_payload = {
        "transformed_data": transformed_flat,
        "substitutions_applied": etl_result.get("substitutions_applied") or [],
        "etl_errors": etl_errors_payload,
        "sheets_processed": etl_result.get("sheets_processed") or [],
        "preview": {
            "would_create": etl_result.get("would_create") or {},
            "would_fail": etl_result.get("would_fail") or {},
            "total_rows": etl_result.get("total_rows") or 0,
        },
    }
    # ETL mirror of the template dry-run gate: any python / database /
    # substitution error flips commit to ``awaiting_resolve`` so the
    # operator has to abort + reupload. The frontend already renders
    # the per-row detail from ``etl_errors``.
    detected_issues.extend(_emit_etl_dry_run_failure_issues(etl_errors_payload))
    session.open_issues = detected_issues
    if issue_mod.has_blocking_issues(detected_issues):
        session.status = ImportSession.STATUS_AWAITING_RESOLVE
    else:
        session.status = ImportSession.STATUS_READY
    # Clear the internal requested_rule_id marker if it was still there
    # (lookup succeeded path).
    if "_requested_rule_id" in (session.config or {}):
        session.config = {
            k: v for k, v in (session.config or {}).items()
            if k != "_requested_rule_id"
        }
    session.save(update_fields=[
        "parsed_payload", "open_issues", "status", "config", "updated_at",
    ])
    return session


def analyze_etl(
    *,
    company_id: int,
    user,
    file_bytes: bytes,
    file_name: str,
    transformation_rule_id: Optional[int] = None,
    config: Optional[Dict[str, Any]] = None,
) -> ImportSession:
    """Create an ETL-mode ``ImportSession`` and populate it by running
    ``ETLPipelineService.execute(commit=False)`` under the hood.

    Captures the service's outputs — transformed data, substitutions
    applied, substitution errors, python/database errors — into the
    session's ``parsed_payload`` + ``open_issues`` so the operator
    can review diagnostics without re-uploading.

    Synchronous — runs in-process. The v2 HTTP views use
    :func:`analyze_etl_async` which hands the body off to Celery.

    Phase 3 issue detection:
      * ``erp_id_conflict``  — same as template mode, scoped to
        ``transformed_data['Transaction']``.
      * ``missing_etl_parameter`` — rule declares auto-JE columns that
        aren't present in transformed rows.
      * ETL service errors (python_errors / database_errors /
        substitution_errors) pass through as ``parsed_payload['etl_errors']``
        so the frontend's existing error-bucket renderers keep working.
    """
    session = _create_etl_session(
        company_id=company_id, user=user, file_bytes=file_bytes,
        file_name=file_name, transformation_rule_id=transformation_rule_id,
        config=config,
    )
    return _run_analyze_etl(session)


def analyze_etl_async(
    *,
    company_id: int,
    user,
    file_bytes: bytes,
    file_name: str,
    transformation_rule_id: Optional[int] = None,
    config: Optional[Dict[str, Any]] = None,
) -> ImportSession:
    """Create the ETL session in ``ANALYZING`` state and enqueue the worker.

    Returns immediately; the frontend polls ``GET /sessions/<id>/``
    until the session leaves ``analyzing``. Eager mode runs inline.
    """
    session = _create_etl_session(
        company_id=company_id, user=user, file_bytes=file_bytes,
        file_name=file_name, transformation_rule_id=transformation_rule_id,
        config=config,
    )
    from .tasks import analyze_session_task
    analyze_session_task.delay(session.pk)
    session.refresh_from_db()
    return session


# --- public: resolve -------------------------------------------------------


class ResolveNotApplicable(Exception):
    """Raised when ``resolve_session`` is called on a session whose status
    doesn't accept resolutions. View layer maps to 409.
    """


def _redetect_issues(session: ImportSession) -> List[Dict[str, Any]]:
    """Re-run the detectors relevant to ``session.mode`` over the current
    ``parsed_payload``. Called after every resolve-batch so an operator
    who partially resolves a group still sees the remaining issues.

    Mirrors the detector calls in ``analyze_template`` / ``analyze_etl``.
    When a new detector is added, add it here too or the resolve cycle
    will happily close out a session that an analyze would have blocked.
    """
    detected: List[Dict[str, Any]] = []
    company_id = session.company_id
    if session.mode == ImportSession.MODE_ETL:
        transformed = (session.parsed_payload or {}).get("transformed_data") or {}
        if "Transaction" in transformed:
            detected.extend(_detect_transaction_erp_id_conflicts(
                transformed["Transaction"], "Transaction",
            ))
        auto_je = (session.config or {}).get("auto_create_journal_entries")
        detected.extend(_detect_missing_etl_parameters(
            auto_je, transformed, rule=session.transformation_rule,
        ))
        for model_name, rows in transformed.items():
            detected.extend(_detect_bad_date_format(rows, model_name))
            detected.extend(_detect_negative_amounts(
                rows, model_name, rule=session.transformation_rule,
            ))
            detected.extend(_detect_unmatched_references(
                rows, model_name, company_id=company_id,
            ))
    else:
        sheets = (session.parsed_payload or {}).get("sheets") or {}
        for sheet_name, rows in sheets.items():
            if sheet_name == "Transaction":
                detected.extend(_detect_transaction_erp_id_conflicts(
                    rows, sheet_name,
                ))
            detected.extend(_detect_bad_date_format(rows, sheet_name))
            detected.extend(_detect_unmatched_references(
                rows, sheet_name, company_id=company_id,
            ))
    return detected


def resolve_session(
    session: ImportSession,
    resolutions: List[Dict[str, Any]],
) -> ImportSession:
    """Apply a batch of operator-provided resolutions to the session.

    Each resolution is ``{issue_id, action, params}``. For each one we:
      * find the matching issue in ``session.open_issues``,
      * dispatch to the per-action handler in ``resolve_handlers``,
      * record the result on ``session.resolutions``,
      * re-run detection on the (now-mutated) parsed_payload.

    Once all applicable resolutions have run we recompute the session
    status: ``ready`` if no blocking issues remain, otherwise the
    session stays in ``awaiting_resolve``. An ``abort`` resolution
    short-circuits the batch and flips the session to ``error``.

    Raises:
      * ``ResolveNotApplicable`` — session is terminal/committing.
      * ``resolve_handlers.ResolutionError`` — malformed params /
        unknown action / unknown issue_id (view translates to 400).
    """
    if session.is_terminal():
        raise ResolveNotApplicable(
            f"session #{session.pk} is terminal ({session.status}); "
            f"cannot resolve"
        )
    if session.status == ImportSession.STATUS_COMMITTING:
        raise ResolveNotApplicable(
            f"session #{session.pk} is committing; resolve not allowed"
        )

    if not isinstance(resolutions, list):
        raise _resolve_handlers.ResolutionError(
            "`resolutions` must be a list"
        )

    now_iso = timezone.now().isoformat()
    open_issues_by_id = {
        i.get("issue_id"): i for i in (session.open_issues or [])
    }
    new_resolutions: List[Dict[str, Any]] = []
    aborted = False
    abort_info: Optional[Dict[str, Any]] = None

    for res in resolutions:
        if not isinstance(res, dict):
            raise _resolve_handlers.ResolutionError(
                "each resolution must be an object"
            )
        issue_id = res.get("issue_id")
        action = res.get("action")
        params = res.get("params") or {}
        if not issue_id:
            raise _resolve_handlers.ResolutionError(
                "resolution is missing issue_id"
            )
        if not action:
            raise _resolve_handlers.ResolutionError(
                "resolution is missing action"
            )
        issue = open_issues_by_id.get(issue_id)
        if issue is None:
            raise _resolve_handlers.ResolutionError(
                f"no open issue with issue_id={issue_id!r}"
            )
        result = _resolve_handlers.apply_resolution(
            session, issue, action, params,
        )
        record = {
            "issue_id": issue_id,
            "action": action,
            "params": params,
            "result": result,
            "resolved_at": now_iso,
        }
        new_resolutions.append(record)
        if result.get("abort"):
            aborted = True
            abort_info = result
            break

    # Persist the resolution records BEFORE any status transition so the
    # audit trail is complete even if redetection or save fails.
    session.resolutions = list(session.resolutions or []) + new_resolutions

    if aborted:
        session.status = ImportSession.STATUS_ERROR
        session.result = {
            "error": "operator aborted",
            "stage": "resolve",
            "abort": abort_info,
        }
        session.save(update_fields=[
            "resolutions", "status", "result", "updated_at",
        ])
        return session

    # Re-detect issues against the now-mutated parsed_payload. Replaces
    # the open_issues list wholesale — any issue that's still valid will
    # be re-emitted; any issue the operator resolved won't re-appear.
    redetected = _redetect_issues(session)
    session.open_issues = redetected

    if issue_mod.has_blocking_issues(redetected):
        session.status = ImportSession.STATUS_AWAITING_RESOLVE
    else:
        session.status = ImportSession.STATUS_READY

    session.save(update_fields=[
        "parsed_payload", "open_issues", "resolutions",
        "staged_substitution_rules", "status", "updated_at",
    ])
    return session


# --- public: discard -------------------------------------------------------


def discard_session(session: ImportSession) -> ImportSession:
    """Mark a session as discarded and clear its file bytes.

    Idempotent: discarding an already-discarded session is a no-op.
    Terminal sessions (committed / error) cannot be discarded — they
    stay in their terminal state for audit.
    """
    if session.is_terminal():
        return session
    session.status = ImportSession.STATUS_DISCARDED
    session.file_bytes = None
    session.save(update_fields=["status", "file_bytes", "updated_at"])
    return session
