"""Weekly activity digest — xlsx roll-up emailed to the platform admin.

Runs on a Celery Beat schedule (see ``nord_backend/settings.py``),
also exposed as:

  * a management command (``python manage.py send_activity_digest``)
  * a DRF action (``POST /api/admin/activity/digest/run/``) for the
    "Run now" button on the admin page.

The digest compares the current 7-day window to the prior 7 days so
the admin can see *movement*, not just levels. Slower / noisier than
the live dashboards on purpose — the goal is "what changed while I
was looking away," not a real-time feed.
"""

from __future__ import annotations

import logging
from datetime import datetime, timedelta
from io import BytesIO
from typing import Any

from celery import shared_task
from django.conf import settings
from django.core.mail import EmailMessage
from django.contrib.auth import get_user_model
from django.db.models import Count, Sum
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.models import UserActivityEvent, UserActivitySession
from core.services.activity_friction import compute_friction
from core.services.activity_funnels import compute_all_funnels


log = logging.getLogger(__name__)


_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_BOLD = Font(bold=True)
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _ms_to_label(ms: int | None) -> str:
    if not ms or ms <= 0:
        return "—"
    s = ms // 1000
    if s < 60:
        return f"{s}s"
    m, s = divmod(s, 60)
    if m < 60:
        return f"{m}m {s}s" if s else f"{m}m"
    h, m = divmod(m, 60)
    return f"{h}h {m}m" if m else f"{h}h"


def _pct_delta(curr: float | int | None, prev: float | int | None) -> str:
    """Human-readable WoW change. Returns ``""`` when we can't compare."""
    if prev in (None, 0) or curr is None:
        return ""
    diff_pct = (float(curr) - float(prev)) / float(prev) * 100.0
    arrow = "▲" if diff_pct > 0 else "▼" if diff_pct < 0 else "•"
    return f"{arrow} {abs(diff_pct):.0f}%"


def _write_header(ws, labels: list[str]) -> None:
    for i, lbl in enumerate(labels, start=1):
        cell = ws.cell(row=1, column=i, value=lbl)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws, max_cols: int, max_sample_rows: int = 50) -> None:
    for c in range(1, max_cols + 1):
        width = 10
        for r in range(1, min(ws.max_row, max_sample_rows) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            width = max(width, min(48, len(str(v)) + 2))
        ws.column_dimensions[get_column_letter(c)].width = width


# ---------------------------------------------------------------- data shaping


def _collect_windowed(days: int = 7) -> dict[str, Any]:
    """Pull aggregated numbers for the window, plus a paired window
    immediately before it for WoW deltas."""
    now = timezone.now()
    since = now - timedelta(days=days)
    prev_since = now - timedelta(days=days * 2)

    def _window_totals(lo, hi) -> dict[str, int]:
        qs = UserActivityEvent.objects.filter(created_at__gte=lo, created_at__lt=hi)
        hb = qs.filter(kind=UserActivityEvent.KIND_HEARTBEAT)
        return {
            "events": qs.count(),
            "focused_ms": hb.aggregate(v=Sum("duration_ms"))["v"] or 0,
            "distinct_users": qs.values("user_id").distinct().count(),
            "errors": qs.filter(kind=UserActivityEvent.KIND_ERROR).count(),
            "sessions": UserActivitySession.objects.filter(started_at__gte=lo, started_at__lt=hi).count(),
        }

    current = _window_totals(since, now)
    previous = _window_totals(prev_since, since)

    # Time-by-area heatmap (this week only — WoW deltas per cell
    # would balloon the sheet; aggregates per user and per area are
    # more actionable).
    by_user_area = list(
        UserActivityEvent.objects
        .filter(created_at__gte=since, kind=UserActivityEvent.KIND_HEARTBEAT)
        .values("user_id", "user__username", "area")
        .annotate(focused_ms=Sum("duration_ms"), events=Count("id"))
        .order_by("-focused_ms")
    )

    # Funnels + friction (compute helpers are the same ones the
    # live dashboards use — single source of truth).
    funnels = compute_all_funnels(days=days)
    friction = compute_friction(days=days)

    # Users: inactive in the last window + new users created
    User = get_user_model()
    active_user_ids = set(
        UserActivityEvent.objects.filter(created_at__gte=since).values_list("user_id", flat=True).distinct()
    )
    all_users = list(
        User.objects.filter(is_active=True).values("id", "username", "email", "last_login", "date_joined")
    )
    inactive = [u for u in all_users if u["id"] not in active_user_ids]
    newly_joined = [u for u in all_users if u["date_joined"] and u["date_joined"] >= since]

    return {
        "now": now,
        "since": since,
        "prev_since": prev_since,
        "days": days,
        "current": current,
        "previous": previous,
        "by_user_area": by_user_area,
        "funnels": funnels,
        "friction": friction,
        "inactive_users": inactive,
        "new_users": newly_joined,
    }


# ---------------------------------------------------------------- workbook


def build_digest_workbook(data: dict[str, Any]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    _build_summary_sheet(wb, data)
    _build_time_by_area_sheet(wb, data)
    _build_funnels_sheet(wb, data)
    _build_friction_sheet(wb, data)
    _build_users_sheet(wb, data)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_summary_sheet(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Resumo")
    now = data["now"]
    since = data["since"]
    curr = data["current"]
    prev = data["previous"]

    ws["A1"] = "Digest semanal — Atividade da plataforma"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Janela: {since.date().isoformat()} → {now.date().isoformat()}  ({data['days']}d)"
    ws["A2"].font = Font(italic=True, color="6B7280")

    # KPI block
    row = 4
    ws.cell(row=row, column=1, value="Métrica").font = _HEADER_FONT
    ws.cell(row=row, column=2, value="Esta janela").font = _HEADER_FONT
    ws.cell(row=row, column=3, value="Janela anterior").font = _HEADER_FONT
    ws.cell(row=row, column=4, value="Variação").font = _HEADER_FONT
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = _HEADER_FILL

    kpis = [
        ("Eventos totais", curr["events"], prev["events"]),
        ("Tempo focado", _ms_to_label(curr["focused_ms"]), _ms_to_label(prev["focused_ms"])),
        ("Usuários ativos", curr["distinct_users"], prev["distinct_users"]),
        ("Sessões", curr["sessions"], prev["sessions"]),
        ("Erros", curr["errors"], prev["errors"]),
    ]
    for i, (lbl, a, b) in enumerate(kpis, start=1):
        ws.cell(row=row + i, column=1, value=lbl).font = _BOLD
        ws.cell(row=row + i, column=2, value=a)
        ws.cell(row=row + i, column=3, value=b)
        # For the label/text rows (focused_ms) _pct_delta() gets a
        # string and bails out — not worth detecting numeric fields
        # here, the empty string is a fine cue.
        delta = _pct_delta(a if isinstance(a, (int, float)) else None, b if isinstance(b, (int, float)) else None)
        ws.cell(row=row + i, column=4, value=delta)

    # Top friction snippets — one-line highlights to read on phone.
    row = row + len(kpis) + 3
    ws.cell(row=row, column=1, value="Principais sinais de fricção").font = _BOLD
    ws.cell(row=row, column=1).fill = _HEADER_FILL
    ws.cell(row=row, column=1).font = _HEADER_FONT
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=4)
    row += 1
    bullets: list[str] = []
    fr = data["friction"]
    if fr["slow_actions"]:
        a = fr["slow_actions"][0]
        bullets.append(f"Ação mais lenta: {a['action']} @ {a['area']} — p95 {a['p95_ms']}ms (n={a['samples']})")
    if fr["repeat_errors"]:
        e = fr["repeat_errors"][0]
        bullets.append(f"Maior cadeia de erros: {e['username']} em {e['area']} — {e['errors']} erros")
    if fr["back_and_forth"]:
        b = fr["back_and_forth"][0]
        bullets.append(f"Ciclo mais comum: {b['from_area']} ↔ {b['to_area']} — {b['count']}×")
    if fr["long_dwell_no_action"]:
        d = fr["long_dwell_no_action"][0]
        bullets.append(f"Maior travamento: {d['username']} em {d['area']} — {_ms_to_label(d['focused_ms'])}")
    if not bullets:
        bullets.append("Sem sinais detectados na janela.")
    for line in bullets:
        ws.cell(row=row, column=1, value=line)
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=4)
        row += 1

    # Worst funnel drop-offs.
    row += 1
    ws.cell(row=row, column=1, value="Maiores quedas em funis").font = _HEADER_FONT
    ws.cell(row=row, column=1).fill = _HEADER_FILL
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=4)
    row += 1
    drops: list[tuple[str, float]] = []
    for f in data["funnels"]:
        for s in f["steps"]:
            if s["dropoff_pct"] is not None:
                drops.append((f"{f['label']} · {s['label']}", float(s["dropoff_pct"])))
    drops.sort(key=lambda x: x[1], reverse=True)
    for label, pct in drops[:5]:
        ws.cell(row=row, column=1, value=f"{label} — queda de {pct:.1f}%")
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=4)
        row += 1

    _autosize(ws, 4)


def _build_time_by_area_sheet(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Tempo por área")
    _write_header(ws, ["Usuário", "Área", "Eventos", "Tempo focado (ms)", "Tempo focado"])
    for i, row in enumerate(data["by_user_area"], start=2):
        ws.cell(row=i, column=1, value=row["user__username"])
        ws.cell(row=i, column=2, value=row["area"])
        ws.cell(row=i, column=3, value=row["events"])
        ws.cell(row=i, column=4, value=int(row["focused_ms"] or 0))
        ws.cell(row=i, column=5, value=_ms_to_label(row["focused_ms"] or 0))
    _autosize(ws, 5)


def _build_funnels_sheet(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Funis")
    _write_header(ws, [
        "Funil", "Passo", "Sessões", "Queda %", "p50 ms", "p95 ms", "Amostras",
    ])
    row = 2
    for f in data["funnels"]:
        for s in f["steps"]:
            t = s.get("timing_from_previous") or {}
            ws.cell(row=row, column=1, value=f["label"])
            ws.cell(row=row, column=2, value=s["label"])
            ws.cell(row=row, column=3, value=s["reached"])
            ws.cell(row=row, column=4, value=s["dropoff_pct"])
            ws.cell(row=row, column=5, value=t.get("p50_ms"))
            ws.cell(row=row, column=6, value=t.get("p95_ms"))
            ws.cell(row=row, column=7, value=t.get("samples"))
            row += 1
    _autosize(ws, 7)


def _build_friction_sheet(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Fricção")
    fr = data["friction"]
    row = 1

    def _section(title: str, headers: list[str], rows: list[list[Any]]) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = _HEADER_FONT
        ws.cell(row=row, column=1).fill = _HEADER_FILL
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=max(1, len(headers)))
        row += 1
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = _BOLD
        row += 1
        for r in rows:
            for i, v in enumerate(r, start=1):
                ws.cell(row=row, column=i, value=v)
            row += 1
        row += 1  # spacer between sections

    _section(
        "Ações lentas (p95)",
        ["Ação", "Área", "n", "p50 ms", "p95 ms", "máx ms"],
        [[a["action"], a["area"], a["samples"], a["p50_ms"], a["p95_ms"], a["max_ms"]]
         for a in fr["slow_actions"]],
    )
    _section(
        "Cadeias de erro",
        ["Usuário", "Área", "Erros", "Início", "Fim"],
        [[e["username"], e["area"], e["errors"],
          e["first_at"].isoformat() if hasattr(e["first_at"], "isoformat") else str(e["first_at"]),
          e["last_at"].isoformat() if hasattr(e["last_at"], "isoformat") else str(e["last_at"])]
         for e in fr["repeat_errors"]],
    )
    _section(
        "Navegação cíclica",
        ["De", "Para", "Ocorrências", "Amostra de usuários"],
        [[b["from_area"], b["to_area"], b["count"],
          ", ".join(u["username"] for u in b["sample_users"][:5])]
         for b in fr["back_and_forth"]],
    )
    _section(
        "Travamento suspeito",
        ["Usuário", "Área", "Tempo focado"],
        [[d["username"], d["area"], _ms_to_label(d["focused_ms"])]
         for d in fr["long_dwell_no_action"]],
    )
    _autosize(ws, 6)


def _build_users_sheet(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Usuários")
    _write_header(ws, ["Seção", "Usuário", "E-mail", "Último login", "Cadastro"])
    row = 2

    def _add(section: str, users: list[dict]) -> None:
        nonlocal row
        for u in users:
            ws.cell(row=row, column=1, value=section)
            ws.cell(row=row, column=2, value=u["username"])
            ws.cell(row=row, column=3, value=u.get("email") or "")
            ws.cell(row=row, column=4, value=u["last_login"].isoformat() if u.get("last_login") else "")
            ws.cell(row=row, column=5, value=u["date_joined"].isoformat() if u.get("date_joined") else "")
            row += 1

    _add("Novos esta janela", data["new_users"])
    _add("Inativos (sem eventos)", data["inactive_users"])
    _autosize(ws, 5)


# ---------------------------------------------------------------- recipient


def _recipient() -> str | None:
    """Where the digest lands. Priority:
       1. ``ADMIN_DIGEST_EMAIL`` env var (explicit),
       2. dcaffaro user's email,
       3. None (caller logs a warning).
    """
    env = getattr(settings, "ADMIN_DIGEST_EMAIL", None) or \
        __import__("os").environ.get("ADMIN_DIGEST_EMAIL")
    if env:
        return env
    User = get_user_model()
    u = User.objects.filter(username="dcaffaro").first()
    return u.email if u else None


# ---------------------------------------------------------------- task


@shared_task(bind=True, name="core.tasks_activity_digest.send_weekly_digest")
def send_weekly_digest(self, *, days: int = 7, dry_run: bool = False,
                       recipient: str | None = None) -> dict[str, Any]:
    """Build + email the activity digest.

    ``dry_run=True`` returns the payload stats without sending — useful
    for the management command's ``--dry-run`` flag and unit tests.
    """
    data = _collect_windowed(days=days)
    xlsx_bytes = build_digest_workbook(data)

    to = recipient or _recipient()
    subject = f"[Nord] Digest semanal de atividade — {data['now'].date().isoformat()}"
    summary = (
        f"Janela: {data['since'].date()} → {data['now'].date()} ({data['days']}d)\n"
        f"Eventos: {data['current']['events']} "
        f"(WoW {_pct_delta(data['current']['events'], data['previous']['events'])})\n"
        f"Usuários ativos: {data['current']['distinct_users']}\n"
        f"Erros: {data['current']['errors']}\n\n"
        "Detalhes no anexo XLSX. Atalhos:\n"
        "  • /admin/activity          — heatmap\n"
        "  • /admin/activity/funnels  — funis\n"
        "  • /admin/activity/friction — sinais de fricção\n"
    )

    if dry_run:
        return {"sent": False, "recipient": to, "xlsx_bytes": len(xlsx_bytes), "subject": subject}

    if not to:
        log.warning("send_weekly_digest: no recipient resolved, skipping email")
        return {"sent": False, "reason": "no_recipient", "xlsx_bytes": len(xlsx_bytes)}

    msg = EmailMessage(
        subject=subject,
        body=summary,
        from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
        to=[to],
    )
    filename = f"nord-digest-{data['now'].date().isoformat()}.xlsx"
    msg.attach(filename, xlsx_bytes, _XLSX_MIME)
    try:
        msg.send(fail_silently=False)
    except Exception as e:
        # Log + re-raise so Celery records the failure; the beat
        # schedule will try again next week. We don't want to bury
        # an SMTP misconfiguration silently.
        log.exception("send_weekly_digest failed to send: %s", e)
        raise

    return {"sent": True, "recipient": to, "xlsx_bytes": len(xlsx_bytes), "filename": filename}
