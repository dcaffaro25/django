"""XLSX export for the new report engine.

Takes a :class:`ReportCalculator` result (or a persisted ``ReportInstance.result``)
and renders a multi-period workbook. Independent of the legacy
``detailed_statement_excel`` because the new result shape is period-map-per-line
rather than the tree-per-section that the legacy engine emits.

Sheets:

1. **Report** — the main multi-period table. Columns: Label + one per period.
   Indentation + bold applied via cell styling; variance % columns get
   a "%" number format.
2. **Memory** — per-block calculation provenance (accounts contributing,
   method, etc.). Used for audit / the future "explain this number" UI.
3. **Warnings** — only created if the result has warnings.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_BOLD_FONT = Font(bold=True)


def build_xlsx(result: Dict[str, Any], name: str = "Demonstrativo") -> bytes:
    """Render ``result`` to an XLSX byte string."""
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    _build_report_sheet(wb, result, name)
    _build_memory_sheet(wb, result)
    if result.get("warnings"):
        _build_warnings_sheet(wb, result["warnings"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_report_sheet(wb: Workbook, result: Dict[str, Any], name: str) -> None:
    ws = wb.create_sheet(_safe_sheet_name(name or "Report"))

    periods: List[dict] = result.get("periods", [])
    lines: List[dict] = result.get("lines", [])

    # Title row
    template = result.get("template") or {}
    ws.cell(row=1, column=1, value=template.get("name", name)).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=_format_report_type(template.get("report_type", "")))
    ws.cell(row=2, column=1).font = Font(italic=True, color="6B7280")

    header_row = 4
    ws.cell(row=header_row, column=1, value="Linha").font = _HEADER_FONT
    ws.cell(row=header_row, column=1).fill = _HEADER_FILL
    for idx, p in enumerate(periods, start=2):
        cell = ws.cell(row=header_row, column=idx, value=p.get("label") or p.get("id"))
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="right")

    # Data rows
    row = header_row + 1
    for line in lines:
        if line["type"] == "spacer":
            row += 1
            continue

        label = line.get("label") or line.get("id")
        indent = int(line.get("indent") or line.get("depth") or 0)
        label_cell = ws.cell(row=row, column=1, value=("  " * indent) + label)
        if line.get("bold") or line["type"] in ("subtotal", "total", "header", "section"):
            label_cell.font = _BOLD_FONT

        for idx, p in enumerate(periods, start=2):
            pid = p["id"]
            value = line.get("values", {}).get(pid)
            cell = ws.cell(row=row, column=idx, value=value)
            cell.alignment = Alignment(horizontal="right")
            if line.get("bold") or line["type"] in ("subtotal", "total"):
                cell.font = _BOLD_FONT
            if p["type"] in ("variance_pct", "variance_pp"):
                cell.number_format = '0.00" %"'
            else:
                cell.number_format = "#,##0.00"
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 45
    for idx in range(2, len(periods) + 2):
        ws.column_dimensions[get_column_letter(idx)].width = 16


def _build_memory_sheet(wb: Workbook, result: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Memory")
    headers = ["Block", "Period", "Method", "Account count", "Account IDs", "Raw total"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    row = 2
    for line in result.get("lines", []):
        mem = line.get("memory") or {}
        for period_id, entry in mem.items():
            if not entry:
                continue
            ws.cell(row=row, column=1, value=line.get("id"))
            ws.cell(row=row, column=2, value=period_id)
            ws.cell(row=row, column=3, value=entry.get("calc_method"))
            ws.cell(row=row, column=4, value=entry.get("account_count"))
            ws.cell(row=row, column=5, value=", ".join(str(a) for a in (entry.get("account_ids") or [])))
            ws.cell(row=row, column=6, value=entry.get("raw_total"))
            row += 1

    for col, w in [(1, 24), (2, 14), (3, 20), (4, 10), (5, 40), (6, 16)]:
        ws.column_dimensions[get_column_letter(col)].width = w


def _build_warnings_sheet(wb: Workbook, warnings: List[dict]) -> None:
    ws = wb.create_sheet("Warnings")
    for col, h in enumerate(["Level", "Block", "Message"], start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    for idx, w in enumerate(warnings, start=2):
        ws.cell(row=idx, column=1, value=w.get("level"))
        ws.cell(row=idx, column=2, value=w.get("block_id"))
        ws.cell(row=idx, column=3, value=w.get("message"))

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 80


def _format_report_type(rt: str) -> str:
    labels = {
        "income_statement": "Demonstração de Resultado",
        "balance_sheet": "Balanço Patrimonial",
        "cash_flow": "Fluxo de Caixa",
        "trial_balance": "Balancete",
        "general_ledger": "Razão",
        "custom": "Relatório Personalizado",
    }
    return labels.get(rt, rt.replace("_", " ").title())


def _safe_sheet_name(name: str) -> str:
    """openpyxl restricts sheet names to 31 chars and disallows :\\/?*[]."""
    for ch in r":\\/?*[]":
        name = name.replace(ch, "")
    return (name or "Report")[:31]
