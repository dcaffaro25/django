"""
Excel export for detailed financial statements (income statement, balance sheet, cash flow).

Produces a comprehensive workbook with:
- Report summary and detail (flattened hierarchy)
- Calculation memory (parameters, formulas, section breakdown)
- Raw data (AccountBalanceHistory rows used in calculations)
"""

import io
import base64
from datetime import date
from decimal import Decimal
from typing import Dict, List, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def _flatten_report_tree(nodes: List[Dict], depth: int = 0) -> List[Dict]:
    """Flatten hierarchical report nodes into rows with depth, code, name, path, balance."""
    rows = []
    for node in nodes or []:
        rows.append({
            'depth': depth,
            'id': node.get('id'),
            'account_code': node.get('account_code') or '',
            'name': node.get('name') or '',
            'path': node.get('path') or '',
            'balance': node.get('balance'),
            'is_leaf': node.get('is_leaf', False),
        })
        if node.get('children'):
            rows.extend(_flatten_report_tree(node['children'], depth + 1))
    return rows


def _flatten_report_tree_with_section(
    nodes: List[Dict], section_label: str, depth: int = 0
) -> List[Dict]:
    """Flatten report nodes with section label for hierarchy + JE sheet."""
    rows = []
    for node in nodes or []:
        path = (node.get('path') or '').strip() or (node.get('name') or '').strip()
        report_line = f"{section_label} > {path}" if path else section_label
        rows.append({
            'section': section_label,
            'depth': depth,
            'id': node.get('id'),
            'account_code': node.get('account_code') or '',
            'name': node.get('name') or '',
            'path': node.get('path') or '',
            'report_line': report_line,
            'balance': node.get('balance'),
            'is_leaf': node.get('is_leaf', False),
        })
        if node.get('children'):
            rows.extend(
                _flatten_report_tree_with_section(
                    node['children'], section_label, depth + 1
                )
            )
    return rows


def _build_hierarchy_with_je_rows(
    report: Dict[str, Any],
    report_type: str,
    journal_entry_rows: List[Dict[str, Any]],
    template_section_label: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    Build pivot-style rows: each report line (with level), then the journal entries
    that compose that line. JEs are grouped by account_id and attached under the
    report row for that account.
    When template_section_label is set (template-based report), use it instead of
    Revenue/Costs/Expenses or Assets/Liabilities/Equity for the single section with data.
    """
    if report_type == 'income_statement':
        section_headers = _section_headers_income()
        if template_section_label:
            section_headers = [(template_section_label, 'expenses')]
    elif report_type == 'balance_sheet':
        section_headers = _section_headers_balance_sheet()
        if template_section_label:
            section_headers = [(template_section_label, 'assets')]
    else:
        section_headers = _section_headers_cash_flow()
        if template_section_label:
            section_headers = [(template_section_label, 'operating')]

    # Group JEs by account_id for lookup
    je_by_account: Dict[Any, List[Dict]] = {}
    for je in journal_entry_rows or []:
        acc_id = je.get('account_id')
        if acc_id is not None:
            je_by_account.setdefault(acc_id, []).append(je)

    out: List[Dict[str, Any]] = []
    for section_label, key in section_headers:
        nodes = report.get(key) or []
        flat = _flatten_report_tree_with_section(nodes, section_label)
        for r in flat:
            # One row for the report line (row_type = 'Report Line')
            out.append({
                'report_section': r['section'],
                'level': r['depth'],
                'report_line': r['report_line'],
                'account_id': r.get('id'),
                'account_code': r.get('account_code'),
                'account_name': r.get('name'),
                'line_balance': r.get('balance'),
                'row_type': 'Report Line',
                'journal_entry_id': None,
                'transaction_id': None,
                'date': None,
                'transaction_date': None,
                'description': None,
                'debit_amount': None,
                'credit_amount': None,
                'state': None,
                'is_reconciled': None,
            })
            # Rows for each JE that belongs to this account (leaf rows with account_id)
            acc_id = r.get('id')
            if acc_id is not None:
                for je in je_by_account.get(acc_id, []):
                    out.append({
                        'report_section': r['section'],
                        'level': r['depth'],
                        'report_line': r['report_line'],
                        'account_id': je.get('account_id'),
                        'account_code': je.get('account_code'),
                        'account_name': je.get('account_name'),
                        'line_balance': r.get('balance'),
                        'row_type': 'Journal Entry',
                        'journal_entry_id': je.get('journal_entry_id'),
                        'transaction_id': je.get('transaction_id'),
                        'date': je.get('date'),
                        'transaction_date': je.get('transaction_date'),
                        'description': je.get('description') or je.get('transaction_description'),
                        'debit_amount': je.get('debit_amount'),
                        'credit_amount': je.get('credit_amount'),
                        'state': je.get('state'),
                        'is_reconciled': je.get('is_reconciled'),
                    })
    return out


def _report_line_to_section_and_line(report_lines_str: Optional[str]) -> tuple:
    """Split 'Section > Path' into (Report Section, Report Line). For pivot-friendly JE sheet."""
    if not report_lines_str or not str(report_lines_str).strip():
        return ('', '')
    s = str(report_lines_str).strip()
    if ' > ' in s:
        part = s.split(' > ', 1)
        return (part[0].strip(), s)
    return (s, s)


def _collect_calculation_memory_income(report: Dict, request_params: Dict, balance_type: str) -> List[Dict]:
    """Build calculation memory rows for income statement."""
    rows = []
    totals = report.get('totals') or {}
    # Formulas
    rows.append({'section': 'Formulas', 'description': 'Total Revenue = sum of all Revenue section line items', 'value': totals.get('total_revenue')})
    rows.append({'section': 'Formulas', 'description': 'Total Costs (COGS) = sum of all Cost of Goods Sold section line items', 'value': totals.get('total_costs')})
    rows.append({'section': 'Formulas', 'description': 'Gross Profit = Total Revenue - Total Costs', 'value': totals.get('gross_profit')})
    rows.append({'section': 'Formulas', 'description': 'Total Expenses = sum of all Expenses section line items', 'value': totals.get('total_expenses')})
    rows.append({'section': 'Formulas', 'description': 'Net Income = Gross Profit - Total Expenses', 'value': totals.get('net_income')})
    rows.append({'section': 'Parameters', 'description': 'Balance type used for movements', 'value': balance_type})
    rows.append({'section': 'Parameters', 'description': 'Period', 'value': f"{report.get('start_date')} to {report.get('end_date')}"})
    rows.append({'section': 'Parameters', 'description': 'Currency', 'value': (report.get('currency') or {}).get('code') or report.get('currency')})
    return rows


def _collect_calculation_memory_balance_sheet(report: Dict, request_params: Dict, balance_type: str) -> List[Dict]:
    """Build calculation memory rows for balance sheet."""
    rows = []
    totals = report.get('totals') or {}
    rows.append({'section': 'Formulas', 'description': 'Total Assets = sum of all Assets section line items', 'value': totals.get('total_assets')})
    rows.append({'section': 'Formulas', 'description': 'Total Liabilities = sum of all Liabilities section line items', 'value': totals.get('total_liabilities')})
    rows.append({'section': 'Formulas', 'description': 'Total Equity = sum of all Equity section line items', 'value': totals.get('total_equity')})
    rows.append({'section': 'Formulas', 'description': 'Total Liabilities and Equity = Total Liabilities + Total Equity', 'value': totals.get('total_liabilities_and_equity')})
    rows.append({'section': 'Formulas', 'description': 'Check: Total Assets = Total Liabilities and Equity', 'value': 'OK' if abs((totals.get('total_assets') or 0) - (totals.get('total_liabilities_and_equity') or 0)) < 0.01 else 'Mismatch'})
    rows.append({'section': 'Parameters', 'description': 'Balance type used', 'value': balance_type})
    rows.append({'section': 'Parameters', 'description': 'As of date', 'value': report.get('as_of_date')})
    rows.append({'section': 'Parameters', 'description': 'Currency', 'value': (report.get('currency') or {}).get('code') or report.get('currency')})
    return rows


def _collect_calculation_memory_cash_flow(report: Dict, request_params: Dict, balance_type: str) -> List[Dict]:
    """Build calculation memory rows for cash flow."""
    rows = []
    totals = report.get('totals') or {}
    rows.append({'section': 'Formulas', 'description': 'Total Operating = sum of Operating Activities section', 'value': totals.get('total_operating')})
    rows.append({'section': 'Formulas', 'description': 'Total Investing = sum of Investing Activities section', 'value': totals.get('total_investing')})
    rows.append({'section': 'Formulas', 'description': 'Total Financing = sum of Financing Activities section', 'value': totals.get('total_financing')})
    rows.append({'section': 'Formulas', 'description': 'Net Cash Flow = Total Operating + Total Investing + Total Financing', 'value': totals.get('net_cash_flow')})
    rows.append({'section': 'Parameters', 'description': 'Balance type used', 'value': balance_type})
    rows.append({'section': 'Parameters', 'description': 'Period', 'value': f"{report.get('start_date')} to {report.get('end_date')}"})
    rows.append({'section': 'Parameters', 'description': 'Currency', 'value': (report.get('currency') or {}).get('code') if report.get('currency') else report.get('currency')})
    return rows


def _collect_calculation_memory_template(
    report: Dict, request_params: Dict, balance_type: str, report_type: str
) -> List[Dict]:
    """Calculation memory for template-based reports (no Revenue/COGS/Expenses wording)."""
    rows = []
    totals = report.get('totals') or {}
    if report_type == 'income_statement':
        rows.append({'section': 'Parameters', 'description': 'Statement total (net income)', 'value': totals.get('net_income')})
    elif report_type == 'balance_sheet':
        rows.append({'section': 'Parameters', 'description': 'Statement total (liabilities and equity)', 'value': totals.get('total_liabilities_and_equity')})
    else:
        rows.append({'section': 'Parameters', 'description': 'Statement total (net cash flow)', 'value': totals.get('net_cash_flow')})
    rows.append({'section': 'Parameters', 'description': 'Balance type used', 'value': balance_type})
    if report.get('start_date') and report.get('end_date'):
        rows.append({'section': 'Parameters', 'description': 'Period', 'value': f"{report.get('start_date')} to {report.get('end_date')}"})
    if report.get('as_of_date'):
        rows.append({'section': 'Parameters', 'description': 'As of date', 'value': report.get('as_of_date')})
    rows.append({'section': 'Parameters', 'description': 'Currency', 'value': (report.get('currency') or {}).get('code') or (report.get('currency') or {}).get('name') or ''})
    return rows


def _style_header(ws, row_num: int, num_cols: int):
    """Apply header style to a row."""
    thin = Side(style='thin')
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


def _section_headers_income() -> List[tuple]:
    """Section name and key for income statement."""
    return [
        ('Revenue', 'revenues'),
        ('Cost of Goods Sold', 'costs'),
        ('Expenses', 'expenses'),
    ]


def _section_headers_balance_sheet() -> List[tuple]:
    return [
        ('Assets', 'assets'),
        ('Liabilities', 'liabilities'),
        ('Equity', 'equity'),
    ]


def _section_headers_cash_flow() -> List[tuple]:
    return [
        ('Operating Activities', 'operating'),
        ('Investing Activities', 'investing'),
        ('Financing Activities', 'financing'),
    ]


def build_detailed_statement_excel(
    report: Dict[str, Any],
    report_type: str,
    request_params: Dict[str, Any],
    raw_history_rows: List[Dict[str, Any]],
    balance_type: str = 'posted',
    journal_entry_rows: Optional[List[Dict[str, Any]]] = None,
    template_section_label: Optional[str] = None,
    template_lines: Optional[List[Dict[str, Any]]] = None,
) -> bytes:
    """
    Build an Excel workbook for a detailed financial statement.

    report_type: 'income_statement' | 'balance_sheet' | 'cash_flow'
    request_params: the POST body (parent_ids, dates, balance_type, etc.)
    raw_history_rows: list of dicts with keys account_id, account_code, account_name, year, month,
                     posted_total_debit, posted_total_credit, bank_reconciled_*, all_*,
                     balance_type_used, net_movement_used (or ending_balance_used for balance sheet)
    journal_entry_rows: optional list of journal entry dicts (id, transaction_id, date, description,
                        account_id, account_code, account_name, debit_amount, credit_amount, state, is_reconciled,
                        report_lines: semicolon-separated list of report line descriptions where the JE was considered)
    template_section_label: when set (template-based report), use this as the section label instead of
                            Revenue/Costs/Expenses or Assets/Liabilities/Equity for the single section with data.
    template_lines: when set (template-based report), list of dicts with line_number, indent_level, label, is_bold,
                    balance, account_ids. Used to build a "Statement Lines (Debug)" sheet matching markdown wording.
    """
    journal_entry_rows = journal_entry_rows or []
    template_lines = template_lines or []
    wb = Workbook()
    wb.remove(wb.active)

    # ----- Sheet 1: Report Summary -----
    ws_summary = wb.create_sheet('Report Summary', 0)
    if report_type == 'income_statement':
        section_headers = _section_headers_income()
        if template_section_label:
            section_headers = [(template_section_label, 'expenses')]
        total_keys = ['total_revenue', 'total_costs', 'gross_profit', 'total_expenses', 'net_income']
    elif report_type == 'balance_sheet':
        section_headers = _section_headers_balance_sheet()
        if template_section_label:
            section_headers = [(template_section_label, 'assets')]
        total_keys = ['total_assets', 'total_liabilities', 'total_equity', 'total_liabilities_and_equity']
    else:
        section_headers = _section_headers_cash_flow()
        if template_section_label:
            section_headers = [(template_section_label, 'operating')]
        total_keys = ['total_operating', 'total_investing', 'total_financing', 'net_cash_flow']

    totals = report.get('totals') or {}
    row = 1
    ws_summary.cell(row=row, column=1, value='Section')
    ws_summary.cell(row=row, column=2, value='Total')
    _style_header(ws_summary, row, 2)
    row += 1
    for label, key in section_headers:
        ws_summary.cell(row=row, column=1, value=label)
        section_nodes = report.get(key) or []
        section_total = _sum_tree(section_nodes)
        ws_summary.cell(row=row, column=2, value=float(section_total) if section_total is not None else None)
        row += 1
    ws_summary.cell(row=row, column=1, value='Totals')
    _style_header(ws_summary, row, 2)
    row += 1
    if template_section_label:
        # Template-based: do not use Revenue/COGS/Expenses labels; show only statement total
        total_val = totals.get('net_income') if report_type == 'income_statement' else (
            totals.get('total_liabilities_and_equity') if report_type == 'balance_sheet' else totals.get('net_cash_flow')
        )
        ws_summary.cell(row=row, column=1, value='Statement Total')
        ws_summary.cell(row=row, column=2, value=total_val)
    else:
        for k in total_keys:
            ws_summary.cell(row=row, column=1, value=k.replace('_', ' ').title())
            ws_summary.cell(row=row, column=2, value=totals.get(k))
            row += 1
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 18

    # ----- Sheet 2: Report Detail (flattened hierarchy) -----
    ws_detail = wb.create_sheet('Report Detail', 1)
    ws_detail.cell(row=1, column=1, value='Section')
    ws_detail.cell(row=1, column=2, value='Depth')
    ws_detail.cell(row=1, column=3, value='Account ID')
    ws_detail.cell(row=1, column=4, value='Account Code')
    ws_detail.cell(row=1, column=5, value='Name')
    ws_detail.cell(row=1, column=6, value='Path')
    ws_detail.cell(row=1, column=7, value='Balance')
    ws_detail.cell(row=1, column=8, value='Is Leaf')
    _style_header(ws_detail, 1, 8)
    row = 2
    for section_label, key in section_headers:
        nodes = report.get(key) or []
        flat = _flatten_report_tree(nodes)
        for r in flat:
            ws_detail.cell(row=row, column=1, value=section_label)
            ws_detail.cell(row=row, column=2, value=r['depth'])
            ws_detail.cell(row=row, column=3, value=r.get('id'))
            ws_detail.cell(row=row, column=4, value=r.get('account_code'))
            ws_detail.cell(row=row, column=5, value=r.get('name'))
            ws_detail.cell(row=row, column=6, value=r.get('path'))
            ws_detail.cell(row=row, column=7, value=r.get('balance'))
            ws_detail.cell(row=row, column=8, value=r.get('is_leaf'))
            row += 1
    for c in range(1, 9):
        ws_detail.column_dimensions[get_column_letter(c)].width = 18 if c in (5, 6) else 14

    sheet_idx = 2
    # ----- Statement Lines (Debug): same line names as markdown, with debug detail -----
    if template_section_label and template_lines:
        ws_debug = wb.create_sheet('Statement Lines (Debug)', sheet_idx)
        # Same structure as markdown: Line | Label | Balance; plus debug: Indent Level | Label (indented) | Is Bold | Account IDs
        debug_headers = ['Line Number', 'Indent Level', 'Label', 'Label (indented)', 'Is Bold', 'Balance', 'Account IDs']
        for col, h in enumerate(debug_headers, 1):
            ws_debug.cell(row=1, column=col, value=h)
        _style_header(ws_debug, 1, len(debug_headers))
        for r_idx, line_data in enumerate(template_lines, start=2):
            line_num = line_data.get('line_number')
            indent = line_data.get('indent_level', 0) or 0
            label = (line_data.get('label') or '').strip()
            label_indented = ('    ' * indent) + label  # same visual as markdown (4 spaces per level)
            is_bold = line_data.get('is_bold', False)
            balance = line_data.get('balance')
            account_ids = line_data.get('account_ids') or []
            account_ids_str = ', '.join(str(a) for a in account_ids) if isinstance(account_ids, list) else str(account_ids)
            ws_debug.cell(row=r_idx, column=1, value=line_num)
            ws_debug.cell(row=r_idx, column=2, value=indent)
            ws_debug.cell(row=r_idx, column=3, value=label)
            ws_debug.cell(row=r_idx, column=4, value=label_indented)
            ws_debug.cell(row=r_idx, column=5, value=is_bold)
            if balance is not None:
                ws_debug.cell(row=r_idx, column=6, value=float(balance) if isinstance(balance, (Decimal, float)) else balance)
            else:
                ws_debug.cell(row=r_idx, column=6, value=None)
            ws_debug.cell(row=r_idx, column=7, value=account_ids_str)
        for c in range(1, len(debug_headers) + 1):
            ws_debug.column_dimensions[get_column_letter(c)].width = 18 if c in (6, 7) else 14
        ws_debug.column_dimensions['C'].width = 45
        ws_debug.column_dimensions['D'].width = 50
        sheet_idx += 1

    # ----- Report Hierarchy with Journal Entries (pivot-style) -----
    ws_hierarchy_je = wb.create_sheet('Report Hierarchy with Journal Entries', sheet_idx)
    hierarchy_je_rows = _build_hierarchy_with_je_rows(
        report, report_type, journal_entry_rows,
        template_section_label=template_section_label,
    )
    hierarchy_headers = [
        'Report Section', 'Level', 'Report Line', 'Account ID', 'Account Code', 'Account Name',
        'Line Balance', 'Row Type', 'Journal Entry ID', 'Transaction ID', 'Date',
        'Transaction Date', 'Description', 'Debit Amount', 'Credit Amount', 'State', 'Is Reconciled',
    ]
    for col, h in enumerate(hierarchy_headers, 1):
        ws_hierarchy_je.cell(row=1, column=col, value=h.replace('_', ' ').title())
    _style_header(ws_hierarchy_je, 1, len(hierarchy_headers))
    for r_idx, row_data in enumerate(hierarchy_je_rows, start=2):
        for c_idx, key in enumerate(
            [
                'report_section', 'level', 'report_line', 'account_id', 'account_code',
                'account_name', 'line_balance', 'row_type', 'journal_entry_id',
                'transaction_id', 'date', 'transaction_date', 'description',
                'debit_amount', 'credit_amount', 'state', 'is_reconciled',
            ],
            1,
        ):
            val = row_data.get(key)
            if isinstance(val, (Decimal, float)):
                val = float(val) if isinstance(val, Decimal) else val
            ws_hierarchy_je.cell(row=r_idx, column=c_idx, value=val)
    for c in range(1, len(hierarchy_headers) + 1):
        ws_hierarchy_je.column_dimensions[get_column_letter(c)].width = 16
    ws_hierarchy_je.column_dimensions['C'].width = 45
    ws_hierarchy_je.column_dimensions['N'].width = 40
    sheet_idx += 1

    # ----- Calculation Memory -----
    ws_calc = wb.create_sheet('Calculation Memory', sheet_idx)
    if template_section_label:
        # Template-based: use neutral wording (no Revenue/COGS/Expenses)
        calc_rows = _collect_calculation_memory_template(report, request_params, balance_type, report_type)
    elif report_type == 'income_statement':
        calc_rows = _collect_calculation_memory_income(report, request_params, balance_type)
    elif report_type == 'balance_sheet':
        calc_rows = _collect_calculation_memory_balance_sheet(report, request_params, balance_type)
    else:
        calc_rows = _collect_calculation_memory_cash_flow(report, request_params, balance_type)
    ws_calc.cell(row=1, column=1, value='Section')
    ws_calc.cell(row=1, column=2, value='Description')
    ws_calc.cell(row=1, column=3, value='Value')
    _style_header(ws_calc, 1, 3)
    for i, r in enumerate(calc_rows, start=2):
        ws_calc.cell(row=i, column=1, value=r.get('section'))
        ws_calc.cell(row=i, column=2, value=r.get('description'))
        val = r.get('value')
        if isinstance(val, (Decimal, float)):
            ws_calc.cell(row=i, column=3, value=float(val))
        else:
            ws_calc.cell(row=i, column=3, value=str(val) if val is not None else '')
    ws_calc.column_dimensions['A'].width = 18
    ws_calc.column_dimensions['B'].width = 55
    ws_calc.column_dimensions['C'].width = 22
    sheet_idx += 1

    # ----- Request Parameters -----
    ws_params = wb.create_sheet('Request Parameters', sheet_idx)
    ws_params.cell(row=1, column=1, value='Parameter')
    ws_params.cell(row=1, column=2, value='Value')
    _style_header(ws_params, 1, 2)
    row = 2
    for k, v in sorted(request_params.items()):
        ws_params.cell(row=row, column=1, value=k)
        if isinstance(v, (list, tuple)):
            ws_params.cell(row=row, column=2, value=', '.join(str(x) for x in v))
        else:
            ws_params.cell(row=row, column=2, value=str(v) if v is not None else '')
        row += 1
    ws_params.column_dimensions['A'].width = 28
    ws_params.column_dimensions['B'].width = 50

    # ----- Sheet 6: Raw Data - Balance History -----
    ws_raw = wb.create_sheet('Raw Data - Balance History', 5)
    if raw_history_rows:
        headers = list(raw_history_rows[0].keys())
        for col, h in enumerate(headers, 1):
            ws_raw.cell(row=1, column=col, value=h)
        _style_header(ws_raw, 1, len(headers))
        for r_idx, row_data in enumerate(raw_history_rows, start=2):
            for c_idx, h in enumerate(headers, 1):
                val = row_data.get(h)
                if isinstance(val, Decimal):
                    val = float(val)
                ws_raw.cell(row=r_idx, column=c_idx, value=val)
        for c in range(1, len(headers) + 1):
            ws_raw.column_dimensions[get_column_letter(c)].width = 14
    else:
        ws_raw.cell(row=1, column=1, value='No raw balance history rows (leaf accounts may have no history in period).')
    ws_raw.column_dimensions['A'].width = 50
    sheet_idx += 1

    # ----- Journal Entries (pivot-friendly: Report Section, Report Line first) -----
    ws_je = wb.create_sheet('Journal Entries', sheet_idx)
    if journal_entry_rows:
        # Pivot-friendly: Report Section and Report Line first so Excel Pivot Table can group by them
        je_headers = [
            'report_section', 'report_line', 'journal_entry_id', 'transaction_id', 'date',
            'transaction_date', 'description', 'account_id', 'account_code', 'account_name',
            'debit_amount', 'credit_amount', 'state', 'is_reconciled',
        ]
        display_headers = [
            'Report Section', 'Report Line', 'Journal Entry ID', 'Transaction ID', 'Date',
            'Transaction Date', 'Description', 'Account ID', 'Account Code', 'Account Name',
            'Debit Amount', 'Credit Amount', 'State', 'Is Reconciled',
        ]
        for col, h in enumerate(display_headers, 1):
            ws_je.cell(row=1, column=col, value=h)
        _style_header(ws_je, 1, len(display_headers))
        for r_idx, row_data in enumerate(journal_entry_rows, start=2):
            report_lines_str = row_data.get('report_lines')
            section, line = _report_line_to_section_and_line(report_lines_str)
            out_row = {
                'report_section': section,
                'report_line': line or report_lines_str,
                'journal_entry_id': row_data.get('journal_entry_id'),
                'transaction_id': row_data.get('transaction_id'),
                'date': row_data.get('date'),
                'transaction_date': row_data.get('transaction_date'),
                'description': row_data.get('description') or row_data.get('transaction_description'),
                'account_id': row_data.get('account_id'),
                'account_code': row_data.get('account_code'),
                'account_name': row_data.get('account_name'),
                'debit_amount': row_data.get('debit_amount'),
                'credit_amount': row_data.get('credit_amount'),
                'state': row_data.get('state'),
                'is_reconciled': row_data.get('is_reconciled'),
            }
            for c_idx, key in enumerate(je_headers, 1):
                val = out_row.get(key)
                if isinstance(val, (Decimal, float)):
                    val = float(val) if isinstance(val, Decimal) else val
                ws_je.cell(row=r_idx, column=c_idx, value=val)
        for c in range(1, len(display_headers) + 1):
            ws_je.column_dimensions[get_column_letter(c)].width = 16
        ws_je.column_dimensions['B'].width = 50
        ws_je.column_dimensions['G'].width = 40
    else:
        ws_je.cell(row=1, column=1, value='No journal entries in period for the report accounts (or filter excluded them).')
    ws_je.column_dimensions['A'].width = 22
    sheet_idx += 1

    # ----- Leaf Account Summary (per-account totals from raw data) -----
    ws_leaf = wb.create_sheet('Leaf Account Summary', sheet_idx)
    if raw_history_rows:
        def _dec(x):
            return Decimal(str(x)) if x is not None else Decimal('0')
        # Aggregate by account: sum debit, credit, net_movement (use balance_type columns or _used fields)
        by_account: Dict[Any, Dict] = {}
        for r in raw_history_rows:
            acc_id = r.get('account_id')
            key = (acc_id, r.get('account_code'), r.get('account_name'))
            if key not in by_account:
                by_account[key] = {
                    'account_id': acc_id,
                    'account_code': r.get('account_code'),
                    'account_name': r.get('account_name'),
                    'total_debit': Decimal('0'),
                    'total_credit': Decimal('0'),
                    'net_movement': Decimal('0'),
                    'months_count': 0,
                }
            debit = r.get('debit_used') or r.get('posted_total_debit')
            credit = r.get('credit_used') or r.get('posted_total_credit')
            net = r.get('net_movement_used') or r.get('net_movement')
            by_account[key]['total_debit'] += _dec(debit)
            by_account[key]['total_credit'] += _dec(credit)
            by_account[key]['net_movement'] += _dec(net)
            by_account[key]['months_count'] += 1
        leaf_headers = ['Account ID', 'Account Code', 'Account Name', 'Total Debit', 'Total Credit', 'Net Movement', 'Months']
        for col, h in enumerate(leaf_headers, 1):
            ws_leaf.cell(row=1, column=col, value=h)
        _style_header(ws_leaf, 1, len(leaf_headers))
        for r_idx, (_, data) in enumerate(sorted(by_account.items(), key=lambda x: (str(x[1].get('account_code') or ''))), start=2):
            ws_leaf.cell(row=r_idx, column=1, value=data.get('account_id'))
            ws_leaf.cell(row=r_idx, column=2, value=data.get('account_code'))
            ws_leaf.cell(row=r_idx, column=3, value=data.get('account_name'))
            ws_leaf.cell(row=r_idx, column=4, value=float(data['total_debit']))
            ws_leaf.cell(row=r_idx, column=5, value=float(data['total_credit']))
            ws_leaf.cell(row=r_idx, column=6, value=float(data['net_movement']))
            ws_leaf.cell(row=r_idx, column=7, value=data['months_count'])
        for c in range(1, 8):
            ws_leaf.column_dimensions[get_column_letter(c)].width = 16
    else:
        ws_leaf.cell(row=1, column=1, value='No leaf account data (no raw history in period).')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _sum_tree(nodes: List[Dict]) -> Optional[float]:
    """Recursively sum balance of tree nodes."""
    total = None
    for node in nodes or []:
        b = node.get('balance')
        if b is not None:
            total = (total or 0) + float(b)
        if node.get('children'):
            child_sum = _sum_tree(node['children'])
            if child_sum is not None:
                total = (total or 0) + child_sum
    return total


def build_detailed_statement_excel_base64(
    report: Dict[str, Any],
    report_type: str,
    request_params: Dict[str, Any],
    raw_history_rows: List[Dict[str, Any]],
    balance_type: str = 'posted',
    journal_entry_rows: Optional[List[Dict[str, Any]]] = None,
    template_section_label: Optional[str] = None,
    template_lines: Optional[List[Dict[str, Any]]] = None,
) -> str:
    """Same as build_detailed_statement_excel but returns base64-encoded string."""
    data = build_detailed_statement_excel(
        report=report,
        report_type=report_type,
        request_params=request_params,
        raw_history_rows=raw_history_rows,
        balance_type=balance_type,
        journal_entry_rows=journal_entry_rows,
        template_section_label=template_section_label,
        template_lines=template_lines,
    )
    return base64.b64encode(data).decode('ascii')
