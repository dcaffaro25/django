"""Generic queryset → XLSX exporter for tabular list endpoints.

Renders every row in an already-filtered queryset (no pagination) through
its DRF serializer, then dumps it into a single-sheet workbook. Nested
dicts/lists are flattened to strings so the result stays spreadsheet-ish.

Mounted on each list-style ViewSet as ``@action(detail=False) export_xlsx``
so every table gets a "Download XLSX" that reflects the same filters the
caller is currently looking at — not just the visible page.
"""

from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Iterable

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _flatten_value(v: Any) -> Any:
    """Make a value spreadsheet-safe.

    Excel handles str/number/bool/None/date natively. Everything else
    (nested dicts, lists, Decimals) gets coerced — Decimals to float for
    numeric cells; structured types to compact JSON so the row stays
    single-line.
    """
    if v is None:
        return None
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (str, int, float, bool, date, datetime)):
        return v
    if isinstance(v, (list, tuple, set)):
        # Lists of scalars → comma-separated; lists of dicts → JSON
        if all(isinstance(x, (str, int, float, bool)) or x is None for x in v):
            return ", ".join("" if x is None else str(x) for x in v)
        return json.dumps(list(v), ensure_ascii=False, default=str)
    if isinstance(v, dict):
        return json.dumps(v, ensure_ascii=False, default=str)
    return str(v)


def _collect_columns(rows: list[dict]) -> list[str]:
    """Preserve first-row key order, then append any keys that only appear later."""
    seen: dict[str, None] = {}
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen[key] = None
    return list(seen.keys())


def serialized_rows_to_xlsx(
    rows: Iterable[dict],
    *,
    sheet_name: str = "Data",
    filename: str = "export.xlsx",
) -> HttpResponse:
    """Render already-serialized rows (list of dicts) as an xlsx download."""
    wb = Workbook()
    ws = wb.active
    # openpyxl enforces a 31-char sheet name limit; trim safely.
    ws.title = (sheet_name or "Data")[:31]

    rows_list = list(rows)
    columns = _collect_columns(rows_list) if rows_list else []

    # Header row — frozen so large exports stay readable when scrolled.
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"

    for r_idx, row in enumerate(rows_list, start=2):
        for c_idx, key in enumerate(columns, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_flatten_value(row.get(key)))

    # Modest auto-width: don't scan every row on huge exports — sample the
    # first 200 for header + sample-based column width.
    sample = rows_list[:200]
    for c_idx, key in enumerate(columns, start=1):
        width = len(str(key))
        for row in sample:
            raw = _flatten_value(row.get(key))
            if raw is None:
                continue
            width = max(width, min(60, len(str(raw))))
        ws.column_dimensions[get_column_letter(c_idx)].width = max(8, width + 2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(buf.read(), content_type=_XLSX_MIME)
    # RFC 5987: allow UTF-8 filenames for Portuguese names like
    # "plano_de_contas.xlsx". The plain `filename=` fallback strips non-ASCII.
    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"; '
        f"filename*=UTF-8''{filename}"
    )
    return response


def queryset_to_xlsx(
    queryset,
    serializer_class,
    *,
    context: dict | None = None,
    filename: str = "export.xlsx",
    sheet_name: str = "Data",
) -> HttpResponse:
    """Serialize an already-filtered queryset and return an xlsx download.

    The queryset is NOT paginated — callers are expected to have already
    applied `filter_queryset()` (and nothing else). This is the whole point
    of the export: "give me the full filtered result, not just the page
    I'm looking at."
    """
    serializer = serializer_class(queryset, many=True, context=context or {})
    # `serializer.data` may be a ReturnList (list subclass) — fine for iteration.
    return serialized_rows_to_xlsx(
        serializer.data, sheet_name=sheet_name, filename=filename
    )
