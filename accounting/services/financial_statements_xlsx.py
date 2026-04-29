"""Excel exporter for the Demonstrativos page.

Takes the payload produced by ``compute_financial_statements`` and
emits a 4-sheet ``.xlsx``:

  1. **DRE** — formula-driven income statement. Subtotal rows
     (Receita Líquida, Lucro Bruto, EBIT, LAIR, Lucro Líquido) use
     real Excel formulas referencing the lines above, so an operator
     can edit a category cell and see totals recompute.
  2. **Balanço** — Ativo / Passivo + PL columns with formula totals.
  3. **DFC** — direct method, FCO/FCI/FCF sections + per-category
     sub-lines, formula-driven section totals + net change.
  4. **Memória de cálculo** — per-account drill-down for every
     category in DRE, Balanço and DFC. One row per account, sorted
     by category and absolute amount.

Number formatting: ``#,##0.00;(#,##0.00);"-"`` (BR accounting). Sheet
metadata (period, entity, basis, includes-pending) ships in a header
block so the file is self-describing when an auditor opens it next
quarter.
"""
from __future__ import annotations

import io
from decimal import Decimal
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


_NUMBER_FORMAT = '#,##0.00;(#,##0.00);"-"'

_BORDER_THIN = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_BOLD = Font(bold=True)
_SUBTLE = Font(color="6B7280", size=9, italic=True)


def _line(
    ws: Worksheet, row: int, label: str, value=None, *,
    bold: bool = False, indent: int = 0, formula: Optional[str] = None,
):
    """Write a (label, value-or-formula) row. Indent shifts the label
    cell visually; ``bold=True`` styles the row as a subtotal."""
    label_cell = ws.cell(row=row, column=1, value=("    " * indent) + label)
    val_cell = ws.cell(row=row, column=2)
    if formula is not None:
        val_cell.value = formula
    elif value is not None:
        val_cell.value = float(value) if isinstance(value, Decimal) else value
    val_cell.number_format = _NUMBER_FORMAT
    val_cell.alignment = Alignment(horizontal="right")
    if bold:
        label_cell.font = _BOLD
        val_cell.font = _BOLD


def _header_block(ws: Worksheet, title: str, payload: dict):
    """3-row header: title, period subtitle, blank row."""
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    period = payload.get("period") or {}
    df = period.get("date_from") or "—"
    dt = period.get("date_to") or "—"
    basis = "Caixa" if payload.get("basis") == "cash" else "Competência"
    inc = "incl. pendentes" if payload.get("include_pending") else "somente posted"
    ws.cell(
        row=2, column=1,
        value=f"Período: {df} a {dt} · Regime: {basis} · {inc}",
    ).font = _SUBTLE


def _categorise(payload: dict) -> Dict[str, dict]:
    """Index ``payload['categories']`` by key, parse ``amount`` into
    Decimal once so the writers don't repeat that work."""
    out: Dict[str, dict] = {}
    for c in payload.get("categories") or []:
        try:
            amount = Decimal(c.get("amount") or "0")
        except Exception:
            amount = Decimal("0")
        out[c["key"]] = {
            "key": c["key"],
            "label": c.get("label") or c["key"],
            "amount": amount,
            "accounts": c.get("accounts") or [],
        }
    return out


def _write_dre(ws: Worksheet, payload: dict):
    """13-row DRE with formula subtotals.

    Layout (column A = label, column B = value):
        4  Receita Bruta              (data)
        5  (-) Deduções               (data)
        6  Receita Líquida            =B4+B5
        7  (-) Custos                 (data)
        8  Lucro Bruto                =B6+B7
        9  (-) Despesas Operacionais  (data)
       10  EBIT                       =B8+B9
       11  (+) Receitas Financeiras   (data)
       12  (-) Despesas Financeiras   (data)
       13  (+/-) Outras               (data)
       14  Resultado Financeiro       =B11+B12+B13
       15  LAIR                       =B10+B14
       16  (-) IRPJ + CSLL            (data)
       17  Lucro Líquido              =B15+B16
    """
    _header_block(ws, "Demonstração do Resultado (DRE)", payload)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.cell(row=3, column=1, value="Linha").font = _BOLD
    ws.cell(row=3, column=2, value=payload.get("currency") or "BRL").font = _BOLD

    cats = _categorise(payload)
    g = lambda k: cats.get(k, {}).get("amount", Decimal("0"))

    _line(ws, 4,  "Receita Bruta",                       g("receita_bruta"),     bold=True)
    _line(ws, 5,  "(-) Deduções da Receita",             g("deducao_receita"),   indent=1)
    _line(ws, 6,  "Receita Líquida",                     formula="=B4+B5",       bold=True)
    _line(ws, 7,  "(-) Custos",                          g("custo"),             indent=1)
    _line(ws, 8,  "Lucro Bruto",                         formula="=B6+B7",       bold=True)
    _line(ws, 9,  "(-) Despesas Operacionais",           g("despesa_operacional"), indent=1)
    _line(ws, 10, "EBIT (Lucro Operacional)",            formula="=B8+B9",       bold=True)
    _line(ws, 11, "(+) Receitas Financeiras",            g("receita_financeira"), indent=1)
    _line(ws, 12, "(-) Despesas Financeiras",            g("despesa_financeira"), indent=1)
    _line(ws, 13, "(+/-) Outras Receitas/Despesas",      g("outras_receitas"),   indent=1)
    _line(ws, 14, "Resultado Financeiro",                formula="=B11+B12+B13", indent=1)
    _line(ws, 15, "LAIR (Lucro antes IR)",               formula="=B10+B14",     bold=True)
    _line(ws, 16, "(-) IRPJ + CSLL",                     g("imposto_sobre_lucro"), indent=1)
    _line(ws, 17, "Lucro Líquido do Exercício",          formula="=B15+B16",     bold=True)


def _write_balanco(ws: Worksheet, payload: dict):
    """Balance sheet: Ativo column on the left, Passivo + PL on the right.
    Both totals are formula-driven, plus a "Diferença" check row."""
    _header_block(ws, "Balanço Patrimonial", payload)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18

    cats = _categorise(payload)
    g = lambda k: cats.get(k, {}).get("amount", Decimal("0"))

    ws.cell(row=3, column=1, value="ATIVO").font = _BOLD
    _line(ws, 4, "Ativo Circulante",        g("ativo_circulante"))
    _line(ws, 5, "Ativo Não Circulante",    g("ativo_nao_circulante"))
    _line(ws, 6, "Total do Ativo",          formula="=B4+B5",   bold=True)

    ws.cell(row=8, column=1, value="PASSIVO + PATRIMÔNIO LÍQUIDO").font = _BOLD
    _line(ws, 9,  "Passivo Circulante",     g("passivo_circulante"))
    _line(ws, 10, "Passivo Não Circulante", g("passivo_nao_circulante"))
    _line(ws, 11, "Patrimônio Líquido",     g("patrimonio_liquido"))
    _line(ws, 12, "Total Passivo + PL",     formula="=B9+B10+B11", bold=True)

    _line(ws, 14, "Diferença (Ativo − Passivo+PL)", formula="=B6-B12", bold=True)


def _write_dfc(ws: Worksheet, payload: dict):
    """Direct method DFC. One section per FCO/FCI/FCF block, with
    sub-line categories indented and a formula-driven section
    subtotal. Final row computes the net change."""
    _header_block(ws, "Demonstração do Fluxo de Caixa (DFC)", payload)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18

    cashflow = payload.get("cashflow") or {}
    by_section_totals = cashflow.get("by_section") or {}
    by_category = cashflow.get("by_category") or []

    # Bucket categories per section preserving backend order.
    section_rows: Dict[str, List[dict]] = {
        "operacional": [],
        "investimento": [],
        "financiamento": [],
        "no_section": [],
    }
    for c in by_category:
        section_rows.setdefault(c.get("section") or "no_section", []).append(c)

    section_titles = {
        "operacional": "FCO — Atividades Operacionais",
        "investimento": "FCI — Atividades de Investimento",
        "financiamento": "FCF — Atividades de Financiamento",
        "no_section": "Não classificadas",
    }

    row = 3
    ws.cell(row=row, column=1, value="Atividade").font = _BOLD
    ws.cell(row=row, column=2, value=payload.get("currency") or "BRL").font = _BOLD
    row += 1

    section_subtotal_rows: List[int] = []
    for section in ("operacional", "investimento", "financiamento", "no_section"):
        rows = section_rows.get(section) or []
        if not rows:
            continue
        # section header
        section_header_row = row
        ws.cell(row=row, column=1, value=section_titles[section]).font = _BOLD
        row += 1
        # category lines
        category_value_rows: List[int] = []
        for c in rows:
            try:
                amount = Decimal(c.get("amount") or "0")
            except Exception:
                amount = Decimal("0")
            _line(ws, row, c.get("label") or c.get("key") or "?", amount, indent=1)
            category_value_rows.append(row)
            row += 1
        # subtotal
        if category_value_rows:
            sum_range = f"B{category_value_rows[0]}:B{category_value_rows[-1]}"
            _line(
                ws, row, f"Subtotal {section_titles[section]}",
                formula=f"=SUM({sum_range})", bold=True,
            )
            section_subtotal_rows.append(row)
            row += 1
        row += 1  # blank between sections

    # Net change = sum of section subtotals
    if section_subtotal_rows:
        formula = "=" + "+".join(f"B{r}" for r in section_subtotal_rows)
        _line(ws, row, "Variação Líquida do Caixa", formula=formula, bold=True)


def _write_memoria(ws: Worksheet, payload: dict):
    """Per-account drill-down. One row per (statement, category, account).
    Includes DRE/Balanço categories AND DFC sub-categories so the
    auditor can trace any total back to the source accounts in one
    sheet."""
    ws.cell(row=1, column=1, value="Memória de Cálculo").font = Font(bold=True, size=13)
    ws.cell(
        row=2, column=1,
        value=(
            "Cada linha do DRE / Balanço / DFC é a soma das contas listadas "
            "abaixo, agrupadas pela categoria (DRE/Balanço usa "
            "``effective_category``; DFC usa ``effective_cashflow_category``)."
        ),
    ).font = _SUBTLE

    headers = ["Demonstrativo", "Categoria", "Conta (id)", "Conta (nome)", "Valor"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = _BORDER_THIN
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 48
    ws.column_dimensions["E"].width = 18

    row = 5
    # DRE / Balanço categories
    for c in payload.get("categories") or []:
        is_balanco = c["key"] in {
            "ativo_circulante", "ativo_nao_circulante",
            "passivo_circulante", "passivo_nao_circulante",
            "patrimonio_liquido",
        }
        statement = "Balanço" if is_balanco else "DRE"
        for a in c.get("accounts") or []:
            ws.cell(row=row, column=1, value=statement)
            ws.cell(row=row, column=2, value=c.get("label") or c["key"])
            ws.cell(row=row, column=3, value=a.get("id"))
            ws.cell(row=row, column=4, value=a.get("name") or "")
            try:
                v = float(Decimal(a.get("amount") or "0"))
            except Exception:
                v = 0.0
            vc = ws.cell(row=row, column=5, value=v)
            vc.number_format = _NUMBER_FORMAT
            vc.alignment = Alignment(horizontal="right")
            row += 1

    # DFC sub-categories
    cashflow = payload.get("cashflow") or {}
    for c in cashflow.get("by_category") or []:
        for a in c.get("accounts") or []:
            ws.cell(row=row, column=1, value="DFC")
            ws.cell(row=row, column=2, value=c.get("label") or c["key"])
            ws.cell(row=row, column=3, value=a.get("id"))
            ws.cell(row=row, column=4, value=a.get("name") or "")
            try:
                v = float(Decimal(a.get("amount") or "0"))
            except Exception:
                v = 0.0
            vc = ws.cell(row=row, column=5, value=v)
            vc.number_format = _NUMBER_FORMAT
            vc.alignment = Alignment(horizontal="right")
            row += 1


def export_financial_statements_xlsx(payload: dict) -> bytes:
    """Build the workbook for ``payload`` (output of
    ``compute_financial_statements``) and return raw .xlsx bytes."""
    wb = Workbook()
    # Default sheet → DRE
    ws_dre = wb.active
    ws_dre.title = "DRE"
    _write_dre(ws_dre, payload)

    ws_bp = wb.create_sheet("Balanço")
    _write_balanco(ws_bp, payload)

    ws_dfc = wb.create_sheet("DFC")
    _write_dfc(ws_dfc, payload)

    ws_mem = wb.create_sheet("Memória de Cálculo")
    _write_memoria(ws_mem, payload)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
