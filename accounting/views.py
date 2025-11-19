# NORD/accounting/views.py
from __future__ import annotations
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import viewsets, status, response
from rest_framework.decorators import action, api_view
from django.db.models import Q, Sum, Count, F, Value as V
from django.db.models.functions import TruncDate, Coalesce, Cast
from django.db import transaction
from django.utils import timezone
from datetime import timedelta
from itertools import product, combinations
from multitenancy.api_utils import generic_bulk_create, generic_bulk_update, generic_bulk_delete
from multitenancy.utils import resolve_tenant
from multitenancy.models import CustomUser, Company, Entity
from multitenancy.mixins import ScopedQuerysetMixin
from .models import (Currency, Account, Transaction, JournalEntry, Rule, CostCenter, Bank, BankAccount, BankTransaction, Reconciliation, CostCenter,ReconciliationTask, ReconciliationConfig)
from .serializers import (CurrencySerializer, AccountSerializer, TransactionSerializer, CostCenterSerializer, JournalEntrySerializer, JournalEntryListSerializer, RuleSerializer, BankSerializer, BankAccountSerializer, BankTransactionSerializer, ReconciliationSerializer, TransactionListSerializer,ReconciliationTaskSerializer,ReconciliationConfigSerializer)
from .services.transaction_service import *
from .utils import update_journal_entries_and_transaction_flags, parse_ofx_text, decode_ofx_content, generate_ofx_transaction_hash, find_book_combos
from datetime import datetime
import pandas as pd
from decimal import Decimal
from django.db.models import DecimalField, DateField
from django.db.utils import OperationalError
from itertools import product
from networkx.algorithms import bipartite
import networkx as nx
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from django.http import HttpResponse
from bisect import bisect_left, bisect_right
from .tasks import match_many_to_many_task
#from celery.task.control import inspect
from nord_backend.celery import app
from multitenancy.api_utils import _to_bool
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters as drf_filters
from .filters import BankTransactionFilter, TransactionFilter
from django.db import transaction as db_tx
import uuid
import os
from django.db import transaction as db_tx
from rest_framework.decorators import action
from rest_framework.response import Response
from accounting.models import BankTransaction, JournalEntry, Reconciliation
from accounting.services.bank_structs import ensure_pending_bank_structs, ensure_gl_account_for_bank

import logging

from datetime import date
from django.db.models import Prefetch
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status

from accounting.models import Reconciliation, BankTransaction, JournalEntry

# accounting/views_embeddings.py

import os
import time
import requests
from typing import List, Dict, Any, Optional

from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework import status

import logging
from django.db.models import CharField
from django.db.models.functions import Cast, Concat
from celery import current_app
from celery.result import AsyncResult
from django.db.models import Q
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response
from rest_framework import status
from rest_framework.views import APIView

from .models import Account, BankTransaction, Transaction
from .serializers import (
    StartEmbeddingBackfillSerializer,
    TaskIdSerializer,
    TaskStatusSerializer,
)
from .services.embedding_client import EmbeddingClient, _embed_url
from .tasks import generate_missing_embeddings, recalc_unposted_flags_task

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.conf import settings

#from .permissions import HasEmbeddingsApiKey  # or AllowAny for testing if you prefer
from .serializers import EmbedTestSerializer, BackfillSerializer, EmbeddingSearchSerializer
from django.core.cache import cache
from core.models import Job
from core.serializers import JobSerializer
from pgvector.django import CosineDistance

from core.chat.retrieval import _vec_stats, _snippet, _strip_accents, json_safe

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q
from .models import ReconciliationPipeline
from .serializers import (
    ReconciliationPipelineSerializer,
    ResolvedReconciliationPipelineSerializer,
)
import re
from accounting.utils import _normalize_digits, _normalize_raw_digits
from rest_framework import permissions




def _mean_date(dates):
    """Return the average (mean) of a list of date objects, or None."""
    ds = [d for d in dates if d]
    if not ds:
        return None
    avg_ord = sum(d.toordinal() for d in ds) / len(ds)
    # round to nearest day for stability
    return date.fromordinal(int(round(avg_ord)))

log = logging.getLogger("recon")  # or logging.getLogger(__name__)

def _dbg(tag, **k):  # keep logs consistent with the service
    parts = " ".join(f"{kk}={vv}" for kk, vv in k.items())
    log.debug("[%s] %s", tag, parts)


def _info(tag, **k):
    parts = " ".join(f"{kk}={vv}" for kk, vv in k.items())
    log.info("[%s] %s", tag, parts)


def _warn(tag, **k):
    parts = " ".join(f"{kk}={vv}" for kk, vv in k.items())
    log.warning("[%s] %s", tag, parts)

# Currency ViewSet
class CurrencyViewSet(viewsets.ModelViewSet):
    queryset = Currency.objects.all()
    serializer_class = CurrencySerializer
    
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]
    
    @action(methods=['post'], detail=False)
    def bulk_create(self, request, *args, **kwargs):
        return generic_bulk_create(self, request.data)

    @action(methods=['put'], detail=False)
    def bulk_update(self, request, *args, **kwargs):
        return generic_bulk_update(self, request.data)

    @action(methods=['delete'], detail=False)
    def bulk_delete(self, request, *args, **kwargs):
        ids = request.data  # Assuming request.data is a list of IDs
        return generic_bulk_delete(self, ids)

# Account ViewSet
class CostCenterViewSet(ScopedQuerysetMixin, viewsets.ModelViewSet):
    queryset = CostCenter.objects.all()
    serializer_class = CostCenterSerializer
    
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]
    
    @action(methods=['post'], detail=False)
    def bulk_create(self, request, *args, **kwargs):
        return generic_bulk_create(self, request.data)

    @action(methods=['put'], detail=False)
    def bulk_update(self, request, *args, **kwargs):
        return generic_bulk_update(self, request.data)

    @action(methods=['delete'], detail=False)
    def bulk_delete(self, request, *args, **kwargs):
        ids = request.data  # Assuming request.data is a list of IDs
        return generic_bulk_delete(self, ids)

# Account ViewSet
class AccountViewSet(ScopedQuerysetMixin, viewsets.ModelViewSet):
    queryset = Account.objects.all()
    serializer_class = AccountSerializer
    
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]
    
    @action(methods=['post'], detail=False)
    def bulk_create(self, request, *args, **kwargs):
        return generic_bulk_create(self, request.data)

    @action(methods=['put'], detail=False)
    def bulk_update(self, request, *args, **kwargs):
        return generic_bulk_update(self, request.data)

    @action(methods=['delete'], detail=False)
    def bulk_delete(self, request, *args, **kwargs):
        ids = request.data  # Assuming request.data is a list of IDs
        return generic_bulk_delete(self, ids)

class AccountSummaryView(APIView):
    
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, *args, **kwargs):
        company_id = request.query_params.get('company_id')
        entity_id = request.query_params.get('entity_id')
        min_depth = int(request.query_params.get('min_depth', 1))
        include_pending = request.query_params.get('include_pending', 'false').lower() == 'true'
        beginning_date = request.query_params.get('beginning_date')
        end_date = request.query_params.get('end_date')

        if not company_id:
            return Response({"error": "Company ID is required"}, status=status.HTTP_400_BAD_REQUEST)

        accounts_with_balances = Account.get_accounts_summary(
            company_id, entity_id, min_depth, include_pending, beginning_date, end_date
        )
        data = [{'account_code': account.account_code, 'balance': balance} for account, balance in accounts_with_balances]
        return Response(data)

# Account ViewSet
class ReconciliationViewSet(ScopedQuerysetMixin, viewsets.ModelViewSet):
    queryset = Reconciliation.objects.all()
    serializer_class = ReconciliationSerializer
    
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]
    
    # Bulk operations
    @action(methods=['post'], detail=False)
    def bulk_create(self, request, *args, **kwargs):
        return generic_bulk_create(self, request.data)

    @action(methods=['put'], detail=False)
    def bulk_update(self, request, *args, **kwargs):
        return generic_bulk_update(self, request.data)

    @action(methods=['delete'], detail=False)
    def bulk_delete(self, request, *args, **kwargs):
        #data = request.data
        #bank_filters = data.get("bank_filters", {})
        print(request.data)
        ids = request.data if isinstance(request.data, list) else []
    
        if not ids:
            return Response(
                {'detail': 'Provide a non-empty list of IDs to delete.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return generic_bulk_delete(self, ids)
    
    @action(methods=['get'], detail=False, url_path='summaries')
    def summaries(self, request, *args, **kwargs):
        """
        Return reconciliations in a compact, 'matches-like' format.
    
        Query params:
          - status: comma-separated list of statuses (default: matched,approved)
        """
        status_param = request.query_params.get("status", "matched,approved")
        wanted_status = [s.strip() for s in status_param.split(",") if s.strip()]
    
        qs = self.filter_queryset(
            self.get_queryset().filter(status__in=wanted_status)
        )
    
        # Prefetch to avoid N+1 queries
        bank_qs = BankTransaction.objects.select_related(
            "bank_account", "bank_account__entity", "currency"
        )
        book_qs = JournalEntry.objects.select_related(
            "account", "account__bank_account",
            "transaction", "transaction__entity", "transaction__currency"
        )
    
        qs = qs.prefetch_related(
            Prefetch("bank_transactions", queryset=bank_qs),
            Prefetch("journal_entries", queryset=book_qs),
        )
    
        # Pagination support
        page = self.paginate_queryset(qs)
        recs = page if page is not None else qs
    
        results = []
    
        def get_bank_transaction_summary(bank_items):
            lines = []
            for tx in bank_items:
                lines.append(
                    f"ID: {tx.id}, Date: {tx.date}, Amount: {tx.amount}, Desc: {tx.description}"
                )
            return "\n".join(lines)
    
        def get_journal_entries_summary(book_items):
            lines = []
            for entry in book_items:
                account_code = entry.account.account_code if entry.account_id else "N/A"
                account_name = entry.account.name if entry.account_id else "N/A"
                direction = "DEBIT" if entry.debit_amount else "CREDIT"
                eff = entry.get_effective_amount()
                desc = entry.transaction.description if entry.transaction_id else ""
                lines.append(
                    f"ID: {entry.transaction.id if entry.transaction_id else entry.id}, "
                    f"Date: {entry.date}, JE: {direction} {eff} - "
                    f"({account_code}) {account_name}, Desc: {desc}"
                )
            return "\n".join(lines)
    
        for rec in recs:
            banks = list(rec.bank_transactions.all())
            books = list(rec.journal_entries.all())
    
            bank_ids = [b.id for b in banks]
            book_ids = [j.id for j in books]
    
            # --- amounts and sums ---
            bank_sum = Decimal("0")
            bank_amounts: list[float] = []
            bank_dates: list[date] = []
    
            for b in banks:
                if b.amount is not None:
                    bank_sum += b.amount
                    bank_amounts.append(float(b.amount))
                if b.date:
                    bank_dates.append(b.date)
    
            book_sum = Decimal("0")
            book_amounts: list[float] = []
            book_dates: list[date] = []
    
            for je in books:
                eff = je.get_effective_amount()
                if eff is not None:
                    book_sum += eff
                    book_amounts.append(float(eff))
                d = je.date or (je.transaction.date if je.transaction_id and je.transaction else None)
                if d:
                    book_dates.append(d)
    
            # signed difference (bank - book)
            difference = bank_sum - book_sum
    
            # average dates (per side)
            bank_avg = _mean_date(bank_dates)
            book_avg = _mean_date(book_dates)
    
            # overall min/max date considering both bank and books
            all_dates = bank_dates + book_dates
            min_date = min(all_dates) if all_dates else None
            max_date = max(all_dates) if all_dates else None
    
            # descriptions (here we use the detailed multiline summaries)
            bank_description = get_bank_transaction_summary(banks)
            book_description = get_journal_entries_summary(books)
    
            results.append({
                "reconciliation_id": rec.id,
                "bank_ids": bank_ids,
                "book_ids": book_ids,
                "bank_description": bank_description,
                "book_description": book_description,
                "bank_sum_value": float(bank_sum),
                "book_sum_value": float(book_sum),
                "difference": float(difference),
                "bank_amounts": bank_amounts,
                "book_amounts": book_amounts,
                "bank_avg_date": bank_avg.isoformat() if bank_avg else None,
                "book_avg_date": book_avg.isoformat() if book_avg else None,
                "min_date": min_date.isoformat() if min_date else None,
                "max_date": max_date.isoformat() if max_date else None,
                "reference": rec.reference,
                "notes": rec.notes,
            })
    
        if page is not None:
            return self.get_paginated_response(results)
        return Response(results, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=["get"], url_path="export-unreconciled-report")
    def export_unreconciled_report(self, request, *args, **kwargs):
        """
        Export an Excel file with:
          - Report: summary KPIs (counts, totals, date stats, description patterns)
          - Banks_to_reconcile: all unreconciled bank transactions
          - Books_unmatched: all unreconciled journal entries

        Optional query params:
          - company_id (if you are not using tenant_id for scoping)
          - date_from (YYYY-MM-DD)
          - date_to   (YYYY-MM-DD)
          - bank_account_ids=1,2,3 (optional filter)
        """
        # --- 1. Resolve scope (company/tenant) ---
        company_id = request.query_params.get("company_id")
        if not company_id:
            tenant_id = request.query_params.get("tenant_id")
            if not tenant_id:
                return Response(
                    {"detail": "Provide company_id or tenant_id"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            from multitenancy.utils import resolve_tenant
            company_id = resolve_tenant(tenant_id).id

        # Date filters
        date_from_str = request.query_params.get("date_from")
        date_to_str = request.query_params.get("date_to")
        date_from = date_to = None
        try:
            if date_from_str:
                date_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
            if date_to_str:
                date_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
        except ValueError:
            return Response(
                {"detail": "Invalid date_from/date_to, use YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        bank_account_ids_raw = request.query_params.get("bank_account_ids")
        bank_account_ids = []
        if bank_account_ids_raw:
            bank_account_ids = [
                int(x) for x in bank_account_ids_raw.split(",") if x.strip().isdigit()
            ]

        # --- 2. Query unreconciled banks & books ---
        bank_qs = BankTransaction.objects.filter(company_id=company_id).exclude(
            reconciliations__status__in=["matched", "approved"]
        )
        if date_from:
            bank_qs = bank_qs.filter(date__gte=date_from)
        if date_to:
            bank_qs = bank_qs.filter(date__lte=date_to)
        if bank_account_ids:
            bank_qs = bank_qs.filter(bank_account_id__in=bank_account_ids)

        # Unmatched journal entries linked to bank accounts
        book_qs = (
            JournalEntry.objects.filter(company_id=company_id)
            .exclude(reconciliations__status__in=["matched", "approved"])
            .filter(account__bank_account__isnull=False)
            .select_related("transaction", "account", "account__bank_account")
        )
        if date_from:
            book_qs = book_qs.filter(
                Q(date__gte=date_from) | Q(transaction__date__gte=date_from)
            )
        if date_to:
            book_qs = book_qs.filter(
                Q(date__lte=date_to) | Q(transaction__date__lte=date_to)
            )

        bank_rows = list(
            bank_qs.select_related("bank_account", "currency").order_by("date", "id")
        )
        book_rows = list(book_qs.order_by("date", "id"))

        # --- 3. Compute high-level stats ---
        total_banks = len(bank_rows)
        total_books = len(book_rows)

        total_bank_amount = (
            sum((bt.amount or Decimal("0")) for bt in bank_rows)
            if bank_rows
            else Decimal("0")
        )
        total_book_amount = (
            sum((je.get_effective_amount() or Decimal("0")) for je in book_rows)
            if book_rows
            else Decimal("0")
        )

        bank_dates = [bt.date for bt in bank_rows if bt.date]
        book_dates = [
            (je.date or (je.transaction.date if je.transaction else None))
            for je in book_rows
        ]
        bank_dates = [d for d in bank_dates if d]
        book_dates = [d for d in book_dates if d]

        bank_min_date = min(bank_dates) if bank_dates else None
        bank_max_date = max(bank_dates) if bank_dates else None
        book_min_date = min(book_dates) if book_dates else None
        book_max_date = max(book_dates) if book_dates else None

        # Totals by currency
        bank_by_curr: dict[str, Decimal] = {}
        for bt in bank_rows:
            code = getattr(bt.currency, "code", None) or "N/A"
            bank_by_curr.setdefault(code, Decimal("0"))
            bank_by_curr[code] += bt.amount or Decimal("0")

        book_by_curr: dict[str, Decimal] = {}
        for je in book_rows:
            code_obj = getattr(je.transaction, "currency", None)
            ccode = getattr(code_obj, "code", None) if code_obj else "N/A"
            book_amt = je.get_effective_amount() or Decimal("0")
            book_by_curr.setdefault(ccode, Decimal("0"))
            book_by_curr[ccode] += book_amt

        # Simple description patterns: top 10 cleaned descriptions on each side
        bank_desc_counter = Counter()
        for bt in bank_rows:
            bank_desc_counter[clean_description_for_embedding(bt.description)] += 1

        book_desc_counter = Counter()
        for je in book_rows:
            book_desc_counter[clean_description_for_embedding(
                getattr(je.transaction, "description", "") or ""
            )] += 1

        top_bank_desc = bank_desc_counter.most_common(10)
        top_book_desc = book_desc_counter.most_common(10)

        # --- 4. Build Excel workbook ---
        wb = Workbook()
        ws_report = wb.active
        ws_report.title = "Report"

        ws_report["A1"] = "Reconciliation Unreconciled Report"
        ws_report["A2"] = f"Company ID: {company_id}"
        ws_report["A3"] = f"Generated at: {datetime.utcnow().isoformat()}"

        row = 5
        ws_report.cell(row=row, column=1, value="Banks - total records")
        ws_report.cell(row=row, column=2, value=total_banks)
        row += 1
        ws_report.cell(row=row, column=1, value="Banks - total amount")
        ws_report.cell(row=row, column=2, value=float(total_bank_amount))
        row += 1

        ws_report.cell(row=row, column=1, value="Books - total records")
        ws_report.cell(row=row, column=2, value=total_books)
        row += 1
        ws_report.cell(row=row, column=1, value="Books - total amount")
        ws_report.cell(row=row, column=2, value=float(total_book_amount))
        row += 2

        ws_report.cell(row=row, column=1, value="Bank dates (min)")
        ws_report.cell(row=row, column=2, value=str(bank_min_date) if bank_min_date else "")
        row += 1
        ws_report.cell(row=row, column=1, value="Bank dates (max)")
        ws_report.cell(row=row, column=2, value=str(bank_max_date) if bank_max_date else "")
        row += 1
        ws_report.cell(row=row, column=1, value="Book dates (min)")
        ws_report.cell(row=row, column=2, value=str(book_min_date) if book_min_date else "")
        row += 1
        ws_report.cell(row=row, column=1, value="Book dates (max)")
        ws_report.cell(row=row, column=2, value=str(book_max_date) if book_max_date else "")
        row += 2

        ws_report.cell(row=row, column=1, value="Bank totals by currency")
        row += 1
        for code, amt in bank_by_curr.items():
            ws_report.cell(row=row, column=1, value=code)
            ws_report.cell(row=row, column=2, value=float(amt))
            row += 1

        row += 2
        ws_report.cell(row=row, column=1, value="Book totals by currency")
        row += 1
        for code, amt in book_by_curr.items():
            ws_report.cell(row=row, column=1, value=code)
            ws_report.cell(row=row, column=2, value=float(amt))
            row += 1

        row += 2
        ws_report.cell(row=row, column=1, value="Top bank description patterns (cleaned)")
        row += 1
        ws_report.cell(row=row, column=1, value="Description")
        ws_report.cell(row=row, column=2, value="Count")
        row += 1
        for desc, cnt in top_bank_desc:
            ws_report.cell(row=row, column=1, value=desc)
            ws_report.cell(row=row, column=2, value=cnt)
            row += 1

        row += 2
        ws_report.cell(row=row, column=1, value="Top book description patterns (cleaned)")
        row += 1
        ws_report.cell(row=row, column=1, value="Description")
        ws_report.cell(row=row, column=2, value="Count")
        row += 1
        for desc, cnt in top_book_desc:
            ws_report.cell(row=row, column=1, value=desc)
            ws_report.cell(row=row, column=2, value=cnt)
            row += 1

        # --- 5. Banks_to_reconcile sheet ---
        ws_banks = wb.create_sheet(title="Banks_to_reconcile")
        bank_headers = [
            "BankTransactionID",
            "BankAccountID",
            "BankAccountName",
            "Date",
            "Currency",
            "Amount",
            "OriginalDescription",
            "CleanDescriptionForEmbedding",
            "SuggestedNewJournalAccount",
            "SuggestedCostCenter",
            "SuggestedCounterparty",
            "UserNotes",
        ]
        ws_banks.append(bank_headers)

        for bt in bank_rows:
            cur_code = getattr(bt.currency, "code", None)
            acct = getattr(bt, "bank_account", None)
            acct_name = getattr(acct, "name", "") if acct else ""
            cleaned = clean_description_for_embedding(bt.description or "")
            ws_banks.append(
                [
                    bt.id,
                    bt.bank_account_id,
                    acct_name,
                    bt.date.isoformat() if bt.date else "",
                    cur_code,
                    float(bt.amount or 0),
                    bt.description or "",
                    cleaned,
                    "",
                    "",
                    "",
                    "",
                ]
            )

        for col_idx, col in enumerate(ws_banks.columns, start=1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws_banks.column_dimensions[col_letter].width = min(max_len + 2, 60)

        # --- 6. Books_unmatched sheet ---
        ws_books = wb.create_sheet(title="Books_unmatched")
        book_headers = [
            "JournalEntryID",
            "TransactionID",
            "AccountID",
            "AccountName",
            "BankAccountID",
            "Date",
            "Currency",
            "EffectiveAmount",
            "OriginalDescription",
            "CleanDescriptionForEmbedding",
            "UserNotes",
        ]
        ws_books.append(book_headers)

        for je in book_rows:
            tx = getattr(je, "transaction", None)
            acct = getattr(je, "account", None)
            bank_acct = getattr(acct, "bank_account", None) if acct else None
            cur_code_obj = getattr(tx, "currency", None)
            cur_code_val = getattr(cur_code_obj, "code", None) if cur_code_obj else None
            eff_amt = je.get_effective_amount() or Decimal("0")
            desc = getattr(tx, "description", "") or ""
            cleaned = clean_description_for_embedding(desc)
            d = je.date or (tx.date if tx else None)
            ws_books.append(
                [
                    je.id,
                    je.transaction_id,
                    je.account_id,
                    getattr(acct, "name", "") if acct else "",
                    getattr(bank_acct, "id", None) if bank_acct else None,
                    d.isoformat() if d else "",
                    cur_code_val,
                    float(eff_amt),
                    desc,
                    cleaned,
                    "",
                ]
            )

        for col_idx, col in enumerate(ws_books.columns, start=1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws_books.column_dimensions[col_letter].width = min(max_len + 2, 60)

        # --- 7. Return as HTTP response ---
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"reconciliation_unreconciled_{company_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
    
    
# Transaction ViewSet
class TransactionViewSet(ScopedQuerysetMixin, viewsets.ModelViewSet):
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]
    #queryset = Transaction.objects.select_related('company').prefetch_related('journal_entries')
    queryset = (
        Transaction.objects
        .select_related('company', 'currency', 'entity')
        .prefetch_related(
            'journal_entries',
            # add more if needed
        )
    )
    
    filter_backends = [DjangoFilterBackend, drf_filters.SearchFilter, drf_filters.OrderingFilter]
    filterset_class = TransactionFilter
    search_fields = ["description", "entity__name", "journal_entries__account__name"]
    ordering_fields = ["date", "amount", "id", "created_at"]
    ordering = ["-date", "-id"]
    
    def get_serializer_class(self):
        if self.action == 'list' or 'unmatched':
            return TransactionListSerializer
        return TransactionSerializer
    
    @action(detail=False, methods=['get'], url_path='unmatched')
    def unmatched(self, request, tenant_id=None):
        """
        Returns transactions that still have unreconciled journal entries.
        Optional query parameters:
        - company_id: filter by company
        - date_from / date_to: filter by transaction date
        """
        # Base queryset: only transactions with journal entries tied to a bank account
        qs = Transaction.objects.filter(journal_entries__account__bank_account__isnull=False)

        # Apply optional filters
        company_id = request.query_params.get("tenant_id")
        if company_id:
            qs = qs.filter(company__id=company_id)

        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)

        # Exclude any transaction where all bank-side JEs are reconciled (matched or approved)
        matched_recon_tx_ids = Reconciliation.objects.filter(
            status__in=['matched', 'approved'],
            journal_entries__account__bank_account__isnull=False
        ).values_list('journal_entries__transaction_id', flat=True)

        qs_unmatched = qs.exclude(id__in=matched_recon_tx_ids).distinct()

        serializer_class = self.get_serializer_class()  # use list vs detail serializer
        serializer = serializer_class(qs_unmatched, many=True)
        return Response(serializer.data)
    
    
    
    # Bulk operations
    @action(methods=['post'], detail=False)
    def bulk_create(self, request, *args, **kwargs):
        return generic_bulk_create(self, request.data)

    @action(methods=['put'], detail=False)
    def bulk_update(self, request, *args, **kwargs):
        return generic_bulk_update(self, request.data)

    @action(methods=['delete'], detail=False)
    def bulk_delete(self, request, *args, **kwargs):
        ids = request.data
        return generic_bulk_delete(self, ids)

    # Post a transaction
    @action(detail=True, methods=['post'])
    def post(self, request, pk=None):
        transaction = self.get_object()
        try:
            post_transaction(transaction)
            return Response({'status': 'Transaction posted successfully'})
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    # Unpost a transaction
    @action(detail=True, methods=['post'])
    def unpost(self, request, pk=None):
        transaction = self.get_object()
        try:
            unpost_transaction(transaction)
            return Response({'status': 'Transaction unposted successfully'})
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    
    # Cancel a transaction
    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        transaction = self.get_object()
        if transaction.state == 'posted' and not request.user.is_superuser:
            return Response({'error': 'Only superusers can cancel posted transactions'}, status=status.HTTP_403_FORBIDDEN)

        cancel_transaction(transaction)
        return Response({'status': 'Transaction canceled successfully'})
    
    @action(detail=False, methods=['post'], url_path='recalc-unposted-flags-task')
    def recalc_unposted_flags_task(self, request, tenant_id=None, *args, **kwargs):
        """
        Enqueue a Celery task that recomputes is_balanced/is_reconciled flags
        on all unposted transactions.  Returns the Celery task ID.
        """
        async_result = recalc_unposted_flags_task.delay()
        return Response({
            "task_id": async_result.id,
            "status": "queued"
        })
    
    # Automatically create a balancing journal entry
    @action(detail=True, methods=['post'])
    def create_balancing_entry(self, request, pk=None):
        transaction = self.get_object()
        account_id = request.data.get('account_id')
        if not account_id:
            return Response({'error': 'Account ID is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            balancing_account = Account.objects.get(pk=account_id, company=transaction.company)
        except Account.DoesNotExist:
            return Response({'error': 'Invalid account ID'}, status=status.HTTP_400_BAD_REQUEST)

        journal_entry = create_balancing_journal_entry(transaction, balancing_account)
        if not journal_entry:
            return Response({'message': 'Transaction is already balanced'})

        return Response({'status': 'Balancing journal entry created', 'entry': JournalEntrySerializer(journal_entry).data})

    # Filter transactions by company, status, and balance
    @action(detail=False, methods=['get'])
    def filtered(self, request, tenant_id=None):
        """
        Filters transactions based on query parameters and tenant_id.
        """
        # Extract query parameters
        #company_id = request.query_params.get('company_id')
        status_filter = request.query_params.get('status')
        min_balance = request.query_params.get('min_balance')
        max_balance = request.query_params.get('max_balance')

        # Base queryset
        queryset = self.get_queryset().filter(company__subdomain=tenant_id)  # Use tenant_id here

        # Apply additional filters
        #if company_id:
        #    queryset = queryset.filter(company_id=company_id)
        if status_filter:
            if status_filter != "Todos" and status_filter != "*" and status_filter != "":
                queryset = queryset.filter(state=status_filter)
        if min_balance is not None:
            queryset = queryset.annotate(balance=Sum('journal_entries__debit_amount') - Sum('journal_entries__credit_amount'))
            queryset = queryset.filter(balance__gte=min_balance)
        if max_balance is not None:
            queryset = queryset.annotate(balance=Sum('journal_entries__debit_amount') - Sum('journal_entries__credit_amount'))
            queryset = queryset.filter(balance__lte=max_balance)

        # Serialize and return results
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    

class JournalEntryViewSet(ScopedQuerysetMixin, viewsets.ModelViewSet):
    #queryset = JournalEntry.objects.all()
    queryset = (
        JournalEntry.objects
        .select_related('company', 'account', 'transaction')
        # If you also need cost_center or any other FK, include it too
    )
    serializer_class = JournalEntrySerializer
    
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]
    
    
    def get_serializer_class(self):
        if self.action in ['list', 'unmatched']:
            return JournalEntryListSerializer
        return JournalEntrySerializer
    
    # Bulk operations
    @action(methods=['post'], detail=False)
    def bulk_create(self, request, *args, **kwargs):
        return generic_bulk_create(self, request.data)

    @action(methods=['put'], detail=False)
    def bulk_update(self, request, *args, **kwargs):
        return generic_bulk_update(self, request.data)

    @action(methods=['delete'], detail=False)
    def bulk_delete(self, request, *args, **kwargs):
        ids = request.data  # Assuming request.data is a list of IDs
        return generic_bulk_delete(self, ids)
    
    @action(detail=False, methods=['get'], url_path='unmatched')
    def unmatched(self, request, tenant_id=None):
        """
        Returns transactions that still have unreconciled journal entries.
        Optional query parameters:
        - company_id: filter by company
        - date_from / date_to: filter by transaction date
        """
        # Base queryset: only transactions with journal entries tied to a bank account
        
        qs = JournalEntry.objects.filter(account__bank_account__isnull=False)
        qs = qs.select_related('transaction', 'account', 'company').prefetch_related('reconciliations')
        # Apply optional filters
        company_id = request.query_params.get("tenant_id")
        if company_id:
            qs = qs.filter(company__id=company_id)

        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)

        # Exclude any transaction where all bank-side JEs are reconciled (matched or approved)
        matched_je_ids = (
            Reconciliation.objects
            .filter(status__in=['matched', 'approved'])
            .values_list('journal_entries__id', flat=True)
        )
        
        qs_unmatched = qs.exclude(id__in=matched_je_ids).distinct()
        serializer_class = self.get_serializer_class()  # use list vs detail serializer
        serializer = serializer_class(qs_unmatched, many=True)
        return Response(serializer.data)
    
    # Get journal entries by transaction
    @action(detail=False, methods=['get'])
    def by_transaction(self, request):
        transaction_id = request.query_params.get('transaction_id')
        if not transaction_id:
            return Response({'error': 'Transaction ID is required'}, status=status.HTTP_400_BAD_REQUEST)

        journal_entries = self.queryset.filter(transaction_id=transaction_id)
        serializer = self.get_serializer(journal_entries, many=True)
        return Response(serializer.data)

    # Filter journal entries by status or company
    @action(detail=False, methods=['get'])
    def filtered(self, request):
        status_filter = request.query_params.get('status', None)
        company_id = request.query_params.get('tenant_id', None)

        filters = Q()
        if status_filter:
            filters &= Q(state=status_filter)
        if company_id:
            filters &= Q(transaction__company_id=company_id)

        journal_entries = self.queryset.filter(filters)
        serializer = self.get_serializer(journal_entries, many=True)
        return Response(serializer.data)

    
# Rule ViewSet
class RuleViewSet(ScopedQuerysetMixin, viewsets.ModelViewSet):
    queryset = Rule.objects.all()
    serializer_class = RuleSerializer
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]
        
    @action(methods=['post'], detail=False)
    def bulk_create(self, request, *args, **kwargs):
        return generic_bulk_create(self, request.data)

    @action(methods=['put'], detail=False)
    def bulk_update(self, request, *args, **kwargs):
        return generic_bulk_update(self, request.data)

    @action(methods=['delete'], detail=False)
    def bulk_delete(self, request, *args, **kwargs):
        ids = request.data  # Assuming request.data is a list of IDs
        return generic_bulk_delete(self, ids)
    
# Bank ViewSet
class BankViewSet(viewsets.ModelViewSet):
    queryset = Bank.objects.all()
    serializer_class = BankSerializer
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]
        
    @action(methods=['post'], detail=False)
    def bulk_create(self, request, *args, **kwargs):
        return generic_bulk_create(self, request.data)

    @action(methods=['put'], detail=False)
    def bulk_update(self, request, *args, **kwargs):
        return generic_bulk_update(self, request.data)

    @action(methods=['delete'], detail=False)
    def bulk_delete(self, request, *args, **kwargs):
        ids = request.data  # Assuming request.data is a list of IDs
        return generic_bulk_delete(self, ids)



# BankAccount ViewSet
class BankAccountViewSet(ScopedQuerysetMixin, viewsets.ModelViewSet):
    queryset = BankAccount.objects.all()
    serializer_class = BankAccountSerializer
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]
        
    @action(methods=['post'], detail=False)
    def bulk_create(self, request, *args, **kwargs):
        return generic_bulk_create(self, request.data)

    @action(methods=['put'], detail=False)
    def bulk_update(self, request, *args, **kwargs):
        return generic_bulk_update(self, request.data)

    @action(methods=['delete'], detail=False)
    def bulk_delete(self, request, *args, **kwargs):
        ids = request.data  # Assuming request.data is a list of IDs
        return generic_bulk_delete(self, ids)
    
# BankTransaction ViewSet
class BankTransactionViewSet(ScopedQuerysetMixin, viewsets.ModelViewSet):
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]
        
    queryset = (
        BankTransaction.objects
        .select_related("bank_account", "bank_account__entity", "currency")  # add "bank_account__bank" if you render it
        .order_by("-date", "-id")
    )
    serializer_class = BankTransactionSerializer

    filter_backends = [DjangoFilterBackend, drf_filters.SearchFilter, drf_filters.OrderingFilter]
    filterset_class = BankTransactionFilter
    search_fields = [
        "description",
        "reference_number",
        "bank_account__name",
        "bank_account__account_number",
        "bank_account__entity__name",  # <-- fixed
    ]
    ordering_fields = ["date", "amount", "id", "created_at"]
    ordering = ["-date", "-id"]

    # If your mixin needs to know how to scope by entity:
    entity_lookup = "bank_account__entity"  # <-- only if your ScopedQuerysetMixin uses this

    def get_queryset(self):
        qs = super().get_queryset()

        # /entities/<entity_id>/... or ?entity_id=...
        entity_id = self.kwargs.get("entity_id") or self.request.query_params.get("entity_id")
        if entity_id:
            qs = qs.filter(bank_account__entity_id=entity_id)

        bank_account_id = self.request.query_params.get("bank_account")
        if bank_account_id:
            qs = qs.filter(bank_account_id=bank_account_id)

        # accept ?status=... or ?status=PENDING,STARTED
        status_param = self.request.query_params.get("status")
        if status_param:
            if "," in status_param:
                qs = qs.filter(status__in=[s.strip() for s in status_param.split(",") if s.strip()])
            else:
                qs = qs.filter(status=status_param)

        return qs
    
    @action(methods=['post'], detail=False)
    def bulk_create(self, request, *args, **kwargs):
        return generic_bulk_create(self, request.data)

    @action(methods=['put'], detail=False)
    def bulk_update(self, request, *args, **kwargs):
        return generic_bulk_update(self, request.data)

    @action(methods=['delete'], detail=False)
    def bulk_delete(self, request, *args, **kwargs):
        ids = request.data  # Assuming request.data is a list of IDs
        return generic_bulk_delete(self, ids)
    
    def _scan_ofx_files(self, files_data):
        """
        Scan OFX files and classify each transaction as duplicate or pending.
        This helper does not insert anything into the database; it only inspects
        the files to inform the user which transactions are duplicates.
    
        Returns a list of dictionaries like:
        {
          "filename": <optional filename>,
          "bank": {... bank lookup info ...},
          "account": {... account lookup info ...},
          "inserted": 0,     # always zero in scan
          "duplicates": <num dupes>,
          "duplicate_ratio": <float 0–1>,
          "warning": <str or None>,
          "transactions": [
            {
              "date": "YYYY-MM-DD",
              "amount": 123.45,
              "transaction_type": "...",
              "description": "...",
              "tx_hash": "...",
              "status": "duplicate" | "pending"
            },
            ...
          ],
        }
        """
        results = []
        # local helper to strip non-digits / leading zeros
        def _digits_no_lz(v):
            import re as _re
            s = _re.sub(r"\D", "", (v or ""))
            return s.lstrip("0")
    
        for idx, file_item in enumerate(files_data):
            summary = {
                "index": idx,
                "filename": file_item.get("name"),
                "inserted": 0,
                "duplicates": 0,
                "duplicate_ratio": 0.0,
                "warning": None,
                "transactions": [],
            }
            # decode OFX payload
            ofx_content = decode_ofx_content(file_item)
            if not ofx_content:
                summary["warning"] = "No valid ofx_text or base64Data found."
                results.append(summary)
                continue
            # parse OFX
            try:
                parsed = parse_ofx_text(ofx_content)
            except Exception as exc:
                summary["warning"] = f"Failed to parse OFX: {exc}"
                results.append(summary)
                continue
    
            bank_code = parsed.get("bank_code")
            account_id = parsed.get("account_id")
            txns = parsed.get("transactions", [])
            # find bank/account for context (like in original import_ofx)
            bank_obj = Bank.objects.filter(bank_code=bank_code).first() if bank_code else None
            # default fields
            summary["bank"] = {
                "result": "Success" if bank_obj else "Error",
                "message": ("Bank found." if bank_obj else f"Bank code '{bank_code}' not found."),
                "value": BankSerializer(bank_obj).data if bank_obj else bank_code,
            }
            account_obj = None
            if bank_obj:
                # match by exact account number or by branch+account combination
                qs = BankAccount.objects.filter(bank__bank_code=bank_code)
                account_obj = qs.filter(account_number=account_id).first()
                if not account_obj and account_id:
                    norm = _digits_no_lz(account_id)
                    account_obj = (
                        qs.annotate(
                            branch_s=Cast("branch_id", CharField()),
                            acct_s=Cast("account_number", CharField()),
                            ba_concat=Concat("branch_s", "acct_s", output_field=CharField()),
                        )
                        .filter(ba_concat=norm)
                        .first()
                    )
            summary["account"] = {
                "result": "Success" if account_obj else "Error",
                "message": ("BankAccount found." if account_obj else f"BankAccount '{account_id}' not found."),
                "value": BankAccountSerializer(account_obj).data if account_obj else account_id,
            }
    
            dup_count = 0
            for tx in txns:
                tx_date_str = tx.get("date")
                try:
                    tx_date = datetime.strptime(tx_date_str, "%Y-%m-%d").date()
                except Exception:
                    tx_date = None
                tx_amount = tx.get("amount", 0.0)
                tx_type = tx.get("transaction_type", "")
                tx_desc = tx.get("memo", "")
                # build tx_hash using same fields as in create
                tx_hash = generate_ofx_transaction_hash(
                    date_str=tx_date_str or "",
                    amount=tx_amount,
                    transaction_type=tx_type,
                    memo=tx_desc,
                    bank_number=bank_code,
                    account_number=account_id,
                )
                exists = BankTransaction.objects.filter(tx_hash=tx_hash).exists()
                if exists:
                    dup_count += 1
                summary["transactions"].append({
                    "date": tx_date_str,
                    "amount": tx_amount,
                    "transaction_type": tx_type,
                    "description": tx_desc,
                    "tx_hash": tx_hash,
                    "status": "duplicate" if exists else "pending",
                })
    
            total = len(txns)
            summary["duplicates"] = dup_count
            summary["duplicate_ratio"] = (dup_count / total) if total else 0.0
            if summary["duplicate_ratio"] > 0.8:
                summary["warning"] = "More than 80% of transactions are duplicates; this file was likely imported before."
            results.append(summary)
    
        return results
    
    @action(detail=False, methods=['post'])
    def import_ofx(self, request, *args, **kwargs):
        """
        Analyze OFX files and return per-file summaries.
        No database writes happen here.  Each transaction is labeled as
        "duplicate" if its tx_hash is already present in the database, otherwise
        "pending".  A file-level warning is included if most transactions
        appear to be duplicates.  The result is purely informational.
        """
        files_data = request.data.get("files")
        if not files_data or not isinstance(files_data, list):
            return Response({"error": "Please provide 'files' as a list."},
                            status=status.HTTP_400_BAD_REQUEST)
    
        import_results = self._scan_ofx_files(files_data)
        return Response({"import_results": import_results}, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['post'])
    def import_ofx_transactions(self, request, *args, **kwargs):
        """
        Persist OFX transactions according to an import policy:
          - 'records' (default): Import all files but only create DB rows for
            transactions marked 'pending' (non-duplicates).  Files with high
            duplicate ratios are still processed, but duplicates are skipped.
          - 'files': Skip any file whose duplicate ratio exceeds 80% and only
            insert non-duplicate records from the remaining files.
    
        The endpoint accepts the same 'files' list as import_ofx.  It returns
        the scan summary (with duplicates marked) plus insertion counts and a
        warning if entire files were skipped due to the 'files' policy.
        """
        files_data = request.data.get("files")
        if not files_data or not isinstance(files_data, list):
            return Response({"error": "Please provide 'files' as a list."},
                            status=status.HTTP_400_BAD_REQUEST)
    
        import_policy = request.data.get("import_policy", "files").strip().lower()
        if import_policy not in ("records", "files"):
            return Response({"error": "import_policy must be 'records' or 'files'"},
                            status=status.HTTP_400_BAD_REQUEST)
    
        # First scan files to classify duplicates
        scanned = self._scan_ofx_files(files_data)
        duplicate_threshold = 0.80
        results = []
    
        for file_summary in scanned:
            # Make a copy so we can append insertion results
            out = dict(file_summary)
            out["inserted"] = 0  # track new inserts
            
            # If policy='files' and duplicate_ratio > threshold, skip the file
            if (
                import_policy == "files"
                and file_summary.get("duplicate_ratio", 0) > duplicate_threshold
            ):
                out["warning"] = (
                    out.get("warning")
                    or "File skipped because more than 80% of transactions were duplicates."
                )
                results.append(out)
                continue

            # Otherwise, insert non-duplicate transactions
            tx_infos = file_summary.get("transactions", []) or []

            # Extract raw bank/account info from scan summary
            bank_value = file_summary.get("bank", {}).get("value")
            account_value = file_summary.get("account", {}).get("value")

            raw_bank_code = (
                bank_value.get("bank_code")
                if isinstance(bank_value, dict)
                else None
            )
            raw_account_num = (
                account_value.get("account_number")
                if isinstance(account_value, dict)
                else None
            )

            # Normalize codes
            bank_code_norm = _normalize_digits(raw_bank_code)        # '0237' -> '237'
            bank_code_raw = _normalize_raw_digits(raw_bank_code)     # '0237' -> '0237'
            account_num_norm = _normalize_raw_digits(raw_account_num)  # keep all significant digits

            # Find Bank
            bank_obj = None
            if bank_code_norm or bank_code_raw:
                bank_code_candidates = {c for c in (bank_code_norm, bank_code_raw) if c}
                bank_obj = Bank.objects.filter(bank_code__in=bank_code_candidates).first()

            # Find BankAccount
            bank_acct_obj = None
            if bank_obj and account_num_norm:
                qs = BankAccount.objects.filter(bank=bank_obj)

                # First try direct account_number match
                bank_acct_obj = qs.filter(account_number=account_num_norm).first()

                if not bank_acct_obj:
                    # Fallback: branch + account concatenation (normalized)
                    norm = re.sub(r"\D", "", account_num_norm)
                    bank_acct_obj = (
                        qs.annotate(
                            branch_s=Cast("branch_id", CharField()),
                            acct_s=Cast("account_number", CharField()),
                            ba_concat=Concat("branch_s", "acct_s", output_field=CharField()),
                        )
                        .filter(ba_concat=norm)
                        .first()
                    )

            # derive currency and company_id
            currency_obj = None
            company_id = None
            if bank_acct_obj:
                currency_obj = getattr(bank_acct_obj, "currency", None)
                company_id = getattr(bank_acct_obj, "company_id", None)
                if company_id is None and getattr(bank_acct_obj, "entity", None):
                    company_id = getattr(bank_acct_obj.entity, "company_id", None)
            elif bank_obj:
                currency_obj = getattr(bank_obj, "default_currency", None)

            # Insert transactions
            with db_tx.atomic():
                for tx_info in tx_infos:
                    if tx_info.get("status") != "pending":
                        continue

                    raw_date = tx_info.get("date")
                    parsed_date = None
                    if raw_date:
                        try:
                            parsed_date = datetime.strptime(raw_date, "%Y-%m-%d").date()
                        except Exception:
                            parsed_date = None

                    tx_amount = tx_info.get("amount")
                    tx_desc = tx_info.get("description")
                    tx_hash = tx_info.get("tx_hash")

                    try:
                        BankTransaction.objects.create(
                            company_id=company_id,
                            bank_account=bank_acct_obj,
                            date=parsed_date,
                            amount=Decimal(str(tx_amount)) if tx_amount is not None else None,
                            description=tx_desc,
                            currency=currency_obj,
                            status="pending",
                            tx_hash=tx_hash,
                        )
                        out["inserted"] += 1
                        tx_info["status"] = "inserted"
                    except Exception as ex:
                        tx_info["status"] = f"error: {ex}"

            results.append(out)

        return Response({"results": results}, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def unreconciled(self, request, tenant_id):
        unreconciled_transactions = BankTransaction.objects.filter(reconciliations__isnull=True)
        serializer = BankTransactionSerializer(unreconciled_transactions, many=True)
        return response.Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def finalize_reconciliation_matches(self, request, *args, **kwargs):
        """
        Finalize a batch of reconciliation matches (optionally creating an adjustment on one side).
    
        Optimised version:
          - no giant transaction; each DB write uses autocommit
          - batch pre-load of BankTransaction / JournalEntry
          - no global select_for_update() locks
          - single bulk_update() for promotion
        """
        request_id = str(uuid.uuid4())
    
        data = request.data
        matches = data.get("matches", [])
        adjustment_side = data.get("adjustment_side", "none")  # "bank" | "journal" | "none"
        reference = data.get("reference", "")
        notes = data.get("notes", "")
    
        _info(
            "finalize_begin",
            request_id=request_id,
            n_matches=len(matches),
            adjustment_side=adjustment_side,
            reference=reference or "",
            notes_len=len(notes or ""),
        )
    
        created_records: list[dict] = []
        problems: list[dict] = []
    
        # ------------------------------------------------------------------
        # 1) Pre-collect all IDs for batching
        # ------------------------------------------------------------------
        all_bank_ids: set[int] = set()
        all_journal_ids: set[int] = set()
        for m in matches:
            raw_bank = m.get("bank_transaction_ids", []) or []
            raw_journal = m.get("journal_entry_ids", []) or []
            all_bank_ids.update(int(x) for x in raw_bank if x is not None)
            all_journal_ids.update(int(x) for x in raw_journal if x is not None)
    
        if not all_bank_ids and not all_journal_ids:
            return Response({"created": [], "problems": [{"reason": "no_ids"}]})
    
        # ------------------------------------------------------------------
        # 2) Pre-load already reconciled IDs in TWO queries
        # ------------------------------------------------------------------
        already_bank_ids = set(
            BankTransaction.objects.filter(
                id__in=all_bank_ids,
                reconciliations__status__in=["matched", "approved"],
            ).values_list("id", flat=True)
        )
        already_journal_ids = set(
            JournalEntry.objects.filter(
                id__in=all_journal_ids,
                reconciliations__status__in=["matched", "approved"],
            ).values_list("id", flat=True)
        )
    
        _dbg(
            "preload_already",
            request_id=request_id,
            n_bank_ids=len(all_bank_ids),
            n_journal_ids=len(all_journal_ids),
            already_banks=len(already_bank_ids),
            already_journals=len(already_journal_ids),
        )
    
        # ------------------------------------------------------------------
        # 3) Pre-load all referenced Bank / Journal rows (NO locking)
        # ------------------------------------------------------------------
        bank_map: dict[int, BankTransaction] = BankTransaction.objects.in_bulk(all_bank_ids)
    
        journal_map: dict[int, JournalEntry] = {
            je.id: je
            for je in JournalEntry.objects.filter(id__in=all_journal_ids)
            .select_related("transaction", "account", "account__bank_account", "cost_center")
        }
    
        _dbg(
            "preload_maps",
            request_id=request_id,
            banks_loaded=len(bank_map),
            journals_loaded=len(journal_map),
        )
    
        # ------------------------------------------------------------------
        # 4) Caches for expensive helpers and per-batch usage tracking
        # ------------------------------------------------------------------
        pending_structs_cache: dict[tuple[int, int], tuple] = {}   # (company, currency) -> (pending_ba, pending_gl)
        gl_cache: dict[tuple[int, int], object] = {}               # (company, bank_account_id) -> gl_account
    
        # journal entries to bulk-update at the end
        jes_to_update: list[JournalEntry] = []
    
        # avoid reusing same IDs in the same batch
        used_bank_ids_batch: set[int] = set()
        used_journal_ids_batch: set[int] = set()
    
        # ------------------------------------------------------------------
        # 5) Process each match
        # ------------------------------------------------------------------
        for idx, match in enumerate(matches, start=1):
            try:
                raw_bank = match.get("bank_transaction_ids", []) or []
                raw_journal = match.get("journal_entry_ids", []) or []
                bank_ids = list({int(x) for x in raw_bank if x is not None})
                journal_ids = list({int(x) for x in raw_journal if x is not None})
    
                if not bank_ids or not journal_ids:
                    problems.append({"reason": "empty_ids", "match": match})
                    _warn("match_skip_empty_ids", request_id=request_id, i=idx)
                    continue
    
                # Skip if used already in this batch
                if any(bid in used_bank_ids_batch for bid in bank_ids) or any(
                    jid in used_journal_ids_batch for jid in journal_ids
                ):
                    problems.append(
                        {
                            "reason": "overlap_in_batch",
                            "bank_ids": bank_ids,
                            "journal_ids": journal_ids,
                        }
                    )
                    _warn("match_skip_overlap_batch", request_id=request_id, i=idx)
                    continue
    
                # Fast in-memory "already reconciled" check
                if any(bid in already_bank_ids for bid in bank_ids) or any(
                    jid in already_journal_ids for jid in journal_ids
                ):
                    problems.append(
                        {
                            "reason": "already_reconciled",
                            "bank_ids": bank_ids,
                            "journal_ids": journal_ids,
                        }
                    )
                    _warn("match_skip_already_reconciled", request_id=request_id, i=idx)
                    continue
    
                # Get objects from pre-loaded maps
                bank_txs = [bank_map.get(bid) for bid in bank_ids if bid in bank_map]
                journal_entries = [
                    journal_map.get(jid) for jid in journal_ids if jid in journal_map
                ]
    
                bank_txs = [bt for bt in bank_txs if bt is not None]
                journal_entries = [je for je in journal_entries if je is not None]
    
                if not bank_txs or not journal_entries:
                    problems.append(
                        {
                            "reason": "not_found",
                            "bank_ids": bank_ids,
                            "journal_ids": journal_ids,
                        }
                    )
                    _warn("match_skip_not_found", request_id=request_id, i=idx)
                    continue
    
                company_set = {bt.company_id for bt in bank_txs} | {
                    je.company_id for je in journal_entries
                }
                currency_set = {bt.currency_id for bt in bank_txs} | {
                    je.transaction.currency_id
                    for je in journal_entries
                    if je.transaction_id
                }
    
                if len(company_set) != 1:
                    problems.append(
                        {
                            "reason": "mixed_company",
                            "bank_ids": bank_ids,
                            "journal_ids": journal_ids,
                        }
                    )
                    _warn(
                        "match_skip_mixed_company",
                        request_id=request_id,
                        i=idx,
                        companies=list(company_set),
                    )
                    continue
                company_id = next(iter(company_set))
    
                if len(currency_set) != 1:
                    problems.append(
                        {
                            "reason": "mixed_currency",
                            "bank_ids": bank_ids,
                            "journal_ids": journal_ids,
                        }
                    )
                    _warn(
                        "match_skip_mixed_currency",
                        request_id=request_id,
                        i=idx,
                        currencies=list(currency_set),
                    )
                    continue
                currency_id = next(iter(currency_set))
    
                # Cached pending structs per (company, currency)
                pend_key = (company_id, currency_id)
                if pend_key in pending_structs_cache:
                    pending_ba, pending_gl = pending_structs_cache[pend_key]
                else:
                    pending_ba, pending_gl = ensure_pending_bank_structs(
                        company_id, currency_id=currency_id
                    )
                    pending_structs_cache[pend_key] = (pending_ba, pending_gl)
    
                sum_bank = sum((bt.amount for bt in bank_txs), Decimal("0"))
                sum_journal = sum(
                    ((je.get_effective_amount() or Decimal("0")) for je in journal_entries),
                    Decimal("0"),
                )
                diff = sum_bank - sum_journal
    
                bank_account_set = {bt.bank_account_id for bt in bank_txs if bt.bank_account_id}
    
                # Adjustment logic (same as before, just no savepoints)
                if adjustment_side != "none" and diff != Decimal("0"):
                    if adjustment_side == "bank":
                        if len(bank_account_set) != 1:
                            problems.append(
                                {
                                    "reason": "cannot_adjust_bank_with_multiple_accounts",
                                    "bank_account_ids": list(bank_account_set),
                                    "bank_ids": bank_ids,
                                    "journal_ids": journal_ids,
                                }
                            )
                            _warn(
                                "adjust_skip_multi_ba",
                                request_id=request_id,
                                i=idx,
                                bank_accounts=list(bank_account_set),
                            )
                        else:
                            adjustment_amount = sum_journal - sum_bank
                            bt0 = bank_txs[0]
                            _dbg(
                                "adjust_bank_create",
                                request_id=request_id,
                                i=idx,
                                amount=str(adjustment_amount),
                            )
                            adj_bt = BankTransaction.objects.create(
                                company_id=company_id,
                                bank_account=bt0.bank_account,
                                date=bt0.date,
                                currency=bt0.currency,
                                amount=adjustment_amount,
                                description="Adjustment record for reconciliation",
                                status="pending",
                                tx_hash="adjustment_for_rec",
                            )
                            bank_txs.append(adj_bt)
                            sum_bank += adjustment_amount
    
                    elif adjustment_side == "journal":
                        adjustment_amount = sum_bank - sum_journal
                        je0 = journal_entries[0]
                        debit_amount = adjustment_amount if adjustment_amount > 0 else None
                        credit_amount = -adjustment_amount if adjustment_amount < 0 else None
                        _dbg(
                            "adjust_journal_create",
                            request_id=request_id,
                            i=idx,
                            amount=str(adjustment_amount),
                            debit=str(debit_amount or 0),
                            credit=str(credit_amount or 0),
                        )
                        adj_je = JournalEntry.objects.create(
                            company_id=company_id,
                            transaction=je0.transaction,
                            account=(
                                pending_gl
                                if (
                                    je0.account is None
                                    or getattr(je0.account, "bank_account_id", None) is None
                                )
                                else je0.account
                            ),
                            cost_center=je0.cost_center,
                            debit_amount=debit_amount,
                            credit_amount=credit_amount,
                            state="pending",
                            date=je0.date or je0.transaction.date,
                            bank_designation_pending=True,
                        )
                        journal_entries.append(adj_je)
                        sum_journal += adjustment_amount
    
                final_diff = sum_bank - sum_journal
                rec_status = "matched" if final_diff == Decimal("0") else "pending"
    
                bank_ids_used = [x.id for x in bank_txs]
                journal_ids_used = [x.id for x in journal_entries]
    
                bank_ids_str = ", ".join(str(x) for x in bank_ids_used)
                journal_ids_str = ", ".join(str(x) for x in journal_ids_used)
                combined_notes = (
                    f"{notes}\n"
                    f"Bank IDs: {bank_ids_str}\n"
                    f"Journal IDs: {journal_ids_str}\n"
                    f"Difference: {final_diff}"
                )
    
                rec = Reconciliation.objects.create(
                    company_id=company_id,
                    status=rec_status,
                    reference=reference,
                    notes=combined_notes,
                )
                rec.bank_transactions.set(bank_txs)
                rec.journal_entries.set(journal_entries)
    
                created_records.append(
                    {
                        "reconciliation_id": rec.id,
                        "status": rec_status,
                        "bank_ids_used": bank_ids_used,
                        "journal_ids_used": journal_ids_used,
                    }
                )
    
                _info(
                    "reconciliation_created",
                    request_id=request_id,
                    i=idx,
                    reconciliation_id=rec.id,
                    status=rec_status,
                    n_bank=len(bank_txs),
                    n_journal=len(journal_entries),
                )
    
                # Mark used in this batch (so they won't be reused)
                used_bank_ids_batch.update(bank_ids_used)
                used_journal_ids_batch.update(journal_ids_used)
    
                # Promotion: collect changes for bulk_update
                if len(bank_account_set) == 1:
                    target_ba_id = next(iter(bank_account_set))
    
                    gl_key = (company_id, target_ba_id)
                    if gl_key in gl_cache:
                        target_gl = gl_cache[gl_key]
                    else:
                        target_ba = next(
                            bt.bank_account
                            for bt in bank_txs
                            if bt.bank_account_id == target_ba_id
                        )
                        target_gl = ensure_gl_account_for_bank(company_id, target_ba)
                        gl_cache[gl_key] = target_gl
    
                    changed = 0
                    for je in journal_entries:
                        acct_ba_id = (
                            getattr(je.account, "bank_account_id", None)
                            if je.account_id
                            else None
                        )
                        if (
                            acct_ba_id == pending_ba.id
                            or je.account_id is None
                            or getattr(je, "bank_designation_pending", False)
                        ):
                            if je.account_id != target_gl.id:
                                je.account_id = target_gl.id
                            if hasattr(je, "bank_designation_pending") and je.bank_designation_pending:
                                je.bank_designation_pending = False
                            jes_to_update.append(je)
                            changed += 1
                    _dbg(
                        "promotion_done",
                        request_id=request_id,
                        i=idx,
                        reassigned_lines=changed,
                        target_gl_id=target_gl.id,
                    )
                else:
                    problems.append(
                        {
                            "reason": "multiple_bank_accounts_in_match",
                            "bank_account_ids": list(bank_account_set),
                            "bank_ids": bank_ids,
                            "journal_ids": journal_ids,
                        }
                    )
                    _warn(
                        "promotion_skipped_multi_ba",
                        request_id=request_id,
                        i=idx,
                        bank_accounts=list(bank_account_set),
                    )
    
            except Exception as outer_e:
                problems.append(
                    {"reason": "outer_exception", "error": str(outer_e), "match": match}
                )
                _warn(
                    "match_outer_exception",
                    request_id=request_id,
                    i=idx,
                    error=str(outer_e),
                )
    
        # ------------------------------------------------------------------
        # 6) Bulk-update all changed journal entries (promotion)
        # ------------------------------------------------------------------
        if jes_to_update:
            unique_jes = {je.id: je for je in jes_to_update}.values()
            JournalEntry.objects.bulk_update(
                list(unique_jes), ["account", "bank_designation_pending"]
            )
    
        _info(
            "finalize_end",
            request_id=request_id,
            created=len(created_records),
            problems=len(problems),
        )
    
        return Response({"created": created_records, "problems": problems})

    
    @action(detail=False, methods=['post'])
    def finalize_reconciliation_matches_legacy(self, request, *args, **kwargs):
        """
        Finalize a batch of reconciliation matches (optionally creating an adjustment on one side).
        """
        request_id = str(uuid.uuid4())
    
        data = request.data
        matches = data.get("matches", [])
        adjustment_side = data.get("adjustment_side", "none")  # "bank" | "journal" | "none"
        reference = data.get("reference", "")
        notes = data.get("notes", "")
    
        _info("finalize_begin",
              request_id=request_id,
              n_matches=len(matches),
              adjustment_side=adjustment_side,
              reference=reference or "",
              notes_len=len(notes or ""))
    
        # CHANGED: richer output container
        created_records = []
        problems = []
        
        # ------------------------------------------------------------------
        # 1) Pre-collect all IDs for batching
        # ------------------------------------------------------------------
        all_bank_ids: set[int] = set()
        all_journal_ids: set[int] = set()
        for m in matches:
            raw_bank = m.get("bank_transaction_ids", []) or []
            raw_journal = m.get("journal_entry_ids", []) or []
            all_bank_ids.update(int(x) for x in raw_bank if x is not None)
            all_journal_ids.update(int(x) for x in raw_journal if x is not None)
    
        if not all_bank_ids and not all_journal_ids:
            return Response({"created": [], "problems": [{"reason": "no_ids"}]})
        
        with db_tx.atomic():
            
            # ------------------------------------------------------------------
            # 2) Pre-load already reconciled IDs in TWO queries
            # ------------------------------------------------------------------
            already_bank_ids = set(
                BankTransaction.objects.filter(
                    id__in=all_bank_ids,
                    reconciliations__status__in=["matched", "approved"],
                ).values_list("id", flat=True)
            )
            already_journal_ids = set(
                JournalEntry.objects.filter(
                    id__in=all_journal_ids,
                    reconciliations__status__in=["matched", "approved"],
                ).values_list("id", flat=True)
            )
    
            _dbg(
                "preload_already",
                request_id=request_id,
                n_bank_ids=len(all_bank_ids),
                n_journal_ids=len(all_journal_ids),
                already_banks=len(already_bank_ids),
                already_journals=len(already_journal_ids),
            )
    
            # ------------------------------------------------------------------
            # 3) Pre-load all referenced Bank / Journal rows and lock them
            # ------------------------------------------------------------------
            bank_map: dict[int, BankTransaction] = {
                bt.id: bt
                for bt in BankTransaction.objects.filter(id__in=all_bank_ids)
                .select_for_update(of=("self",))
            }
            journal_map: dict[int, JournalEntry] = {
                je.id: je
                for je in JournalEntry.objects.filter(id__in=all_journal_ids)
                .select_related("transaction", "account", "account__bank_account", "cost_center")
                .select_for_update(of=("self",))
            }
    
            _dbg(
                "preload_maps",
                request_id=request_id,
                banks_loaded=len(bank_map),
                journals_loaded=len(journal_map),
            )
    
            # ------------------------------------------------------------------
            # 4) Caches for expensive helpers
            # ------------------------------------------------------------------
            pending_structs_cache: dict[tuple[int, int], tuple] = {}
            gl_cache: dict[tuple[int, int], object] = {}
    
            # Collect JEs that need to be updated for promotion (do bulk_update later)
            jes_to_update: list[JournalEntry] = []
            je_update_fields = {"account", "bank_designation_pending"}
    
            # ------------------------------------------------------------------
            # 5) Process each match
            # ------------------------------------------------------------------
            
            for idx, match in enumerate(matches, start=1):
                try:
                    raw_bank = match.get("bank_transaction_ids", []) or []
                    raw_journal = match.get("journal_entry_ids", []) or []
                    bank_ids = list({int(x) for x in raw_bank if x is not None})
                    journal_ids = list({int(x) for x in raw_journal if x is not None})
    
                    #_dbg("match_recv",request_id=request_id,i=idx,bank_ids=bank_ids,journal_ids=journal_ids,)
    
                    if not bank_ids or not journal_ids:
                        problems.append({"reason": "empty_ids", "match": match})
                        _warn("match_skip_empty_ids", request_id=request_id, i=idx)
                        continue
    
                    # Fast in-memory "already reconciled" check
                    if any(bid in already_bank_ids for bid in bank_ids) or any(
                        jid in already_journal_ids for jid in journal_ids
                    ):
                        problems.append(
                            {
                                "reason": "already_reconciled",
                                "bank_ids": bank_ids,
                                "journal_ids": journal_ids,
                            }
                        )
                        _warn("match_skip_already_reconciled", request_id=request_id, i=idx)
                        continue
    
                    # Get locked objects from pre-loaded maps
                    bank_txs = [bank_map.get(bid) for bid in bank_ids if bid in bank_map]
                    journal_entries = [
                        journal_map.get(jid) for jid in journal_ids if jid in journal_map
                    ]
    
                    bank_txs = [bt for bt in bank_txs if bt is not None]
                    journal_entries = [je for je in journal_entries if je is not None]
    
                    #_dbg("match_locked_counts",request_id=request_id,i=idx,bank_locked=len(bank_txs),journal_locked=len(journal_entries),)
    
                    if not bank_txs or not journal_entries:
                        problems.append(
                            {
                                "reason": "not_found",
                                "bank_ids": bank_ids,
                                "journal_ids": journal_ids,
                            }
                        )
                        _warn("match_skip_not_found", request_id=request_id, i=idx)
                        continue
    
                    company_set = {bt.company_id for bt in bank_txs} | {
                        je.company_id for je in journal_entries
                    }
                    currency_set = {bt.currency_id for bt in bank_txs} | {
                        je.transaction.currency_id
                        for je in journal_entries
                        if je.transaction_id
                    }
    
                    #_dbg("match_sets",request_id=request_id,i=idx,companies=list(company_set),currencies=list(currency_set))
    
                    if len(company_set) != 1:
                        problems.append(
                            {
                                "reason": "mixed_company",
                                "bank_ids": bank_ids,
                                "journal_ids": journal_ids,
                            }
                        )
                        _warn(
                            "match_skip_mixed_company",
                            request_id=request_id,
                            i=idx,
                            companies=list(company_set),
                        )
                        continue
                    company_id = next(iter(company_set))
    
                    if len(currency_set) != 1:
                        problems.append(
                            {
                                "reason": "mixed_currency",
                                "bank_ids": bank_ids,
                                "journal_ids": journal_ids,
                            }
                        )
                        _warn(
                            "match_skip_mixed_currency",
                            request_id=request_id,
                            i=idx,
                            currencies=list(currency_set),
                        )
                        continue
                    currency_id = next(iter(currency_set))
    
                    # Cached pending structs per (company, currency)
                    pend_key = (company_id, currency_id)
                    if pend_key in pending_structs_cache:
                        pending_ba, pending_gl = pending_structs_cache[pend_key]
                    else:
                        pending_ba, pending_gl = ensure_pending_bank_structs(
                            company_id, currency_id=currency_id
                        )
                        pending_structs_cache[pend_key] = (pending_ba, pending_gl)
    
                    #_dbg("pending_structs",request_id=request_id,i=idx,pending_ba_id=pending_ba.id,pending_gl_id=pending_gl.id)
    
                    sum_bank = sum((bt.amount for bt in bank_txs), Decimal("0"))
                    sum_journal = sum(
                        ((je.get_effective_amount() or Decimal("0")) for je in journal_entries),
                        Decimal("0"),
                    )
                    diff = sum_bank - sum_journal
    
                    #_dbg("totals_initial",request_id=request_id,i=idx,sum_bank=str(sum_bank),sum_journal=str(sum_journal), diff=str(diff))
    
                    bank_account_set = {
                        bt.bank_account_id for bt in bank_txs if bt.bank_account_id
                    }
    
                    # Adjustment logic (unchanged, just using cached structs)
                    if adjustment_side != "none" and diff != Decimal("0"):
                        if adjustment_side == "bank":
                            if len(bank_account_set) != 1:
                                problems.append(
                                    {
                                        "reason": "cannot_adjust_bank_with_multiple_accounts",
                                        "bank_account_ids": list(bank_account_set),
                                        "bank_ids": bank_ids,
                                        "journal_ids": journal_ids,
                                    }
                                )
                                _warn(
                                    "adjust_skip_multi_ba",
                                    request_id=request_id,
                                    i=idx,
                                    bank_accounts=list(bank_account_set),
                                )
                            else:
                                adjustment_amount = sum_journal - sum_bank
                                bt0 = bank_txs[0]
                                _dbg(
                                    "adjust_bank_create",
                                    request_id=request_id,
                                    i=idx,
                                    amount=str(adjustment_amount),
                                )
                                adj_bt = BankTransaction.objects.create(
                                    company_id=company_id,
                                    bank_account=bt0.bank_account,
                                    date=bt0.date,
                                    currency=bt0.currency,
                                    amount=adjustment_amount,
                                    description="Adjustment record for reconciliation",
                                    status="pending",
                                    tx_hash="adjustment_for_rec",
                                )
                                bank_txs.append(adj_bt)
                                sum_bank += adjustment_amount
    
                        elif adjustment_side == "journal":
                            adjustment_amount = sum_bank - sum_journal
                            je0 = journal_entries[0]
                            debit_amount = adjustment_amount if adjustment_amount > 0 else None
                            credit_amount = (
                                -adjustment_amount if adjustment_amount < 0 else None
                            )
                            _dbg(
                                "adjust_journal_create",
                                request_id=request_id,
                                i=idx,
                                amount=str(adjustment_amount),
                                debit=str(debit_amount or 0),
                                credit=str(credit_amount or 0),
                            )
                            adj_je = JournalEntry.objects.create(
                                company_id=company_id,
                                transaction=je0.transaction,
                                account=(
                                    pending_gl
                                    if (
                                        je0.account is None
                                        or getattr(je0.account, "bank_account_id", None)
                                        is None
                                    )
                                    else je0.account
                                ),
                                cost_center=je0.cost_center,
                                debit_amount=debit_amount,
                                credit_amount=credit_amount,
                                state="pending",
                                date=je0.date or je0.transaction.date,
                                bank_designation_pending=True,
                            )
                            journal_entries.append(adj_je)
                            sum_journal += adjustment_amount
    
                    final_diff = sum_bank - sum_journal
                    rec_status = "matched" if final_diff == Decimal("0") else "pending"
    
                    #_dbg("totals_final",request_id=request_id,i=idx,sum_bank=str(sum_bank),sum_journal=str(sum_journal),final_diff=str(final_diff),rec_status=rec_status)
    
                    bank_ids_used = [x.id for x in bank_txs]
                    journal_ids_used = [x.id for x in journal_entries]
    
                    bank_ids_str = ", ".join(str(x) for x in bank_ids_used)
                    journal_ids_str = ", ".join(str(x) for x in journal_ids_used)
                    combined_notes = (
                        f"{notes}\n"
                        f"Bank IDs: {bank_ids_str}\n"
                        f"Journal IDs: {journal_ids_str}\n"
                        f"Difference: {final_diff}"
                    )
    
                    rec = Reconciliation.objects.create(
                        company_id=company_id,
                        status=rec_status,
                        reference=reference,
                        notes=combined_notes,
                    )
                    rec.bank_transactions.set(bank_txs)
                    rec.journal_entries.set(journal_entries)
    
                    created_records.append(
                        {
                            "reconciliation_id": rec.id,
                            "status": rec_status,
                            "bank_ids_used": bank_ids_used,
                            "journal_ids_used": journal_ids_used,
                        }
                    )
    
                    #_info("reconciliation_created",request_id=request_id,i=idx,reconciliation_id=rec.id,status=rec_status,n_bank=len(bank_txs),n_journal=len(journal_entries))
    
                    # Promotion (collect changes for bulk_update)
                    if len(bank_account_set) == 1:
                        target_ba_id = next(iter(bank_account_set))
    
                        gl_key = (company_id, target_ba_id)
                        if gl_key in gl_cache:
                            target_gl = gl_cache[gl_key]
                        else:
                            target_ba = next(
                                bt.bank_account
                                for bt in bank_txs
                                if bt.bank_account_id == target_ba_id
                            )
                            target_gl = ensure_gl_account_for_bank(company_id, target_ba)
                            gl_cache[gl_key] = target_gl
    
                        changed = 0
                        for je in journal_entries:
                            acct_ba_id = (
                                getattr(je.account, "bank_account_id", None)
                                if je.account_id
                                else None
                            )
                            if (
                                acct_ba_id == pending_ba.id
                                or je.account_id is None
                                or getattr(je, "bank_designation_pending", False)
                            ):
                                if je.account_id != target_gl.id:
                                    je.account_id = target_gl.id
                                if hasattr(je, "bank_designation_pending") and je.bank_designation_pending:
                                    je.bank_designation_pending = False
                                jes_to_update.append(je)
                                changed += 1
                        #_dbg("promotion_done",request_id=request_id,i=idx,reassigned_lines=changed,target_gl_id=target_gl.id)
                    else:
                        problems.append(
                            {
                                "reason": "multiple_bank_accounts_in_match",
                                "bank_account_ids": list(bank_account_set),
                                "bank_ids": bank_ids,
                                "journal_ids": journal_ids,
                            }
                        )
                        _warn(
                            "promotion_skipped_multi_ba",
                            request_id=request_id,
                            i=idx,
                            bank_accounts=list(bank_account_set),
                        )
    
                except Exception as outer_e:
                    problems.append(
                        {"reason": "outer_exception", "error": str(outer_e), "match": match}
                    )
                    _warn(
                        "match_outer_exception",
                        request_id=request_id,
                        i=idx,
                        error=str(outer_e),
                    )
    
            # ------------------------------------------------------------------
            # 6) Bulk-update all changed journal entries (promotion)
            # ------------------------------------------------------------------
            if jes_to_update:
                # remove duplicates
                unique_jes = {je.id: je for je in jes_to_update}.values()
                JournalEntry.objects.bulk_update(
                    list(unique_jes), ["account", "bank_designation_pending"]
                )
    
        _info(
            "finalize_end",
            request_id=request_id,
            created=len(created_records),
            problems=len(problems),
            created_records=created_records,
        )
    
        return Response({"created": created_records, "problems": problems})
    
    @action(detail=False, methods=['get'])
    def reconciliation_status(self, request, tenant_id=None):
        """
        Poll reconciliation task status/result
        """
        task_id = request.query_params.get("task_id")
        if not task_id:
            return Response({"error": "task_id is required"}, status=400)

        res = AsyncResult(task_id)
        return Response({
            "task_id": task_id,
            "status": res.status,
            "result": res.result if res.ready() else None
        })
    
    
    
class UnreconciledDashboardView(APIView):
    """
    Dashboard endpoint providing aggregated metrics on unreconciled records.
    
    A record is considered unreconciled if it does not have a related Reconciliation
    with status 'matched' or 'approved'.
    
    Returns overall and daily aggregates.
    """
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]
        
    def get(self, request, tenant_id=None):
        try:
            # --- BANK TRANSACTIONS ---
            # Exclude reconciled records and those with missing dates.
            bank_qs = BankTransaction.objects.exclude(
                reconciliations__status__in=['matched', 'approved']
            ).filter(date__isnull=False).exclude(date='')
            
            if tenant_id:
                bank_qs = bank_qs.filter(company__subdomain=tenant_id)
            
            bank_overall = bank_qs.aggregate(
                count=Count('id'),
                total=Coalesce(Sum('amount'), V(0, output_field=DecimalField()))
            )
            
            bank_daily = bank_qs.annotate(day=TruncDate('date')).values('day').annotate(
                count=Count('id'),
                total=Coalesce(Sum('amount'), V(0, output_field=DecimalField()))
            ).order_by('day')
            
            # --- JOURNAL ENTRIES ---
            journal_qs = JournalEntry.objects.exclude(
                reconciliations__status__in=['matched', 'approved']
            ).filter(transaction__date__isnull=False).exclude(transaction__date='')
            
            if tenant_id:
                journal_qs = journal_qs.filter(transaction__company__subdomain=tenant_id)
            
            journal_overall = journal_qs.aggregate(
                count=Count('id'),
                total=Coalesce(Sum(F('debit_amount') - F('credit_amount')), V(0, output_field=DecimalField()))
            )
            
            journal_daily = journal_qs.annotate(day=TruncDate('transaction__date')).values('day').annotate(
                count=Count('id'),
                total=Coalesce(Sum(F('debit_amount') - F('credit_amount')), V(0, output_field=DecimalField()))
            ).order_by('day')
            
            response_data = {
                "bank_transactions": {
                    "overall": bank_overall,
                    "daily": list(bank_daily)
                },
                "journal_entries": {
                    "overall": journal_overall,
                    "daily": list(journal_daily)
                }
            }
            
            return Response(response_data)
        
        except OperationalError as e:
            return Response({"error": "Database error: " + str(e)}, status=500)
        except Exception as e:
            return Response({"error": str(e)}, status=500)
    
    
class ReconciliationTaskViewSet(ScopedQuerysetMixin, viewsets.ModelViewSet):
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]
        
    queryset = ReconciliationTask.objects.all().order_by("-created_at")
    serializer_class = ReconciliationTaskSerializer
    
    def get_queryset(self):
        qs = ReconciliationTask.objects.all().order_by("-created_at")
        # Accept tenant from router path or ?tenant_id=...
        tenant_id = self.kwargs.get("tenant_id") or self.request.query_params.get("tenant_id")
        if tenant_id:
            qs = qs.filter(tenant_id=tenant_id)
        return qs
    
    @action(detail=False, methods=["post"])
    def start(self, request, tenant_id=None):
        """
        Start reconciliation as a background task.

        Accepts payload keys:
          - bank_ids: optional list of BankTransaction IDs
          - book_ids: optional list of journal Transaction IDs
          - config_id: ID of a ReconciliationConfig (run single stage)
          - pipeline_id: ID of a ReconciliationPipeline (run multi-stage)
          - auto_match_100: if true, auto-persist matches with confidence==1.0
        """
        data = request.data
        auto_match_100 = _to_bool(data.get("auto_match_100", False))

        config_id = data.get("config_id")
        pipeline_id = data.get("pipeline_id")

        cfg = None
        pipe = None
        soft_limit = None

        if config_id:
            try:
                cfg = ReconciliationConfig.objects.get(id=config_id)
                soft_limit = getattr(cfg, "soft_time_limit_seconds", None)
            except ReconciliationConfig.DoesNotExist:
                cfg = None

        if pipeline_id:
            try:
                pipe = ReconciliationPipeline.objects.get(id=pipeline_id)
                # Pipeline-level soft limit overrides config-level if both are set
                soft_limit = getattr(pipe, "soft_time_limit_seconds", None)
            except ReconciliationPipeline.DoesNotExist:
                pipe = None

        task_obj = ReconciliationTask.objects.create(
            task_id=uuid.uuid4(),
            tenant_id=tenant_id,
            parameters=data,
            status="queued",
            config=cfg,
            pipeline=pipe,
            config_name=cfg.name if cfg else "",
            pipeline_name=pipe.name if pipe else "",
            soft_time_limit_seconds=soft_limit,
        )

        async_result = match_many_to_many_task.delay(task_obj.id, data, tenant_id, auto_match_100)

        task_obj.task_id = async_result.id
        task_obj.save(update_fields=["task_id", "updated_at"])

        return Response({
            "message": "Task enqueued",
            "task_id": async_result.id,
            "db_id": task_obj.id,
        })
    
    @action(detail=True, methods=["get"], url_path="fresh-suggestions")
    def fresh_suggestions(self, request, pk=None, tenant_id=None):
        """
        Return only the suggestions from this task that are still 'fresh':
        i.e., none of their bank_ids or journal_entries_ids are already used
        in a matched/approved reconciliation for this tenant.
        """
        task = self.get_object()

        result = task.result or {}
        suggestions = result.get("suggestions") or []
        if not suggestions:
            return Response({"count": 0, "suggestions": []})

        # Company scope from tenant
        company = resolve_tenant(tenant_id)
        company_id = company.id

        # Sets of already reconciled IDs for this company
        used_bank_ids = set(
            BankTransaction.objects.filter(
                company_id=company_id,
                reconciliations__status__in=["matched", "approved"],
            ).values_list("id", flat=True)
        )
        used_book_ids = set(
            JournalEntry.objects.filter(
                company_id=company_id,
                reconciliations__status__in=["matched", "approved"],
            ).values_list("id", flat=True)
        )

        # Optional filters
        try:
            min_conf = float(request.query_params.get("min_confidence", "0"))
        except ValueError:
            min_conf = 0.0
        try:
            limit = int(request.query_params.get("limit", "1000"))
        except ValueError:
            limit = 1000

        fresh = []
        for s in suggestions:
            if s.get("confidence_score", 0) < min_conf:
                continue
            b_ids = set(s.get("bank_ids", []))
            j_ids = set(s.get("journal_entries_ids", []))
            # Skip if any id already reconciled
            if b_ids & used_bank_ids:
                continue
            if j_ids & used_book_ids:
                continue
            fresh.append(s)
            if len(fresh) >= limit:
                break

        return Response(
            {"count": len(fresh), "suggestions": fresh},
            status=status.HTTP_200_OK,
        )
    
    @action(detail=True, methods=["get"])
    def status(self, request, pk=None, tenant_id=None):
        """
        Get status/result of a task by DB ID
        """
        task = self.get_object()
        serializer = self.get_serializer(task)
        return Response(serializer.data)
    
    @action(detail=True, methods=["post"])
    def cancel(self, request, pk=None, tenant_id=None):
        """
        Cancel a reconciliation task.

        - Marks the DB task as 'cancelled' (with optional reason).
        - Sends a Celery revoke() for the underlying task_id (best-effort).
        """
        task = self.get_object()

        # If it's already in a terminal state, don't "cancel" again
        if task.status in ["completed", "failed", "cancelled"]:
            return Response(
                {
                    "detail": "Task is already finished and cannot be cancelled.",
                    "status": task.status,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        reason = (request.data.get("reason") or "Cancelled by user").strip()

        # Best-effort Celery revoke (may or may not actually kill a running task)
        if task.task_id:
            try:
                current_app.control.revoke(task.task_id, terminate=True)
            except Exception as e:
                log.warning(
                    "Failed to revoke Celery task %s for ReconciliationTask %s: %s",
                    task.task_id, task.id, e
                )

        task.status = "cancelled"
        task.error_message = reason
        task.save(update_fields=["status", "error_message", "updated_at"])

        serializer = self.get_serializer(task)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=["get"])
    def queued(self, request, tenant_id=None):
        """
        Returns both:
        1. DB-persisted tasks (filterable by tenant_id, status)
        2. Live Celery queue info (active/reserved/scheduled)
        Filters:
          - ?tenant_id=foo
          - ?last_n=100   (last 100 tasks)
          - ?hours_ago=6  (tasks from the last 6 hours)
        """
        
        tenant_filter = request.query_params.get("tenant_id")
        status_filter = request.query_params.get("status")
        last_n = request.query_params.get("last_n")
        hours_ago = request.query_params.get("hours_ago")
        
        # ---- DB tasks ----
        qs = ReconciliationTask.objects.all().order_by("-created_at")
        if tenant_filter:
            qs = qs.filter(tenant_id=tenant_filter)
        if status_filter:
            qs = qs.filter(status=status_filter)
        
        if hours_ago:
            try:
                raw = str(hours_ago).lower()
                if raw.endswith("d"):
                    hours = int(raw[:-1]) * 24
                elif raw.endswith("h"):
                    hours = int(raw[:-1])
                else:
                    hours = int(raw)  # fallback if pure number
        
                cutoff = timezone.now() - timedelta(hours=hours)
                qs = qs.filter(created_at__gte=cutoff)
            except ValueError:
                pass
    
        if last_n:
            try:
                last_n = int(last_n)
                qs = qs.order_by("-created_at")[:last_n]
            except ValueError:
                pass
    
        
        db_tasks = self.get_serializer(qs, many=True).data

        # ---- Celery live tasks ----
        try:
            i = app.control.inspect()
            live_info = {
                "active": i.active() or {},
                "reserved": i.reserved() or {},
                "scheduled": i.scheduled() or {}
            }
        except Exception as e:
            live_info = {"error": str(e)}

        return Response({
            "db_tasks": db_tasks,
            "celery_live": live_info
        })
    
    @action(detail=False, methods=["get"])
    def task_counts(self, request, tenant_id=None):
        """
        Lightweight endpoint to return counts of tasks by status.
        Filters:
          - ?tenant_id=foo
          - ?last_n=100   (last 100 tasks)
          - ?hours_ago=6  (tasks from the last 6 hours)
        """
        tenant_filter = tenant_id
        last_n = request.query_params.get("last_n")
        hours_ago = request.query_params.get("hours_ago")
    
        qs = ReconciliationTask.objects.all()
    
        if tenant_filter:
            qs = qs.filter(tenant_id=tenant_filter)
    
        if hours_ago:
            try:
                raw = str(hours_ago).lower()
                if raw.endswith("d"):
                    hours = int(raw[:-1]) * 24
                elif raw.endswith("h"):
                    hours = int(raw[:-1])
                else:
                    hours = int(raw)  # fallback if pure number
        
                cutoff = timezone.now() - timedelta(hours=hours)
                qs = qs.filter(created_at__gte=cutoff)
            except ValueError:
                pass
    
        if last_n:
            try:
                last_n = int(last_n)
                qs = qs.order_by("-created_at")[:last_n]
            except ValueError:
                pass
    
        counts = qs.values("status").annotate(total=Count("id"))
        status_map = {row["status"]: row["total"] for row in counts}
    
        return Response({
            "running": status_map.get("running", 0),
            "completed": status_map.get("completed", 0),
            "queued": status_map.get("queued", 0),
            "failed": status_map.get("failed", 0),
        })

class ReconciliationConfigViewSet2(viewsets.ModelViewSet):
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]
        
    queryset = ReconciliationConfig.objects.all()
    serializer_class = ReconciliationConfigSerializer

    def get_queryset(self):
        user = self.request.user
        company_id = self.request.query_params.get("company_id")

        qs = super().get_queryset()

        return qs.filter(
            models.Q(scope="global") |
            models.Q(scope="company", company_id=company_id) |
            models.Q(scope="user", user=user)
        )

class ReconciliationConfigViewSet(viewsets.ModelViewSet):
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]
        
    queryset = ReconciliationConfig.objects.all()
    serializer_class = ReconciliationConfigSerializer

    @action(detail=False, methods=["get"])
    def resolved(self, request, *args, **kwargs):
        """
        Return all configs available to the current user:
        - Global
        - Company
        - User
        - Company+User
        """
        user = request.user
        company_id = request.query_params.get("company_id")

        qs = ReconciliationConfig.objects.filter(
            Q(scope="global")
            | Q(scope="company", company_id=company_id)
            | Q(scope="user", user=user)
            | Q(scope="company_user", company_id=company_id, user=user)
        )

        serializer = ResolvedReconciliationConfigSerializer(qs, many=True)
        return Response(serializer.data)

class ReconciliationPipelineViewSet(viewsets.ModelViewSet):
    """
    API endpoint for managing reconciliation pipelines.  Pipelines bundle
    together multiple ReconciliationConfig 'recipes' in a specific order,
    with optional per‑stage overrides.
    """
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]
        
    queryset = ReconciliationPipeline.objects.all().order_by("-updated_at")
    serializer_class = ReconciliationPipelineSerializer

    @action(detail=False, methods=["get"])
    def resolved(self, request, *args, **kwargs):
        """
        Return all pipeline definitions available to the current user:
        - Global
        - Company
        - User
        - Company+User
        """
        user = request.user
        company_id = request.query_params.get("company_id")
        qs = ReconciliationPipeline.objects.filter(
            Q(scope="global")
            | Q(scope="company", company_id=company_id)
            | Q(scope="user", user=user)
            | Q(scope="company_user", company_id=company_id, user=user)
        )
        serializer = ResolvedReconciliationPipelineSerializer(qs, many=True)
        return Response(serializer.data)

class EmbeddingHealthView(APIView):
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]

    def get(self, request, tenant_id=None):
        client = EmbeddingClient()
        t0 = time.perf_counter()
        try:
            vecs = client.embed_texts(["health check ok"])
            latency_ms = int((time.perf_counter() - t0) * 1000)
            dim = len(vecs[0]) if vecs and isinstance(vecs[0], list) else 0
            ok = (dim == settings.EMBED_DIM)
            return Response(
                {
                    "ok": ok,
                    "dim": dim,
                    "latency_ms": latency_ms,
                    "endpoint": _embed_url(),
                    "model": client.model,
                    "used_internal": bool(settings.EMBED_INTERNAL_HOST),
                },
                status=status.HTTP_200_OK if ok else status.HTTP_503_SERVICE_UNAVAILABLE,
            )
        except Exception as e:
            latency_ms = int((time.perf_counter() - t0) * 1000)
            return Response(
                {
                    "ok": False,
                    "error": str(e),
                    "latency_ms": latency_ms,
                    "endpoint": _embed_url(),
                    "model": settings.EMBED_MODEL,
                    "used_internal": bool(settings.EMBED_INTERNAL_HOST),
                },
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

class EmbeddingMissingCountsView(APIView):
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]

    def get(self, request, tenant_id=None):
        tx  = Transaction.objects.filter(description_embedding__isnull=True).count()
        btx = BankTransaction.objects.filter(description_embedding__isnull=True).count()
        acc = Account.objects.filter(account_description_embedding__isnull=True).count()
        return Response(
            {
                "transactions_missing": tx,
                "bank_transactions_missing": btx,
                "accounts_missing": acc,
                "total_missing": tx + btx + acc,
            },
            status=status.HTTP_200_OK,
        )

class EmbeddingBackfillView(APIView):
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]

    def post(self, request, tenant_id=None):
        # you can keep your serializer; here we focus on headers + optional pre-totals
        per_model_limit = int(request.data.get("per_model_limit") or 2000)
        client_opts = request.data.get("client_opts") or {}

        # enqueue with helpful headers
        headers = {
            "job_kind": "embeddings",
            "tenant_id": getattr(request.user, "company_id", None),
            "user_id": getattr(request.user, "id", None),
        }
        task = generate_missing_embeddings.apply_async(
            kwargs={"per_model_limit": per_model_limit, "client_opts": client_opts},
            headers=headers,
        )

        # (optional) prime totals right away, so UI has a target immediately
        try:
            totals = {
                "transactions": Transaction.objects.filter(description_embedding__isnull=True)[:per_model_limit].count(),
                "bank_transactions": BankTransaction.objects.filter(description_embedding__isnull=True)[:per_model_limit].count(),
                "accounts": Account.objects.filter(account_description_embedding__isnull=True)[:per_model_limit].count(),
            }
            Job.objects.filter(task_id=task.id).update(
                total=sum(totals.values()),
                by_category={"totals": totals, "done": {"transactions": 0, "bank_transactions": 0, "accounts": 0}},
            )
        except Exception:
            pass

        return Response({"task_id": task.id, "state": task.state, "mode": "async"}, status=status.HTTP_202_ACCEPTED)

STATE_MAP = {
    "PENDING":  "PENDING",
    "RECEIVED": "RECEIVED",
    "SENT":     "SENT",
    "STARTED":  "STARTED",
    "PROGRESS": "PROGRESS",   # our custom state while updating meta
    "RETRY":    "RETRY",
    "SUCCESS":  "SUCCESS",
    "FAILURE":  "FAILURE",
    "REVOKED":  "REVOKED",
}

EMBED_KIND_DEFAULT = "embeddings.backfill"

ACTIVE_STATES   = {
    STATE_MAP["PENDING"], STATE_MAP["RECEIVED"], STATE_MAP["SENT"],
    STATE_MAP["STARTED"], STATE_MAP["PROGRESS"], STATE_MAP["RETRY"],
}
FINISHED_STATES = {
    STATE_MAP["SUCCESS"], STATE_MAP["FAILURE"], STATE_MAP["REVOKED"],
}
def _friendly_state(celery_state: str) -> str:
    # Normalize to your canonical state names
    return STATE_MAP.get((celery_state or "PENDING").upper(), STATE_MAP["PENDING"])

def _parse_json_field(value: Any) -> Optional[Dict[str, Any]]:
    if value is None:
        return None
    if isinstance(value, (dict, list)):
        return value
    if isinstance(value, str) and value:
        try:
            return json.loads(value)
        except Exception:
            return {"raw": value}
    return None

def _runtime_seconds(start: Optional[datetime], end: Optional[datetime]) -> Optional[float]:
    if not start or not end:
        return None
    return (end - start).total_seconds()

def _job_to_dict(j: Job) -> Dict[str, Any]:
    return {
        "task_id": j.task_id,
        "state": j.state,                       # already one of your STATE_MAP values
        "status_friendly": j.state,
        "kind": j.kind,
        "queue": j.queue,
        "worker": j.worker,
        "created_at": j.created_at,
        "enqueued_at": j.enqueued_at,
        "started_at": j.started_at,
        "finished_at": j.finished_at,
        "runtime_s": _runtime_seconds(j.started_at or j.enqueued_at or j.created_at, j.finished_at or now()),
        "retries": j.retries,
        "max_retries": j.max_retries,
        "total": j.total,
        "done": j.done,
        "percent": j.percent,
        "by_category": j.by_category,
        "result": j.result if j.state == STATE_MAP["SUCCESS"] else None,
        "error": j.error if j.state in (STATE_MAP["FAILURE"], STATE_MAP["REVOKED"]) else None,
        "meta": j.meta,
        "tenant_id": j.tenant_id,
    }

class EmbeddingJobsListView(APIView):
    """
    GET /api/embeddings/jobs/?limit=25&status=any&kind=embeddings&include_active=1
    """
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]

    def get(self, request, tenant_id=None):
        limit = int(request.query_params.get("limit", 25))
        status_filter = (request.query_params.get("status") or "any").upper()
        kind = request.query_params.get("kind") or "embeddings"
        include_active = request.query_params.get("include_active") in ("1", "true", "yes")

        qs = Job.objects.filter(kind=kind).order_by("-created_at")

        if status_filter != "ANY":
            qs = qs.filter(state=status_filter)

        finished = list(qs[:limit])
        data = {"count": len(finished), "finished": JobSerializer(finished, many=True).data}

        if include_active:
            active_states = ("PENDING", "SENT", "RECEIVED", "STARTED", "RETRY", "PROGRESS")
            active = Job.objects.filter(kind=kind, state__in=active_states).order_by("-created_at")
            data["active"] = JobSerializer(active, many=True).data
            data["active_count"] = active.count()

        return Response(data, status=status.HTTP_200_OK)


class EmbeddingTaskStatusView(APIView):
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]

    def get(self, request, task_id, tenant_id=None):
        try:
            job = Job.objects.get(task_id=task_id)
        except Job.DoesNotExist:
            # still return Celery state, if any
            ar = AsyncResult(task_id)
            return Response({
                "task_id": task_id,
                "state": ar.state or "PENDING",
                "status": STATE_MAP.get(ar.state or "PENDING", "PENDING"),
                "ready": ar.ready(),
                "successful": ar.successful() if ar.ready() else False,
            }, status=status.HTTP_200_OK)

        # overlay live state from Celery (if different)
        ar = AsyncResult(task_id)
        live_state = ar.state or job.state
        info = ar.info if isinstance(ar.info, dict) else {}
        payload = JobSerializer(job).data
        payload.update({
            "state": live_state,
            "status": STATE_MAP.get(live_state, live_state),
            "ready": ar.ready(),
            "successful": ar.successful() if ar.ready() else (job.state == "SUCCESS"),
            "progress": {
                "totals": (job.by_category or {}).get("totals"),
                "done": (job.by_category or {}).get("done"),
                "remaining": (job.by_category or {}).get("remaining"),
                "done_all": (job.by_category or {}).get("done_all"),
                "remaining_all": (job.by_category or {}).get("remaining_all"),
            },
        })
        return Response(payload, status=status.HTTP_200_OK)

class EmbeddingTaskCancelView(APIView):
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]

    def post(self, request, task_id, tenant_id=None):
        current_app.control.revoke(task_id, terminate=True)
        Job.objects.filter(task_id=task_id).update(
            state=STATE_MAP["REVOKED"],
            finished_at=now(),
            error="revoked",
        )
        return Response({"task_id": task_id, "revoked": True}, status=status.HTTP_200_OK)

class EmbeddingsTestView(APIView):
    """
    POST /api/embeddings/test/
    { "texts": ["hello", "banana"] }  # optional overrides also supported
    """
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]

    def post(self, request, tenant_id=None):
        ser = EmbedTestSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        d = ser.validated_data

        client = EmbeddingClient(
            base_url=d.get("base_url"),
            path=d.get("path"),
            model=d.get("model"),
            timeout_s=d.get("timeout_s"),
            dim=d.get("dim"),
            api_key=d.get("api_key"),
            num_thread=d.get("num_thread"),
            keep_alive=d.get("keep_alive"),
        )
        vecs = client.embed_texts(d["texts"])
        return Response(
            {"count": len(vecs), "dim": len(vecs[0]) if vecs else 0, "embeddings": vecs, "endpoint": client.url},
            status=status.HTTP_200_OK,
        )
    
def _search_qs(model, vec, field_name: str, k: int, values: list[str]):
    qs = (model.objects
          .filter(**{f"{field_name}__isnull": False})
          .annotate(score=CosineDistance(field_name, vec))
          .order_by("score")[:k])
    rows = list(qs.values(*values, "score"))
    # Attach similarity = 1 - distance (pgvector cosine distance = 1 - cosine similarity)
    for r in rows:
        dist = float(r["score"])
        r["similarity"] = 1.0 - dist
        r["score"] = dist
    return rows


class EmbeddingsSearchView(APIView):
    if settings.AUTH_OFF:
        permission_classes = []
    else:
        permission_classes = [permissions.IsAuthenticated]

    def post(self, request, tenant_id=None):
        body = request.data or {}
        query = (body.get("query") or "").strip()
        k_each = int(body.get("k_each", 8))
        min_similarity = float(body.get("min_similarity", 0.10))
        debug_mode = body.get("debug") in (True, "true", "1") or request.query_params.get("debug") in ("1", "true", "yes")

        # Build embedding (also try an accent-less variant to diagnose)
        emb = EmbeddingClient()
        qvec = []
        qvec_noacc = []
        timings = {}
        try:
            import time
            t0 = time.perf_counter()
            qvec = emb.embed_one(query or " ")
            timings["embed"] = int((time.perf_counter() - t0) * 1000)

            if query and query != _strip_accents(query):
                qvec_noacc = emb.embed_one(_strip_accents(query))
        except Exception as e:
            log.exception("embed error")
            return Response({"ok": False, "error": str(e)}, status=502)

        # Corpus coverage
        tx_n  = Transaction.objects.filter(description_embedding__isnull=False).count()
        btx_n = BankTransaction.objects.filter(description_embedding__isnull=False).count()
        acc_n = Account.objects.filter(account_description_embedding__isnull=False).count()

        # TopK (raw)
        def _search_all(vec):
            rows = {}
            t1 = time.perf_counter()
            rows["transactions"] = _search_qs(
                Transaction, vec, "description_embedding", k_each,
                ["id", "description", "amount", "date"]
            )
            rows["bank_transactions"] = _search_qs(
                BankTransaction, vec, "description_embedding", k_each,
                ["id", "description", "amount", "date"]
            )
            rows["accounts"] = _search_qs(
                Account, vec, "account_description_embedding", k_each,
                ["id", "name", "description"]
            )
            timings["search"] = int((time.perf_counter() - t1) * 1000)
            return rows

        hits_raw = _search_all(qvec)
        # Also compute with accent-less query (for comparison) if we produced it
        hits_raw_noacc = _search_all(qvec_noacc) if qvec_noacc else None

        # Filter by min_similarity
        def _apply_threshold(rows):
            out = {}
            for bucket, items in rows.items():
                out[bucket] = [
                    {**r, "description": _snippet(r.get("description") or r.get("name"))}
                    for r in items if r["similarity"] >= min_similarity
                ]
            return out

        hits = _apply_threshold(hits_raw)
        hits_noacc = _apply_threshold(hits_raw_noacc) if hits_raw_noacc else None

        # Log key debug lines
        log.info(
            "emb.search q='%s' len=%d qstats=%s tx=%d btx=%d acc=%d min_sim=%.3f k=%d",
            _snippet(query, 120), len(qvec), _vec_stats(qvec), tx_n, btx_n, acc_n, min_similarity, k_each
        )
        if qvec_noacc:
            log.info("emb.search noacc='%s' qstats=%s", _strip_accents(query), _vec_stats(qvec_noacc))

        # Show handful of raw neighbors per bucket (even if below threshold)
        for bucket in ("transactions", "bank_transactions", "accounts"):
            sample = hits_raw[bucket][:3]
            log.debug("raw.%s: %s", bucket, [
                {
                    "id": r["id"],
                    "sim": round(r["similarity"], 4),
                    "dist": round(r["score"], 4),
                    "text": _snippet(r.get("description") or r.get("name"), 100)
                } for r in sample
            ])

        # Optional lexical sanity check (does DB contain the term?)
        lexical = None
        if debug_mode and query:
            term = query if len(query) >= 3 else None
            if term:
                lexical = {
                    "tx_like":  list(Transaction.objects.filter(description__icontains=term).values("id","description")[:3]),
                    "btx_like": list(BankTransaction.objects.filter(description__icontains=term).values("id","description")[:3]),
                    "acc_like": list(Account.objects.filter(Q(name__icontains=term) | Q(description__icontains=term)).values("id","name","description")[:3]),
                }

        # Build response
        resp = {
            "ok": True,
            "meta": {
                "query": query,
                "k_each": k_each,
                "min_similarity": min_similarity,
                "model": emb.model,
                "timings_ms": timings,
                "counts": {"transactions": tx_n, "bank_transactions": btx_n, "accounts": acc_n},
            },
            "transactions": hits["transactions"],
            "bank_transactions": hits["bank_transactions"],
            "accounts": hits["accounts"],
        }

        if debug_mode:
            resp["debug"] = {
                "query_stats": _vec_stats(qvec),
                "query_noacc_stats": _vec_stats(qvec_noacc) if qvec_noacc else None,
                "raw_top": {
                    "transactions": [
                        {"id": r["id"], "sim": r["similarity"], "dist": r["score"], "text": _snippet(r.get("description") or "", 120)}
                        for r in hits_raw["transactions"][:5]
                    ],
                    "bank_transactions": [
                        {"id": r["id"], "sim": r["similarity"], "dist": r["score"], "text": _snippet(r.get("description") or "", 120)}
                        for r in hits_raw["bank_transactions"][:5]
                    ],
                    "accounts": [
                        {"id": r["id"], "sim": r["similarity"], "dist": r["score"], "text": _snippet(r.get("name") or r.get("description") or "", 120)}
                        for r in hits_raw["accounts"][:5]
                    ],
                },
                "lexical_samples": lexical,
                "note": "similarity = 1 - cosine_distance (pgvector). If everything is < threshold, try lowering min_similarity.",
            }
            if hits_noacc is not None:
                resp["debug"]["accentless_top"] = {
                    "transactions": hits_noacc["transactions"][:5],
                    "bank_transactions": hits_noacc["bank_transactions"][:5],
                    "accounts": hits_noacc["accounts"][:5],
                }

        return Response(resp, status=status.HTTP_200_OK)