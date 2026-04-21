"""End-to-end backend scenario runner for the new report engine.

Invoked as: `python scripts/test_reports_e2e.py` from the repo root.
"""

import os
import sys
from pathlib import Path

# Allow running the script directly (not via django manage): prepend the
# repo root to sys.path so `nord_backend.settings` is importable.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import django  # noqa: E402
from unittest.mock import MagicMock  # noqa: E402
from decimal import Decimal  # noqa: E402
from io import BytesIO  # noqa: E402

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "nord_backend.settings")
django.setup()

from rest_framework.test import APIRequestFactory, force_authenticate  # noqa: E402
from accounting.reports.views import (  # noqa: E402
    CalculateViewSet, ExportViewSet, AiStub,
)
from accounting.reports.services.document_schema import validate_document  # noqa: E402
from accounting.services.financial_statement_service import FinancialStatementGenerator  # noqa: E402
from accounting.reports.services import intelligence  # noqa: E402


def _stub_resolver(self, sel):
    if sel is None:
        return []
    if (
        getattr(sel, "account_ids", None)
        or getattr(sel, "code_prefix", None)
        or getattr(sel, "path_contains", None)
    ):
        return [MagicMock(id=1, account_direction=1, is_active=True)]
    return []


def main():
    intelligence.AccountResolver.resolve = _stub_resolver
    FinancialStatementGenerator._calc_net_movement = MagicMock(return_value=Decimal("1200"))
    FinancialStatementGenerator._calc_ending_balance = MagicMock(return_value=Decimal("5000"))

    factory = APIRequestFactory()
    user = MagicMock(is_authenticated=True, is_superuser=True)
    tenant = MagicMock(id=1)

    doc = {
        "name": "DRE E2E",
        "report_type": "income_statement",
        "defaults": {"calculation_method": "net_movement", "sign_policy": "natural"},
        "blocks": [
            {"type": "section", "id": "revenue", "label": "Receita", "children": [
                {"type": "line", "id": "sales", "label": "Vendas",
                 "accounts": {"code_prefix": "4.01", "include_descendants": True}},
                {"type": "line", "id": "services", "label": "Servicos",
                 "accounts": {"code_prefix": "4.02"}},
                {"type": "subtotal", "id": "rev_total", "label": "Total",
                 "formula": "sum(children)"},
            ]},
            {"type": "spacer", "id": "sp1"},
            {"type": "section", "id": "costs", "label": "Custos",
             "defaults": {"sign_policy": "invert"}, "children": [
                {"type": "line", "id": "cogs", "label": "CMV",
                 "accounts": {"code_prefix": "5"}},
            ]},
            {"type": "total", "id": "net_income", "label": "Lucro Liquido",
             "formula": "rev_total - cogs", "bold": True},
        ],
    }

    # --- 1. Pydantic round-trip ---
    print("=== 1. pydantic round-trip ===")
    validate_document(doc)
    print("  OK: full nested doc + defaults cascade")

    # --- 2. /calculate/ multi-period with variance ---
    print("=== 2. /calculate/ YoY + variance ===")
    def period_aware(accounts, start, end, include_pending):
        return Decimal("1200") if start.year == 2025 else Decimal("1000")
    FinancialStatementGenerator._calc_net_movement = MagicMock(side_effect=period_aware)

    periods = [
        {"id": "cur", "label": "2025", "type": "range", "start": "2025-01-01", "end": "2025-12-31"},
        {"id": "prev", "label": "2024", "type": "range", "start": "2024-01-01", "end": "2024-12-31"},
        {"id": "var_abs", "label": "delta", "type": "variance_abs", "base": "prev", "compare": "cur"},
        {"id": "var_pct", "label": "pct", "type": "variance_pct", "base": "prev", "compare": "cur"},
    ]
    req = factory.post("/api/reports/calculate/", {"template": doc, "periods": periods}, format="json")
    force_authenticate(req, user=user); req.tenant = tenant
    resp = CalculateViewSet.as_view({"post": "create"})(req)
    print(f"  status: {resp.status_code}, lines: {len(resp.data['lines'])}, periods: {len(resp.data['periods'])}")
    by_id = {l["id"]: l for l in resp.data["lines"]}
    print(f"  sales.cur = {by_id['sales']['values']['cur']}  (1200)")
    print(f"  sales.var_abs = {by_id['sales']['values']['var_abs']}  (200)")
    print(f"  sales.var_pct = {by_id['sales']['values']['var_pct']}  (20.0)")
    print(f"  rev_total.cur = {by_id['rev_total']['values']['cur']}  (2400)")
    print(f"  cogs.cur = {by_id['cogs']['values']['cur']}  (-1200 after invert)")
    print(f"  net_income.cur = {by_id['net_income']['values']['cur']}  (formula: rev_total - cogs)")
    assert resp.status_code == 200
    assert by_id["sales"]["values"]["var_abs"] == 200.0
    assert abs(by_id["sales"]["values"]["var_pct"] - 20.0) < 0.01
    print("  PASS")

    # --- 3. incompatible period type ---
    print("=== 3. BP with range → 400 ===")
    bs_doc = {
        "name": "BP", "report_type": "balance_sheet",
        "defaults": {"calculation_method": "ending_balance"},
        "blocks": [{"type": "line", "id": "assets", "label": "Ativo", "accounts": {"code_prefix": "1"}}],
    }
    req = factory.post("/api/reports/calculate/", {"template": bs_doc,
        "periods": [{"id": "x", "label": "x", "type": "range", "start": "2025-01-01", "end": "2025-12-31"}]},
        format="json")
    force_authenticate(req, user=user); req.tenant = tenant
    resp = CalculateViewSet.as_view({"post": "create"})(req)
    print(f"  status: {resp.status_code} (expected 400)")
    assert resp.status_code == 400
    print("  PASS")

    # --- 4. /export/xlsx/ stateless ---
    print("=== 4. /export/xlsx/ ===")
    FinancialStatementGenerator._calc_net_movement = MagicMock(return_value=Decimal("500"))
    req = factory.post("/api/reports/calculate/", {"template": doc, "periods": periods}, format="json")
    force_authenticate(req, user=user); req.tenant = tenant
    result = CalculateViewSet.as_view({"post": "create"})(req).data
    req2 = factory.post("/api/reports/export/xlsx/", {"result": result, "name": "E2E"}, format="json")
    force_authenticate(req2, user=user); req2.tenant = tenant
    r = ExportViewSet.as_view({"post": "xlsx"})(req2)
    print(f"  status: {r.status_code}, bytes: {len(r.content)}")
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(r.content))
    print(f"  sheets: {wb.sheetnames}")
    report_sheet = wb[wb.sheetnames[0]]
    print(f"  header row: {[c.value for c in report_sheet[4]]}")
    assert r.status_code == 200 and "Memory" in wb.sheetnames
    print("  PASS")

    # --- 5. /export/pdf/ → 501 ---
    print("=== 5. /export/pdf/ 501 ===")
    req3 = factory.post("/api/reports/export/pdf/", {"result": result, "name": "E2E"}, format="json")
    force_authenticate(req3, user=user); req3.tenant = tenant
    r = ExportViewSet.as_view({"post": "pdf"})(req3)
    print(f"  status: {r.status_code}  error: {str(r.data.get('error',''))[:60]}")
    assert r.status_code == 501
    print("  PASS")

    # --- 6. AI endpoints → 400 without key ---
    print("=== 6. AI endpoints graceful w/o API key ===")
    os.environ.pop("OPEN_AI_API_KEY", None); os.environ.pop("OPENAI_API_KEY", None)
    os.environ.pop("ANTHROPIC_API_KEY", None)
    from django.conf import settings
    for a in ("OPEN_AI_API_KEY", "OPENAI_API_KEY", "ANTHROPIC_API_KEY"):
        if hasattr(settings, a):
            setattr(settings, a, None)

    # Three endpoints should 400 with an error_type='ai_error' payload when
    # no API key is configured. `explain` is intentionally different: it
    # always returns 200 with a _coded_explanation fallback so users get
    # *something* useful even when the AI is unavailable.
    fail_cases = [
        ("generate_template", {"report_type": "income_statement"}),
        ("refine", {"action": "normalize_labels", "document": doc}),
        ("chat", {"messages": [{"role": "user", "content": "hi"}], "document": doc}),
    ]
    for action, payload in fail_cases:
        req = factory.post(f"/api/reports/ai/{action}/", payload, format="json")
        force_authenticate(req, user=user); req.tenant = tenant
        view = AiStub.as_view({"post": action})
        r = view(req)
        err = r.data.get("error", "") if hasattr(r.data, "get") else str(r.data)
        print(f"  {action}: status={r.status_code}  error='{err[:50]}...'")
        assert r.status_code == 400, f"{action} expected 400 got {r.status_code}"

    # `explain` fallback path
    req = factory.post("/api/reports/ai/explain/", {
        "document": doc, "result": result, "block_id": "sales", "period_id": "cur",
    }, format="json")
    force_authenticate(req, user=user); req.tenant = tenant
    r = AiStub.as_view({"post": "explain"})(req)
    print(f"  explain: status={r.status_code}  text='{r.data.get('text','')[:60]}...'")
    assert r.status_code == 200 and r.data.get("text"), \
        f"explain should fall back to coded text; got status={r.status_code}"
    print("  PASS (3 AI endpoints return 400, explain falls back to coded 200)")

    # --- 7. explain coded fallback ---
    print("=== 7. explain coded fallback ===")
    from accounting.reports.services.ai_assistant import _coded_explanation
    text = _coded_explanation(
        block_info={"id": "sales", "type": "line", "label": "Vendas",
                    "calculation_method": "net_movement"},
        period={"id": "cur", "label": "2025", "type": "range"},
        value=1200.0,
        memory={"account_ids": [1, 2, 3], "calc_method": "net_movement"},
        accounts=[
            {"id": 1, "account_code": "4.01.01", "name": "Venda de produtos"},
            {"id": 2, "account_code": "4.01.02", "name": "Venda de servicos"},
        ],
    )
    print(f"  text: {text[:180]}")
    assert "Vendas" in text and "4.01.01" in text and "net_movement" in text
    print("  PASS")

    # --- 8. formula parser rejects unsafe syntax ---
    print("=== 8. client-agnostic formula safety (python side) ===")
    from accounting.reports.services.intelligence import FormulaEvaluator, FormulaError
    ev = FormulaEvaluator(block_values={"a": Decimal("10"), "b": Decimal("5")})
    assert ev.evaluate("a + b") == Decimal("15")
    for bad in ("a ** 2", "__import__('os')", "a[0]", "a.attr"):
        try:
            ev.evaluate(bad)
            print(f"  FAIL: '{bad}' should have been rejected"); sys.exit(1)
        except (FormulaError, SyntaxError) as e:
            print(f"  rejected '{bad}': {str(e)[:40]}")
    print("  PASS")

    # --- 9. Cycle detection (via document_schema flatten + deps) ---
    # The cycle detector lives in the frontend; backend only runs the
    # formula evaluator. Our runtime cycle safety comes from the multi-pass
    # settle in the calculator bailing on divergence. Verify it doesn't
    # infinite-loop on a cycle.
    print("=== 9. calculator settles on cyclic formulas without hanging ===")
    from accounting.reports.services.calculator import ReportCalculator
    cyclic_doc = {
        "name": "Cycle",
        "report_type": "income_statement",
        "blocks": [
            {"type": "total", "id": "a", "label": "A", "formula": "b + 1"},
            {"type": "total", "id": "b", "label": "B", "formula": "a + 1"},
        ],
    }
    calc = ReportCalculator(company_id=1)
    calc._legacy._calc_net_movement = MagicMock(return_value=Decimal("0"))
    calc._legacy._calc_ending_balance = MagicMock(return_value=Decimal("0"))
    calc._resolver.resolve = lambda sel: []  # type: ignore[assignment]
    res = calc.calculate(
        document=cyclic_doc,
        periods=[{"id": "p", "label": "p", "type": "range",
                  "start": "2025-01-01", "end": "2025-12-31"}],
    )
    print(f"  bailed with {len(res['warnings'])} warnings, lines: {len(res['lines'])}")
    print("  PASS (no hang)")

    print()
    print("=== ALL E2E SCENARIOS PASSED ===")


if __name__ == "__main__":
    main()
